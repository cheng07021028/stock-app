







from __future__ import annotations

import streamlit as st

from godpick_factor_schema import enrich_dataframe, ensure_factor_columns, V72_FACTOR_FIELDS

# >>> APP_AUTH_GUARD_V84
try:
    from app_auth import require_login
    require_login()
except Exception as _auth_e:
    st.error(f"登入系統載入失敗：{_auth_e}")
    st.stop()


try:
    from godpick_persistence_service import (
        load_watchlist_permanent,
        save_watchlist_permanent,
        load_records_permanent,
        save_records_sync_fast,
        upsert_records_authority_fast,
        records_authority_status,
        github_config as durable_github_config,
        firebase_configured as durable_firebase_configured,
    )
except Exception:
    load_watchlist_permanent = None
    save_watchlist_permanent = None
    load_records_permanent = None
    save_records_sync_fast = None
    upsert_records_authority_fast = None
    records_authority_status = None
    durable_github_config = None
    durable_firebase_configured = None
# <<< APP_AUTH_GUARD_V84

# pages/7_股神推薦.py
# -*- coding: utf-8 -*-

from datetime import date, timedelta, datetime
from zoneinfo import ZoneInfo
from typing import Any
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import copy
import json
import base64
import io
import hashlib

import pandas as pd
import numpy as np
import requests
import streamlit as st
try:
    import firebase_admin
    from firebase_admin import credentials, firestore
except Exception:
    firebase_admin = None
    credentials = None
    firestore = None


try:
    from godpick_column_schema import (
        UNIFIED_RECOMMEND_DISPLAY_COLUMNS,
        UNIFIED_MANAGEMENT_COLUMNS as SHARED_UNIFIED_MANAGEMENT_COLUMNS,
        normalize_godpick_dataframe,
        unified_display_columns,
        prune_empty_recommendation_columns,
        dedupe_keep_order as shared_dedupe_keep_order,
    )
except Exception:
    UNIFIED_RECOMMEND_DISPLAY_COLUMNS = []
    SHARED_UNIFIED_MANAGEMENT_COLUMNS = []
    normalize_godpick_dataframe = None
    unified_display_columns = None
    prune_empty_recommendation_columns = None
    shared_dedupe_keep_order = None

from utils import (
    compute_radar_scores,
    compute_signal_snapshot,
    compute_support_resistance_snapshot,
    format_number,
    get_all_code_name_map,
    get_history_data,
    get_normalized_watchlist,
    clear_history_disk_cache,
    get_history_disk_cache_stats,
    inject_pro_theme,
    render_pro_hero,
    render_pro_info_card,
    render_pro_kpi_row,
    render_pro_section,
)

try:
    from utils import get_history_data_debug
except Exception:
    get_history_data_debug = None

try:
    from stock_master_service import load_stock_master
except Exception:
    load_stock_master = None

try:
    from godpick_night_strategy import enrich_night_strategy, NIGHT_COLUMNS as GODPICK_NIGHT_COLUMNS
except Exception:
    enrich_night_strategy = None
    GODPICK_NIGHT_COLUMNS = []

try:
    from official_factor_service import merge_official_factors, load_factor_cache
except Exception:
    merge_official_factors = None
    load_factor_cache = None

try:
    from godpick_performance_feedback import (
        PERFORMANCE_FEEDBACK_VERSION,
        FEEDBACK_COLUMNS as GODPICK_PERFORMANCE_FEEDBACK_COLUMNS,
        apply_performance_feedback,
        load_godpick_performance_profile,
        performance_feedback_summary,
    )
except Exception:
    PERFORMANCE_FEEDBACK_VERSION = "performance_feedback_unavailable"
    GODPICK_PERFORMANCE_FEEDBACK_COLUMNS = []
    apply_performance_feedback = None
    load_godpick_performance_profile = None
    performance_feedback_summary = None


try:
    from godpick_calibration_sample_service import (
        CALIBRATION_SAMPLE_VERSION,
        assess_individual_sample_quality,
        save_calibration_samples,
    )
except Exception:
    CALIBRATION_SAMPLE_VERSION = "calibration_sample_unavailable"
    assess_individual_sample_quality = None
    save_calibration_samples = None


try:
    from godpick_decision_engine import (
        DECISION_ENGINE_VERSION as GODPICK_DECISION_ENGINE_VERSION,
        DECISION_ENGINE_COLUMNS as GODPICK_DECISION_ENGINE_COLUMNS,
        DECISION_ROLE_VALUES as GODPICK_DECISION_ROLE_VALUES,
        apply_godpick_decision_engine,
    )
except Exception:
    GODPICK_DECISION_ENGINE_VERSION = "decision_engine_unavailable"
    GODPICK_DECISION_ENGINE_COLUMNS = []
    GODPICK_DECISION_ROLE_VALUES = []
    apply_godpick_decision_engine = None


try:
    from godpick_execution_governance import (
        EXECUTION_GOVERNANCE_VERSION,
        apply_scan_quality_to_frame,
        build_action_table,
        build_candidate_diagnosis,
        build_engine_diagnostic_table,
        build_scan_quality_report,
        canonicalize_final_partition,
        govern_recommend_list,
        report_allows_formal_action,
    )
except Exception:
    EXECUTION_GOVERNANCE_VERSION = "execution_governance_unavailable"
    apply_scan_quality_to_frame = None
    build_action_table = None
    build_candidate_diagnosis = None
    build_engine_diagnostic_table = None
    build_scan_quality_report = None
    canonicalize_final_partition = None
    govern_recommend_list = None
    report_allows_formal_action = None

try:
    from godpick_official_release_timing import evaluate_twse_t86_release_timing
except Exception:
    evaluate_twse_t86_release_timing = None

try:
    from godpick_learning_system import (
        LEARNING_SYSTEM_VERSION,
        MODEL_VERSION as GODPICK_AI_MODEL_VERSION,
        LEARNING_COLUMNS as GODPICK_LEARNING_COLUMNS,
        apply_daily_learning_overlay,
        build_learning_summary,
        load_learning_state,
        save_learning_run,
    )
except Exception:
    LEARNING_SYSTEM_VERSION = "learning_system_unavailable"
    GODPICK_AI_MODEL_VERSION = "learning_model_unavailable"
    GODPICK_LEARNING_COLUMNS = []
    apply_daily_learning_overlay = None
    build_learning_summary = None
    load_learning_state = None
    save_learning_run = None

try:
    from godpick_super_ai_engine import (
        SUPER_AI_VERSION, SUPER_AI_COLUMNS, apply_super_ai_engine,
    )
    from godpick_super_ai_experience import (
        save_super_ai_run, load_super_ai_experience_profile, refresh_super_ai_experience_profile,
    )
    from godpick_t1_trade_truth import (
        refresh_t1_trade_truth, refresh_t1_truth_async, load_t1_truth_rows,
        load_t1_truth_summary, load_probability_calibration,
    )
    from godpick_durability_service import audit_core_durability
except Exception:
    SUPER_AI_VERSION = "super_ai_unavailable"
    SUPER_AI_COLUMNS = []
    apply_super_ai_engine = None
    save_super_ai_run = None
    load_super_ai_experience_profile = None
    refresh_super_ai_experience_profile = None
    refresh_t1_trade_truth = None
    refresh_t1_truth_async = None
    load_t1_truth_rows = None
    load_t1_truth_summary = None
    load_probability_calibration = None
    audit_core_durability = None

try:
    from godpick_v188_cache_guard import (
        V189_CACHE_GUARD_VERSION,
        inspect_v188_decision_frame,
        repair_v188_decision_frame,
    )
except Exception:
    V189_CACHE_GUARD_VERSION = "v189_cache_guard_unavailable"
    inspect_v188_decision_frame = None
    repair_v188_decision_frame = None

try:
    from godpick_full_market_discovery import (
        FULL_MARKET_DISCOVERY_VERSION,
        SECTOR_SHRINKAGE_VERSION,
        evaluate_legacy_soft_gates,
        apply_sector_bayesian_shrinkage,
    )
except Exception:
    FULL_MARKET_DISCOVERY_VERSION = "full_market_discovery_unavailable"
    SECTOR_SHRINKAGE_VERSION = "sector_shrinkage_unavailable"
    evaluate_legacy_soft_gates = None
    apply_sector_bayesian_shrinkage = None

try:
    from godpick_recommendation_rotation import (
        ROTATION_GUARD_VERSION,
        save_rotation_snapshot,
        rotation_diagnostics,
    )
except Exception:
    ROTATION_GUARD_VERSION = "rotation_guard_unavailable"
    save_rotation_snapshot = None
    rotation_diagnostics = None


STATE_FIX_VERSION = "widget_state_final_v4_verified_no_direct_rec_record_codes_20260425"
DUPLICATE_CONFIRM_VERSION = "duplicate_confirm_v1_20260425"
PRELAUNCH_789_VERSION = "prelaunch_789_v1_20260425"
MACRO_LINK_VERSION = "macro_link_v37_market_session_effect_bridge_20260429"
WEIGHT_STATE_FIX_VERSION = "weight_widget_state_fix_v1_20260427"
GOD_DECISION_ENGINE_VERSION = "god_decision_engine_v5_20260427"
SCAN_SETTINGS_PERSIST_VERSION = "scan_settings_apply_reset_v1_20260427"
SCAN_SETTINGS_WIDGET_FIX_VERSION = "scan_settings_widget_state_fix_v1_20260427"
SCAN_SETTINGS_AUTOSAVE_VERSION = "scan_settings_autosave_reload_fix_v1_20260427"
PAGE07_SPEED_FIX_VERSION = "page07_v189_v188_final_cache_guard_20260812"
OPPORTUNITY_MODE_VERSION = "low_pullback_retest_v1_20260428"
SECTOR_FLOW_VERSION = "sector_flow_rotation_v1_20260428"
OVERNIGHT_GLOBAL_BRIDGE_VERSION = "overnight_global_bridge_v74_taifex_fallback_20260430"
NIGHT_NEXT_ENTRY_VERSION = "night_next_entry_v109_official_factor_cache_20260513"
PAGE_TITLE = "股神推薦 VNext｜績效回饋校正版"
PFX = "godpick_"

HISTORY_DEBUG_EAGER = False  # False: 只有抓不到歷史資料時才補跑 debug，避免每檔雙重抓取拖慢速度
PROGRESS_UPDATE_EVERY = 100  # V35：再降低前端重繪頻率，避免 Streamlit 每檔刷新拖慢
SCAN_MAX_WORKERS = 8          # V48.3：資料源穩定優先；配合每執行緒 Session / Yahoo 併發閘門，避免 429 大量漏股

# V48.2：資料品質統計必須把「K線成功」與「通過推薦前置篩選」分開。
# signal/risk/prelaunch/trade_filtered 都已成功取得並完成 K 線分析，不能誤算成無有效K線。
_KLINE_VALID_STATUSES = {"ok", "signal_filtered", "risk_filtered", "prelaunch_filtered", "trade_filtered"}
_RETRYABLE_SCAN_STATUSES = {"no_history", "analysis_error", "future_exception"}
V22_CHECKPOINT_EVERY = 500    # V35：降低寫入斷點頻率，避免 JSON I/O 拖慢掃描
GODPICK_SCAN_CHECKPOINT_FILE = "godpick_scan_checkpoint.json"
HISTORY_DEBUG_ON_FAIL = False  # V35：掃描中失敗股票不再即時跑慢速 debug，失敗原因彙總到除錯摘要
V180_DISABLE_DUPLICATE_HISTORY_RETRY = True  # get_history_data 已含 HTTP retry + Yahoo/TWSE/TPEx fallback；外層不可再整套重跑


# v26 欄位統一：讓 7_股神推薦匯出 / 匯入 8 / 匯入 10 使用共用欄位集合。
try:
    if UNIFIED_RECOMMEND_DISPLAY_COLUMNS:
        GODPICK_RECORD_COLUMNS = shared_dedupe_keep_order((GODPICK_RECORD_COLUMNS or []) + list(UNIFIED_RECOMMEND_DISPLAY_COLUMNS)) if shared_dedupe_keep_order else list(dict.fromkeys((GODPICK_RECORD_COLUMNS or []) + list(UNIFIED_RECOMMEND_DISPLAY_COLUMNS)))
except Exception:
    pass

GODPICK_DEFAULT_SCORE_WEIGHTS = {
    "市場環境": 10,
    "技術結構": 15,
    "起漲前兆": 20,
    "類股熱度": 15,
    "自動因子": 10,
    "交易可行": 10,
    "型態突破": 12,
    "爆發力": 8,
}

# 執行推薦時會把已套用權重複製到這裡，避免 ThreadPool 內直接讀 widget 狀態造成不穩。
GODPICK_ACTIVE_SCORE_WEIGHTS = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()


GODPICK_SETTINGS_FILE = "godpick_user_settings.json"
GODPICK_COLUMN_ORDER_FILE = "godpick_column_orders.json"  # v72：欄位順序獨立保存，避免權重設定/GitHub 舊值覆蓋
GODPICK_LATEST_FILE = "godpick_latest_recommendations.json"
GODPICK_LATEST_ANCHOR_FILE = "godpick_latest_run_anchor.json"  # V185：小型永久錨點，避免重新部署後回退舊快照
GODPICK_LIST_FILE = "godpick_recommend_list.json"
MACRO_MODE_BRIDGE_FILE = "macro_mode_bridge.json"
MARKET_SNAPSHOT_FILE = "market_snapshot.json"
OFFICIAL_FACTORS_CACHE_FILE = "official_factors_cache.json"
RECOMMENDATION_READINESS_FILE = "godpick_recommendation_readiness.json"


GODPICK_RECORD_COLUMNS = [
    "record_id",
    "股票代號",
    "股票名稱",
    "市場別",
    "類別",
    "官方因子總分",
    "官方資料完整度",
    "官方因子資料狀態",
    "官方資料日期",
    "外資近5日買賣超",
    "投信近5日買賣超",
    "三大法人近5日合計",
    "法人連買天數",
    "法人籌碼官方分數",
    "月營收YoY%",
    "月營收MoM%",
    "累計營收YoY%",
    "營收成長官方分數",
    "PBR股價淨值比",
    "股利殖利率%",
    "官方估值風險分數",
    "官方基本面成長分數",
    "官方因子更新時間",
    "官方因子資料源",
    "推薦模式",
    "推薦型態",
    "機會型態",
    "低檔位置分數",
    "拉回承接分數",
    "支撐回測分數",
    "止跌轉強分數",
    "機會股分數",
    "機會股說明",
    "進場時機",
    "進場時機分數",
    "建議動作",
    "等待條件",
    "近端支撐",
    "主要支撐",
    "近端壓力",
    "突破確認價",
    "停損參考",
    "操作區間",
    "風險報酬比_決策",
    "追高風險分數_決策",
    "追高風險等級",
    "是否建議追價",
    "風險扣分原因",
    "決策說明",
    "推薦等級",
    "推薦總分",
    "股神實戰總分",
    "Alpha選股潛力分",
    "Entry進場買點分",
    "Risk風控安全分",
    "Feedback績效校正分",
    "選股潛力分",
    "進場買點分",
    "風控安全分",
    "績效校正分",
    "績效校正說明",
    "新買點分級",
    "推薦角色",
    "過熱原因",
    "建議動作",
    "建議倉位",
    "小量試單建議",
    "加碼條件",
    "失效條件",
    "失效條件_績效回饋",
    "績效回饋建議",
    "績效樣本數",
    "績效回饋版本",
    "決策版本",
    "推薦用途",
    "買進分數",
    "是否可直接買進",
    "盤中確認條件",
    "專業決策摘要",
    "上漲機率估計%",
    "上漲機率等級",
    "上漲機率信心",
    "上漲機率說明",
    "上漲機率因子明細",
    "夜間股神總分",
    "隔日實戰排序分",
    "隔日進場分數",
    "波段潛力分數",
    "技術趨勢分數",
    "量價動能分數",
    "法人籌碼分數",
    "大戶鎖碼分數",
    "基本面成長分數",
    "營收成長分數",
    "EPS成長分數",
    "估值風險分數",
    "PER本益比",
    "估算EPS",
    "外資近1日買賣超",
    "投信近1日買賣超",
    "三大法人近1日合計",
    "法人買超占量比%",
    "資料完整度",
    "進場型態_隔日",
    "隔日建議動作",
    "預估進場點",
    "回測承接價",
    "突破確認價_隔日",
    "停損價_隔日",
    "第一壓力價",
    "觀察週期",
    "夜間股神建議",
    "隔日作戰策略",
    "夜間風險提醒",

    "買點狀態",
    "進場型態",
    "高分禁買旗標",
    "高分禁買原因",
    "實戰買點分數",
    "支撐距離%",
    "壓力空間%",
    "近5日漲幅%",
    "長上影風險",
    "實戰操作建議",
    "V76買點防呆版本",
    "大盤參考等級",
    "大盤可參考分數",
    "大盤加權分",
    "大盤風險濾網",
    "大盤推薦權重",
    "大盤降權原因",
    "大盤操作風格",
    "大盤市場廣度分數",
    "大盤量價確認分數",
    "大盤權值支撐分數",
    "大盤推薦同步分數",
    "大盤資料日期",
    "大盤橋接分數",
    "大盤橋接狀態",
    "大盤橋接加權",
    "大盤橋接風控",
    "大盤橋接策略",
    "大盤橋接更新時間",
    "大盤交易時段",
    "大盤交易時段可用",
    "大盤資料品質",
    "大盤影響加減分",
    "大盤影響說明",
    "大盤資料診斷摘要",
    "隔夜風控分數",
    "隔夜風險等級",
    "隔夜偏向",
    "隔夜解讀",
    "隔夜資料品質",
    "台指夜盤資料來源",
    "台指夜盤備援說明",
    "台指夜盤漲跌",
    "NASDAQ漲跌%",
    "S&P500漲跌%",
    "道瓊漲跌%",
    "費半漲跌%",
    "Nasdaq期貨偏向",
    "S&P期貨偏向",
    "匯率風險等級",
    "隔日大盤預測日期",
    "隔日大盤方向",
    "隔日大盤分數",
    "隔日大盤信心",
    "隔日上漲機率%",
    "隔日震盪機率%",
    "隔日下跌機率%",
    "隔日預估漲跌%",
    "隔日大盤預測加減分",
    "隔日大盤權重校正",
    "隔日建議總部位上限%",
    "隔日偏好選股風格",
    "隔日應避免風格",
    "隔日大盤預測理由",
    "股神決策模式",
    "股神進場建議",
    "建議部位%",
    "建議倉位%",
    "建議投入等級",
    "分批策略",
    "第一筆進場%",
    "第二筆加碼條件",
    "停利策略",
    "停損策略",
    "最大風險%",
    "資金風險說明",
    "單檔風險等級",
    "族群集中警示",
    "組合配置建議",
    "大盤策略模式",
    "大盤多空分數",
    "推薦積極度係數",
    "適合推薦型態",
    "大盤策略建議",
    "大盤風控建議",
    "市場策略調整說明",
    "動態建議倉位%",

    "風險報酬比",
    "追價風險分",
    "停損距離%",
    "目標報酬%",
    "不建議買進原因",
    "最佳操作劇本",
    "大盤情境調權說明",
    "大盤情境分桶",
    "推薦分層",
    "隔日操作建議",
    "失效價位",
    "轉弱條件",
    "買點分級",
    "風險說明",
    "股神推論邏輯",
    "權重設定",
    "推薦分桶",
    "起漲等級",
    "信心等級",
    "買點劇本",
    "失效條件",
    "假突破風險",
    "過熱風險",
    "3日追蹤預留",
    "5日追蹤預留",
    "10日追蹤預留",
    "20日追蹤預留",
    "技術結構分數",
    "起漲前兆分數",
    "飆股起漲分數",
    "起漲摘要",
    "交易可行分數",
    "類股熱度分數",
    "強勢族群等級",
    "族群資金流分數",
    "族群輪動狀態",
    "同族群強勢比例",
    "同族群推薦密度",
    "同族群平均量能分",
    "族群策略建議",
    "族群資金流說明",
    "同類股領先幅度",
    "是否領先同類股",
    "推薦標籤",
    "推薦理由摘要",
    "K線驗證標記",
    "推薦日價格",
    "推薦日支撐壓力摘要",
    "K線查詢參數",
    "K線檢視提示",
    "推薦價格",
    "停損價",
    "賣出目標1",
    "賣出目標2",
    "推薦日期",
    "推薦時間",
    "建立時間",
    "更新時間",
    "目前狀態",
    "是否已實際買進",
    "實際買進價",
    "實際賣出價",
    "實際報酬%",
    "最新價",
    "最新更新時間",
    "損益金額",
    "損益幅%",
    "是否達停損",
    "是否達目標1",
    "是否達目標2",
    "持有天數",
    "模式績效標籤",
    "備註",
]

try:
    _phase1_record_cols = list(GODPICK_RECORD_COLUMNS) + list(GODPICK_DECISION_ENGINE_COLUMNS or [])
    GODPICK_RECORD_COLUMNS = list(dict.fromkeys(_phase1_record_cols))
except Exception:
    pass

# V159：原程式在 GODPICK_RECORD_COLUMNS 尚未宣告前就嘗試合併共用欄位，
# 因例外被靜默略過，造成正式分區、R1-M、進場可執行性及 K 線新鮮度在寫入紀錄時被截掉。
try:
    GODPICK_RECORD_COLUMNS = list(dict.fromkeys(
        list(GODPICK_RECORD_COLUMNS) + list(UNIFIED_RECOMMEND_DISPLAY_COLUMNS or [])
    ))
except Exception:
    pass


# =========================================================
# 基礎工具
# =========================================================

# >>> V72_FACTOR_ENRICH_HELPER
def _v72_enrich_recommendation_df_safe(df):
    try:
        return enrich_dataframe(df)
    except Exception:
        try:
            return ensure_factor_columns(df)
        except Exception:
            return df
# <<< V72_FACTOR_ENRICH_HELPER

def _k(key: str) -> str:
    return f"{PFX}{key}"


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).strip()


def _safe_float(v: Any, default=None):
    try:
        if pd.isna(v):
            return default
    except Exception:
        pass
    try:
        return float(v)
    except Exception:
        return default


def _normalize_code(v: Any) -> str:
    text = _safe_str(v)
    if not text:
        return ""
    if text.isdigit():
        return text
    digits = "".join(ch for ch in text if ch.isdigit())
    if 4 <= len(digits) <= 6:
        return digits
    return text


def _normalize_category(v: Any) -> str:
    text = _safe_str(v)
    if not text:
        return ""
    return text.replace("　", " ").strip()


def _score_clip(v: float, low: float = 0.0, high: float = 100.0) -> float:
    return max(low, min(high, v))


def _ensure_radar_dict(radar_obj: Any) -> dict[str, Any]:
    if radar_obj is None:
        radar_obj = {}
    elif not isinstance(radar_obj, dict):
        try:
            radar_obj = dict(radar_obj)
        except Exception:
            radar_obj = {}

    base = {
        "trend": _safe_float(radar_obj.get("trend"), 50) or 50,
        "momentum": _safe_float(radar_obj.get("momentum"), 50) or 50,
        "volume": _safe_float(radar_obj.get("volume"), 50) or 50,
        "position": _safe_float(radar_obj.get("position"), 50) or 50,
        "structure": _safe_float(radar_obj.get("structure"), 50) or 50,
        "summary": _safe_str(radar_obj.get("summary")) or "",
    }
    for k, v in radar_obj.items():
        if k not in base:
            base[k] = v
    return base



# V59：上漲機率估計工具
# 說明：此為「條件機率估計」，不是保證上漲機率。會用推薦總分、起漲/技術/交易/類股、大盤、風險與資料完整度綜合換算。
def _estimate_upside_probability_row(row: Any) -> dict[str, Any]:
    try:
        score = _safe_float(row.get("推薦總分"), 0) or 0
        tech = _safe_float(row.get("技術結構分數"), 50) or 50
        launch = _safe_float(row.get("起漲前兆分數"), 50) or 50
        hot = _safe_float(row.get("類股熱度分數"), 50) or 50
        trade = _safe_float(row.get("交易可行分數"), 50) or 50
        breakout = _safe_float(row.get("型態突破分數"), 50) or 50
        burst = _safe_float(row.get("爆發力分數"), 50) or 50
        market = _safe_float(row.get("大盤橋接分數"), row.get("大盤可參考分數"))
        if market is None:
            market = _safe_float(row.get("市場環境分數"), 50) or 50

        # 風險欄位方向統一：數值越高代表追價/波動風險越高，因此使用 safety_score 加分。
        risk_candidates = [
            row.get("追價風險分"),
            row.get("追高風險分數_決策"),
            row.get("追價風險分數"),
            row.get("風險分數"),
        ]
        risk = None
        for rv in risk_candidates:
            val = _safe_float(rv)
            if val is not None:
                risk = val
                break
        if risk is None:
            risk = 50.0
        safety = _score_clip(100 - risk)

        confidence = 100.0
        notes = []
        if _safe_float(row.get("最新價")) is None and _safe_float(row.get("推薦價格")) is None:
            confidence -= 18
            notes.append("即時價缺漏")
        if _safe_float(row.get("技術結構分數")) is None:
            confidence -= 12
            notes.append("技術分缺漏")
        if _safe_float(row.get("起漲前兆分數")) is None:
            confidence -= 12
            notes.append("起漲分缺漏")
        if not _safe_str(row.get("大盤橋接狀態")) and not _safe_str(row.get("大盤資料品質")):
            confidence -= 8
            notes.append("大盤橋接不足")
        if _safe_str(row.get("大盤交易時段可用")) in ["False", "false", "否", "不可用"]:
            confidence -= 6
            notes.append("非即時交易時段")
        confidence = _score_clip(confidence)

        prob = 50.0
        prob += (score - 60) * 0.42
        prob += (launch - 50) * 0.10
        prob += (tech - 50) * 0.08
        prob += (trade - 50) * 0.07
        prob += (hot - 50) * 0.06
        prob += (breakout - 50) * 0.05
        prob += (burst - 50) * 0.04
        prob += (market - 50) * 0.06
        prob += (safety - 50) * 0.08

        grade = _safe_str(row.get("推薦等級"))
        buy_grade = _safe_str(row.get("買點分級"))
        if grade in ["股神級", "S", "S級"]:
            prob += 2.5
        elif grade in ["強烈關注", "A+", "A級"]:
            prob += 1.5
        if buy_grade in ["A+", "A", "S", "S級"]:
            prob += 2.0
        elif buy_grade in ["C", "D"]:
            prob -= 3.0

        shrink = confidence / 100.0
        prob = 50 + (prob - 50) * shrink
        prob = round(_score_clip(prob, 25, 82), 1)

        if prob >= 68 and confidence >= 80:
            level = "高"
        elif prob >= 58 and confidence >= 65:
            level = "中高"
        elif prob >= 52:
            level = "中"
        elif prob >= 45:
            level = "偏低"
        else:
            level = "低"

        if confidence >= 85:
            conf_label = "高"
        elif confidence >= 70:
            conf_label = "中高"
        elif confidence >= 55:
            conf_label = "中"
        else:
            conf_label = "低"

        detail = {
            "推薦總分": round(score, 2),
            "技術結構": round(tech, 2),
            "起漲前兆": round(launch, 2),
            "類股熱度": round(hot, 2),
            "交易可行": round(trade, 2),
            "型態突破": round(breakout, 2),
            "爆發力": round(burst, 2),
            "大盤": round(market, 2),
            "風險分數": round(risk, 2),
            "安全分數": round(safety, 2),
            "資料信心": round(confidence, 2),
        }
        reason = f"條件機率估計 {prob}%：依推薦總分、起漲/技術/交易/類股、大盤與風險折扣換算；非保證上漲機率。"
        if notes:
            reason += " 資料信心下修：" + "、".join(notes[:4]) + "。"
        return {
            "上漲機率估計%": prob,
            "上漲機率等級": level,
            "上漲機率信心": conf_label,
            "上漲機率說明": reason,
            "上漲機率因子明細": json.dumps(detail, ensure_ascii=False),
        }
    except Exception as e:
        return {
            "上漲機率估計%": None,
            "上漲機率等級": "無法估計",
            "上漲機率信心": "低",
            "上漲機率說明": f"上漲機率估計失敗：{e}",
            "上漲機率因子明細": "{}",
        }


def _score_band(v: Any) -> str:
    x = _safe_float(v, 0) or 0
    if x >= 90:
        return "極強"
    if x >= 80:
        return "偏強"
    if x >= 70:
        return "可用"
    if x >= 60:
        return "觀察"
    return "保守"


def _build_pattern_breakout_scores(df: pd.DataFrame, sr_snapshot: dict, signal_snapshot: dict) -> dict[str, Any]:
    if df is None or df.empty:
        return {"型態名稱": "資料不足", "型態突破分數": 0.0, "突破風險": "資料不足"}

    last = df.iloc[-1]
    close_now = _safe_float(last.get("收盤價"), 0) or 0
    ma20 = _safe_float(last.get("MA20"))
    ma60 = _safe_float(last.get("MA60"))
    vol5 = _safe_float(last.get("VOL5"))
    vol20 = _safe_float(last.get("VOL20"))
    ret5 = _safe_float(last.get("RET5"), 0) or 0
    res20 = _safe_float(sr_snapshot.get("res_20"))
    sup20 = _safe_float(sr_snapshot.get("sup_20"))

    score = 45.0
    pattern_name = "整理中"
    risk_text = "正常"

    if close_now and res20 not in [None, 0]:
        dist = ((res20 - close_now) / res20) * 100
        if -1.5 <= dist <= 1.5:
            score += 28
            pattern_name = "平台整理突破"
        elif 1.5 < dist <= 4.5:
            score += 18
            pattern_name = "箱型整理待突破"
        elif dist < -1.5:
            score += 12
            pattern_name = "已突破觀察"

    if ma20 not in [None, 0] and ma60 not in [None, 0]:
        if close_now >= ma20 >= ma60:
            score += 16
        elif close_now >= ma20:
            score += 8

    if vol5 not in [None, 0] and vol20 not in [None, 0]:
        vr = vol5 / vol20
        if vr >= 1.6:
            score += 18
        elif vr >= 1.2:
            score += 10
        elif vr < 0.8:
            score -= 5

    if ret5 > 12:
        score -= 8
        risk_text = "短線偏熱"
    if ret5 > 20:
        score -= 12
        risk_text = "短線過熱"

    if sup20 not in [None, 0] and close_now < sup20:
        score -= 10
        risk_text = "跌破支撐"

    return {
        "型態名稱": pattern_name,
        "型態突破分數": _score_clip(score),
        "突破風險": risk_text,
    }


def _build_burst_scores(df: pd.DataFrame) -> dict[str, Any]:
    if df is None or df.empty:
        return {"爆發力分數": 0.0, "爆發等級": "資料不足"}

    last = df.iloc[-1]
    ret5 = _safe_float(last.get("RET5"), 0) or 0
    ret20 = _safe_float(last.get("RET20"), 0) or 0
    vol5 = _safe_float(last.get("VOL5"))
    vol20 = _safe_float(last.get("VOL20"))
    close_now = _safe_float(last.get("收盤價"), 0) or 0
    atr14 = _safe_float(last.get("ATR14"))

    score = 48.0
    if ret5 > 5:
        score += 12
    if ret5 > 8:
        score += 10
    if ret20 > 12:
        score += 8
    if vol5 not in [None, 0] and vol20 not in [None, 0]:
        vr = vol5 / vol20
        if vr >= 1.8:
            score += 14
        elif vr >= 1.3:
            score += 8
    if close_now not in [None, 0] and atr14 not in [None, 0]:
        atr_pct = atr14 / close_now * 100
        if 2.2 <= atr_pct <= 6.0:
            score += 8
        elif atr_pct > 8.5:
            score -= 8

    score = _score_clip(score)
    if score >= 85:
        level = "高爆發"
    elif score >= 72:
        level = "中強勢"
    elif score >= 60:
        level = "觀察"
    else:
        level = "普通"
    return {"爆發力分數": score, "爆發等級": level}


def _last_num(df: pd.DataFrame, col: str, default: float | None = None) -> float | None:
    try:
        if df is None or df.empty or col not in df.columns:
            return default
        return _safe_float(df[col].iloc[-1], default)
    except Exception:
        return default


def _recent_low_high(df: pd.DataFrame, days: int = 60) -> tuple[float | None, float | None]:
    try:
        w = df.tail(days)
        low = pd.to_numeric(w.get("最低價"), errors="coerce").dropna()
        high = pd.to_numeric(w.get("最高價"), errors="coerce").dropna()
        return (float(low.min()) if not low.empty else None, float(high.max()) if not high.empty else None)
    except Exception:
        return None, None


def _vol_ratio(df: pd.DataFrame, short_col: str = "VOL5", long_col: str = "VOL20") -> float | None:
    v1 = _last_num(df, short_col)
    v2 = _last_num(df, long_col)
    if v1 not in [None, 0] and v2 not in [None, 0]:
        return float(v1) / float(v2)
    return None


def _build_opportunity_scores(df: pd.DataFrame, sr_snapshot: dict, signal_snapshot: dict, radar: dict) -> dict[str, Any]:
    """低檔 / 拉回 / 回測機會股評分。"""
    if df is None or df.empty or len(df) < 30:
        return {"低檔位置分數": 0.0, "拉回承接分數": 0.0, "支撐回測分數": 0.0, "止跌轉強分數": 0.0, "機會股分數": 0.0, "機會型態": "資料不足", "推薦型態": "資料不足", "機會股說明": "歷史資料不足，無法判斷低檔/拉回/回測機會", "追高風險分_機會": 80.0}

    last = df.iloc[-1]
    close_now = _safe_float(last.get("收盤價"), 0) or 0
    open_now = _safe_float(last.get("開盤價"), close_now) or close_now
    ma5 = _safe_float(last.get("MA5"))
    ma10 = _safe_float(last.get("MA10"))
    ma20 = _safe_float(last.get("MA20"))
    ma60 = _safe_float(last.get("MA60"))
    ret3 = _safe_float(last.get("RET3"), 0) or 0
    ret5 = _safe_float(last.get("RET5"), 0) or 0
    ret20 = _safe_float(last.get("RET20"), 0) or 0
    sup20 = _safe_float(sr_snapshot.get("sup_20"))
    sup60 = _safe_float(sr_snapshot.get("sup_60"))
    res20 = _safe_float(sr_snapshot.get("res_20"))
    vr = _vol_ratio(df)

    low60, high60 = _recent_low_high(df, 60)
    low120, high120 = _recent_low_high(df, 120)
    low_dist60 = ((close_now - low60) / low60 * 100) if close_now and low60 not in [None, 0] else None
    high_drawdown60 = ((close_now - high60) / high60 * 100) if close_now and high60 not in [None, 0] else None
    low_dist120 = ((close_now - low120) / low120 * 100) if close_now and low120 not in [None, 0] else None

    low_score = 42.0
    if low_dist60 is not None:
        if low_dist60 <= 8: low_score += 30
        elif low_dist60 <= 15: low_score += 20
        elif low_dist60 <= 25: low_score += 10
        else: low_score -= min(18, (low_dist60 - 25) * 0.35)
    if low_dist120 is not None and low_dist120 <= 18: low_score += 8
    if high_drawdown60 is not None:
        if -28 <= high_drawdown60 <= -8: low_score += 12
        elif high_drawdown60 > -3: low_score -= 10
    if ret20 > 18: low_score -= 18
    elif -8 <= ret20 <= 8: low_score += 8
    if ma20 not in [None, 0] and close_now:
        dist_ma20 = (close_now - ma20) / ma20 * 100
        if -4 <= dist_ma20 <= 5: low_score += 8
        elif dist_ma20 < -10: low_score -= 8

    pullback_score = 38.0
    if ma20 not in [None, 0] and ma60 not in [None, 0]:
        if ma20 >= ma60: pullback_score += 18
        if close_now >= ma60: pullback_score += 10
        else: pullback_score -= 14
    if ma20 not in [None, 0] and close_now:
        dist_ma20 = (close_now - ma20) / ma20 * 100
        if -3 <= dist_ma20 <= 4.5: pullback_score += 26
        elif -7 <= dist_ma20 < -3: pullback_score += 12
        elif dist_ma20 > 10: pullback_score -= 16
    if vr is not None:
        if 0.55 <= vr <= 1.10: pullback_score += 10
        elif vr >= 1.8 and ret5 < 0: pullback_score -= 12
    if -10 <= ret20 <= 12 and ret5 <= 6: pullback_score += 7
    if close_now >= open_now and ret3 >= -2: pullback_score += 7

    retest_score = 40.0
    support_candidates = [x for x in [sup20, sup60, ma20, ma60] if x not in [None, 0]]
    nearest_support = None
    if close_now and support_candidates:
        nearest_support = min(support_candidates, key=lambda x: abs(close_now - x) / x)
        support_dist = (close_now - nearest_support) / nearest_support * 100
        if -1.5 <= support_dist <= 4.5: retest_score += 28
        elif 4.5 < support_dist <= 8: retest_score += 12
        elif support_dist < -3: retest_score -= 14
    if res20 not in [None, 0] and close_now:
        dist_res = (close_now - res20) / res20 * 100
        if -3 <= dist_res <= 3: retest_score += 15
    if ma20 not in [None, 0] and ma60 not in [None, 0] and ma20 >= ma60: retest_score += 8
    if ret5 > 12: retest_score -= 12

    rebound_score = 40.0
    if close_now >= open_now: rebound_score += 10
    if ret3 > 0: rebound_score += 10
    if ret5 > -3: rebound_score += 8
    if ma5 not in [None, 0] and ma10 not in [None, 0] and close_now >= ma5:
        rebound_score += 8
        if ma5 >= ma10: rebound_score += 8
    if vr is not None:
        if 0.9 <= vr <= 1.7: rebound_score += 10
        elif vr > 2.4 and ret5 < 0: rebound_score -= 10
    sig = _safe_float(signal_snapshot.get("score"), 0) or 0
    rebound_score += min(12, max(0, sig * 2.2))

    chase_risk = 35.0
    if ret5 > 8: chase_risk += 18
    if ret20 > 18: chase_risk += 22
    if low_dist60 is not None and low_dist60 > 35: chase_risk += 12
    if ma20 not in [None, 0] and close_now:
        dist_ma20 = (close_now - ma20) / ma20 * 100
        if dist_ma20 > 10: chase_risk += 16
        elif -3 <= dist_ma20 <= 5: chase_risk -= 8
    if vr is not None and vr > 2.2 and ret5 > 5: chase_risk += 12

    low_score = _score_clip(low_score)
    pullback_score = _score_clip(pullback_score)
    retest_score = _score_clip(retest_score)
    rebound_score = _score_clip(rebound_score)
    chase_risk = _score_clip(chase_risk)
    opportunity_score = _score_clip(low_score * 0.28 + pullback_score * 0.24 + retest_score * 0.24 + rebound_score * 0.20 + max(0, 100 - chase_risk) * 0.04)

    candidates = [("低檔轉強", low_score), ("拉回承接", pullback_score), ("回測支撐", retest_score), ("止跌反彈", rebound_score)]
    candidates.sort(key=lambda x: x[1], reverse=True)
    opportunity_type = candidates[0][0]
    if chase_risk >= 75 and opportunity_score >= 65:
        opportunity_type = f"{opportunity_type}｜但勿追高"

    reason_parts = []
    if low_score >= 70: reason_parts.append("位置接近低檔")
    if pullback_score >= 70: reason_parts.append("拉回均線承接")
    if retest_score >= 70: reason_parts.append("回測支撐不破")
    if rebound_score >= 70: reason_parts.append("止跌轉強")
    if chase_risk >= 72: reason_parts.append("追高風險偏高，等回測")
    elif chase_risk <= 55: reason_parts.append("追高風險相對低")
    if nearest_support not in [None, 0]: reason_parts.append(f"鄰近支撐 {nearest_support:.2f}")
    if not reason_parts: reason_parts.append("低檔/拉回條件普通，列觀察")

    return {
        "低檔位置分數": round(low_score, 2),
        "拉回承接分數": round(pullback_score, 2),
        "支撐回測分數": round(retest_score, 2),
        "止跌轉強分數": round(rebound_score, 2),
        "機會股分數": round(opportunity_score, 2),
        "機會型態": opportunity_type,
        "推薦型態": opportunity_type,
        "機會股說明": "、".join(reason_parts[:6]),
        "追高風險分_機會": round(chase_risk, 2),
    }



def _pct_distance(price: Any, base: Any) -> float | None:
    p = _safe_float(price)
    b = _safe_float(base)
    if p in [None, 0] or b in [None, 0]:
        return None
    try:
        return (float(p) - float(b)) / float(b) * 100
    except Exception:
        return None


def _build_entry_decision_scores(
    df: pd.DataFrame,
    sr_snapshot: dict,
    opportunity_info: dict,
    trade_plan: dict,
    trade_feasibility: dict,
) -> dict[str, Any]:
    """V10 股神進場決策引擎：把推薦股票轉成可操作的進場/等待/停損策略。"""
    if df is None or df.empty or len(df) < 25:
        return {
            "進場時機": "資料不足",
            "進場時機分數": 0.0,
            "建議動作": "暫不判斷",
            "等待條件": "歷史資料不足，先不要追價",
            "近端支撐": None,
            "主要支撐": None,
            "近端壓力": None,
            "突破確認價": None,
            "停損參考": None,
            "操作區間": "",
            "風險報酬比_決策": None,
            "追高風險分數_決策": 80.0,
            "追高風險等級": "高",
            "是否建議追價": "否",
            "風險扣分原因": "歷史資料不足",
            "決策說明": "資料不足時不建議追價，先等待資料恢復。",
        }

    last = df.iloc[-1]
    close_now = _safe_float(last.get("收盤價"), 0) or 0
    ma5 = _safe_float(last.get("MA5"))
    ma10 = _safe_float(last.get("MA10"))
    ma20 = _safe_float(last.get("MA20"))
    ma60 = _safe_float(last.get("MA60"))
    ret3 = _safe_float(last.get("RET3"), 0) or 0
    ret5 = _safe_float(last.get("RET5"), 0) or 0
    ret20 = _safe_float(last.get("RET20"), 0) or 0
    rsi = _safe_float(last.get("RSI14"), _safe_float(last.get("RSI")))
    vr = _vol_ratio(df)

    sup20 = _safe_float(sr_snapshot.get("sup_20"))
    sup60 = _safe_float(sr_snapshot.get("sup_60"))
    res20 = _safe_float(sr_snapshot.get("res_20"))
    res60 = _safe_float(sr_snapshot.get("res_60"))

    support_candidates = [x for x in [sup20, ma20, sup60, ma60] if x not in [None, 0]]
    resistance_candidates = [x for x in [res20, res60] if x not in [None, 0]]
    near_support = None
    main_support = None
    near_resistance = None
    if close_now and support_candidates:
        below_or_near = [x for x in support_candidates if x <= close_now * 1.03]
        near_support = max(below_or_near) if below_or_near else min(support_candidates, key=lambda x: abs(close_now - x))
        main_support = min(support_candidates)
    if close_now and resistance_candidates:
        above_or_near = [x for x in resistance_candidates if x >= close_now * 0.98]
        near_resistance = min(above_or_near) if above_or_near else max(resistance_candidates)

    breakout_price = _safe_float(trade_plan.get("breakout_buy"), near_resistance)
    stop_ref = _safe_float(trade_plan.get("stop_price"))
    if stop_ref in [None, 0] and near_support not in [None, 0]:
        stop_ref = near_support * 0.975

    pullback_buy = _safe_float(trade_plan.get("pullback_buy"))
    zone_low = None
    zone_high = None
    zone_vals = [x for x in [pullback_buy, near_support, close_now] if x not in [None, 0]]
    if zone_vals:
        zone_low = min(zone_vals)
        zone_high = max(zone_vals)
    operation_zone = ""
    if zone_low not in [None, 0] and zone_high not in [None, 0]:
        operation_zone = f"{zone_low:.2f} ~ {zone_high:.2f}"

    support_dist = _pct_distance(close_now, near_support)
    ma20_dist = _pct_distance(close_now, ma20)
    pressure_space = None
    if close_now not in [None, 0] and near_resistance not in [None, 0]:
        pressure_space = (near_resistance - close_now) / close_now * 100

    # 追高風險：越高越不適合追價
    chase_risk = _safe_float(opportunity_info.get("追高風險分_機會"), 35) or 35
    risk_reasons = []
    if ret5 > 8:
        chase_risk += 14
        risk_reasons.append("5日漲幅偏大")
    if ret20 > 18:
        chase_risk += 18
        risk_reasons.append("20日漲幅偏大")
    if ma20_dist is not None and ma20_dist > 10:
        chase_risk += 16
        risk_reasons.append("股價離月線過遠")
    if rsi is not None and rsi >= 72:
        chase_risk += 12
        risk_reasons.append("RSI過熱")
    if vr is not None and vr > 2.4 and ret5 > 5:
        chase_risk += 10
        risk_reasons.append("放量急漲")
    if pressure_space is not None and pressure_space < 4:
        chase_risk += 10
        risk_reasons.append("接近壓力區")
    if support_dist is not None and -1.5 <= support_dist <= 5:
        chase_risk -= 10
    chase_risk = _score_clip(chase_risk)

    low_score = _safe_float(opportunity_info.get("低檔位置分數"), 0) or 0
    pullback_score = _safe_float(opportunity_info.get("拉回承接分數"), 0) or 0
    retest_score = _safe_float(opportunity_info.get("支撐回測分數"), 0) or 0
    rebound_score = _safe_float(opportunity_info.get("止跌轉強分數"), 0) or 0
    trade_score = _safe_float(trade_feasibility.get("交易可行分數"), 50) or 50
    rr_trade = _safe_float(trade_plan.get("rr1"), _safe_float(trade_plan.get("rr2")))

    entry_score = 45.0
    entry_score += low_score * 0.10
    entry_score += pullback_score * 0.16
    entry_score += retest_score * 0.18
    entry_score += rebound_score * 0.14
    entry_score += trade_score * 0.12
    if support_dist is not None:
        if -1.5 <= support_dist <= 4.0:
            entry_score += 12
        elif 4.0 < support_dist <= 8.0:
            entry_score += 6
        elif support_dist > 12:
            entry_score -= 8
    if pressure_space is not None:
        if pressure_space >= 10:
            entry_score += 8
        elif pressure_space < 4:
            entry_score -= 8
    if rr_trade is not None:
        if rr_trade >= 2.0:
            entry_score += 10
        elif rr_trade >= 1.4:
            entry_score += 5
        elif rr_trade < 1.0:
            entry_score -= 8
    entry_score -= max(0, chase_risk - 55) * 0.35
    entry_score = _score_clip(entry_score)

    if chase_risk >= 78:
        chase_level = "高"
    elif chase_risk >= 62:
        chase_level = "中"
    else:
        chase_level = "低"

    if entry_score >= 82 and chase_risk <= 62:
        timing = "可分批進場"
        action = "小量分批，嚴守停損"
    elif entry_score >= 72 and chase_risk <= 72:
        timing = "接近可進場"
        action = "觀察承接，等紅K或量能確認"
    elif retest_score >= 70 and support_dist is not None and support_dist <= 5:
        timing = "等支撐確認"
        action = "支撐不破再分批，跌破停損"
    elif pullback_score >= 70:
        timing = "等拉回承接"
        action = "等靠近均線或支撐後觀察承接"
    elif chase_risk >= 75:
        timing = "不宜追高"
        action = "等拉回，不追價"
    else:
        timing = "觀察等待"
        action = "等待突破、支撐或量能確認"

    wait_parts = []
    if chase_risk >= 70:
        wait_parts.append("等追高風險下降")
    if near_support not in [None, 0]:
        wait_parts.append(f"守住支撐 {near_support:.2f}")
    if breakout_price not in [None, 0]:
        wait_parts.append(f"突破 {breakout_price:.2f} 轉強")
    if vr is None or vr < 0.8:
        wait_parts.append("等量能確認")
    if not wait_parts:
        wait_parts.append("等紅K續強與風險報酬維持")

    rr_decision = None
    if close_now not in [None, 0] and stop_ref not in [None, 0] and near_resistance not in [None, 0]:
        downside = max(0.01, close_now - stop_ref)
        upside = max(0.0, near_resistance - close_now)
        rr_decision = upside / downside if downside else None
    if rr_decision is None:
        rr_decision = rr_trade

    should_chase = "是" if (entry_score >= 80 and chase_risk <= 58 and pressure_space is not None and pressure_space >= 8) else "否"
    if should_chase == "否" and timing in ["可分批進場", "接近可進場"]:
        should_chase = "不追價，可分批"

    decision_note = f"{timing}；{action}。"
    if risk_reasons:
        decision_note += " 風險：" + "、".join(risk_reasons[:4]) + "。"
    if near_support not in [None, 0] or near_resistance not in [None, 0]:
        decision_note += f" 支撐/壓力參考：{format_number(near_support,2)} / {format_number(near_resistance,2)}。"

    return {
        "進場時機": timing,
        "進場時機分數": round(entry_score, 2),
        "建議動作": action,
        "等待條件": "、".join(wait_parts[:5]),
        "近端支撐": round(near_support, 2) if near_support not in [None, 0] else None,
        "主要支撐": round(main_support, 2) if main_support not in [None, 0] else None,
        "近端壓力": round(near_resistance, 2) if near_resistance not in [None, 0] else None,
        "突破確認價": round(breakout_price, 2) if breakout_price not in [None, 0] else None,
        "停損參考": round(stop_ref, 2) if stop_ref not in [None, 0] else None,
        "操作區間": operation_zone,
        "風險報酬比_決策": round(rr_decision, 2) if rr_decision not in [None, 0] else None,
        "追高風險分數_決策": round(chase_risk, 2),
        "追高風險等級": chase_level,
        "是否建議追價": should_chase,
        "風險扣分原因": "、".join(risk_reasons[:6]) if risk_reasons else "無明顯追高扣分",
        "決策說明": decision_note,
    }

def _is_opportunity_mode(mode: str) -> bool:
    text = _safe_str(mode)
    return any(k in text for k in ["低檔", "拉回", "回測", "機會", "保守低風險"])


def _build_entry_zone_text(pullback_buy: Any, breakout_buy: Any) -> str:
    pb = _safe_float(pullback_buy)
    bb = _safe_float(breakout_buy)
    if pb not in [None] and bb not in [None]:
        low = min(pb, bb)
        high = max(pb, bb)
        return f"{format_number(low, 2)} ~ {format_number(high, 2)}"
    if pb is not None:
        return format_number(pb, 2)
    if bb is not None:
        return format_number(bb, 2)
    return ""


def _build_market_environment(base_df: pd.DataFrame) -> dict[str, Any]:
    if base_df is None or base_df.empty:
        return {"score": 50.0, "label": "中性", "summary": "無市場樣本"}

    ret_mean = pd.to_numeric(base_df.get("區間漲跌幅%"), errors="coerce").fillna(0).mean()
    signal_mean = pd.to_numeric(base_df.get("訊號分數"), errors="coerce").fillna(0).mean()
    prelaunch_mean = pd.to_numeric(base_df.get("起漲前兆分數"), errors="coerce").fillna(0).mean()
    positive_ratio = (pd.to_numeric(base_df.get("區間漲跌幅%"), errors="coerce").fillna(0) > 0).mean()

    score = 50.0
    score += max(min(ret_mean * 0.9, 14), -14)
    score += max(min(signal_mean * 5.5, 18), -18)
    score += max(min((prelaunch_mean - 50) * 0.35, 12), -12)
    score += max(min((positive_ratio - 0.5) * 60, 10), -10)
    score = _score_clip(score)

    if score >= 80:
        label = "市場順風"
    elif score >= 68:
        label = "市場偏多"
    elif score >= 55:
        label = "市場中性偏多"
    elif score >= 45:
        label = "市場中性"
    else:
        label = "市場逆風"

    summary = f"{label}｜平均漲幅 {ret_mean:.2f}%｜正報酬占比 {positive_ratio*100:.1f}%"
    return {"score": score, "label": label, "summary": summary}




# =========================================================
# 永久設定 / 本輪推薦結果保存
# =========================================================
def _safe_json_read_local(path: str, default):
    try:
        p = Path(path)
        if not p.is_absolute():
            p = Path(__file__).resolve().parent.parent / p
        if not p.exists():
            return copy.deepcopy(default)
        with open(p, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
        return data if data is not None else copy.deepcopy(default)
    except Exception:
        return copy.deepcopy(default)


def _safe_json_write_local(path: str, payload) -> tuple[bool, str]:
    """原子寫入專案根目錄，並立即回讀驗證，避免 Reboot/工作目錄差異。"""
    try:
        target = Path(path)
        if not target.is_absolute():
            target = Path(__file__).resolve().parent.parent / target
        target.parent.mkdir(parents=True, exist_ok=True)
        tmp = target.with_suffix(target.suffix + ".tmp")
        tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        tmp.replace(target)
        verify = json.loads(target.read_text(encoding="utf-8"))
        if not isinstance(verify, type(payload)):
            raise ValueError("回讀型別不一致")
        return True, f"已原子寫入並回讀驗證：{target}"
    except Exception as e:
        return False, f"本機寫入失敗：{path} / {e}"


def _generic_github_file_config(path_name: str) -> dict[str, str]:
    return {
        "token": _safe_str(st.secrets.get("GITHUB_TOKEN", "")),
        "owner": _safe_str(st.secrets.get("GITHUB_REPO_OWNER", "cheng07021028")),
        "repo": _safe_str(st.secrets.get("GITHUB_REPO_NAME", "stock-app")),
        "branch": _safe_str(st.secrets.get("GITHUB_REPO_BRANCH", "main")) or "main",
        "path": path_name,
    }


def _read_json_from_github_path(path_name: str, default):
    cfg = _generic_github_file_config(path_name)
    token = cfg["token"]
    if not token:
        return copy.deepcopy(default), "未設定 GITHUB_TOKEN"

    try:
        resp = requests.get(
            _github_contents_url(cfg["owner"], cfg["repo"], cfg["path"]),
            headers=_github_headers(token),
            params={"ref": cfg["branch"]},
            timeout=20,
        )
        if resp.status_code == 404:
            return copy.deepcopy(default), ""
        if resp.status_code != 200:
            return copy.deepcopy(default), f"讀取 GitHub {path_name} 失敗：{resp.status_code}"

        content = resp.json().get("content", "")
        if not content:
            return copy.deepcopy(default), ""
        payload = json.loads(base64.b64decode(content).decode("utf-8"))
        return payload, ""
    except Exception as e:
        return copy.deepcopy(default), f"讀取 GitHub {path_name} 例外：{e}"


def _write_json_to_github_path_sync(cfg: dict[str, str], payload) -> tuple[bool, str]:
    """真正的 GitHub GET/PUT；只在背景執行，避免 Streamlit 按鈕等待網路。"""
    token = _safe_str(cfg.get("token"))
    path_name = _safe_str(cfg.get("path"))
    if not token:
        return False, "未設定 GITHUB_TOKEN"

    sha = ""
    try:
        resp = requests.get(
            _github_contents_url(cfg["owner"], cfg["repo"], path_name),
            headers=_github_headers(token),
            params={"ref": cfg["branch"]},
            timeout=12,
        )
        if resp.status_code == 200:
            sha = _safe_str(resp.json().get("sha"))
        elif resp.status_code != 404:
            return False, f"讀取 GitHub SHA 失敗：{resp.status_code}"
    except Exception as e:
        return False, f"讀取 GitHub SHA 例外：{e}"

    body = {
        "message": f"update {path_name} at {_now_text()}",
        "content": base64.b64encode(json.dumps(payload, ensure_ascii=False, indent=2, default=str).encode("utf-8")).decode("utf-8"),
        "branch": cfg["branch"],
    }
    if sha:
        body["sha"] = sha

    try:
        resp = requests.put(
            _github_contents_url(cfg["owner"], cfg["repo"], path_name),
            headers=_github_headers(token),
            json=body,
            timeout=20,
        )
        if resp.status_code in (200, 201):
            return True, f"已寫入 GitHub：{path_name}"
        return False, f"GitHub 寫入 {path_name} 失敗：{resp.status_code} / {resp.text[:300]}"
    except Exception as e:
        return False, f"GitHub 寫入 {path_name} 例外：{e}"


@st.cache_resource(show_spinner=False)
def _page07_github_write_executor():
    return ThreadPoolExecutor(max_workers=1, thread_name_prefix="godpick-page07-github")


@st.cache_resource(show_spinner=False)
def _page07_record_authority_executor_v181():
    # 推薦紀錄權威恢復可能需讀 GitHub/Firestore 20MB 級資料；絕不能再卡住
    # 「最後結果運算」與畫面呈現。單執行緒維持寫入順序，business key upsert 防重。
    return ThreadPoolExecutor(max_workers=1, thread_name_prefix="godpick-page07-records")


def _v181_background_record_upsert(rows: list[dict[str, Any]]) -> dict[str, Any]:
    try:
        if not callable(upsert_records_authority_fast):
            return {"ok": False, "message": "權威增量服務未載入", "count": 0}
        report, stats = upsert_records_authority_fast(
            copy.deepcopy(rows),
            reason="07 V181 推薦完成背景權威紀錄",
        )
        return {
            "ok": bool(getattr(report, "permanent_ok", False)),
            "message": "；".join([str(x) for x in (report.messages() if hasattr(report, "messages") else [])][:6]),
            "count": int(stats.get("changed", 0) or 0),
            "added": int(stats.get("added", 0) or 0),
            "updated": int(stats.get("updated", 0) or 0),
        }
    except Exception as exc:
        return {"ok": False, "message": f"V181背景權威紀錄例外：{exc}", "count": 0}


def _write_json_to_github_path(path_name: str, payload) -> tuple[bool, str]:
    """V164：本機先完成，GitHub 改為單執行緒背景排隊。"""
    cfg = _generic_github_file_config(path_name)
    if not _safe_str(cfg.get("token")):
        return False, "未設定 GITHUB_TOKEN"
    try:
        _page07_github_write_executor().submit(
            _write_json_to_github_path_sync,
            dict(cfg),
            copy.deepcopy(payload),
        )
        return True, f"GitHub 背景同步已排程：{path_name}"
    except Exception as e:
        return False, f"GitHub 背景同步排程失敗：{e}"



def _payload_authority_stamp_v185(payload: Any) -> tuple[str, str]:
    if not isinstance(payload, dict):
        return "", ""
    date_text = _safe_str(
        payload.get("recommendation_date") or payload.get("saved_at") or payload.get("kline_date")
    )
    try:
        ts = pd.to_datetime(date_text, errors="coerce")
        date_key = ts.strftime("%Y-%m-%d") if pd.notna(ts) else date_text[:10]
    except Exception:
        date_key = date_text[:10]
    saved_key = _safe_str(payload.get("saved_at")).replace("T", " ")[:26]
    return date_key, saved_key


@st.cache_data(show_spinner=False, ttl=30)
def _load_latest_recommendation_authority_v185() -> tuple[dict[str, Any], list[str]]:
    """Load the newest recommendation authority without GitHub-first rollback.

    V184 and earlier could read the repository's bundled 2026-07-09 JSON before a
    newer local/Firestore copy.  V185 uses a small durable run anchor first, then
    restores the full snapshot only when needed.  The actual recommendation
    business timestamp, not deployment/file time, decides which payload is newer.
    """
    local_full = _safe_json_read_local(GODPICK_LATEST_FILE, {})
    details: list[str] = []
    anchor: dict[str, Any] = {}
    try:
        from godpick_persistence_service import load_named_json_permanent
        anchor_raw, anchor_details = load_named_json_permanent(
            GODPICK_LATEST_ANCHOR_FILE, {}, firestore_doc="godpick_latest_run_anchor"
        )
        if isinstance(anchor_raw, dict):
            anchor = anchor_raw
        details.extend([f"錨點｜{x}" for x in (anchor_details or [])])
    except Exception as exc:
        details.append(f"錨點永久層讀取例外：{exc}")

    local_stamp = _payload_authority_stamp_v185(local_full)
    anchor_stamp = _payload_authority_stamp_v185(anchor)
    full_payload = local_full if isinstance(local_full, dict) else {}

    # If no V185 anchor exists yet, or the anchor proves the repo/local full pack
    # is older, restore the large full snapshot from the durable authority layer.
    if not anchor_stamp[0] or anchor_stamp > local_stamp:
        try:
            from godpick_persistence_service import load_named_json_permanent
            remote_full, full_details = load_named_json_permanent(
                GODPICK_LATEST_FILE, {}, firestore_doc="godpick_latest_recommendations"
            )
            if isinstance(remote_full, dict) and remote_full:
                full_payload = remote_full
            details.extend([f"完整快照｜{x}" for x in (full_details or [])])
        except Exception as exc:
            details.append(f"完整快照永久層讀取例外：{exc}")

    full_stamp = _payload_authority_stamp_v185(full_payload)
    # A successfully persisted anchor is intentionally sufficient to prevent the
    # UI from lying that the latest run is still 7/9 while the large candidate
    # pack is still syncing.  It contains the actionable recommendation rows and
    # enough scan metadata to reconstruct the latest page safely.
    if anchor_stamp > full_stamp:
        recovered = dict(anchor)
        recovered.setdefault("recommendations", [])
        recovered.setdefault("candidate_diagnosis", [])
        recovered.setdefault("category_strength", [])
        recovered.setdefault("hot_pick", [])
        recovered["authority_recovery"] = "V185 durable run anchor"
        recovered["full_snapshot_pending_or_older"] = True
        details.insert(0, f"V185權威：永久錨點較新（{anchor_stamp[1] or anchor_stamp[0]}），禁止回退舊完整快照。")
        return recovered, details

    if isinstance(full_payload, dict) and full_payload:
        details.insert(0, f"V185權威：完整推薦快照（{full_stamp[1] or full_stamp[0] or '日期未驗證'}）。")
        return full_payload, details
    if isinstance(anchor, dict) and anchor:
        details.insert(0, "V185權威：僅有永久錨點。")
        return anchor, details
    return {}, details



def _settings_ts_value(payload: dict[str, Any]) -> datetime:
    raw = _safe_str(payload.get("updated_at")) or _safe_str(payload.get("weight_settings_updated_at"))
    try:
        return datetime.strptime(raw[:19], "%Y-%m-%d %H:%M:%S")
    except Exception:
        return datetime.min


def _weight_stamp_from_payload(payload: dict[str, Any]) -> str:
    if not isinstance(payload, dict):
        return ""
    # 優先使用權重專屬時間，避免欄位設定 / 掃描設定更新誤判成權重更新。
    return _safe_str(payload.get("weight_settings_updated_at")) or _safe_str(payload.get("updated_at"))


def _load_persistent_settings(local_first: bool = True) -> dict[str, Any]:
    """讀取股神推薦永久設定。

    V143 重點：14_股神權重校正剛套用後，7_股神推薦優先讀本機 JSON，
    避免 GitHub API 尚未同步或舊快取造成必須 Ctrl+F5 才會更新。一般情況仍可用
    updated_at / weight_settings_updated_at 比較 GitHub 與本機的新舊。
    """
    default_payload = {
        "original_default_weights": GODPICK_DEFAULT_SCORE_WEIGHTS.copy(),
        "applied_weights": GODPICK_DEFAULT_SCORE_WEIGHTS.copy(),
        "score_weights": GODPICK_DEFAULT_SCORE_WEIGHTS.copy(),
        "column_orders": {},
        "scan_settings": {},
        "updated_at": "",
        "weight_settings_updated_at": "",
        "weight_update_seq": 0,
        "weight_source": "",
        "last_weight_calibration_profile": "",
        "version": "godpick_v5_persistent_settings",
    }

    local_payload = _safe_json_read_local(GODPICK_SETTINGS_FILE, {})
    github_payload: dict[str, Any] = {}
    github_msg = ""
    if local_first and isinstance(local_payload, dict) and local_payload:
        github_msg = "本次重新載入採本機優先，未等待 GitHub。"
    else:
        github_payload, github_msg = _read_json_from_github_path(GODPICK_SETTINGS_FILE, {})

    candidates: list[tuple[str, dict[str, Any]]] = []
    if isinstance(github_payload, dict) and github_payload:
        candidates.append(("github", github_payload))
    if isinstance(local_payload, dict) and local_payload:
        candidates.append(("local", local_payload))

    if not candidates:
        payload = default_payload.copy()
    elif local_first and isinstance(local_payload, dict) and local_payload:
        payload = local_payload
    elif len(candidates) == 1:
        payload = candidates[0][1]
    else:
        def _candidate_ts(item: tuple[str, dict[str, Any]]):
            source, data = item
            ts = _settings_ts_value(data)
            if ts == datetime.min and source == "local":
                return datetime.max
            return ts

        payload = sorted(candidates, key=_candidate_ts, reverse=True)[0][1]

    payload = {**default_payload, **payload}
    raw_weights = payload.get("applied_weights")
    if not isinstance(raw_weights, dict):
        raw_weights = payload.get("score_weights")
    payload["applied_weights"] = _normalize_weight_map(raw_weights)
    payload["score_weights"] = payload["applied_weights"].copy()
    if not _safe_str(payload.get("weight_settings_updated_at")):
        payload["weight_settings_updated_at"] = _safe_str(payload.get("updated_at"))
    try:
        payload["weight_update_seq"] = int(payload.get("weight_update_seq", 0) or 0)
    except Exception:
        payload["weight_update_seq"] = 0
    if not isinstance(payload.get("column_orders"), dict):
        payload["column_orders"] = {}
    if not isinstance(payload.get("scan_settings"), dict):
        payload["scan_settings"] = {}
    st.session_state[_k("persistent_settings_source_detail")] = f"GitHub: {github_msg}｜本機設定: {'有' if isinstance(local_payload, dict) and local_payload else '無'}"
    return payload


def _save_persistent_settings(applied_weights: dict[str, int]) -> tuple[bool, list[str]]:
    old_payload = _load_persistent_settings(local_first=True)
    now = _now_text()
    try:
        old_seq = int(old_payload.get("weight_update_seq", 0) or 0) if isinstance(old_payload, dict) else 0
    except Exception:
        old_seq = 0
    normalized = _normalize_weight_map(applied_weights)
    payload = {
        "original_default_weights": GODPICK_DEFAULT_SCORE_WEIGHTS.copy(),
        "applied_weights": normalized,
        "score_weights": normalized.copy(),
        "column_orders": old_payload.get("column_orders", {}) if isinstance(old_payload, dict) else {},
        "scan_settings": old_payload.get("scan_settings", {}) if isinstance(old_payload, dict) else {},
        "updated_at": now,
        "weight_settings_updated_at": now,
        "weight_update_seq": old_seq + 1,
        "weight_source": "7_股神推薦",
        "last_weight_calibration_profile": "7_股神推薦手動套用",
        "version": "godpick_v143_weight_reload_sync",
    }
    local_ok, local_msg = _safe_json_write_local(GODPICK_SETTINGS_FILE, payload)
    github_ok, github_msg = _write_json_to_github_path(GODPICK_SETTINGS_FILE, payload)
    return (local_ok or github_ok), [local_msg, github_msg]

def _apply_persisted_weights_to_state(payload: dict[str, Any] | None = None, *, force_widget_sync: bool = True) -> tuple[bool, str]:
    """V143：把 godpick_user_settings.json 的 applied_weights 立即套回 7_股神推薦。"""
    if payload is None:
        payload = _load_persistent_settings(local_first=True)
    if not isinstance(payload, dict):
        return False, "設定檔格式異常，無法重新載入權重。"
    raw_weights = payload.get("applied_weights") if isinstance(payload.get("applied_weights"), dict) else payload.get("score_weights")
    weights = _normalize_weight_map(raw_weights)
    weight_stamp = _weight_stamp_from_payload(payload)
    profile = _safe_str(payload.get("last_weight_calibration_profile") or payload.get("weight_profile") or payload.get("weight_source") or "")
    try:
        weight_seq = int(payload.get("weight_update_seq", 0) or 0)
    except Exception:
        weight_seq = 0

    st.session_state[_k("score_weights")] = weights.copy()
    st.session_state[_k("score_weights_edit")] = weights.copy()
    st.session_state[_k("weight_settings_loaded_at")] = weight_stamp
    st.session_state[_k("weight_settings_loaded_seq")] = weight_seq
    st.session_state[_k("weight_settings_loaded_profile")] = profile

    if force_widget_sync:
        for _name, _val in weights.items():
            st.session_state[_k(f"weight_edit_{_name}")] = int(_val)

    detail = " / ".join([f"{k}{v}%" for k, v in weights.items()])
    suffix = f"｜來源/組合：{profile}" if profile else ""
    stamp = f"｜權重時間：{weight_stamp}" if weight_stamp else ""
    seq_text = f"｜版本序號：{weight_seq}" if weight_seq else ""
    return True, f"已重新載入 14_股神權重校正套用權重：{detail}{suffix}{stamp}{seq_text}"


def _maybe_auto_reload_weight_settings() -> None:
    """V143：偵測 14_股神權重校正已更新時，自動同步權重到本頁 session_state。"""
    payload = _load_persistent_settings(local_first=True)
    weight_stamp = _weight_stamp_from_payload(payload) if isinstance(payload, dict) else ""
    loaded_at = _safe_str(st.session_state.get(_k("weight_settings_loaded_at")))
    try:
        weight_seq = int(payload.get("weight_update_seq", 0) or 0) if isinstance(payload, dict) else 0
    except Exception:
        weight_seq = 0
    try:
        loaded_seq = int(st.session_state.get(_k("weight_settings_loaded_seq"), 0) or 0)
    except Exception:
        loaded_seq = 0

    should_reload = False
    if not st.session_state.get(_k("weight_settings_auto_loaded_once"), False):
        should_reload = True
        st.session_state[_k("weight_settings_auto_loaded_once")] = True
    elif weight_seq and weight_seq != loaded_seq:
        should_reload = True
    elif weight_stamp and weight_stamp != loaded_at:
        should_reload = True

    if should_reload:
        ok, msg = _apply_persisted_weights_to_state(payload, force_widget_sync=True)
        if ok and ((weight_stamp and loaded_at and weight_stamp != loaded_at) or (weight_seq and weight_seq != loaded_seq)):
            st.session_state[_k("weight_reload_notice")] = "偵測到 14_股神權重校正已更新，已自動重新載入權重。"


def _load_column_order_shadow_payload() -> dict[str, Any]:
    """v72：欄位順序獨立檔，避免套用後被 GitHub 舊設定覆蓋。"""
    payload = _safe_json_read_local(GODPICK_COLUMN_ORDER_FILE, {})
    return payload if isinstance(payload, dict) else {}


def _save_column_order_shadow_payload(payload: dict[str, Any]) -> tuple[bool, str]:
    if not isinstance(payload, dict):
        payload = {}
    payload["updated_at"] = _now_text()
    payload["version"] = "godpick_column_order_v72_stable"
    return _safe_json_write_local(GODPICK_COLUMN_ORDER_FILE, payload)


def _load_persistent_column_order(name: str) -> list[str]:
    shadow = _load_column_order_shadow_payload()
    shadow_orders = shadow.get("column_orders", {}) if isinstance(shadow, dict) else {}
    val = shadow_orders.get(name, []) if isinstance(shadow_orders, dict) else []
    if isinstance(val, list) and val:
        return [str(x) for x in val if str(x)]

    payload = _load_persistent_settings()
    orders = payload.get("column_orders", {}) if isinstance(payload, dict) else {}
    val = orders.get(name, []) if isinstance(orders, dict) else []
    return [str(x) for x in val if str(x)] if isinstance(val, list) else []


def _save_persistent_column_order(name: str, order: list[str]) -> tuple[bool, list[str]]:
    clean_order = [str(x) for x in order if str(x)]

    shadow = _load_column_order_shadow_payload()
    shadow_orders = shadow.get("column_orders", {}) if isinstance(shadow, dict) else {}
    if not isinstance(shadow_orders, dict):
        shadow_orders = {}
    shadow_orders[name] = clean_order
    shadow["column_orders"] = shadow_orders
    shadow_ok, shadow_msg = _save_column_order_shadow_payload(shadow)

    payload = _load_persistent_settings()
    orders = payload.get("column_orders", {}) if isinstance(payload, dict) else {}
    if not isinstance(orders, dict):
        orders = {}
    orders[name] = clean_order
    payload["column_orders"] = orders
    payload["applied_weights"] = _normalize_weight_map(payload.get("applied_weights", GODPICK_DEFAULT_SCORE_WEIGHTS))
    payload["updated_at"] = _now_text()
    payload["version"] = "godpick_v72_column_order_stable"
    local_ok, local_msg = _safe_json_write_local(GODPICK_SETTINGS_FILE, payload)
    github_ok, github_msg = _write_json_to_github_path(GODPICK_SETTINGS_FILE, payload)
    return (shadow_ok or local_ok or github_ok), [shadow_msg, local_msg, github_msg]


def _df_to_records_for_json(df: pd.DataFrame) -> list[dict[str, Any]]:
    if df is None or df.empty:
        return []
    clean = df.copy()
    for col in clean.columns:
        if pd.api.types.is_datetime64_any_dtype(clean[col]):
            clean[col] = clean[col].astype(str)
    return json.loads(clean.to_json(orient="records", force_ascii=False, date_format="iso"))


def _records_to_df_for_json(records: list[dict[str, Any]]) -> pd.DataFrame:
    if not isinstance(records, list) or not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def _ensure_v92_night_compat_df(df: pd.DataFrame | None, *, source: str = "") -> pd.DataFrame:
    """V92：舊推薦快取 / 舊推薦清單相容。

    V90 之後新增夜間隔日欄位，但使用者可能已經有舊的
    godpick_latest_recommendations.json 或 session_state。
    這裡只在缺欄時做一次快速補齊，避免後續 data_editor / 欄位管理 / 本輪精華推薦 KeyError。
    若 enrich_night_strategy 可用，會用既有價量欄位推估隔日作戰策略；否則只補空欄。
    """
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    if df.empty:
        return df.copy()

    out = df.copy()
    # 先用共用欄位標準化補別名與去重，避免舊資料欄位名不一致。
    try:
        if callable(normalize_godpick_dataframe):
            out = normalize_godpick_dataframe(out, add_missing=False, clean_none=True)
    except Exception:
        pass

    night_cols = list(GODPICK_NIGHT_COLUMNS or [])
    need_night = bool(night_cols) and any(c not in out.columns for c in night_cols)
    if need_night and callable(enrich_night_strategy):
        try:
            out = enrich_night_strategy(out)
        except Exception as e:
            for c in night_cols:
                if c not in out.columns:
                    out[c] = ""
            out["資料完整度"] = out.get("資料完整度", "")
            mask = out["資料完整度"].astype(str).str.strip().eq("") if "資料完整度" in out.columns else pd.Series([True] * len(out), index=out.index)
            out.loc[mask, "資料完整度"] = f"舊快取補欄；夜間策略未重算：{e}"
    elif need_night:
        for c in night_cols:
            if c not in out.columns:
                out[c] = ""
        if "資料完整度" not in out.columns:
            out["資料完整度"] = "舊快取補欄；夜間策略模組未載入"

    # V92：幾個跨頁必用欄位一定存在。
    for c in ["股票代號", "股票名稱", "推薦總分", "推薦等級", "資料完整度"]:
        if c not in out.columns:
            out[c] = ""
    return out


def _conditional_reference_rows(df: pd.DataFrame | None, *, max_rows: int = 8) -> pd.DataFrame:
    """Build a small, explicitly non-buy conditional reference list.

    This fallback is used only when the formal/A-/R1 action list is empty.  It
    never promotes exclusion, high-risk, over-heated or low-liquidity stocks.
    Every returned row remains a *reference* candidate and is marked as
    "未觸發不可買" so the UI can still provide useful names without fabricating a
    formal recommendation.
    """
    work = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(df or [])
    if work.empty:
        return work
    if callable(canonicalize_final_partition):
        try:
            work = canonicalize_final_partition(work)
        except Exception:
            pass

    def _text_series(col: str) -> pd.Series:
        if col not in work.columns:
            return pd.Series([""] * len(work), index=work.index, dtype="object")
        return work[col].fillna("").astype(str).str.strip()

    def _num_series(col: str, default: float = 0.0) -> pd.Series:
        if col not in work.columns:
            return pd.Series([float(default)] * len(work), index=work.index, dtype="float64")
        return pd.to_numeric(work[col], errors="coerce").fillna(float(default))

    bucket = _text_series("正式推薦分區")
    role_blob = (
        _text_series("正式推薦資格") + "｜" + _text_series("操作許可") + "｜"
        + _text_series("正式推薦排除原因") + "｜" + _text_series("真禁買原因") + "｜"
        + _text_series("硬否決原因") + "｜" + _text_series("過熱原因") + "｜"
        + _text_series("流動性等級") + "｜" + _text_series("冷門股警示")
    )
    hard_block = (
        bucket.isin(["正式排除清單", "高風險雷達觀察"])
        | role_blob.str.contains("BLOCK|禁止買進|過熱禁買|假強排除|低流動性排除|冷門禁追|極低量", na=False)
    )

    amount = _num_series("成交額百萬")
    avg_amount = _num_series("20日均成交額百萬")
    volume = _num_series("最新成交量_張")
    avg_volume = _num_series("20日均量_張")
    liquid = amount.ge(80) | avg_amount.ge(80) | volume.ge(1000) | avg_volume.ge(1000)

    latest_price = _num_series("最新價")
    strength = pd.concat(
        [
            _num_series("候選強度分"),
            _num_series("推薦總分"),
            _num_series("股神實戰總分"),
            _num_series("Alpha選股潛力分"),
        ],
        axis=1,
    ).max(axis=1)
    operability = _num_series("可操作分")
    entry = pd.concat([_num_series("Entry進場買點分"), _num_series("進場買點分")], axis=1).max(axis=1)
    risk = pd.concat([_num_series("Risk風控安全分"), _num_series("風控安全分")], axis=1).max(axis=1)
    rr = pd.concat([_num_series("實戰風險報酬比"), _num_series("風險報酬比"), _num_series("風險報酬比_決策")], axis=1).max(axis=1)
    chase = pd.concat([_num_series("追價風險分"), _num_series("追高風險分數_決策")], axis=1).max(axis=1)
    rise5 = _num_series("近5日漲幅%")
    rise20 = _num_series("近20日漲幅%")

    reference_bucket = bucket.isin(["盤中雷達追蹤", "早期潛伏觀察", "不可直接買觀察"])
    strict_mask = (
        reference_bucket & ~hard_block & liquid & latest_price.gt(0)
        & strength.ge(60) & operability.ge(48) & entry.ge(48) & risk.ge(48)
        & chase.le(72) & rr.ge(0.80) & rise5.le(18) & rise20.le(35)
    )
    candidates = work.loc[strict_mask].copy()

    # A second, still conservative pass prevents a completely blank page when
    # one optional score is slightly below the strict line.  Hard vetoes,
    # liquidity, price and over-heating checks remain mandatory.
    if candidates.empty:
        relaxed_mask = (
            reference_bucket & ~hard_block & liquid & latest_price.gt(0)
            & strength.ge(58) & operability.ge(42) & entry.ge(42) & risk.ge(42)
            & chase.le(76) & rr.ge(0.50) & rise5.le(20) & rise20.le(40)
        )
        candidates = work.loc[relaxed_mask].copy()

    if candidates.empty:
        return candidates

    # Build an execution-oriented rank; this is only for ordering conditional
    # references and never upgrades their formal recommendation permission.
    idx = candidates.index
    cand_strength = strength.loc[idx]
    cand_op = operability.loc[idx]
    cand_entry = entry.loc[idx]
    cand_risk = risk.loc[idx]
    cand_rr = rr.loc[idx].clip(lower=0, upper=4)
    cand_chase = chase.loc[idx]
    candidates["__conditional_reference_score"] = (
        cand_strength * 0.30
        + cand_op * 0.25
        + cand_entry * 0.15
        + cand_risk * 0.15
        + (cand_rr * 25).clip(upper=100) * 0.10
        + (100 - cand_chase).clip(lower=0, upper=100) * 0.05
    ).round(2)
    candidates = candidates.sort_values("__conditional_reference_score", ascending=False, kind="mergesort")

    # Avoid filling the fallback with one single industry/theme.
    selected_indices: list[Any] = []
    sector_counts: dict[str, int] = {}
    for row_idx, row in candidates.iterrows():
        sector = _safe_str(row.get("類別")) or _safe_str(row.get("產業")) or "未分類"
        if sector_counts.get(sector, 0) >= 2:
            continue
        selected_indices.append(row_idx)
        sector_counts[sector] = sector_counts.get(sector, 0) + 1
        if len(selected_indices) >= max(1, int(max_rows)):
            break
    if not selected_indices:
        selected_indices = candidates.index[: max(1, int(max_rows))].tolist()
    candidates = candidates.loc[selected_indices].copy()

    def _reference_reason(row: pd.Series) -> str:
        parts = []
        op_v = _safe_float(row.get("可操作分"), 0) or 0
        entry_v = max(_safe_float(row.get("Entry進場買點分"), 0) or 0, _safe_float(row.get("進場買點分"), 0) or 0)
        risk_v = max(_safe_float(row.get("Risk風控安全分"), 0) or 0, _safe_float(row.get("風控安全分"), 0) or 0)
        rr_v = max(_safe_float(row.get("實戰風險報酬比"), 0) or 0, _safe_float(row.get("風險報酬比"), 0) or 0)
        if op_v >= 55:
            parts.append("可操作性尚可")
        if entry_v >= 55:
            parts.append("買點接近確認")
        if risk_v >= 55:
            parts.append("風控條件尚可")
        if rr_v >= 1.2:
            parts.append(f"風報比 {rr_v:.2f}")
        if _safe_float(row.get("成交額百萬"), 0) >= 100:
            parts.append("流動性可交易")
        return "、".join(parts[:4]) or "條件接近，但尚未達正式推薦門檻"

    candidates["推薦用途"] = "條件式參考名單"
    candidates["條件式參考原因"] = candidates.apply(_reference_reason, axis=1)
    candidates["最終操作結論"] = "B｜條件式參考：等待盤中量價、觸發價與守價確認"
    candidates["是否正式推薦"] = "否"
    candidates["操作許可"] = "未觸發不可買｜僅供條件式參考"
    candidates["正式推薦等級"] = "B｜條件式參考"
    candidates["建議倉位上限%"] = 0.0
    candidates["正式推薦動作"] = "只加入觀察；放量站上實戰觸發價並守穩後才重新評估，未觸發不得買進。"
    candidates["條件式參考排序分"] = candidates["__conditional_reference_score"]
    return candidates.drop(columns=["__conditional_reference_score"], errors="ignore").reset_index(drop=True)


def _v191_actionable_tracking_frame(source_df: pd.DataFrame | None) -> tuple[pd.DataFrame, bool, list[str]]:
    """Single-source actionable partition for snapshot/list/Page08 persistence.

    H7 removes the previous split-brain behavior where Page08 selected R1 rows
    from the full candidate diagnosis while the latest snapshot/Page10 list ran
    the already-filtered display frame through governance a second time.  The
    result could be ``08永久紀錄 2`` but ``recommendations=[]`` / Page10=0.
    This helper is the one partition used by both outputs.
    """
    notes: list[str] = []
    report = st.session_state.get(_k("scan_quality_report"), {})
    formal_scan_ok = bool(isinstance(report, dict) and report.get("正式推薦可用", False))
    work = source_df.copy() if isinstance(source_df, pd.DataFrame) else pd.DataFrame(source_df or [])
    if work.empty or "股票代號" not in work.columns:
        return pd.DataFrame(), formal_scan_ok, ["本輪候選為空，沒有可追蹤推薦。"]
    try:
        work = _phase93_single_source_decision_frame(work, work)
    except Exception:
        try:
            work = canonicalize_final_partition(work) if callable(canonicalize_final_partition) else work.copy()
        except Exception:
            work = work.copy()
    if "正式推薦分區" not in work.columns:
        try:
            work = apply_formal_recommendation_engine(work)
        except Exception as exc:
            return pd.DataFrame(), formal_scan_ok, [f"正式分區重算失敗：{exc}"]

    bucket = work.get("正式推薦分區", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str)
    radar = work.get("盤中雷達優先級", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str)
    formal_mask = bucket.isin(["正式下週主推薦", "A-｜準主推薦小量試單"])
    radar_mask = bucket.eq("盤中雷達追蹤") & radar.str.startswith("R1")
    allowed = (formal_mask if formal_scan_ok else pd.Series([False] * len(work), index=work.index)) | radar_mask
    action = work.loc[allowed].copy()
    if action.empty:
        if not formal_scan_ok:
            notes.append("整體掃描未達正式可用；正式/A-不進追蹤清單，僅允許R1研究雷達。")
        notes.append("本輪沒有正式/A-/R1可追蹤名單。")
        return action, formal_scan_ok, notes

    quality_keep: list[bool] = []
    for _, row in action.iterrows():
        if callable(assess_individual_sample_quality):
            try:
                eligible, _, _ = assess_individual_sample_quality(row)
            except Exception:
                eligible = True
        else:
            eligible = True
        quality_keep.append(bool(eligible))
    action = action.loc[pd.Series(quality_keep, index=action.index)].copy()
    if action.empty:
        notes.append("入選分區個股資料品質不足，沒有可追蹤名單。")
        return action, formal_scan_ok, notes
    action["股票代號"] = action["股票代號"].astype(str).map(_normalize_code)
    action = action[action["股票代號"].astype(str).str.strip().ne("")].copy()
    action = action.drop_duplicates(subset=["股票代號"], keep="first")
    notes.append(f"H7單一行動分區：正式/A-/R1 共 {len(action)} 檔。")
    return action.reset_index(drop=True), formal_scan_ok, notes


def _operational_recommendation_rows(df: pd.DataFrame | None, *, refresh_decision: bool = False) -> pd.DataFrame:
    """Return a useful action/reference list without mixing in bad stocks.

    Formal/A- recommendations and R1 core intraday radar are preferred.  When
    none exist, a small conditional-reference fallback is built from the full
    governed candidate pool.  Formal exclusions and high-risk observations are
    never allowed into that fallback.
    """
    work = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(df or [])
    if work.empty:
        return work
    if refresh_decision and callable(apply_godpick_decision_engine):
        try:
            work = apply_godpick_decision_engine(work, None)
        except Exception:
            pass
    if callable(canonicalize_final_partition):
        try:
            work = canonicalize_final_partition(work)
        except Exception:
            pass
    governed = pd.DataFrame()
    if callable(govern_recommend_list):
        try:
            governed = govern_recommend_list(work, include_r1=True)
        except Exception:
            governed = pd.DataFrame()
    if governed.empty and "正式推薦分區" in work.columns:
        bucket = work["正式推薦分區"].fillna("").astype(str)
        radar = work.get("盤中雷達優先級", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str)
        allowed = bucket.isin(["正式下週主推薦", "A-｜準主推薦小量試單"]) | (bucket.eq("盤中雷達追蹤") & radar.str.startswith("R1"))
        governed = work.loc[allowed].copy()
    if governed.empty:
        governed = _conditional_reference_rows(work, max_rows=8)
    if governed.empty:
        return governed
    if "股票代號" in governed.columns:
        governed["股票代號"] = governed["股票代號"].astype(str).map(_normalize_code)
        governed = governed[governed["股票代號"].astype(str).str.strip().ne("")].copy()
        governed = governed.drop_duplicates(subset=["股票代號"], keep="first")
    bucket_order = {"正式下週主推薦": 10, "A-｜準主推薦小量試單": 20, "盤中雷達追蹤": 30}
    governed["__action_order"] = governed.get("正式推薦分區", pd.Series([""] * len(governed), index=governed.index)).map(bucket_order).fillna(80)
    sort_cols = ["__action_order"] + [c for c in ["正式推薦排序分", "條件式參考排序分", "可操作分", "推薦可信度分", "實戰操作品質分", "候選強度分"] if c in governed.columns]
    ascending = [True] + [False] * (len(sort_cols) - 1)
    governed = governed.sort_values(sort_cols, ascending=ascending, kind="mergesort")
    return governed.drop(columns=["__action_order"], errors="ignore").reset_index(drop=True)

def _save_latest_recommendation_pack(rec_df: pd.DataFrame, category_strength_df: pd.DataFrame, hot_pick_df: pd.DataFrame) -> tuple[bool, list[str]]:
    """Persist action results, compact full-candidate diagnosis and scan quality.

    ``godpick_recommend_list.json`` is now governed: only formal/A- and R1 core
    radar rows may enter module 10.  Observation and exclusion rows remain in the
    candidate diagnosis, not in the operational recommendation list.
    """
    candidate_df = st.session_state.get(_k("candidate_diagnosis_store"))
    if not isinstance(candidate_df, pd.DataFrame) or candidate_df.empty:
        try:
            candidate_df = build_candidate_diagnosis(rec_df) if callable(build_candidate_diagnosis) else rec_df.copy()
        except Exception:
            candidate_df = rec_df.copy() if isinstance(rec_df, pd.DataFrame) else pd.DataFrame()
    # H7: latest snapshot + Page10 current list must use the exact same
    # single-source formal/A-/R1 partition as Page08 auto-recording.
    action_source = candidate_df if isinstance(candidate_df, pd.DataFrame) and not candidate_df.empty else rec_df
    action_df, _formal_scan_ok_h7, _action_partition_notes_h7 = _v191_actionable_tracking_frame(action_source)

    scan_report = st.session_state.get(_k("scan_quality_report"), {})
    if not isinstance(scan_report, dict):
        scan_report = {}

    saved_at_now = _now_text()
    candidate_records = _df_to_records_for_json(candidate_df)
    recommendation_records = _df_to_records_for_json(action_df)
    kline_date = _max_row_date_v173(candidate_records + recommendation_records, [
        "本輪市場最新交易日", "K線最後交易日", "行情資料日期", "價格資料日期"
    ])
    execution_context = st.session_state.get(_k("recommend_execution_context_v191"), {})
    if not isinstance(execution_context, dict):
        execution_context = {}
    payload = {
        "saved_at": saved_at_now,
        "recommendation_date": saved_at_now[:10],
        "execution_context": dict(execution_context),
        "execution_owner": _safe_str(execution_context.get("owner")) or "07_股神推薦",
        "execution_trigger": _safe_str(execution_context.get("trigger")) or "手動操作",
        "scan_run_id": _safe_str(st.session_state.get(_k("scan_run_id"))) or f"scan_{saved_at_now.replace(':','').replace(' ','_')}",
        "expected_trade_date": _expected_latest_trade_date_v173().strftime("%Y-%m-%d"),
        "kline_date": kline_date.strftime("%Y-%m-%d") if kline_date is not None else "",
        "weights": _normalize_weight_map(st.session_state.get(_k("score_weights"), GODPICK_DEFAULT_SCORE_WEIGHTS)),
        "recommendations": recommendation_records,
        "candidate_diagnosis": candidate_records,
        "scan_quality": scan_report,
        "category_strength": _df_to_records_for_json(category_strength_df),
        "hot_pick": _df_to_records_for_json(hot_pick_df),
        "execution_governance_version": EXECUTION_GOVERNANCE_VERSION,
        "snapshot_version": "phase104_verified_even_when_zero_formal",
    }
    local_ok, local_msg = _safe_json_write_local(GODPICK_LATEST_FILE, payload)
    verify_payload = _read_project_json_file(GODPICK_LATEST_FILE) if local_ok else {}
    verified = bool(
        isinstance(verify_payload, dict)
        and _safe_str(verify_payload.get("saved_at")) == saved_at_now
        and isinstance(verify_payload.get("candidate_diagnosis"), list)
    )
    if local_ok and not verified:
        local_ok = False
        local_msg = f"快照寫入後回讀驗證失敗：{GODPICK_LATEST_FILE}"
    # V185：完整候選快照可達數 MB，遠端同步仍採背景；但「本輪真的發生過」
    # 不能再只靠背景執行緒。先同步永久保存一份小型 run anchor，內含
    # actionable 名單、掃描品質與K線日期。即使 Streamlit 立刻 rerun/reboot，
    # 頁首與最新推薦也不會再從 8/11 回退到 repo 內的 7/9 舊快照。
    try:
        full_payload_hash = hashlib.sha256(
            json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()
    except Exception:
        full_payload_hash = ""
    anchor_payload = {
        "saved_at": saved_at_now,
        "recommendation_date": saved_at_now[:10],
        "execution_context": payload.get("execution_context", {}),
        "execution_owner": payload.get("execution_owner", "07_股神推薦"),
        "execution_trigger": payload.get("execution_trigger", "手動操作"),
        "scan_run_id": payload.get("scan_run_id", ""),
        "expected_trade_date": payload.get("expected_trade_date", ""),
        "kline_date": payload.get("kline_date", ""),
        "weights": payload.get("weights", {}),
        "recommendations": recommendation_records,
        "category_strength": payload.get("category_strength", []),
        "hot_pick": payload.get("hot_pick", []),
        "scan_quality": scan_report,
        "candidate_count": len(candidate_records),
        "recommendation_count": len(recommendation_records),
        "full_snapshot_payload_hash": full_payload_hash,
        "snapshot_version": "V185_durable_run_anchor",
    }
    try:
        from godpick_durability_service import persist_json_permanent as _persist_anchor_v185
        from godpick_persistence_service import github_config as _anchor_gh_cfg_v185, firebase_configured as _anchor_fs_cfg_v185
        _anchor_remote_configured_v185 = bool(
            _safe_str((_anchor_gh_cfg_v185() or {}).get("token")) or bool(_anchor_fs_cfg_v185())
        )
        anchor_ok, anchor_msg = _persist_anchor_v185(
            GODPICK_LATEST_ANCHOR_FILE, anchor_payload,
            firestore_doc="godpick_latest_run_anchor",
            reason="V185 synchronous recommendation run anchor",
        )
        if not _anchor_remote_configured_v185:
            anchor_ok = False
            anchor_msg = f"{anchor_msg}｜未設定 GitHub/Firebase 遠端永久層；僅本機寫入不得宣稱永久保存"
    except Exception as anchor_exc:
        anchor_ok, anchor_msg = False, f"V185永久錨點保存例外：{anchor_exc}"

    # V184：最新推薦已列入 V183 durability registry；不要再只靠單一路徑
    # GitHub 背景寫入。先本機原子保存，再由 durability outbox 同步
    # runtime-data + Firestore，重啟時才有可稽核的永久化證據。
    try:
        from godpick_durability_service import persist_json_async as _persist_json_async_v184
        durable_ok, durable_msg = _persist_json_async_v184(
            GODPICK_LATEST_FILE, payload, reason="V185 latest recommendation authority"
        )
        github_ok, github_msg = bool(durable_ok), str(durable_msg)
    except Exception:
        github_ok, github_msg = _write_json_to_github_path(GODPICK_LATEST_FILE, payload)

    # H7: do NOT govern ``action_df`` a second time.  Double-governance was the
    # direct cause of Page08 recording R1 rows while Page10 was persisted as [].
    governed_list_df = action_df.copy()
    list_payload = _df_to_records_for_json(governed_list_df)
    if isinstance(list_payload, list):
        fixed_rows = []
        for row in list_payload:
            if not isinstance(row, dict):
                continue
            r = dict(row)
            if not _safe_str(r.get("record_id")):
                r["record_id"] = _create_record_id(
                    _normalize_code(r.get("股票代號")),
                    _safe_str(r.get("推薦日期")) or _now_date_text(),
                    _safe_str(r.get("推薦時間")) or _now_time_text(),
                    _safe_str(r.get("推薦模式")) or "股神推薦",
                )
            r["資料來源"] = GODPICK_LIST_FILE
            r["建立時間"] = _safe_str(r.get("建立時間")) or payload.get("saved_at", _now_text())
            r["更新時間"] = _safe_str(r.get("更新時間")) or payload.get("saved_at", _now_text())
            fixed_rows.append(r)
        list_payload = fixed_rows

    list_local_ok, list_local_msg = _safe_json_write_local(GODPICK_LIST_FILE, list_payload)
    try:
        from godpick_durability_service import persist_json_async as _persist_list_async_v184
        list_github_ok, list_github_msg = _persist_list_async_v184(
            GODPICK_LIST_FILE, list_payload, reason="V185 operational recommendation list"
        )
    except Exception:
        list_github_ok, list_github_msg = _write_json_to_github_path(GODPICK_LIST_FILE, list_payload)

    msgs = [
        local_msg,
        f"V185本輪永久錨點：{'成功' if anchor_ok else '失敗'}｜{anchor_msg}",
        github_msg,
        f"候選診斷保存：{len(candidate_df) if isinstance(candidate_df, pd.DataFrame) else 0} 檔",
        f"本輪快照回讀驗證：{'成功' if verified else '失敗'}｜日期 {saved_at_now[:10]}｜K線 {payload.get('kline_date') or '未驗證'}",
        f"10推薦清單治理後：{len(list_payload) if isinstance(list_payload, list) else 0} 檔",
        *[f"H7行動分區｜{x}" for x in (_action_partition_notes_h7 or [])],
        list_local_msg,
        list_github_msg,
    ]
    st.session_state[_k("latest_recommendation_sync_msgs")] = msgs
    try:
        _project_data_freshness_snapshot_v173.clear()
    except Exception:
        pass
    try:
        _load_latest_recommendation_authority_v185.clear()
    except Exception:
        pass
    # 本輪保存成功的最低標準改為：本機完整快照 + 永久錨點。
    # 遠端大型完整快照可繼續背景同步，但不得把 pending 冒充永久完成。
    return bool(local_ok and anchor_ok), msgs


def _load_latest_recommendation_pack() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    payload, authority_details = _load_latest_recommendation_authority_v185()
    msg = "｜".join(str(x) for x in (authority_details or [])[:4])
    if not isinstance(payload, dict) or not payload:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), msg

    # V191-H3: expose automated-run metadata to the interactive Page07 session.
    # Headless scheduler runs in another namespace; without this bridge a user
    # reopening Page07 could see "請先按開始推薦" even though automation just ran.
    st.session_state[_k("loaded_snapshot_saved_at_v191_h3")] = _safe_str(payload.get("saved_at"))
    st.session_state[_k("loaded_snapshot_execution_owner_v191_h3")] = _safe_str(payload.get("execution_owner"))
    st.session_state[_k("loaded_snapshot_execution_trigger_v191_h3")] = _safe_str(payload.get("execution_trigger"))
    st.session_state[_k("loaded_snapshot_recommendation_count_v191_h3")] = len(payload.get("recommendations") or []) if isinstance(payload.get("recommendations"), list) else 0
    st.session_state[_k("loaded_snapshot_candidate_count_v191_h3")] = len(payload.get("candidate_diagnosis") or []) if isinstance(payload.get("candidate_diagnosis"), list) else int(payload.get("candidate_count") or 0)

    raw_rec_df = _ensure_v92_night_compat_df(
        _records_to_df_for_json(payload.get("recommendations", [])),
        source="latest_recommendations",
    )
    # Re-evaluate old caches locally (no network) so practical stop/RR and the
    # final partition are immediately available after this patch is installed.
    full_rec_df = raw_rec_df.copy()
    if callable(apply_godpick_decision_engine) and not full_rec_df.empty:
        try:
            full_rec_df = apply_godpick_decision_engine(full_rec_df, None)
        except Exception:
            pass
    if callable(canonicalize_final_partition) and not full_rec_df.empty:
        try:
            full_rec_df = canonicalize_final_partition(full_rec_df)
        except Exception:
            pass
    rec_df = _operational_recommendation_rows(full_rec_df, refresh_decision=False)

    cat_df = _records_to_df_for_json(payload.get("category_strength", []))
    hot_df = _ensure_v92_night_compat_df(_records_to_df_for_json(payload.get("hot_pick", [])), source="latest_hot_pick")

    candidate_df = _records_to_df_for_json(payload.get("candidate_diagnosis", []))
    if candidate_df.empty:
        candidate_df = full_rec_df.copy()
    elif callable(apply_godpick_decision_engine):
        try:
            candidate_df = apply_godpick_decision_engine(candidate_df, None)
            if callable(canonicalize_final_partition):
                candidate_df = canonicalize_final_partition(candidate_df)
        except Exception:
            pass
    if isinstance(candidate_df, pd.DataFrame) and not candidate_df.empty:
        try:
            st.session_state[_k("candidate_diagnosis_store")] = (
                build_candidate_diagnosis(candidate_df) if callable(build_candidate_diagnosis) else candidate_df.copy()
            )
        except Exception:
            st.session_state[_k("candidate_diagnosis_store")] = candidate_df.copy()

    scan_report = payload.get("scan_quality", {})
    if not isinstance(scan_report, dict) or not scan_report:
        scan_report = {
            "掃描品質狀態": "舊快取已重新分流｜請更新最新價後再正式執行",
            "掃描品質等級": "legacy_cache",
            "正式推薦可用": False,
            "推薦適用範圍": "舊快取條件式參考",
            "完整候選診斷數": len(candidate_df),
            "正式推薦結果數": len(rec_df),
            "掃描品質說明": "已排除弱勢、正式排除與高風險觀察股；目前清單可作條件式參考，但下單前仍需更新最新價並重新推薦。",
        }
    st.session_state[_k("scan_quality_report")] = scan_report
    return rec_df, cat_df, hot_df, _safe_str(payload.get("saved_at", ""))


def _render_recommend_status_panel(rec_df: pd.DataFrame):
    saved_at = _safe_str(st.session_state.get(_k("result_saved_at"), ""))
    total = 0 if rec_df is None or rec_df.empty else len(rec_df)
    weights = _normalize_weight_map(st.session_state.get(_k("score_weights"), GODPICK_DEFAULT_SCORE_WEIGHTS))

    render_pro_info_card(
        "推薦狀態說明",
        [
            ("目前狀態", "已有本輪推薦結果" if total > 0 else "尚未產生推薦結果", ""),
            ("本輪筆數", total, ""),
            ("保存時間", saved_at or "—", ""),
            ("保存方式", "session_state + JSON 永久記錄", ""),
            ("清除規則", "下一次按開始推薦/重新推薦時覆蓋舊本輪結果", ""),
            ("目前權重", _weight_text(weights), ""),
        ],
        chips=["狀態", "永久記錄", "推薦清單可讀取"],
    )


def _render_weight_dynamic_guide(weights: dict[str, int]):
    weights = _normalize_weight_map(weights)
    top_factors = sorted(weights.items(), key=lambda x: x[1], reverse=True)[:3]
    top_text = "、".join([f"{k}{v}%" for k, v in top_factors])

    render_pro_info_card(
        "推薦條件說明 / 分數解讀",
        [
            ("核心權重", top_text, ""),
            ("分數來源", "推薦總分會依目前已套用權重即時計算；權重改變後，說明與下次推薦分數同步改變。", ""),
            ("85分以上", "高分候選，需同時檢查買點分級與風險說明，不代表無風險追價。", ""),
            ("75~85分", "優先觀察區，適合等待突破確認或回測支撐。", ""),
            ("65~75分", "候補追蹤區，需搭配類股熱度與成交量改善。", ""),
            ("65分以下", "通常不列入主名單，除非起漲補抓名單有特殊結構。", ""),
            ("買點分級", "A+ / A 代表交易條件較完整；B 需等確認；C/D 不建議急追。", ""),
            ("風險解讀", "同時考慮追價風險、停損距離、目標價與交易可行分數。", ""),
        ],
        chips=["動態說明", "依權重更新"],
    )


def _normalize_weight_map(raw: dict[str, Any] | None) -> dict[str, int]:
    base = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()
    if isinstance(raw, dict):
        for k in base.keys():
            try:
                base[k] = int(raw.get(k, base[k]))
            except Exception:
                pass
    # 保護範圍，避免異常值造成分數扭曲
    for k in list(base.keys()):
        base[k] = max(0, min(100, int(base[k])))
    return base


def _weight_total(weights: dict[str, int]) -> int:
    return int(sum(int(v) for v in weights.values()))


def _render_score_weight_panel():
    """股神評分權重控制台：必須總和 100 才能套用，避免誤調造成推薦結果失真。"""
    render_pro_section("股神權重設定", "可調整推薦評分邏輯；只有總和等於 100% 時才能套用。")

    if st.session_state.get(_k("weight_reload_notice")):
        st.success(st.session_state.pop(_k("weight_reload_notice")))

    r1, r2 = st.columns([2, 3])
    with r1:
        reload_from_14 = st.button("重新載入 14_股神權重校正", use_container_width=True, help="14_股神權重校正套用後，按這裡即可讀取最新 godpick_user_settings.json，不必 Ctrl+F5。")
    with r2:
        _loaded_at = _safe_str(st.session_state.get(_k("weight_settings_loaded_at"))) or "尚未載入"
        _profile = _safe_str(st.session_state.get(_k("weight_settings_loaded_profile")))
        st.caption(f"目前權重來源時間：{_loaded_at}" + (f"｜組合：{_profile}" if _profile else ""))

    if reload_from_14:
        payload = _load_persistent_settings(local_first=True)
        ok, msg = _apply_persisted_weights_to_state(payload, force_widget_sync=True)
        st.session_state[_k("weight_reload_notice")] = msg if ok else f"重新載入失敗：{msg}"
        st.rerun()

    if _k("score_weights") not in st.session_state:
        st.session_state[_k("score_weights")] = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()
    if _k("score_weights_edit") not in st.session_state:
        st.session_state[_k("score_weights_edit")] = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()

    # Streamlit 規則：number_input 建立後，不可在同一次 rerun 直接寫入它的 widget key。
    # 所以「恢復原始設定」先寫 pending reset；下一次 rerun、widget 建立前再安全同步。
    if st.session_state.pop(_k("weight_reset_pending"), False):
        st.session_state[_k("score_weights_edit")] = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()
        st.session_state[_k("score_weights")] = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()
        for _name, _val in GODPICK_DEFAULT_SCORE_WEIGHTS.items():
            st.session_state[_k(f"weight_edit_{_name}")] = int(_val)

    edit = _normalize_weight_map(st.session_state.get(_k("score_weights_edit"), GODPICK_DEFAULT_SCORE_WEIGHTS))

    # v25.5：修正權重區塊在 4 欄循環渲染時，第二排與統計/按鈕列視覺交錯，看起來像重複項目。
    # 改成明確兩列，每列 4 個權重欄位，統計與按鈕固定放在最下方。
    weight_keys = list(GODPICK_DEFAULT_SCORE_WEIGHTS.keys())
    weight_rows = [weight_keys[i:i + 4] for i in range(0, len(weight_keys), 4)]

    for row_keys in weight_rows:
        cols = st.columns(4)
        for idx, name in enumerate(row_keys):
            with cols[idx]:
                edit[name] = int(
                    st.number_input(
                        f"{name}%",
                        min_value=0,
                        max_value=100,
                        value=int(edit.get(name, GODPICK_DEFAULT_SCORE_WEIGHTS[name])),
                        step=1,
                        key=_k(f"weight_edit_{name}"),
                    )
                )

    total = _weight_total(edit)
    remain = 100 - total
    st.session_state[_k("score_weights_edit")] = edit

    st.markdown("<div style='height: 0.45rem;'></div>", unsafe_allow_html=True)
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.metric("目前權重總和", f"{total}%")
    with k2:
        st.metric("剩餘權重", f"{remain:+d}%")
    with k3:
        apply_weight = st.button("套用權重", use_container_width=True, type="primary", disabled=(total != 100))
    with k4:
        reset_weight = st.button("恢復原始設定", use_container_width=True)

    if total != 100:
        st.warning("權重總和必須等於 100% 才能套用；目前不能影響推薦結果。")

    if st.session_state.get(_k("weight_reset_msg")):
        st.success(st.session_state.pop(_k("weight_reset_msg")))
        _msgs = st.session_state.pop(_k("weight_reset_msgs"), [])
        if _msgs:
            with st.expander("權重保存明細", expanded=False):
                for msg in _msgs:
                    st.write(f"- {msg}")

    if reset_weight:
        # 不直接寫入 weight_edit_* widget key，避免 StreamlitAPIException。
        # 改為下一輪 rerun 在 number_input 建立前同步。
        st.session_state[_k("weight_reset_pending")] = True
        ok, msgs = _save_persistent_settings(GODPICK_DEFAULT_SCORE_WEIGHTS.copy())
        st.session_state[_k("weight_reset_msg")] = "已恢復原始權重，並永久記錄。" if ok else "已恢復原始權重，但永久記錄失敗。"
        st.session_state[_k("weight_reset_msgs")] = msgs
        st.rerun()

    if apply_weight:
        st.session_state[_k("score_weights")] = edit.copy()
        ok, msgs = _save_persistent_settings(edit.copy())
        if ok:
            st.success("權重已套用並永久記錄。")
        else:
            st.warning("權重已套用，但永久記錄失敗，請查看明細。")
        with st.expander("權重保存明細", expanded=False):
            for msg in msgs:
                st.write(f"- {msg}")

    applied = _normalize_weight_map(st.session_state.get(_k("score_weights"), GODPICK_DEFAULT_SCORE_WEIGHTS))
    st.caption("目前已套用權重：" + _weight_text(applied))
    return applied


def _get_active_weight_map() -> dict[str, int]:
    global GODPICK_ACTIVE_SCORE_WEIGHTS
    return _normalize_weight_map(GODPICK_ACTIVE_SCORE_WEIGHTS)


def _weight_text(weights: dict[str, int] | None = None) -> str:
    weights = _normalize_weight_map(weights or _get_active_weight_map())
    return " / ".join([f"{k}{v}%" for k, v in weights.items()])


def _read_project_json_file(file_name: str) -> dict[str, Any]:
    """v33：安全讀取專案根目錄 JSON；失敗回傳空 dict，不能讓推薦頁空白。"""
    candidates = []
    try:
        candidates.append(Path(__file__).resolve().parent.parent / file_name)
    except Exception:
        pass
    candidates.append(Path.cwd() / file_name)
    candidates.append(Path(file_name))
    seen = set()
    for p in candidates:
        try:
            key = str(p.resolve())
        except Exception:
            key = str(p)
        if key in seen:
            continue
        seen.add(key)
        if not p.exists():
            continue
        try:
            payload = json.loads(p.read_text(encoding="utf-8"))
            return payload if isinstance(payload, dict) else {}
        except Exception:
            continue
    return {}


def _expected_latest_trade_date_v173() -> pd.Timestamp:
    """依台北時間推估目前應有的最新交易日，避免週末/盤前誤報。"""
    now = pd.Timestamp.now(tz="Asia/Taipei")
    today = now.tz_localize(None).normalize()
    after_close = (now.hour, now.minute) >= (14, 30)
    if today.weekday() < 5 and after_close:
        return today
    return (today - pd.tseries.offsets.BDay(1)).normalize()


def _parse_date_v173(value: Any) -> pd.Timestamp | None:
    try:
        # 官方資料日期常以 20260709 的整數保存；直接交給 pandas 會被誤判成 1970 奈秒。
        if isinstance(value, (int, float)) and not pd.isna(value):
            numeric_text = str(int(value))
            if len(numeric_text) == 8:
                value = numeric_text
        ts = pd.to_datetime(value, errors="coerce")
        if pd.isna(ts):
            return None
        if getattr(ts, "tzinfo", None) is not None:
            ts = ts.tz_localize(None)
        return pd.Timestamp(ts).normalize()
    except Exception:
        return None


def _business_lag_v173(newer: pd.Timestamp | None, older: pd.Timestamp | None) -> int:
    if newer is None or older is None or newer <= older:
        return 0
    try:
        return int(len(pd.bdate_range(start=older + pd.Timedelta(days=1), end=newer)))
    except Exception:
        return max(0, int((newer - older).days))


def _max_row_date_v173(rows: list[dict[str, Any]], fields: list[str]) -> pd.Timestamp | None:
    values: list[pd.Timestamp] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        for field in fields:
            ts = _parse_date_v173(row.get(field))
            if ts is not None:
                values.append(ts)
                break
    return max(values) if values else None


@st.cache_data(show_spinner=False, ttl=20)
def _project_data_freshness_snapshot_v173() -> dict[str, Any]:
    """彙整 K 線、大盤、官方因子與推薦保存時間，供頁首警示。"""
    expected = _expected_latest_trade_date_v173()
    latest_payload, _latest_authority_details_v185 = _load_latest_recommendation_authority_v185()
    rows = latest_payload.get("recommendations", []) if isinstance(latest_payload, dict) else []
    if not isinstance(rows, list):
        rows = []
    rows = [x for x in rows if isinstance(x, dict)]
    # 正式清單可能只保留少量可操作股票；完整候選診斷更適合驗證本輪 K 線日期。
    candidate_rows = latest_payload.get("candidate_diagnosis", []) if isinstance(latest_payload, dict) else []
    if not isinstance(candidate_rows, list):
        candidate_rows = []
    candidate_rows = [x for x in candidate_rows if isinstance(x, dict)]
    scan_rows = rows + candidate_rows

    kline_date = _max_row_date_v173(scan_rows, [
        "本輪市場最新交易日", "K線最後交易日", "行情資料日期", "價格資料日期"
    ])
    if kline_date is None and isinstance(latest_payload, dict):
        kline_date = _parse_date_v173(latest_payload.get("kline_date"))
    row_market_date = _max_row_date_v173(scan_rows, ["大盤資料日期", "大盤行情日期", "加權資料日期"])
    row_official_date = _max_row_date_v173(scan_rows, [
        "官方因子資料日期", "官方資料日期", "三大法人資料日期",
        "官方因子更新時間_官方", "官方因子更新時間"
    ])

    market_payload = _read_project_json_file(MARKET_SNAPSHOT_FILE)
    market_date = _parse_date_v173(
        market_payload.get("market_date") or market_payload.get("twse_data_date")
        or market_payload.get("otc_data_date") or market_payload.get("data_date")
        or market_payload.get("updated_at")
    ) if isinstance(market_payload, dict) else None
    market_date = max([x for x in [market_date, row_market_date] if x is not None], default=None)

    # V186：冷啟動先透過 official_factor_service 做「業務日期權威選舉」。
    # 不可直接讀部署包內的舊 JSON，否則 Streamlit Reboot 會把 runtime-data
    # 的新官方因子重新退回 2026-07-11。
    try:
        official_payload = load_factor_cache() if callable(load_factor_cache) else _read_project_json_file(OFFICIAL_FACTORS_CACHE_FILE)
    except Exception:
        official_payload = _read_project_json_file(OFFICIAL_FACTORS_CACHE_FILE)
    official_rows = official_payload.get("records", []) if isinstance(official_payload, dict) else []
    if not isinstance(official_rows, list):
        official_rows = []
    # 新鮮度必須看官方資料內容日期，不可用「快取檔寫入時間」冒充資料日期。
    official_content_date = _max_row_date_v173(official_rows, [
        "官方資料日期", "官方因子資料日期", "三大法人資料日期", "法人資料日期"
    ])
    official_meta = official_payload.get("meta", {}) if isinstance(official_payload, dict) else {}
    if not isinstance(official_meta, dict):
        official_meta = {}
    official_top_date = _parse_date_v173(
        official_payload.get("data_date") or official_meta.get("data_date")
    ) if isinstance(official_payload, dict) else None
    official_date = max(
        [x for x in [official_content_date, official_top_date, row_official_date] if x is not None],
        default=None,
    )
    saved_date = _parse_date_v173(latest_payload.get("saved_at")) if isinstance(latest_payload, dict) else None

    kline_lag = _business_lag_v173(expected, kline_date) if kline_date is not None else 999
    market_ref = kline_date or expected
    market_lag = _business_lag_v173(market_ref, market_date) if market_date is not None else 999
    official_lag = _business_lag_v173(market_ref, official_date) if official_date is not None else 999
    scan_lag = _business_lag_v173(expected, saved_date) if saved_date is not None else 999

    issues: list[str] = []
    warnings: list[str] = []
    infos: list[str] = []
    release_timing: dict[str, Any] = {}
    if callable(evaluate_twse_t86_release_timing):
        try:
            release_timing = evaluate_twse_t86_release_timing(
                market_date=market_ref, official_date=official_date
            )
        except Exception:
            release_timing = {}
    if kline_date is None:
        if saved_date is not None and scan_lag > 0:
            issues.append("舊推薦快照未包含個股K線日期，需在本頁重新推薦")
        else:
            issues.append("個股K線日期未驗證")
    elif kline_lag > 0:
        issues.append(f"個股K線停在{kline_date:%Y-%m-%d}，落後{kline_lag}交易日")
    if market_date is None:
        issues.append("大盤資料日期未驗證")
    elif market_lag > 0:
        issues.append(f"大盤資料停在{market_date:%Y-%m-%d}，落後K線{market_lag}交易日")
    if official_date is None:
        issues.append("官方因子日期未驗證")
    elif official_lag >= 2:
        issues.append(f"官方因子停在{official_date:%Y-%m-%d}，落後K線{official_lag}交易日")
    elif official_lag == 1:
        timing_detail = str(release_timing.get("detail") or "").strip()
        timing_headline = str(release_timing.get("headline") or "").strip()
        timing_msg = f"官方因子為{official_date:%Y-%m-%d}（T-1已驗證）"
        if timing_headline:
            timing_msg += f"｜{timing_headline}"
        if timing_detail:
            timing_msg += f"：{timing_detail}"
        if bool(release_timing.get("t1_is_normal_now")):
            infos.append(timing_msg)
        else:
            warnings.append(timing_msg)
    if saved_date is None:
        issues.append("推薦保存時間未驗證")
    elif scan_lag > 0:
        issues.append(f"最新推薦保存於{saved_date:%Y-%m-%d}，需重新推薦")

    hard_block = bool(
        kline_lag > 0 or market_lag > 0 or official_lag >= 2
        or market_date is None or official_date is None or kline_date is None
    )
    all_messages = issues + warnings
    kline_text = kline_date.strftime("%Y-%m-%d") if kline_date is not None else ""
    if kline_date is not None and kline_date > expected:
        kline_text += "（盤中/當日行情）"
    return {
        "expected_date": expected.strftime("%Y-%m-%d"),
        "kline_date": kline_date.strftime("%Y-%m-%d") if kline_date is not None else "",
        "kline_display": kline_text,
        "market_date": market_date.strftime("%Y-%m-%d") if market_date is not None else "",
        "official_date": official_date.strftime("%Y-%m-%d") if official_date is not None else "",
        "saved_date": saved_date.strftime("%Y-%m-%d") if saved_date is not None else "",
        "kline_lag": kline_lag, "market_lag": market_lag, "official_lag": official_lag, "scan_lag": scan_lag,
        "issues": all_messages,
        "blocking_issues": issues,
        "warnings": warnings,
        "infos": infos,
        "release_timing": release_timing,
        "hard_block": hard_block,
        "ready": not hard_block,
    }


def _render_project_data_freshness_warning_v173() -> dict[str, Any]:
    snapshot = _project_data_freshness_snapshot_v173()
    issues = snapshot.get("issues", []) if isinstance(snapshot, dict) else []
    infos = snapshot.get("infos", []) if isinstance(snapshot, dict) else []
    request_rescan = False
    source_data_ready = bool(
        snapshot.get("market_date") and snapshot.get("official_date")
        and int(snapshot.get("market_lag", 999)) <= 0
        and int(snapshot.get("official_lag", 999)) <= 1
    )
    scan_stale = bool(
        int(snapshot.get("scan_lag", 999)) > 0
        or not snapshot.get("kline_date")
    )
    if issues:
        message = "⚠️ 股神推薦資料時序需注意：" + "；".join(str(x) for x in issues)
        if snapshot.get("hard_block"):
            st.error(message)
            if source_data_ready and scan_stale:
                st.info(
                    "第 17_系統健康檢查已更新前置資料；目前缺的是第 7 頁重新執行選股模型，"
                    "不是再次按第 17 頁。"
                )
            else:
                st.warning(
                    "正式推薦與 A- 準主推薦已暫停升格。請先到第17頁確認『官方因子』不只是抓取成功，"
                    "而是『內容日期驗證通過』；前置資料通過後，本頁可手動重新推薦，中央排程也會直接呼叫本頁推薦流程。"
                )
        else:
            st.warning(message)
        st.caption(
            f"最近完整收盤基準：{snapshot.get('expected_date') or '未驗證'}｜"
            f"K線：{snapshot.get('kline_display') or snapshot.get('kline_date') or '待重新推薦驗證'}｜"
            f"大盤：{snapshot.get('market_date') or '未驗證'}｜"
            f"官方因子：{snapshot.get('official_date') or '未驗證'}｜"
            f"目前推薦保存：{snapshot.get('saved_date') or '未驗證'}"
        )
        if scan_stale:
            request_rescan = st.button(
                "🔄 立即重新推薦（使用第17頁已更新的最新資料）",
                key="page07_freshness_rescan_now_v173_1",
                type="primary",
                use_container_width=True,
                help="會使用目前已保存的推薦條件重新掃描，完成後覆蓋舊推薦快照。",
            )
    elif infos:
        st.info("ℹ️ 官方盤後資料時序正常：" + "；".join(str(x) for x in infos))
        release = snapshot.get("release_timing") if isinstance(snapshot.get("release_timing"), dict) else {}
        if release.get("next_milestone"):
            st.caption(
                f"TWSE逐檔三大法人：18:00產製不含鉅額首版、20:00產製含鉅額完整版｜"
                f"下一時點：{release.get('next_milestone')}｜18:00/20:00為官方檔案產製時點，公開網站/OpenAPI同步可能稍晚。"
            )
        st.caption(
            f"最近完整收盤基準：{snapshot.get('expected_date') or '未驗證'}｜"
            f"K線：{snapshot.get('kline_display') or snapshot.get('kline_date') or '待重新推薦驗證'}｜"
            f"大盤：{snapshot.get('market_date') or '未驗證'}｜"
            f"官方因子：{snapshot.get('official_date') or '未驗證'}｜"
            f"目前推薦保存：{snapshot.get('saved_date') or '未驗證'}"
        )
    else:
        st.success(
            f"資料新鮮度通過：K線、大盤與官方因子符合最近完整收盤／盤中時序基準 {snapshot.get('expected_date')}。"
        )
    snapshot["request_rescan"] = bool(request_rescan)
    return snapshot


def _load_recommendation_readiness_v171() -> dict[str, Any]:
    try:
        path = Path(RECOMMENDATION_READINESS_FILE)
        if not path.exists():
            return {}
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _read_market_snapshot_v33() -> dict[str, Any]:
    """v33：優先讀 01_大盤趨勢 v32 產生的 market_snapshot.json。"""
    snapshot = _read_project_json_file(MARKET_SNAPSHOT_FILE)
    if not isinstance(snapshot, dict):
        return {}
    required = snapshot.get("required_by_godpick")
    if isinstance(required, dict):
        merged = snapshot.copy()
        for k, v in required.items():
            if merged.get(k) in [None, "", []]:
                merged[k] = v
        snapshot = merged
    return snapshot


def _macro_weight_advice_from_snapshot_v33(snapshot: dict[str, Any]) -> str:
    """v33：把 risk_gate / market_score 轉成推薦頁可用的權重建議字串。"""
    gate = _safe_str(snapshot.get("risk_gate"))
    score = _safe_float(snapshot.get("market_score"), 50) or 50
    if gate == "normal":
        return "+5%" if score >= 70 else "+3%"
    if gate == "selective":
        return "0%"
    if gate == "conservative":
        return "-10%"
    if gate == "data_guard":
        return "-15%"
    if score >= 75:
        return "+5%"
    if score >= 60:
        return "+3%"
    if score >= 45:
        return "0%"
    if score >= 30:
        return "-8%"
    return "-15%"


def _macro_bridge_freshness_v121(bridge: dict[str, Any]) -> dict[str, Any]:
    """判斷大盤橋接是否落後；過期資料只能顯示，不得調權或硬封鎖。"""
    if not isinstance(bridge, dict) or not bridge:
        return {"date": "", "lag": 999, "stale": True, "status": "日期未驗證"}
    raw = _safe_str(
        bridge.get("market_date") or bridge.get("twse_data_date") or bridge.get("data_date")
        or bridge.get("updated_at")
    )
    try:
        market_date = pd.to_datetime(raw, errors="coerce")
        if pd.isna(market_date):
            raise ValueError("invalid market date")
        market_date = pd.Timestamp(market_date).tz_localize(None) if getattr(market_date, "tzinfo", None) is not None else pd.Timestamp(market_date)
        market_date = market_date.normalize()
        expected = _expected_latest_trade_date_v173()
        lag = int(len(pd.bdate_range(start=market_date + pd.Timedelta(days=1), end=expected))) if expected > market_date else 0
        stale = lag >= 2
        aligned = lag == 0
        return {
            "date": market_date.strftime("%Y-%m-%d"),
            "lag": lag,
            "stale": stale,
            "aligned": aligned,
            "warning": lag >= 1,
            "status": (
                f"過期｜落後{lag}個交易日，不作硬封鎖" if stale
                else "落後1日｜非最新，正式推薦待同步" if lag == 1
                else "最新/對齊"
            ),
        }
    except Exception:
        return {"date": raw, "lag": 999, "stale": True, "status": "日期未驗證｜不作硬封鎖"}


def _snapshot_to_macro_bridge_v33(snapshot: dict[str, Any]) -> dict[str, Any]:
    """v33：將 market_snapshot.json 正規化成舊版 macro bridge 欄位，保留舊功能不破壞。"""
    if not snapshot:
        return {}
    score = _safe_float(snapshot.get("market_score"), 50) or 50
    trend = _safe_str(snapshot.get("market_trend")) or _macro_reference_grade(score)
    risk_level = _safe_str(snapshot.get("market_risk_level")) or "中"
    gate = _safe_str(snapshot.get("risk_gate")) or "selective"
    position_hint = _safe_str(snapshot.get("position_hint")) or _safe_str(snapshot.get("market_bias"))
    if gate == "normal":
        risk_filter = "正常"
    elif gate == "selective":
        risk_filter = "中性"
    elif gate == "conservative":
        risk_filter = "偏嚴"
    elif gate == "data_guard":
        risk_filter = "嚴格"
    else:
        risk_filter = "中性"
    return {
        "updated_at": _safe_str(snapshot.get("updated_at")),
        "version": _safe_str(snapshot.get("version")) or "market_snapshot_v33_adapter",
        "market_score": score,
        "market_score_raw": snapshot.get("market_score_raw", score),
        "market_state": trend,
        "market_trend": trend,
        "market_risk_level": risk_level,
        "market_bias": _safe_str(snapshot.get("market_bias")),
        "risk_gate": gate,
        "position_hint": position_hint,
        "strategy": position_hint,
        "godpick_weight_advice": _safe_str(snapshot.get("godpick_weight_advice")) or _macro_weight_advice_from_snapshot_v33(snapshot),
        "recommendation_adjustment": snapshot.get("recommendation_adjustment") or snapshot.get("market_bias"),
        "recommendation_bias": {"risk_filter": risk_filter, "position_hint": position_hint, "risk_gate": gate},
        "volume_status": snapshot.get("volume_status"),
        "trend_comment": _safe_str(snapshot.get("trend_comment")),
        "data_quality": _safe_str(snapshot.get("data_quality")),
        "freshness": snapshot.get("freshness"),
        "market_date": _safe_str(snapshot.get("twse_data_date") or snapshot.get("otc_data_date") or snapshot.get("futures_data_date")),
        "twse_change": snapshot.get("twse_change"),
        "twse_change_pct": snapshot.get("twse_change_pct"),
        "otc_change": snapshot.get("otc_change"),
        "otc_change_pct": snapshot.get("otc_change_pct"),
        "futures_change": snapshot.get("futures_change"),
        "futures_change_pct": snapshot.get("futures_change_pct"),
        "mini_futures_change": snapshot.get("mini_futures_change"),
        "mini_futures_change_pct": snapshot.get("mini_futures_change_pct"),
        "data_guard_notes": snapshot.get("data_guard_notes"),
        "market_session": snapshot.get("market_session"),
        "market_session_label": snapshot.get("market_session_label"),
        "market_session_usable": snapshot.get("market_session_usable"),
        "godpick_market_effect": snapshot.get("godpick_market_effect"),
        "overnight_score": snapshot.get("overnight_score"),
        "overnight_risk_level": snapshot.get("overnight_risk_level"),
        "overnight_bias": snapshot.get("overnight_bias"),
        "overnight_comment": snapshot.get("overnight_comment"),
        "night_futures_change": snapshot.get("night_futures_change"),
        "night_futures_change_pct": snapshot.get("night_futures_change_pct"),
        "nasdaq_change_pct": snapshot.get("nasdaq_change_pct"),
        "sp500_change_pct": snapshot.get("sp500_change_pct"),
        "dow_change_pct": snapshot.get("dow_change_pct"),
        "sox_change_pct": snapshot.get("sox_change_pct"),
        "nasdaq_futures_change_pct": snapshot.get("nasdaq_futures_change_pct"),
        "sp500_futures_change_pct": snapshot.get("sp500_futures_change_pct"),
        "us_futures_bias": snapshot.get("us_futures_bias"),
        "fx_risk_level": snapshot.get("fx_risk_level"),
        "overnight_data_quality": snapshot.get("overnight_data_quality"),
        "overnight_updated_at": snapshot.get("overnight_updated_at"),
        "data_diagnostics": snapshot.get("data_diagnostics"),
        "next_day_forecast": snapshot.get("next_day_forecast"),
        "next_day_forecast_version": snapshot.get("next_day_forecast_version"),
        "next_day_forecast_date": snapshot.get("next_day_forecast_date"),
        "next_day_market_direction": snapshot.get("next_day_market_direction"),
        "next_day_market_score": snapshot.get("next_day_market_score"),
        "next_day_confidence": snapshot.get("next_day_confidence"),
        "next_day_confidence_score": snapshot.get("next_day_confidence_score"),
        "next_day_up_probability_pct": snapshot.get("next_day_up_probability_pct"),
        "next_day_flat_probability_pct": snapshot.get("next_day_flat_probability_pct"),
        "next_day_down_probability_pct": snapshot.get("next_day_down_probability_pct"),
        "next_day_expected_return_pct": snapshot.get("next_day_expected_return_pct"),
        "next_day_expected_low": snapshot.get("next_day_expected_low"),
        "next_day_expected_high": snapshot.get("next_day_expected_high"),
        "next_day_data_coverage_pct": snapshot.get("next_day_data_coverage_pct"),
        "next_day_forecast_rationale": snapshot.get("next_day_forecast_rationale"),
        "next_day_godpick_score_delta": snapshot.get("next_day_godpick_score_delta"),
        "next_day_market_weight_delta": snapshot.get("next_day_market_weight_delta"),
        "next_day_position_cap_pct": snapshot.get("next_day_position_cap_pct"),
        "next_day_preferred_style": snapshot.get("next_day_preferred_style"),
        "next_day_avoid_style": snapshot.get("next_day_avoid_style"),
        "next_day_effect_mode": snapshot.get("next_day_effect_mode"),
        "_source": "market_snapshot.json",
    }


def _load_macro_bridge_v33() -> dict[str, Any]:
    """優先 market_snapshot，並標記新鮮度；舊快照不得影響本輪推薦權重。"""
    snapshot_bridge = _snapshot_to_macro_bridge_v33(_read_market_snapshot_v33())
    legacy_bridge = _read_project_json_file(MACRO_MODE_BRIDGE_FILE)
    if not isinstance(legacy_bridge, dict):
        legacy_bridge = {}
    if snapshot_bridge and legacy_bridge:
        result = legacy_bridge.copy()
        result.update({k: v for k, v in snapshot_bridge.items() if v not in [None, "", []]})
    else:
        result = snapshot_bridge or legacy_bridge or {}
    if result:
        fresh = _macro_bridge_freshness_v121(result)
        result["_market_data_date"] = fresh.get("date")
        result["_market_data_business_lag"] = fresh.get("lag")
        result["_market_data_stale"] = fresh.get("stale")
        result["_market_data_freshness_status"] = fresh.get("status")
    return result


def _read_macro_mode_bridge() -> dict[str, Any]:
    """v33：讀取 01_大盤趨勢輸出的 market_snapshot.json / macro_mode_bridge.json。"""
    return _load_macro_bridge_v33()


def _macro_bridge_weight_delta(bridge: dict[str, Any]) -> int:
    if isinstance(bridge, dict) and bridge.get("_market_data_stale"):
        return 0
    raw = _safe_str(bridge.get("godpick_weight_advice"))
    if not raw:
        raw = _macro_weight_advice_from_snapshot_v33(bridge)
    raw = raw.replace("％", "%").replace("＋", "+")
    try:
        return int(float(raw.replace("%", "").replace("+", "").strip()))
    except Exception:
        gate = _safe_str(bridge.get("risk_gate"))
        if gate == "normal":
            return 5
        if gate == "conservative":
            return -10
        if gate == "data_guard":
            return -15
        return 0


def _normalize_int_weight_total(weights: dict[str, int], total: int = 100) -> dict[str, int]:
    keys = list(GODPICK_DEFAULT_SCORE_WEIGHTS.keys())
    out = {k: max(0, int(round(_safe_float(weights.get(k), GODPICK_DEFAULT_SCORE_WEIGHTS.get(k, 0)) or 0))) for k in keys}
    diff = int(total) - sum(out.values())
    # 優先補/扣在市場環境，仍不足再依序調整其他權重。
    order = ["市場環境", "技術結構", "起漲前兆", "類股熱度", "交易可行", "型態突破", "爆發力", "自動因子"]
    guard = 0
    while diff != 0 and guard < 500:
        guard += 1
        changed = False
        for k in order:
            if diff == 0:
                break
            if diff > 0:
                out[k] = out.get(k, 0) + 1
                diff -= 1
                changed = True
            else:
                if out.get(k, 0) > 0:
                    out[k] -= 1
                    diff += 1
                    changed = True
        if not changed:
            break
    return out


def _nextday_forecast_summary_v80(bridge: dict[str, Any]) -> dict[str, Any]:
    """Normalize page-01 next-session forecast and enforce confidence/data guards."""
    if not isinstance(bridge, dict):
        bridge = {}
    nested = bridge.get("next_day_forecast") if isinstance(bridge.get("next_day_forecast"), dict) else {}
    effect = nested.get("godpick_effect") if isinstance(nested.get("godpick_effect"), dict) else {}

    def pick(name: str, fallback: Any = None):
        val = bridge.get(name)
        if val not in [None, "", []]:
            return val
        aliases = {
            "next_day_forecast_date": "forecast_for_date",
            "next_day_market_direction": "direction",
            "next_day_market_score": "direction_score",
            "next_day_confidence": "confidence",
            "next_day_confidence_score": "confidence_score",
            "next_day_up_probability_pct": "up_probability_pct",
            "next_day_flat_probability_pct": "flat_probability_pct",
            "next_day_down_probability_pct": "down_probability_pct",
            "next_day_expected_return_pct": "expected_return_pct",
            "next_day_expected_low": "expected_low",
            "next_day_expected_high": "expected_high",
            "next_day_data_coverage_pct": "data_coverage_pct",
            "next_day_forecast_rationale": "rationale",
        }
        key = aliases.get(name)
        return nested.get(key, fallback) if key else fallback

    confidence = _safe_str(pick("next_day_confidence")) or "低"
    confidence_score = _safe_float(pick("next_day_confidence_score"), 0) or 0
    coverage = _safe_float(pick("next_day_data_coverage_pct"), 0) or 0
    weight_delta = _safe_float(bridge.get("next_day_market_weight_delta"), None)
    if weight_delta is None:
        weight_delta = _safe_float(effect.get("market_weight_delta"), 0) or 0
    score_delta = _safe_float(bridge.get("next_day_godpick_score_delta"), None)
    if score_delta is None:
        score_delta = _safe_float(effect.get("score_delta"), 0) or 0
    # Low confidence/coverage must not influence recommendation weights.
    usable = confidence in {"高", "中"} and confidence_score >= 50 and coverage >= 52
    if not usable:
        weight_delta = 0.0
        score_delta = 0.0
    return {
        "forecast_date": _safe_str(pick("next_day_forecast_date")),
        "direction": _safe_str(pick("next_day_market_direction")) or "資料不足",
        "score": _safe_float(pick("next_day_market_score"), 50) or 50,
        "confidence": confidence,
        "confidence_score": confidence_score,
        "coverage": coverage,
        "up": _safe_float(pick("next_day_up_probability_pct")),
        "flat": _safe_float(pick("next_day_flat_probability_pct")),
        "down": _safe_float(pick("next_day_down_probability_pct")),
        "expected_return": _safe_float(pick("next_day_expected_return_pct")),
        "expected_low": _safe_float(pick("next_day_expected_low")),
        "expected_high": _safe_float(pick("next_day_expected_high")),
        "rationale": _safe_str(pick("next_day_forecast_rationale")),
        "weight_delta": int(round(weight_delta)),
        "score_delta": round(float(score_delta), 1),
        "position_cap": _safe_float(bridge.get("next_day_position_cap_pct"), _safe_float(effect.get("position_cap_pct"))),
        "preferred_style": _safe_str(bridge.get("next_day_preferred_style") or effect.get("preferred_style")),
        "avoid_style": _safe_str(bridge.get("next_day_avoid_style") or effect.get("avoid_style")),
        "effect_mode": _safe_str(bridge.get("next_day_effect_mode") or effect.get("mode")),
        "usable": usable,
        "raw": nested,
    }


def _apply_macro_bridge_to_weights(weights: dict[str, int], bridge: dict[str, Any], enabled: bool = True) -> dict[str, int]:
    """
    v27.3：把大盤橋接檔轉成權重微調。
    +10 / +5：提高市場環境權重，降低追高相關權重。
    -10 / -20：降低市場環境與爆發追價權重，提高交易可行與技術結構防守。
    """
    base = _normalize_weight_map(weights)
    if not enabled or not bridge:
        return base
    if bridge.get("_market_data_stale"):
        return base

    delta = _macro_bridge_weight_delta(bridge)
    out = base.copy()

    if delta > 0:
        add = min(abs(delta), 10)
        out["市場環境"] = out.get("市場環境", 0) + add
        # 大盤偏多時仍避免盲目追高，從爆發力/型態突破小幅挪給大盤。
        take_order = ["爆發力", "型態突破", "自動因子"]
        remain = add
        for k in take_order:
            if remain <= 0:
                break
            take = min(remain, max(0, out.get(k, 0) - 3))
            out[k] = out.get(k, 0) - take
            remain -= take
    elif delta < 0:
        cut = min(abs(delta), 20)
        # 大盤偏弱時，降低大盤鼓勵與爆發追高，轉成防守因子。
        cut_market = min(cut // 2 + cut % 2, max(0, out.get("市場環境", 0) - 3))
        out["市場環境"] = out.get("市場環境", 0) - cut_market
        cut_burst = min(cut - cut_market, max(0, out.get("爆發力", 0) - 2))
        out["爆發力"] = out.get("爆發力", 0) - cut_burst
        out["交易可行"] = out.get("交易可行", 0) + cut_market
        out["技術結構"] = out.get("技術結構", 0) + cut_burst

    # v69：隔夜國際盤只做小幅調權，不壓過個股本身條件。
    overnight = _overnight_effect_summary_v69(bridge)
    od = _safe_float(overnight.get("delta"), 0) or 0
    if od > 0:
        add = min(int(round(abs(od))), 4)
        out["市場環境"] = out.get("市場環境", 0) + add
        take = min(add, max(0, out.get("自動因子", 0) - 3))
        out["自動因子"] = out.get("自動因子", 0) - take
        if add - take > 0:
            out["爆發力"] = max(0, out.get("爆發力", 0) - (add - take))
    elif od < 0:
        cut = min(int(round(abs(od))), 5)
        cut_burst = min(cut, max(0, out.get("爆發力", 0) - 2))
        out["爆發力"] = out.get("爆發力", 0) - cut_burst
        out["交易可行"] = out.get("交易可行", 0) + cut_burst
        if cut - cut_burst > 0:
            cut_pattern = min(cut - cut_burst, max(0, out.get("型態突破", 0) - 3))
            out["型態突破"] = out.get("型態突破", 0) - cut_pattern
            out["技術結構"] = out.get("技術結構", 0) + cut_pattern

    # v80：隔日大盤預測只做小幅、信心感知的權重校正。
    nextday = _nextday_forecast_summary_v80(bridge)
    nd = int(nextday.get("weight_delta", 0) or 0)
    if nd > 0:
        add = min(nd, 4)
        out["市場環境"] = out.get("市場環境", 0) + add
        # 偏多預測增加市場/型態確認，但不提高盲目爆發追價。
        if add >= 2:
            out["型態突破"] = out.get("型態突破", 0) + 1
        remain = add + (1 if add >= 2 else 0)
        for k in ["自動因子", "爆發力"]:
            if remain <= 0:
                break
            take = min(remain, max(0, out.get(k, 0) - 2))
            out[k] = out.get(k, 0) - take
            remain -= take
    elif nd < 0:
        cut = min(abs(nd), 5)
        # 偏空預測不刪逆勢強股；把追價權重轉到交易可行與技術防守。
        for k in ["爆發力", "型態突破"]:
            if cut <= 0:
                break
            take = min(cut, max(0, out.get(k, 0) - 2))
            out[k] = out.get(k, 0) - take
            out["交易可行"] = out.get("交易可行", 0) + (take + 1) // 2
            out["技術結構"] = out.get("技術結構", 0) + take // 2
            cut -= take

    return _normalize_int_weight_total(out, 100)




def _safe_signed_pct_text_v69(v: Any) -> str:
    x = _safe_float(v)
    if x is None:
        return "—"
    return f"{x:+.2f}%"


def _safe_signed_num_text_v69(v: Any) -> str:
    x = _safe_float(v)
    if x is None:
        return "—"
    return f"{x:+.2f}"


def _overnight_effect_summary_v69(bridge: dict[str, Any]) -> dict[str, Any]:
    """v69：把 01 大盤趨勢 v68 的隔夜國際盤欄位轉成推薦頁可用的微調與說明。
    原則：只做小幅加減分與風險提示，不直接刪股票，避免漏掉台股本身剛起漲標的。
    """
    if not isinstance(bridge, dict):
        bridge = {}
    score = _safe_float(bridge.get("overnight_score"), None)
    risk = _safe_str(bridge.get("overnight_risk_level")) or "未標示"
    bias = _safe_str(bridge.get("overnight_bias")) or "未標示"
    comment = _safe_str(bridge.get("overnight_comment"))
    nasdaq = _safe_float(bridge.get("nasdaq_change_pct"), None)
    sox = _safe_float(bridge.get("sox_change_pct"), None)
    night_tx = _safe_float(bridge.get("night_futures_change"), None)
    night_tx_pct = _safe_float(bridge.get("night_futures_change_pct"), None)
    if night_tx is None:
        night_tx = night_tx_pct
    night_src = _safe_str(bridge.get("tw_night_future_source") or bridge.get("night_futures_source") or bridge.get("overnight_source_mode"))
    night_note = _safe_str(bridge.get("tw_night_future_note") or bridge.get("night_futures_note"))
    us_bias = _safe_str(bridge.get("us_futures_bias"))
    fx_risk = _safe_str(bridge.get("fx_risk_level"))

    delta = 0.0
    if score is not None:
        if score >= 75:
            delta += 3.0
        elif score >= 60:
            delta += 1.5
        elif score <= 25:
            delta -= 4.0
        elif score <= 40:
            delta -= 2.0
    if risk in ["高", "偏高", "high", "High"]:
        delta -= 2.0
    elif risk in ["低", "偏低", "low", "Low"]:
        delta += 1.0
    if nasdaq is not None and sox is not None:
        if nasdaq >= 0.6 and sox >= 0.8:
            delta += 1.5
        elif nasdaq <= -0.8 and sox <= -1.0:
            delta -= 2.0
    if night_tx is not None:
        if night_tx >= 80:
            delta += 1.0
        elif night_tx <= -80:
            delta -= 1.5
    delta = round(_score_clip(delta, -6, 5), 1)

    if not comment:
        parts = []
        if score is not None:
            parts.append(f"隔夜分數 {score:.1f}")
        if bias and bias != "未標示":
            parts.append(f"偏向 {bias}")
        if risk and risk != "未標示":
            parts.append(f"風險 {risk}")
        if nasdaq is not None:
            parts.append(f"NASDAQ {nasdaq:+.2f}%")
        if sox is not None:
            parts.append(f"費半 {sox:+.2f}%")
        if night_tx is not None:
            parts.append(f"台指夜盤 {night_tx:+.2f}")
        if us_bias:
            parts.append(f"美期 {us_bias}")
        if fx_risk:
            parts.append(f"匯率風險 {fx_risk}")
        comment = "｜".join(parts) if parts else "隔夜資料不足，僅標示不扣死。"

    return {
        "score": score,
        "risk": risk,
        "bias": bias,
        "comment": comment,
        "delta": delta,
        "nasdaq": nasdaq,
        "sp500": _safe_float(bridge.get("sp500_change_pct"), None),
        "dow": _safe_float(bridge.get("dow_change_pct"), None),
        "sox": sox,
        "night_tx": night_tx,
        "night_tx_pct": night_tx_pct if "night_tx_pct" in locals() else None,
        "night_source": night_src if "night_src" in locals() else "",
        "night_note": night_note if "night_note" in locals() else "",
        "nasdaq_futures": _safe_float(bridge.get("nasdaq_futures_change_pct"), None),
        "sp500_futures": _safe_float(bridge.get("sp500_futures_change_pct"), None),
        "us_bias": us_bias,
        "fx_risk": fx_risk,
        "quality": _safe_str(bridge.get("overnight_data_quality")),
        "updated_at": _safe_str(bridge.get("overnight_updated_at")),
    }

def _market_effect_summary_v37(effect: Any) -> dict[str, Any]:
    """v37: normalize godpick_market_effect from 01 market trend v36."""
    if isinstance(effect, dict):
        delta = _safe_float(effect.get("score_delta") or effect.get("recommend_score_delta") or effect.get("推薦加減分"), 0) or 0
        chase = _safe_str(effect.get("chase_adjustment") or effect.get("追高因子調整"))
        defense = _safe_str(effect.get("defense_adjustment") or effect.get("防守因子調整"))
        style = _safe_str(effect.get("style_bias") or effect.get("市場風格偏向"))
        note = _safe_str(effect.get("effect_note") or effect.get("summary") or effect.get("說明"))
        parts = [p for p in [note, chase, defense, style] if p]
        return {"score_delta": delta, "summary": "｜".join(parts) if parts else "—", "raw": effect}
    txt = _safe_str(effect)
    return {"score_delta": 0, "summary": txt or "—", "raw": effect}


def _market_diagnostics_summary_v37(diag: Any) -> str:
    """v37: compress data_diagnostics into a readable summary."""
    if isinstance(diag, list):
        chunks = []
        for item in diag[:8]:
            if isinstance(item, dict):
                name = _safe_str(item.get("項目") or item.get("name") or item.get("label"))
                ok = item.get("成功") if "成功" in item else item.get("ok")
                status = _safe_str(item.get("狀態") or item.get("status"))
                date = _safe_str(item.get("資料日期") or item.get("data_date"))
                err = _safe_str(item.get("失敗原因") or item.get("error") or item.get("診斷說明"))
                ok_text = "成功" if ok is True else "失敗" if ok is False else (status or "未標示")
                chunks.append(f"{name or '資料源'}:{ok_text}{('/' + date) if date else ''}{('/' + err) if err and ok_text == '失敗' else ''}")
            else:
                chunks.append(_safe_str(item))
        return "｜".join([c for c in chunks if c]) or "—"
    if isinstance(diag, dict):
        chunks = []
        for k, v in list(diag.items())[:8]:
            if isinstance(v, dict):
                ok = v.get("ok")
                date = _safe_str(v.get("data_date") or v.get("資料日期"))
                chunks.append(f"{k}:{'成功' if ok is True else '失敗' if ok is False else '未標示'}{('/' + date) if date else ''}")
            else:
                chunks.append(f"{k}:{_safe_str(v)}")
        return "｜".join(chunks) or "—"
    return _safe_str(diag) or "—"


def _render_macro_bridge_panel(applied_weights: dict[str, int]) -> tuple[dict[str, Any], dict[str, int], bool]:
    """v33：顯示大盤橋接狀態；優先讀 market_snapshot.json，失敗才讀 macro_mode_bridge.json。"""
    bridge = _read_macro_mode_bridge()
    render_pro_section("大盤橋接風控", "讀取 01_大盤趨勢 輸出的 market_snapshot.json / macro_mode_bridge.json，將大盤風控帶入股神推薦。")

    if not bridge:
        st.info("尚未找到 market_snapshot.json 或 macro_mode_bridge.json。請先到 01_大盤趨勢 按『立即寫入股神橋接 / market_snapshot』。")
        return bridge, applied_weights, False

    enabled_key = _k("macro_bridge_enabled")
    if enabled_key not in st.session_state:
        st.session_state[enabled_key] = True

    score = _safe_float(bridge.get("market_score"), 50) or 50
    trend = _safe_str(bridge.get("market_trend") or bridge.get("market_state")) or _macro_reference_grade(score)
    risk_level = _safe_str(bridge.get("market_risk_level")) or _macro_bridge_risk_text(bridge)
    risk_gate = _safe_str(bridge.get("risk_gate")) or "selective"
    weight_advice = _safe_str(bridge.get("godpick_weight_advice")) or _macro_weight_advice_from_snapshot_v33(bridge)
    position_hint = _safe_str(bridge.get("position_hint") or bridge.get("strategy") or bridge.get("market_bias"))
    data_quality = _safe_str(bridge.get("data_quality")) or "未標示"
    market_session = _safe_str(bridge.get("market_session"))
    market_session_label = _safe_str(bridge.get("market_session_label")) or market_session or "未標示"
    market_session_usable = bridge.get("market_session_usable")
    effect_info = _market_effect_summary_v37(bridge.get("godpick_market_effect"))
    diagnostics_summary = _market_diagnostics_summary_v37(bridge.get("data_diagnostics"))
    market_freshness = _macro_bridge_freshness_v121(bridge)
    if market_freshness.get("warning") or market_freshness.get("stale"):
        st.warning(
            f"大盤橋接資料日期 {market_freshness.get('date') or '未驗證'}，"
            f"已落後 {market_freshness.get('lag')} 個交易日。"
            + (
                "系統本輪不套用舊大盤調權/紅燈硬封鎖，改以最新個股K線、主流資金與族群廣度排序。"
                if market_freshness.get("stale") else
                "資料不是最新；仍可保留雷達，但正式推薦與 A- 必須先同步大盤後重新評分。"
            )
        )

    c1, c2, c3, c4, c5, c6 = st.columns([1.0, 1.0, 1.0, 1.0, 1.15, 1.25])
    with c1:
        st.metric("大盤分數", f"{score:.1f}")
    with c2:
        st.metric("趨勢", trend)
    with c3:
        st.metric("風險", risk_level)
    with c4:
        st.metric("風控閘門", risk_gate)
    with c5:
        st.metric("建議加權", weight_advice)
    with c6:
        enabled = st.toggle("套用大盤橋接", value=bool(st.session_state.get(enabled_key, True)), key=enabled_key)

    c7, c8, c9 = st.columns([1.1, 1.1, 2.2])
    with c7:
        st.metric("交易時段", market_session_label)
    with c8:
        st.metric("時段可用", "可用" if market_session_usable is True else "不建議" if market_session_usable is False else "未標示")
    with c9:
        st.metric("大盤影響", f"{effect_info.get('score_delta', 0):+.1f} 分")

    overnight_info = _overnight_effect_summary_v69(bridge)
    c10, c11, c12, c13 = st.columns([1.0, 1.0, 1.0, 2.0])
    with c10:
        st.metric("隔夜分數", "—" if overnight_info.get("score") is None else f"{overnight_info.get('score'):.1f}")
    with c11:
        st.metric("隔夜風險", overnight_info.get("risk") or "未標示")
    with c12:
        st.metric("隔夜加減", f"{overnight_info.get('delta', 0):+.1f} 分")
    with c13:
        st.metric("費半 / Nasdaq", f"{_safe_signed_pct_text_v69(overnight_info.get('sox'))} / {_safe_signed_pct_text_v69(overnight_info.get('nasdaq'))}")

    nextday_info = _nextday_forecast_summary_v80(bridge)
    n1, n2, n3, n4, n5 = st.columns([1.15, 1.0, 1.0, 1.15, 1.3])
    with n1:
        st.metric("隔日大盤預測", nextday_info.get("direction") or "資料不足")
    with n2:
        st.metric("上漲機率", "—" if nextday_info.get("up") is None else f"{nextday_info.get('up'):.1f}%")
    with n3:
        st.metric("下跌機率", "—" if nextday_info.get("down") is None else f"{nextday_info.get('down'):.1f}%")
    with n4:
        st.metric("預測信心", f"{nextday_info.get('confidence')}｜{nextday_info.get('confidence_score', 0):.1f}")
    with n5:
        st.metric("推薦權重校正", f"{nextday_info.get('weight_delta', 0):+d}｜{nextday_info.get('effect_mode') or '中性'}")
    if not nextday_info.get("usable"):
        st.caption("隔日預測目前信心或資料覆蓋不足，系統已自動採資料保護：顯示預測但不影響推薦權重。")

    adjusted = _apply_macro_bridge_to_weights(applied_weights, bridge, enabled=enabled)
    if enabled:
        st.caption("已套用大盤橋接後權重：" + _weight_text(adjusted))
    else:
        st.caption("目前未套用大盤橋接，維持原始權重：" + _weight_text(applied_weights))

    with st.expander("大盤橋接明細 / 對推薦影響", expanded=False):
        st.write(f"**資料來源：** {_safe_str(bridge.get('_source')) or 'macro_mode_bridge.json'}")
        st.write(f"**資料品質：** {data_quality}")
        st.write(f"**交易時段：** {market_session_label}｜可用：{'是' if market_session_usable is True else '否' if market_session_usable is False else '未標示'}")
        st.write(f"**大盤影響：** {effect_info.get('score_delta', 0):+.1f} 分｜{effect_info.get('summary')}")
        st.write(f"**資料診斷：** {diagnostics_summary}")
        st.write(f"**隔夜風控：** {overnight_info.get('risk') or '未標示'}｜{overnight_info.get('bias') or '未標示'}｜加減 {overnight_info.get('delta', 0):+.1f} 分")
        st.write(f"**隔夜解讀：** {overnight_info.get('comment') or '—'}")
        st.write(f"**美盤重點：** Nasdaq {_safe_signed_pct_text_v69(overnight_info.get('nasdaq'))}｜S&P500 {_safe_signed_pct_text_v69(overnight_info.get('sp500'))}｜費半 {_safe_signed_pct_text_v69(overnight_info.get('sox'))}｜台指夜盤 {_safe_signed_num_text_v69(overnight_info.get('night_tx'))}")
        st.write(f"**隔日大盤預測：** {nextday_info.get('direction')}｜上漲 {nextday_info.get('up') if nextday_info.get('up') is not None else '—'}%｜震盪 {nextday_info.get('flat') if nextday_info.get('flat') is not None else '—'}%｜下跌 {nextday_info.get('down') if nextday_info.get('down') is not None else '—'}%")
        st.write(f"**隔日推薦校正：** 權重 {nextday_info.get('weight_delta', 0):+d}｜總部位上限 {nextday_info.get('position_cap') if nextday_info.get('position_cap') is not None else '—'}%｜偏好 {nextday_info.get('preferred_style') or '—'}")
        st.write(f"**隔日預測理由：** {nextday_info.get('rationale') or '—'}")
        st.write(f"**資料日期：** {_safe_str(bridge.get('market_date')) or '—'}")
        st.write(f"**部位建議：** {position_hint or '—'}")
        st.write(f"**大盤解讀：** {_safe_str(bridge.get('trend_comment')) or _safe_str(bridge.get('market_bias')) or '—'}")
        st.json(bridge)

    return bridge, adjusted, enabled



def _macro_bridge_risk_text(bridge: dict[str, Any]) -> str:
    bias = bridge.get("recommendation_bias")
    if isinstance(bias, dict):
        return _safe_str(bias.get("risk_filter")) or "中性"
    return "中性"


def _apply_macro_bridge_columns(df: pd.DataFrame, bridge: dict[str, Any], enabled: bool = True) -> pd.DataFrame:
    """
    v27.4：把大盤橋接狀態寫進推薦結果表，讓完整推薦表、Excel、推薦紀錄都看得到。
    只增加欄位與備註，不重新篩選、不刪股票，避免漏選。
    """
    if df is None or df.empty:
        return df
    x = df.copy()
    if not enabled or not bridge:
        x["大盤橋接分數"] = ""
        x["大盤橋接狀態"] = "未套用"
        x["大盤橋接加權"] = "0%"
        x["大盤橋接風控"] = "未套用"
        x["大盤橋接策略"] = ""
        x["大盤橋接更新時間"] = ""
        x["大盤資料日期"] = ""
        x["大盤資料落後交易日"] = ""
        x["大盤資料新鮮度"] = "未套用"
        x["大盤原始橋接狀態"] = ""
        x["大盤交易時段"] = ""
        x["大盤交易時段可用"] = ""
        x["大盤資料品質"] = ""
        x["大盤影響加減分"] = ""
        x["大盤影響說明"] = ""
        x["大盤資料診斷摘要"] = ""
        x["隔夜風控分數"] = ""
        x["隔夜風險等級"] = ""
        x["隔夜偏向"] = ""
        x["隔夜解讀"] = ""
        x["台指夜盤漲跌"] = ""
        x["NASDAQ漲跌%"] = ""
        x["S&P500漲跌%"] = ""
        x["道瓊漲跌%"] = ""
        x["費半漲跌%"] = ""
        x["Nasdaq期貨偏向"] = ""
        x["S&P期貨偏向"] = ""
        x["匯率風險等級"] = ""
        x["隔日大盤預測日期"] = ""
        x["隔日大盤方向"] = ""
        x["隔日大盤分數"] = ""
        x["隔日大盤信心"] = ""
        x["隔日上漲機率%"] = ""
        x["隔日震盪機率%"] = ""
        x["隔日下跌機率%"] = ""
        x["隔日預估漲跌%"] = ""
        x["隔日大盤預測加減分"] = ""
        x["隔日大盤權重校正"] = ""
        x["隔日建議總部位上限%"] = ""
        x["隔日偏好選股風格"] = ""
        x["隔日應避免風格"] = ""
        x["隔日大盤預測理由"] = ""
        return x

    score = _safe_float(bridge.get("market_score"), 50)
    state = _safe_str(bridge.get("market_trend") or bridge.get("market_state"))
    raw_state = state
    market_freshness = _macro_bridge_freshness_v121(bridge)
    weight = _safe_str(bridge.get("godpick_weight_advice")) or _macro_weight_advice_from_snapshot_v33(bridge)
    risk = _macro_bridge_risk_text(bridge)
    strategy = _safe_str(bridge.get("position_hint") or bridge.get("strategy") or bridge.get("market_bias"))
    updated_at = _safe_str(bridge.get("updated_at"))
    market_session_label = _safe_str(bridge.get("market_session_label")) or _safe_str(bridge.get("market_session"))
    market_session_usable = bridge.get("market_session_usable")
    data_quality = _safe_str(bridge.get("data_quality"))
    effect_info = _market_effect_summary_v37(bridge.get("godpick_market_effect"))
    diagnostics_summary = _market_diagnostics_summary_v37(bridge.get("data_diagnostics"))
    overnight_info = _overnight_effect_summary_v69(bridge)
    nextday_info = _nextday_forecast_summary_v80(bridge)
    if market_freshness.get("stale"):
        score = 50.0
        state = "資料過期｜不作紅燈硬封鎖"
        risk = "資料保護"
        strategy = "舊大盤只顯示；本輪依最新個股K線、主流資金與族群廣度條件式判斷"
        effect_info = {"score_delta": 0.0, "summary": "大盤資料過期，本輪不調整個股分數"}
        nextday_info = dict(nextday_info)
        nextday_info["score_delta"] = 0
        nextday_info["weight_delta"] = 0
        nextday_info["usable"] = False

    x["大盤橋接分數"] = score
    x["大盤橋接狀態"] = state
    x["大盤橋接加權"] = weight
    x["大盤橋接風控"] = risk
    x["大盤橋接策略"] = strategy
    x["大盤橋接更新時間"] = updated_at
    x["大盤資料日期"] = market_freshness.get("date") or _safe_str(bridge.get("market_date"))
    x["大盤資料落後交易日"] = market_freshness.get("lag")
    x["大盤資料新鮮度"] = market_freshness.get("status")
    x["大盤原始橋接狀態"] = raw_state
    x["大盤交易時段"] = market_session_label
    x["大盤交易時段可用"] = "是" if market_session_usable is True else "否" if market_session_usable is False else "未標示"
    x["大盤資料品質"] = data_quality
    x["大盤影響加減分"] = effect_info.get("score_delta", 0)
    x["大盤影響說明"] = effect_info.get("summary")
    x["大盤資料診斷摘要"] = diagnostics_summary
    x["隔夜風控分數"] = overnight_info.get("score")
    x["隔夜風險等級"] = overnight_info.get("risk")
    x["隔夜偏向"] = overnight_info.get("bias")
    x["隔夜解讀"] = overnight_info.get("comment")
    x["隔夜資料品質"] = overnight_info.get("quality")
    x["台指夜盤資料來源"] = overnight_info.get("night_source")
    x["台指夜盤備援說明"] = overnight_info.get("night_note")
    x["台指夜盤漲跌"] = overnight_info.get("night_tx")
    x["NASDAQ漲跌%"] = overnight_info.get("nasdaq")
    x["S&P500漲跌%"] = overnight_info.get("sp500")
    x["道瓊漲跌%"] = overnight_info.get("dow")
    x["費半漲跌%"] = overnight_info.get("sox")
    x["Nasdaq期貨偏向"] = overnight_info.get("us_bias") or _safe_signed_pct_text_v69(overnight_info.get("nasdaq_futures"))
    x["S&P期貨偏向"] = overnight_info.get("us_bias") or _safe_signed_pct_text_v69(overnight_info.get("sp500_futures"))
    x["匯率風險等級"] = overnight_info.get("fx_risk")
    x["隔日大盤預測日期"] = nextday_info.get("forecast_date")
    x["隔日大盤方向"] = nextday_info.get("direction")
    x["隔日大盤分數"] = nextday_info.get("score")
    x["隔日大盤信心"] = f"{nextday_info.get('confidence')}｜{nextday_info.get('confidence_score', 0):.1f}"
    x["隔日上漲機率%"] = nextday_info.get("up")
    x["隔日震盪機率%"] = nextday_info.get("flat")
    x["隔日下跌機率%"] = nextday_info.get("down")
    x["隔日預估漲跌%"] = nextday_info.get("expected_return")
    x["隔日大盤預測加減分"] = nextday_info.get("score_delta")
    x["隔日大盤權重校正"] = nextday_info.get("weight_delta")
    x["隔日建議總部位上限%"] = nextday_info.get("position_cap")
    x["隔日偏好選股風格"] = nextday_info.get("preferred_style")
    x["隔日應避免風格"] = nextday_info.get("avoid_style")
    x["隔日大盤預測理由"] = nextday_info.get("rationale")

    # 同步到原本大盤欄位，讓舊頁面/紀錄頁也可讀。
    if "大盤可參考分數" in x.columns:
        x["大盤可參考分數"] = score
    if "大盤參考等級" in x.columns:
        x["大盤參考等級"] = state
    if "大盤推薦權重" in x.columns:
        x["大盤推薦權重"] = weight
    if "大盤操作風格" in x.columns:
        x["大盤操作風格"] = strategy

    # 大盤偏弱時，不剔除股票，只提醒風控與部位。
    if risk in {"偏嚴", "嚴格"}:
        if "股神進場建議" in x.columns:
            x["股神進場建議"] = x["股神進場建議"].astype(str).map(
                lambda s: (s if s and s != "nan" else "觀察") + "｜大盤風控偏嚴，建議縮小部位"
            )
        if "風險說明" in x.columns:
            x["風險說明"] = x["風險說明"].astype(str).map(
                lambda s: ("" if s in {"nan", "None"} else s) + "｜大盤橋接：風控偏嚴，避免追高。"
            )
    elif risk in {"放寬", "正常"}:
        if "推薦理由摘要" in x.columns:
            x["推薦理由摘要"] = x["推薦理由摘要"].astype(str).map(
                lambda s: ("" if s in {"nan", "None"} else s) + f"｜大盤橋接：{state}，{strategy}。"
            )

    return x


def _apply_official_factor_cache_v109(df: pd.DataFrame | None) -> pd.DataFrame:
    """V109：只讀 official_factors_cache.json，把官方法人/營收/EPS/PER 因子併入推薦結果。

    安全原則：
    - 不在 07 即時連官方網站，避免拖慢推薦。
    - 快取不存在或欄位缺失時只補提示，不中斷頁面。
    - 官方資料完整度 < 60 只顯示，不強行改分；夜間策略會保守採用。
    """
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    if df.empty:
        return df.copy()
    out = df.copy()
    try:
        if callable(merge_official_factors):
            out = merge_official_factors(out)
            if "官方資料完整度" in out.columns:
                comp = pd.to_numeric(out["官方資料完整度"], errors="coerce").fillna(0)
                usable = int((comp >= 60).sum())
                total = int(len(out))
                tag = f"官方因子快取已合併：{usable}/{total} 筆完整度≥60"
            else:
                tag = "官方因子快取已合併，但未取得完整度欄位"
        else:
            tag = "官方因子服務未載入，07 僅使用既有技術量價因子"

        if "資料完整度" not in out.columns:
            out["資料完整度"] = ""
        base = out["資料完整度"].astype(str)
        mask = ~base.str.contains("官方因子快取", na=False)
        out.loc[mask, "資料完整度"] = base[mask].map(lambda x: (x if x and x != "nan" else "夜間資料") + "｜" + tag)
    except Exception as e:
        if "資料完整度" not in out.columns:
            out["資料完整度"] = ""
        try:
            msg = f"官方因子快取合併失敗：{e}"
            out["資料完整度"] = out["資料完整度"].astype(str).map(lambda x: (x if x and x != "nan" else "") + ("｜" if x and x != "nan" else "") + msg)
        except Exception:
            pass
    return out


def _recalc_night_strategy_after_macro_v100(df: pd.DataFrame | None) -> pd.DataFrame:
    """V100：大盤橋接欄位寫入後，重新計算夜間隔日股神欄位。

    V90~V99 的夜間分數會讀取「大盤橋接分數」，但 07 原流程是先算夜間欄位、
    再補大盤橋接欄位，導致夜間隔日分數可能只吃到預設 50 分。
    這裡只在大盤欄位補齊後快速重算夜間層，不改原推薦總分、不刪股票。
    """
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    if df.empty:
        return df.copy()
    out = df.copy()
    try:
        if callable(enrich_night_strategy):
            out = enrich_night_strategy(out)
            if "資料完整度" in out.columns:
                try:
                    macro_state = out.get("大盤橋接狀態", "")
                    macro_score = out.get("大盤橋接分數", "")
                    has_macro = pd.Series(macro_state, index=out.index).astype(str).str.strip().ne("") | pd.Series(macro_score, index=out.index).astype(str).str.strip().ne("")
                except Exception:
                    has_macro = pd.Series([True] * len(out), index=out.index)
                try:
                    base = out["資料完整度"].astype(str)
                    suffix = "｜已套用大盤橋接重算夜間分數"
                    mask = has_macro & ~base.str.contains("大盤橋接重算", na=False)
                    out.loc[mask, "資料完整度"] = base[mask].map(lambda s: (s if s and s != "nan" else "夜間資料") + suffix)
                except Exception:
                    pass
        else:
            for c in list(GODPICK_NIGHT_COLUMNS or []):
                if c not in out.columns:
                    out[c] = ""
            if "資料完整度" not in out.columns:
                out["資料完整度"] = "夜間策略模組未載入"
    except Exception as e:
        for c in list(GODPICK_NIGHT_COLUMNS or []):
            if c not in out.columns:
                out[c] = ""
        if "資料完整度" not in out.columns:
            out["資料完整度"] = ""
        try:
            msg = f"大盤橋接後夜間重算失敗：{e}"
            out["資料完整度"] = out["資料完整度"].astype(str).map(lambda s: (s if s and s != "nan" else "") + ("｜" if s and s != "nan" else "") + msg)
        except Exception:
            pass
    return out


def _derive_buy_point_grade(row: pd.Series) -> str:
    score = _safe_float(row.get("推薦總分"), 0) or 0
    pre = _safe_float(row.get("起漲前兆分數"), 0) or 0
    trade = _safe_float(row.get("交易可行分數"), 0) or 0
    pullback = _safe_float(row.get("拉回買點分數"), 0) or 0
    breakout = _safe_float(row.get("突破買點分數"), 0) or 0
    chase = _safe_float(row.get("追價風險分數"), 50) or 50
    rr = _safe_str(row.get("風險報酬評級"))

    if score >= 88 and pre >= 75 and trade >= 70 and chase <= 65:
        return "A+｜可積極觀察"
    if score >= 80 and trade >= 65 and (pullback >= 65 or breakout >= 65):
        return "A｜優先觀察"
    if score >= 72 and trade >= 55:
        return "B｜等拉回或突破確認"
    if score >= 60:
        return "C｜僅列觀察"
    return "D｜暫不追價"


def _derive_risk_text(row: pd.Series) -> str:
    risk = _safe_float(row.get("風險分數"), 0) or 0
    chase = _safe_float(row.get("追價風險分數"), 0) or 0
    stop_loss = row.get("停損價")
    target1 = row.get("賣出目標1")

    notes = []
    if chase >= 75:
        notes.append("追價風險偏高")
    elif chase >= 60:
        notes.append("追價需控管")
    else:
        notes.append("追價風險可控")

    if risk < 55:
        notes.append("整體風險偏高")
    elif risk < 70:
        notes.append("風險中性")
    else:
        notes.append("風險相對可控")

    if pd.notna(stop_loss):
        notes.append(f"停損 {format_number(stop_loss, 2)}")
    if pd.notna(target1):
        notes.append(f"目標1 {format_number(target1, 2)}")
    return "｜".join(notes)


def _derive_god_reasoning(row: pd.Series) -> str:
    parts = []
    category = _safe_str(row.get("類別"))
    mode = _safe_str(row.get("推薦模式"))
    if category:
        parts.append(f"{category}族群")
    if mode:
        parts.append(mode)
    if _safe_float(row.get("起漲前兆分數"), 0) >= 75:
        parts.append("起漲前兆強")
    if _safe_float(row.get("型態突破分數"), 0) >= 75:
        parts.append(_safe_str(row.get("型態名稱")) or "型態突破")
    if _safe_float(row.get("類股熱度分數"), 0) >= 75:
        parts.append("類股熱度高")
    if _safe_str(row.get("是否領先同類股")) == "是" or row.get("是否領先同類股") is True:
        parts.append("領先同類股")
    if _safe_float(row.get("交易可行分數"), 0) >= 70:
        parts.append("進出場區間清楚")
    if not parts:
        parts.append("條件接近門檻，建議續觀察")
    return "、".join(parts[:7])



def _derive_confidence_level(row: pd.Series) -> str:
    score = _safe_float(row.get("推薦總分"), 0) or 0
    pre = _safe_float(row.get("起漲前兆分數"), 0) or 0
    trade = _safe_float(row.get("交易可行分數"), 0) or 0
    heat = _safe_float(row.get("類股熱度分數"), 0) or 0
    pattern = _safe_float(row.get("型態突破分數"), 0) or 0
    burst = _safe_float(row.get("爆發力分數"), 0) or 0
    hot_risk = _safe_str(row.get("過熱風險"))
    fake_risk = _safe_str(row.get("假突破風險"))

    if score >= 88 and pre >= 75 and trade >= 70 and heat >= 70 and pattern >= 70 and "高" not in hot_risk and "高" not in fake_risk:
        return "S級｜高信心"
    if score >= 80 and pre >= 68 and trade >= 62:
        return "A級｜優先觀察"
    if score >= 72:
        return "B級｜等待確認"
    if score >= 65:
        return "C級｜候補追蹤"
    return "D級｜暫不追價"


def _derive_overheat_risk(row: pd.Series) -> str:
    latest = _safe_float(row.get("最新價"), 0) or 0
    ma20 = _safe_float(row.get("MA20"))
    chase = _safe_float(row.get("追價風險分數"), 0) or 0
    burst = _safe_float(row.get("爆發力分數"), 0) or 0

    dist_ma20 = None
    if latest and ma20 not in [None, 0]:
        dist_ma20 = (latest - ma20) / ma20 * 100

    flags = []
    if dist_ma20 is not None and dist_ma20 >= 18:
        flags.append("離MA20過遠")
    if chase >= 78:
        flags.append("追價風險高")
    if burst >= 88 and chase >= 70:
        flags.append("短線爆發後易震盪")

    if len(flags) >= 2:
        return "高｜" + "、".join(flags)
    if len(flags) == 1:
        return "中｜" + flags[0]
    return "低｜未見明顯過熱"


def _derive_fake_breakout_risk(row: pd.Series) -> str:
    pattern = _safe_float(row.get("型態突破分數"), 0) or 0
    trade = _safe_float(row.get("交易可行分數"), 0) or 0
    volume = _safe_float(row.get("量能啟動分"), 0) or 0
    support = _safe_float(row.get("支撐防守分"), 0) or 0
    breakout = _safe_float(row.get("突破買點分數"), 0) or 0

    flags = []
    if pattern >= 70 and volume < 55:
        flags.append("突破但量能不足")
    if breakout >= 70 and support < 55:
        flags.append("突破後支撐未確認")
    if trade < 55:
        flags.append("交易可行分數偏低")

    if len(flags) >= 2:
        return "高｜" + "、".join(flags)
    if len(flags) == 1:
        return "中｜" + flags[0]
    return "低｜突破結構尚可"



def _derive_prelaunch_grade(row: pd.Series) -> str:
    """起漲等級：依起漲前兆、爆發力、型態突破、量能啟動、交易可行綜合判斷。"""
    pre = _safe_float(row.get("起漲前兆分數"), 0) or 0
    burst = _safe_float(row.get("爆發力分數"), 0) or 0
    pattern = _safe_float(row.get("型態突破分數"), 0) or 0
    vol = _safe_float(row.get("量能啟動分"), 0) or 0
    trade = _safe_float(row.get("交易可行分數"), 0) or 0

    mix = pre * 0.42 + burst * 0.22 + pattern * 0.18 + vol * 0.10 + trade * 0.08

    if mix >= 88:
        return "S｜強烈起漲"
    if mix >= 78:
        return "A｜起漲優先"
    if mix >= 68:
        return "B｜轉強確認"
    if mix >= 55:
        return "C｜初步轉強"
    return "D｜尚未起漲"


def _derive_recommend_bucket(row: pd.Series) -> str:
    score = _safe_float(row.get("推薦總分"), 0) or 0
    pre = _safe_float(row.get("起漲前兆分數"), 0) or 0
    trade = _safe_float(row.get("交易可行分數"), 0) or 0
    heat = _safe_float(row.get("類股熱度分數"), 0) or 0
    pullback = _safe_float(row.get("拉回買點分數"), 0) or 0
    breakout = _safe_float(row.get("突破買點分數"), 0) or 0
    overheat = _safe_str(row.get("過熱風險"))
    fake = _safe_str(row.get("假突破風險"))

    if "高" in overheat:
        return "高分但過熱｜不急追"
    if "高" in fake:
        return "假突破風險｜等確認"
    if score >= 85 and trade >= 70:
        return "立即觀察｜條件完整"
    if pre >= 78 and score >= 75:
        return "剛起漲候選｜優先追蹤"
    if heat >= 75 and score >= 72:
        return "族群領先｜看類股延續"
    if pullback >= breakout and pullback >= 62:
        return "等拉回｜低接觀察"
    if breakout > pullback and breakout >= 62:
        return "等突破｜確認再動"
    return "候補觀察｜等待訊號"


def _derive_trade_script(row: pd.Series) -> str:
    latest = _safe_float(row.get("最新價"))
    pullback = _safe_float(row.get("推薦買點_拉回"))
    breakout = _safe_float(row.get("推薦買點_突破"))
    stop = _safe_float(row.get("停損價"))
    target1 = _safe_float(row.get("賣出目標1"))
    target2 = _safe_float(row.get("賣出目標2"))
    bucket = _safe_str(row.get("推薦分桶"))
    grade = _safe_str(row.get("買點分級"))

    parts = []
    if latest is not None:
        parts.append(f"現價 {format_number(latest, 2)}")
    if pullback is not None:
        parts.append(f"拉回觀察 {format_number(pullback, 2)}")
    if breakout is not None:
        parts.append(f"突破確認 {format_number(breakout, 2)}")
    if stop is not None:
        parts.append(f"失守 {format_number(stop, 2)} 轉弱")
    if target1 is not None:
        parts.append(f"目標1 {format_number(target1, 2)}")
    if target2 is not None:
        parts.append(f"目標2 {format_number(target2, 2)}")

    prefix = bucket or grade or "交易劇本"
    return prefix + "｜" + "｜".join(parts[:7])



def _derive_prelaunch_summary(row: pd.Series) -> str:
    """飆股起漲摘要：把短線爆發因子轉成可讀文字，供 7/8/9 串聯顯示。"""
    score = _safe_float(row.get("飆股起漲分數"), row.get("起漲前兆分數"))
    burst = _safe_float(row.get("爆發力分數"), 0) or 0
    pattern = _safe_float(row.get("型態突破分數"), 0) or 0
    tech = _safe_float(row.get("技術結構分數"), 0) or 0
    trade = _safe_float(row.get("交易可行分數"), 0) or 0
    parts = []

    if score is not None and score >= 90:
        parts.append("接近漲停")
    elif score is not None and score >= 78:
        parts.append("強漲")
    elif score is not None and score >= 68:
        parts.append("明顯上漲")
    elif score is not None and score >= 55:
        parts.append("小漲轉強")

    if burst >= 80:
        parts.append("量能大幅放大")
    elif burst >= 68:
        parts.append("量能轉強")

    if pattern >= 80:
        parts.append("突破20日高")
    elif pattern >= 68:
        parts.append("盤中挑戰20日高")

    if tech >= 70:
        parts.append("站上MA20")
    if trade >= 70:
        parts.append("短均線偏多")

    if not parts:
        return "未見明顯起漲訊號"
    return "、".join(dict.fromkeys(parts))


def _derive_invalid_condition(row: pd.Series) -> str:
    stop = _safe_float(row.get("停損價"))
    support = _safe_float(row.get("推薦買點_拉回"))
    latest = _safe_float(row.get("最新價"))
    fake = _safe_str(row.get("假突破風險"))

    parts = []
    if stop is not None:
        parts.append(f"跌破停損 {format_number(stop, 2)}")
    if support is not None:
        parts.append(f"回測 {format_number(support, 2)} 無法守住")
    if "高" in fake:
        parts.append("突破後量價無法延續")
    if latest is not None:
        parts.append("連續轉弱需降級觀察")
    return "｜".join(parts) if parts else "跌破關鍵支撐或量價轉弱即失效"


def _build_tracking_placeholders(row: pd.Series) -> dict[str, str]:
    code = _normalize_code(row.get("股票代號"))
    rec_date = _safe_str(row.get("推薦日期")) or _now_date_text()
    return {
        "3日追蹤預留": f"{code}｜{rec_date}｜待回填3日最高漲幅/最大回撤/是否觸價",
        "5日追蹤預留": f"{code}｜{rec_date}｜待回填5日最高漲幅/最大回撤/是否觸價",
        "10日追蹤預留": f"{code}｜{rec_date}｜待回填10日最高漲幅/最大回撤/是否觸價",
        "20日追蹤預留": f"{code}｜{rec_date}｜待回填20日最高漲幅/最大回撤/是否觸價",
    }




# =========================================================
# 股神決策引擎 V5：大盤情境調權 / 分層 / 風控 / 劇本
# =========================================================
def _macro_bucket_from_row(row: pd.Series) -> str:
    grade = _safe_str(row.get("大盤參考等級"))
    score = _safe_float(row.get("大盤可參考分數"), 50) or 50
    if grade.startswith("A") or score >= 80:
        return "A｜進攻環境"
    if grade.startswith("B") or score >= 65:
        return "B｜精選偏多"
    if grade.startswith("C") or score >= 50:
        return "C｜震盪控風險"
    return "D｜防守觀望"


def _calc_chase_risk_score(row: pd.Series) -> float:
    price = _safe_float(row.get("最新價"), _safe_float(row.get("推薦價格")))
    pullback = _safe_float(row.get("推薦買點_拉回"), row.get("推薦價格"))
    breakout = _safe_float(row.get("推薦買點_突破"), row.get("推薦價格"))
    pre = _safe_float(row.get("飆股起漲分數"), row.get("起漲前兆分數")) or 0
    overheat = _safe_str(row.get("過熱風險"))
    risk = 35.0
    if price not in [None, 0] and pullback not in [None, 0]:
        dist = (price - pullback) / pullback * 100
        risk += max(0, min(28, dist * 2.3))
    if price not in [None, 0] and breakout not in [None, 0] and price > breakout:
        risk += 8
    if pre >= 90:
        risk += 15
    elif pre >= 78:
        risk += 8
    if "高" in overheat or "過熱" in overheat:
        risk += 15
    return round(_score_clip(risk, 0, 100), 2)


def _calc_trade_risk_reward(row: pd.Series) -> tuple:
    price = _safe_float(row.get("最新價"), _safe_float(row.get("推薦價格")))
    stop = _safe_float(row.get("停損價"))
    target1 = _safe_float(row.get("賣出目標1"))
    if price in [None, 0] or stop in [None, 0] or target1 in [None, 0]:
        return None, None, None
    stop_dist = max(0, (price - stop) / price * 100)
    target_ret = max(0, (target1 - price) / price * 100)
    rr = target_ret / stop_dist if stop_dist > 0 else None
    return (round(rr, 2) if rr is not None else None, round(stop_dist, 2), round(target_ret, 2))


def _derive_position_size(row: pd.Series) -> float:
    score = _safe_float(row.get("推薦總分"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    buy_grade = _safe_str(row.get("買點分級"))
    pos = 0
    if score >= 90:
        pos = 20
    elif score >= 85:
        pos = 15
    elif score >= 78:
        pos = 10
    elif score >= 70:
        pos = 5
    if rr >= 2:
        pos += 5
    elif rr > 0 and rr < 1:
        pos -= 5
    if chase >= 75:
        pos -= 8
    elif chase >= 65:
        pos -= 4
    if macro_bucket.startswith("A"):
        pos += 5
    elif macro_bucket.startswith("C"):
        pos -= 5
    elif macro_bucket.startswith("D"):
        pos -= 10
    if "C" in buy_grade or "D" in buy_grade:
        pos -= 5
    return round(_score_clip(pos, 0, 30), 1)


def _derive_v5_decision_mode(row: pd.Series) -> str:
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    pre = _safe_float(row.get("飆股起漲分數"), row.get("起漲前兆分數")) or 0
    tech = _safe_float(row.get("技術結構分數"), 0) or 0
    heat = _safe_float(row.get("類股熱度分數"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    if macro_bucket.startswith("D"):
        return "逆勢強股防守模式" if pre >= 78 and tech >= 65 else "防守觀望模式"
    if macro_bucket.startswith("C"):
        return "低接確認模式" if chase <= 58 and tech >= 70 else "震盪精選模式"
    if pre >= 78 and heat >= 65:
        return "飆股起漲模式"
    if tech >= 72:
        return "波段順勢模式"
    return "綜合精選模式"


def _derive_entry_advice(row: pd.Series) -> str:
    score = _safe_float(row.get("推薦總分"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    buy_grade = _safe_str(row.get("買點分級"))
    if macro_bucket.startswith("D"):
        return "只允許小部位試單" if score >= 88 and chase < 60 else "不建議進場"
    if chase >= 78:
        return "高分但不急追"
    if score >= 88 and rr >= 1.5 and ("A" in buy_grade or "B" in buy_grade):
        return "可優先觀察進場"
    if score >= 80:
        return "等突破或回測確認"
    if score >= 70:
        return "列入觀察名單"
    return "暫不建議進場"


def _derive_recommend_layer(row: pd.Series) -> str:
    advice = _safe_str(row.get("股神進場建議"))
    mode = _safe_str(row.get("股神決策模式"))
    score = _safe_float(row.get("推薦總分"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    if advice == "可優先觀察進場":
        return "今日可進攻"
    if "逆勢" in mode:
        return "逆勢強股"
    if chase >= 75 and score >= 85:
        return "高分但過熱"
    if macro_bucket.startswith("C") and score >= 80:
        return "等拉回低接"
    if score >= 80:
        return "等突破確認"
    if score >= 70:
        return "觀察不追"
    return "淘汰但接近條件"


def _derive_v5_no_buy_reason(row: pd.Series) -> str:
    reasons = []
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    stop_dist = _safe_float(row.get("停損距離%"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    buy_grade = _safe_str(row.get("買點分級"))
    if macro_bucket.startswith("D"):
        reasons.append("大盤參考等級偏低")
    if chase >= 75:
        reasons.append("追價風險過高")
    if stop_dist >= 8:
        reasons.append("停損距離過大")
    if rr and rr < 1:
        reasons.append("風險報酬比不足")
    if "C" in buy_grade or "D" in buy_grade:
        reasons.append("買點條件尚未完整")
    return "、".join(reasons) if reasons else "未觸發主要否決條件"



# =========================================================
# V76：實戰買點與禁買防呆
# 目的：高分股票不代表立即可買；保留股票但明確標示買點狀態與禁買原因。
# =========================================================
def _v76_calc_support_distance_pct(row: pd.Series) -> float | None:
    price = _safe_float(row.get("最新價"), _safe_float(row.get("推薦價格")))
    support_candidates = [row.get("近端支撐"), row.get("主要支撐"), row.get("推薦買點_拉回"), row.get("停損價")]
    support = None
    for v in support_candidates:
        x = _safe_float(v)
        if x not in [None, 0]:
            support = x
            break
    if price in [None, 0] or support in [None, 0]:
        return None
    return round((price - support) / support * 100, 2)


def _v76_calc_resistance_space_pct(row: pd.Series) -> float | None:
    price = _safe_float(row.get("最新價"), _safe_float(row.get("推薦價格")))
    resistance_candidates = [row.get("近端壓力"), row.get("突破確認價"), row.get("賣出目標1"), row.get("賣出目標2")]
    resistance = None
    for v in resistance_candidates:
        x = _safe_float(v)
        if x not in [None, 0] and (price in [None, 0] or x >= price * 0.98):
            resistance = x
            break
    if price in [None, 0] or resistance in [None, 0]:
        return None
    return round((resistance - price) / price * 100, 2)


def _v76_recent_5d_return(row: pd.Series) -> float | None:
    for c in ["近5日漲幅%", "5日漲跌幅%", "區間漲跌幅%", "RET5"]:
        x = _safe_float(row.get(c))
        if x is not None:
            return round(x, 2)
    return None


def _v76_long_upper_shadow_risk(row: pd.Series) -> str:
    text = " ".join([_safe_str(row.get("假突破風險")), _safe_str(row.get("過熱風險")), _safe_str(row.get("風險說明")), _safe_str(row.get("推薦理由摘要"))])
    if any(k in text for k in ["長上影", "上影", "假突破", "爆量黑K", "爆量長黑"]):
        return "高"
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    ret5 = _v76_recent_5d_return(row)
    if chase >= 80 or (ret5 is not None and ret5 >= 14):
        return "中高"
    if chase >= 68:
        return "中"
    return "低"


def _v76_entry_pattern(row: pd.Series) -> str:
    opportunity = _safe_str(row.get("機會型態"))
    pattern = _safe_str(row.get("型態名稱"))
    pre = _safe_float(row.get("起漲前兆分數"), _safe_float(row.get("飆股起漲分數"), 0)) or 0
    support_dist = _v76_calc_support_distance_pct(row)
    break_score = _safe_float(row.get("型態突破分數"), 0) or 0
    if any(k in opportunity for k in ["低檔", "止跌", "回測", "拉回"]):
        return "低檔起漲 / 支撐回測"
    if support_dist is not None and 0 <= support_dist <= 5.5 and pre >= 60:
        return "回測支撐轉強"
    if break_score >= 72 or any(k in pattern for k in ["突破", "平台", "箱型"]):
        return "突破平台確認"
    if pre >= 78:
        return "轉強確認"
    if _safe_str(row.get("是否領先同類股")) == "是":
        return "族群領先延續"
    return "綜合觀察"


def _v76_no_buy_reasons(row: pd.Series) -> list[str]:
    reasons = []
    score = _safe_float(row.get("推薦總分"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    stop_dist = _safe_float(row.get("停損距離%"), 0) or 0
    support_dist = _v76_calc_support_distance_pct(row)
    resistance_space = _v76_calc_resistance_space_pct(row)
    ret5 = _v76_recent_5d_return(row)
    upper_risk = _v76_long_upper_shadow_risk(row)
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    overnight_risk = _safe_str(row.get("隔夜風險")) or _safe_str(row.get("overnight_risk_level"))
    buy_grade = _safe_str(row.get("買點分級"))
    if score >= 78 and support_dist is not None and support_dist >= 11:
        reasons.append(f"距離支撐過遠{support_dist:.1f}%")
    if score >= 78 and ret5 is not None and ret5 >= 12:
        reasons.append(f"近5日漲幅偏大{ret5:.1f}%")
    if score >= 78 and chase >= 78:
        reasons.append("追價風險過高")
    if rr and rr < 1.25:
        reasons.append(f"R/R不足{rr:.2f}")
    if stop_dist >= 8:
        reasons.append(f"停損距離過大{stop_dist:.1f}%")
    if resistance_space is not None and resistance_space < 3.0:
        reasons.append(f"上方空間不足{resistance_space:.1f}%")
    if upper_risk in ["高", "中高"]:
        reasons.append(f"長上影/假突破風險{upper_risk}")
    if macro_bucket.startswith("D") and score < 88:
        reasons.append("大盤防守環境未達逆勢強股門檻")
    if any(k in overnight_risk for k in ["高", "偏空", "逆風"]):
        reasons.append("隔夜風控偏弱不宜追價")
    if "C" in buy_grade or "D" in buy_grade:
        reasons.append("買點分級偏弱")
    return reasons


def _derive_v76_buy_status(row: pd.Series) -> str:
    reasons = _v76_no_buy_reasons(row)
    score = _safe_float(row.get("推薦總分"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    support_dist = _v76_calc_support_distance_pct(row)
    pre = _safe_float(row.get("起漲前兆分數"), _safe_float(row.get("飆股起漲分數"), 0)) or 0
    break_score = _safe_float(row.get("型態突破分數"), 0) or 0
    if reasons and score >= 78:
        return "高分但暫不追"
    if score >= 88 and chase <= 65 and rr >= 1.5:
        return "立即可小量試單"
    if support_dist is not None and 0 <= support_dist <= 5.5 and score >= 75:
        return "回測可觀察"
    if break_score >= 75 and pre >= 68 and chase <= 72:
        return "突破確認可追蹤"
    if score >= 75:
        return "等拉回 / 等確認"
    return "觀察名單"


def _derive_v76_practical_score(row: pd.Series) -> float:
    score = _safe_float(row.get("推薦總分"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    support_dist = _v76_calc_support_distance_pct(row)
    resistance_space = _v76_calc_resistance_space_pct(row)
    ret5 = _v76_recent_5d_return(row)
    buy_grade = _safe_str(row.get("買點分級"))
    practical = score * 0.55
    practical += max(min((rr - 1.2) * 12, 14), -12) if rr else -4
    practical += max(min((70 - chase) * 0.22, 10), -12)
    if support_dist is not None:
        if 0 <= support_dist <= 5.5:
            practical += 10
        elif support_dist <= 9:
            practical += 3
        elif support_dist >= 13:
            practical -= 10
    if resistance_space is not None:
        if resistance_space >= 8:
            practical += 7
        elif resistance_space < 3:
            practical -= 8
    if ret5 is not None:
        if 0 <= ret5 <= 8:
            practical += 4
        elif ret5 >= 14:
            practical -= 9
    if "A" in buy_grade or "S" in buy_grade:
        practical += 5
    elif "C" in buy_grade or "D" in buy_grade:
        practical -= 6
    return round(_score_clip(practical, 0, 100), 1)


def _derive_v76_action_note(row: pd.Series) -> str:
    status = _safe_str(row.get("買點狀態"))
    reasons = _safe_str(row.get("高分禁買原因"))
    pattern = _safe_str(row.get("進場型態"))
    support = _safe_float(row.get("近端支撐"), _safe_float(row.get("主要支撐")))
    breakout = _safe_float(row.get("突破確認價"), _safe_float(row.get("推薦買點_突破")))
    stop = _safe_float(row.get("停損價"), _safe_float(row.get("停損參考")))
    parts = [status, pattern]
    if reasons:
        parts.append("禁追原因：" + reasons)
    if support:
        parts.append(f"支撐觀察 {support:.2f}")
    if breakout:
        parts.append(f"突破確認 {breakout:.2f}")
    if stop:
        parts.append(f"失效 {stop:.2f}")
    return "｜".join([p for p in parts if p])




# =========================================================
# V144：專業決策層 - 區分「推薦分數」與「買進分數」
# 目的：推薦總分只代表候選強度；買進分數才判斷當下價位是否接近可操作。
# 不刪除原欄位、不硬篩股票，只補上更清楚的實戰判讀。
# =========================================================
def _derive_v144_buy_score(row: pd.Series) -> float:
    base = _safe_float(row.get("實戰買點分數"))
    if base is None:
        base = _derive_v76_practical_score(row)
    rec_score = _safe_float(row.get("推薦總分"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    stop_dist = _safe_float(row.get("停損距離%"), 0) or 0
    reasons = _safe_str(row.get("高分禁買原因"))
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    buy_score = float(base)

    # 推薦分高但買點風險高時，不讓買進分數誤導成「直接買」。
    if rec_score >= 85 and reasons:
        buy_score = min(buy_score, 72.0)
    if chase >= 78:
        buy_score = min(buy_score, 62.0)
    elif chase >= 70:
        buy_score -= 4
    if rr and rr < 1.25:
        buy_score = min(buy_score, 65.0)
    elif rr >= 2.0:
        buy_score += 4
    elif rr >= 1.6:
        buy_score += 2
    if stop_dist >= 8:
        buy_score = min(buy_score, 66.0)
    if macro_bucket.startswith("D") and rec_score < 88:
        buy_score = min(buy_score, 60.0)
    return round(_score_clip(buy_score, 0, 100), 1)


def _derive_v144_recommend_use(row: pd.Series) -> str:
    rec_score = _safe_float(row.get("推薦總分"), 0) or 0
    buy_score = _safe_float(row.get("買進分數"), _safe_float(row.get("實戰買點分數"), 0)) or 0
    no_buy = _safe_str(row.get("高分禁買原因"))
    if rec_score >= 88 and buy_score >= 82 and not no_buy:
        return "隔日作戰候選｜可小量試單"
    if rec_score >= 85 and no_buy:
        return "高分追蹤候選｜等拉回不追高"
    if rec_score >= 80 and buy_score >= 68:
        return "優先追蹤候選｜等突破/回測確認"
    if rec_score >= 70:
        return "觀察名單｜等待訊號補強"
    return "候補觀察｜暫不列主攻"


def _derive_v144_direct_buy_flag(row: pd.Series) -> str:
    buy_score = _safe_float(row.get("買進分數"), 0) or 0
    rec_score = _safe_float(row.get("推薦總分"), 0) or 0
    no_buy = _safe_str(row.get("高分禁買原因"))
    rr = _safe_float(row.get("風險報酬比"), 0) or 0
    if no_buy:
        return "否｜高分但需等拉回/確認"
    if rec_score >= 88 and buy_score >= 82 and rr >= 1.5:
        return "可小量試單｜仍需盤中確認"
    if buy_score >= 72:
        return "待確認｜突破或回測成功再評估"
    return "否｜僅列追蹤"


def _derive_v144_intraday_checks(row: pd.Series) -> str:
    breakout = _safe_float(row.get("突破確認價"), _safe_float(row.get("推薦買點_突破")))
    pullback = _safe_float(row.get("回測承接價"), _safe_float(row.get("推薦買點_拉回")))
    support = _safe_float(row.get("近端支撐"), _safe_float(row.get("主要支撐")))
    stop = _safe_float(row.get("停損價"), _safe_float(row.get("停損參考")))
    parts = []
    parts.append("開盤跳空超過3%不追")
    if breakout:
        parts.append(f"突破{breakout:.2f}需量能延續")
    if pullback:
        parts.append(f"拉回{pullback:.2f}附近守穩才加分")
    elif support:
        parts.append(f"支撐{support:.2f}附近守穩才加分")
    if stop:
        parts.append(f"跌破{stop:.2f}取消作戰")
    parts.append("開高走低或量縮不進場")
    return "｜".join(parts)


def _derive_v144_professional_summary(row: pd.Series) -> str:
    rec_score = _safe_float(row.get("推薦總分"), 0) or 0
    buy_score = _safe_float(row.get("買進分數"), 0) or 0
    use = _safe_str(row.get("推薦用途"))
    direct = _safe_str(row.get("是否可直接買進"))
    rr = _safe_float(row.get("風險報酬比"))
    no_buy = _safe_str(row.get("高分禁買原因"))
    pieces = [f"推薦分數{rec_score:.1f}=候選強度", f"買進分數{buy_score:.1f}=當下買點品質"]
    if rr is not None:
        pieces.append(f"R/R {rr:.2f}")
    if use:
        pieces.append(use)
    if direct:
        pieces.append(direct)
    if no_buy:
        pieces.append("限制：" + no_buy)
    return "｜".join(pieces)


def _apply_v144_professional_decision_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    out["買進分數"] = out.apply(_derive_v144_buy_score, axis=1)
    out["推薦用途"] = out.apply(_derive_v144_recommend_use, axis=1)
    out["是否可直接買進"] = out.apply(_derive_v144_direct_buy_flag, axis=1)
    out["盤中確認條件"] = out.apply(_derive_v144_intraday_checks, axis=1)
    out["專業決策摘要"] = out.apply(_derive_v144_professional_summary, axis=1)
    return out


def _apply_v76_practical_entry_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    out["支撐距離%"] = out.apply(_v76_calc_support_distance_pct, axis=1)
    out["壓力空間%"] = out.apply(_v76_calc_resistance_space_pct, axis=1)
    out["近5日漲幅%"] = out.apply(_v76_recent_5d_return, axis=1)
    out["長上影風險"] = out.apply(_v76_long_upper_shadow_risk, axis=1)
    out["進場型態"] = out.apply(_v76_entry_pattern, axis=1)
    out["高分禁買原因"] = out.apply(lambda r: "、".join(_v76_no_buy_reasons(r)), axis=1)
    out["高分禁買旗標"] = out["高分禁買原因"].fillna("").astype(str).str.strip().map(lambda x: "是" if x else "否")
    out["買點狀態"] = out.apply(_derive_v76_buy_status, axis=1)
    out["實戰買點分數"] = out.apply(_derive_v76_practical_score, axis=1)
    out["實戰操作建議"] = out.apply(_derive_v76_action_note, axis=1)
    out["V76買點防呆版本"] = "v76_practical_entry_guard"
    out = _apply_v144_professional_decision_columns(out)
    return out

def _derive_best_trade_script_v5(row: pd.Series) -> str:
    pullback = _safe_float(row.get("推薦買點_拉回"), row.get("推薦價格"))
    breakout = _safe_float(row.get("推薦買點_突破"), row.get("推薦價格"))
    stop = _safe_float(row.get("停損價"))
    target1 = _safe_float(row.get("賣出目標1"))
    advice = _safe_str(row.get("股神進場建議"))
    parts = [advice]
    if pullback:
        parts.append(f"拉回觀察 {pullback:.2f}")
    if breakout:
        parts.append(f"突破確認 {breakout:.2f}")
    if stop:
        parts.append(f"失效停損 {stop:.2f}")
    if target1:
        parts.append(f"第一目標 {target1:.2f}")
    return "｜".join(parts)


def _derive_next_day_action(row: pd.Series) -> str:
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    pre = _safe_float(row.get("飆股起漲分數"), row.get("起漲前兆分數")) or 0
    if macro_bucket.startswith("D"):
        return "開高不追，僅低量試單或觀望"
    if chase >= 75:
        return "開高不追，等回測支撐"
    if pre >= 80:
        return "若量價續強可追蹤突破確認"
    return "等量價確認後再動作"


def _derive_weak_condition_v5(row: pd.Series) -> str:
    stop = _safe_float(row.get("停損價"))
    parts = []
    if stop:
        parts.append(f"跌破停損 {stop:.2f}")
    parts.append("跌破MA20且量增")
    parts.append("推薦分層轉弱")
    return "、".join(parts)




# =========================================================
# V16 股神風控與資金配置
# =========================================================
def _derive_v16_single_risk_level(row: pd.Series) -> str:
    chase = _safe_float(row.get("追價風險分"), _safe_float(row.get("追高風險分數_決策"), 50)) or 50
    stop_dist = _safe_float(row.get("停損距離%"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), _safe_float(row.get("風險報酬比_決策"), 0)) or 0
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    risk = 35.0
    if chase >= 80:
        risk += 28
    elif chase >= 70:
        risk += 18
    elif chase <= 50:
        risk -= 8
    if stop_dist >= 10:
        risk += 22
    elif stop_dist >= 7:
        risk += 12
    elif 0 < stop_dist <= 4:
        risk -= 6
    if rr and rr < 1:
        risk += 14
    elif rr >= 2:
        risk -= 10
    if macro_bucket.startswith("D"):
        risk += 18
    elif macro_bucket.startswith("C"):
        risk += 8
    risk = _score_clip(risk, 0, 100)
    if risk >= 78:
        return "高風險"
    if risk >= 58:
        return "中風險"
    return "低風險"


def _derive_v16_position(row: pd.Series) -> float:
    base = _safe_float(row.get("建議部位%"), 0) or 0
    score = _safe_float(row.get("推薦總分"), 0) or 0
    rr = _safe_float(row.get("風險報酬比"), _safe_float(row.get("風險報酬比_決策"), 0)) or 0
    risk_level = _derive_v16_single_risk_level(row)
    macro_bucket = _safe_str(row.get("大盤情境分桶"))
    sector_flow = _safe_float(row.get("族群資金流分數"), 50) or 50
    position = base
    if score >= 88 and rr >= 1.5:
        position += 3
    if sector_flow >= 75:
        position += 2
    if risk_level == "高風險":
        position = min(position, 8)
    elif risk_level == "中風險":
        position = min(position, 15)
    else:
        position = min(position, 25)
    if macro_bucket.startswith("D"):
        position = min(position, 5)
    elif macro_bucket.startswith("C"):
        position = min(position, 12)
    return round(_score_clip(position, 0, 25), 1)


def _derive_v16_invest_level(row: pd.Series) -> str:
    pos = _safe_float(row.get("建議倉位%"), _derive_v16_position(row)) or 0
    risk = _safe_str(row.get("單檔風險等級", _derive_v16_single_risk_level(row)))
    if pos >= 20 and risk == "低風險":
        return "高信心配置"
    if pos >= 12:
        return "標準配置"
    if pos > 0:
        return "小部位試單"
    return "暫不投入"


def _derive_v16_first_entry_pct(row: pd.Series) -> float:
    risk = _safe_str(row.get("單檔風險等級", _derive_v16_single_risk_level(row)))
    timing = _safe_str(row.get("進場時機"))
    chase = _safe_float(row.get("追價風險分"), 50) or 50
    if risk == "高風險" or chase >= 75:
        return 30.0
    if "等待" in timing or "觀察" in timing:
        return 40.0
    return 50.0


def _derive_v16_scale_plan(row: pd.Series) -> str:
    pos = _safe_float(row.get("建議倉位%"), _derive_v16_position(row)) or 0
    first = _derive_v16_first_entry_pct(row)
    if pos <= 0:
        return "不進場，等待條件成熟"
    if pos <= 8:
        return f"小部位試單：先投入{first:.0f}%額度，其餘等確認"
    return f"分兩到三筆：第一筆{first:.0f}%額度，確認支撐或突破後再加碼"


def _derive_v16_add_condition(row: pd.Series) -> str:
    breakout = _safe_float(row.get("突破確認價"), _safe_float(row.get("推薦買點_突破")))
    support = _safe_float(row.get("近端支撐"), _safe_float(row.get("停損參考")))
    flow = _safe_float(row.get("族群資金流分數"), 50) or 50
    parts = []
    if breakout:
        parts.append(f"站穩突破確認價 {breakout:.2f}")
    if support:
        parts.append(f"回測支撐 {support:.2f} 不破")
    if flow >= 70:
        parts.append("族群資金流維持偏強")
    else:
        parts.append("量能重新放大且收紅K")
    return "、".join(parts)


def _derive_v16_take_profit(row: pd.Series) -> str:
    t1 = _safe_float(row.get("賣出目標1"), _safe_float(row.get("近端壓力")))
    t2 = _safe_float(row.get("賣出目標2"))
    if t1 and t2:
        return f"目標1 {t1:.2f} 先停利1/3；目標2 {t2:.2f} 再分批出場"
    if t1:
        return f"接近壓力/目標 {t1:.2f} 先減碼，保留獲利部位"
    return "以移動停利為主，跌破短均或量價轉弱先降部位"


def _derive_v16_stop_strategy(row: pd.Series) -> str:
    stop = _safe_float(row.get("停損參考"), _safe_float(row.get("停損價")))
    weak = _safe_str(row.get("轉弱條件"))
    if stop:
        return f"跌破 {stop:.2f} 或{weak if weak else '跌破MA20且量增'}，先停損/減碼"
    return weak if weak else "跌破支撐且量增轉弱，先停損/減碼"


def _derive_v16_max_risk(row: pd.Series) -> float:
    pos = _safe_float(row.get("建議倉位%"), _derive_v16_position(row)) or 0
    stop_dist = _safe_float(row.get("停損距離%"), 0) or 0
    if pos <= 0 or stop_dist <= 0:
        return 0.0
    return round(pos * stop_dist / 100.0, 2)


def _derive_v16_capital_note(row: pd.Series) -> str:
    pos = _safe_float(row.get("建議倉位%"), _derive_v16_position(row)) or 0
    max_risk = _safe_float(row.get("最大風險%"), _derive_v16_max_risk(row)) or 0
    risk = _safe_str(row.get("單檔風險等級", _derive_v16_single_risk_level(row)))
    if pos <= 0:
        return "目前條件不足，不配置資金"
    return f"建議單檔配置{pos:.1f}%，若觸發停損，約影響總資金{max_risk:.2f}%；風險等級：{risk}"


def _derive_v16_sector_warning(row: pd.Series) -> str:
    density = _safe_float(row.get("同族群推薦密度"), 0) or 0
    ratio = _safe_float(row.get("同族群強勢比例"), 0) or 0
    category = _safe_str(row.get("類別")) or "同族群"
    if density >= 35 or ratio >= 70:
        return f"{category} 推薦密度偏高，注意同族群集中風險"
    if density >= 20 or ratio >= 55:
        return f"{category} 有族群聚集，配置不宜過度集中"
    return "族群集中風險正常"


def _derive_v16_portfolio_suggestion(row: pd.Series) -> str:
    risk = _safe_str(row.get("單檔風險等級", _derive_v16_single_risk_level(row)))
    level = _safe_str(row.get("建議投入等級"))
    warning = _safe_str(row.get("族群集中警示"))
    if risk == "高風險":
        return "僅列衛星部位，避免重倉；需嚴格依停損策略執行"
    if "集中" in warning and "偏高" in warning:
        return "同族群持股請分散，擇優配置1~2檔即可"
    if "高信心" in level:
        return "可列核心觀察部位，但仍需分批與停損控管"
    if "標準" in level:
        return "可列標準觀察部位，等待加碼條件成立"
    return "先列追蹤池，等待量價與大盤條件同步"


def _apply_v16_risk_allocation_columns(df: pd.DataFrame) -> pd.DataFrame:
    """V16：補齊風控與資金配置欄位；只做決策輔助，不硬篩股票。"""
    if df is None or df.empty:
        return df
    out = df.copy()
    out["單檔風險等級"] = out.apply(_derive_v16_single_risk_level, axis=1)
    out["建議倉位%"] = out.apply(_derive_v16_position, axis=1)
    out["建議投入等級"] = out.apply(_derive_v16_invest_level, axis=1)
    out["第一筆進場%"] = out.apply(_derive_v16_first_entry_pct, axis=1)
    out["分批策略"] = out.apply(_derive_v16_scale_plan, axis=1)
    out["第二筆加碼條件"] = out.apply(_derive_v16_add_condition, axis=1)
    out["停利策略"] = out.apply(_derive_v16_take_profit, axis=1)
    out["停損策略"] = out.apply(_derive_v16_stop_strategy, axis=1)
    out["最大風險%"] = out.apply(_derive_v16_max_risk, axis=1)
    out["資金風險說明"] = out.apply(_derive_v16_capital_note, axis=1)
    out["族群集中警示"] = out.apply(_derive_v16_sector_warning, axis=1)
    out["組合配置建議"] = out.apply(_derive_v16_portfolio_suggestion, axis=1)
    return out



# =========================
# V17：大盤環境動態策略
# 只做策略加權與說明，不做硬篩選，避免漏掉股票。
# =========================
def _derive_v17_market_score(row: pd.Series) -> float:
    vals = []
    weights = []
    for col, w in [
        ("大盤可參考分數", 0.30),
        ("大盤市場廣度分數", 0.20),
        ("大盤量價確認分數", 0.20),
        ("大盤權值支撐分數", 0.15),
        ("大盤推薦同步分數", 0.15),
    ]:
        v = _safe_float(row.get(col), None)
        if v is not None:
            vals.append(max(0.0, min(100.0, float(v))))
            weights.append(w)
    if vals and sum(weights) > 0:
        return round(sum(v*w for v, w in zip(vals, weights)) / sum(weights), 2)
    mw = _safe_float(row.get("大盤推薦權重"), None)
    if mw is not None:
        return round(max(0.0, min(100.0, 50.0 + float(mw) * 10.0)), 2)
    return 50.0


def _derive_v17_market_mode(row: pd.Series) -> str:
    score = _safe_float(row.get("大盤多空分數"), _derive_v17_market_score(row)) or 50
    risk_txt = _safe_str(row.get("大盤風險濾網"))
    bucket = _safe_str(row.get("大盤情境分桶"))
    if any(k in risk_txt + bucket for k in ["空頭", "高風險", "偏空"]):
        if score < 55:
            return "空頭防守"
    if score >= 75:
        return "多頭進攻"
    if score >= 62:
        return "偏多輪動"
    if score >= 48:
        return "震盪選股"
    if score >= 35:
        return "偏空防守"
    return "空頭防守"


def _derive_v17_aggressiveness(row: pd.Series) -> float:
    mode = _safe_str(row.get("大盤策略模式", _derive_v17_market_mode(row)))
    chase = _safe_str(row.get("追高風險等級"))
    single_risk = _safe_str(row.get("單檔風險等級"))
    opp = _safe_str(row.get("推薦型態")) + _safe_str(row.get("機會型態"))
    base_map = {
        "多頭進攻": 1.18,
        "偏多輪動": 1.08,
        "震盪選股": 0.95,
        "偏空防守": 0.72,
        "空頭防守": 0.52,
    }
    coef = base_map.get(mode, 0.90)
    if any(k in chase for k in ["高", "過熱", "不建議"]):
        coef -= 0.12
    if any(k in single_risk for k in ["高", "極高"]):
        coef -= 0.10
    if any(k in opp for k in ["低檔", "拉回", "回測"]):
        coef += 0.05 if mode in ["震盪選股", "偏多輪動"] else 0.0
    return round(max(0.35, min(1.25, coef)), 2)


def _derive_v17_suitable_types(row: pd.Series) -> str:
    mode = _safe_str(row.get("大盤策略模式", _derive_v17_market_mode(row)))
    if mode == "多頭進攻":
        return "強勢突破、拉回承接、族群領先股"
    if mode == "偏多輪動":
        return "拉回承接、回測支撐、類股輪動剛啟動"
    if mode == "震盪選股":
        return "低檔轉強、回測支撐、量縮整理後轉強"
    if mode == "偏空防守":
        return "保守低風險、低檔止穩、小部位觀察"
    return "現金防守、只追蹤不追價、等待大盤轉強"


def _derive_v17_strategy_note(row: pd.Series) -> str:
    mode = _safe_str(row.get("大盤策略模式", _derive_v17_market_mode(row)))
    score = _safe_float(row.get("大盤多空分數"), 50) or 50
    rec_type = _safe_str(row.get("推薦型態")) or _safe_str(row.get("機會型態"))
    if mode == "多頭進攻":
        return f"大盤分數{score:.1f}，環境偏多，可保留強勢與拉回股；{rec_type}可依風控分批執行。"
    if mode == "偏多輪動":
        return f"大盤分數{score:.1f}，資金輪動機率高，優先看族群資金流與拉回承接。"
    if mode == "震盪選股":
        return f"大盤分數{score:.1f}，不宜全面追價，優先低檔轉強與回測支撐。"
    if mode == "偏空防守":
        return f"大盤分數{score:.1f}，降低倉位，僅保留支撐明確且風險報酬比佳的標的。"
    return f"大盤分數{score:.1f}，防守優先，等待量價與大盤同步轉強。"


def _derive_v17_risk_note(row: pd.Series) -> str:
    mode = _safe_str(row.get("大盤策略模式", _derive_v17_market_mode(row)))
    chase = _safe_str(row.get("追高風險等級")) or "未判定"
    coef = _safe_float(row.get("推薦積極度係數"), _derive_v17_aggressiveness(row)) or 0
    if mode in ["偏空防守", "空頭防守"]:
        return f"{mode}，推薦積極度{coef:.2f}；追高風險{chase}，以小倉位與停損優先。"
    if mode == "震盪選股":
        return f"震盪盤，推薦積極度{coef:.2f}；避免突破失敗，需等待量能確認。"
    return f"{mode}，推薦積極度{coef:.2f}；仍需依停損策略控管單檔風險。"


def _derive_v17_adjust_note(row: pd.Series) -> str:
    base_pos = _safe_float(row.get("建議倉位%"), 0) or 0
    dyn_pos = _safe_float(row.get("動態建議倉位%"), base_pos) or 0
    coef = _safe_float(row.get("推薦積極度係數"), 1) or 1
    return f"原建議倉位{base_pos:.1f}% × 大盤策略係數{coef:.2f} → 動態倉位{dyn_pos:.1f}%。"


def _apply_v17_market_strategy_columns(df: pd.DataFrame) -> pd.DataFrame:
    """V17：依大盤環境產出動態策略；不做硬篩，不改原始推薦名單。"""
    if df is None or df.empty:
        return df
    out = df.copy()
    out["大盤多空分數"] = out.apply(_derive_v17_market_score, axis=1)
    out["大盤策略模式"] = out.apply(_derive_v17_market_mode, axis=1)
    out["推薦積極度係數"] = out.apply(_derive_v17_aggressiveness, axis=1)
    out["適合推薦型態"] = out.apply(_derive_v17_suitable_types, axis=1)
    out["大盤策略建議"] = out.apply(_derive_v17_strategy_note, axis=1)
    out["大盤風控建議"] = out.apply(_derive_v17_risk_note, axis=1)
    base_pos = pd.to_numeric(out.get("建議倉位%", 0), errors="coerce").fillna(0)
    coef = pd.to_numeric(out.get("推薦積極度係數", 1), errors="coerce").fillna(1)
    out["動態建議倉位%"] = (base_pos * coef).clip(lower=0, upper=35).round(1)
    out["市場策略調整說明"] = out.apply(_derive_v17_adjust_note, axis=1)
    return out

def _apply_god_decision_v5_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    out["大盤情境分桶"] = out.apply(_macro_bucket_from_row, axis=1)
    rr_data = out.apply(_calc_trade_risk_reward, axis=1)
    out["風險報酬比"] = [x[0] for x in rr_data]
    out["停損距離%"] = [x[1] for x in rr_data]
    out["目標報酬%"] = [x[2] for x in rr_data]
    out["追價風險分"] = out.apply(_calc_chase_risk_score, axis=1)
    out["股神決策模式"] = out.apply(_derive_v5_decision_mode, axis=1)
    out["股神進場建議"] = out.apply(_derive_entry_advice, axis=1)
    out["建議部位%"] = out.apply(_derive_position_size, axis=1)
    out["推薦分層"] = out.apply(_derive_recommend_layer, axis=1)
    out["不建議買進原因"] = out.apply(_derive_v5_no_buy_reason, axis=1)
    out["最佳操作劇本"] = out.apply(_derive_best_trade_script_v5, axis=1)
    out["隔日操作建議"] = out.apply(_derive_next_day_action, axis=1)
    out["失效價位"] = out["停損價"] if "停損價" in out.columns else ""
    out["轉弱條件"] = out.apply(_derive_weak_condition_v5, axis=1)
    out["大盤情境調權說明"] = out.apply(
        lambda r: f"{_safe_str(r.get('大盤情境分桶'))}｜大盤加權{_safe_float(r.get('大盤加權分'), 0):+.2f}｜{_safe_str(r.get('大盤風險濾網'))}",
        axis=1,
    )
    out = _apply_v16_risk_allocation_columns(out)
    out = _apply_v17_market_strategy_columns(out)
    out = _apply_v76_practical_entry_columns(out)
    return out


# =========================================================
# VNext：股神推薦績效回饋校正版
# 讀取 8_股神推薦紀錄 / godpick_records.json 的歷史績效，新增實戰總分與 A/B/C+/C-/D 分級。
# 不刪除舊欄位；推薦總分仍保留，股神實戰總分用於新版排序與說明。
# =========================================================
def _load_performance_feedback_profile_safe() -> dict[str, Any]:
    cache_key = _k("performance_feedback_profile_cache")
    if cache_key in st.session_state and isinstance(st.session_state.get(cache_key), dict):
        return st.session_state[cache_key]
    if callable(load_godpick_performance_profile):
        try:
            profile = load_godpick_performance_profile("godpick_records.json")
        except Exception as e:
            profile = {"available": False, "message": f"績效回饋載入失敗：{e}", "baseline": {}}
    else:
        profile = {"available": False, "message": "godpick_performance_feedback.py 未載入", "baseline": {}}
    st.session_state[cache_key] = profile
    return profile


def _apply_vnext_performance_feedback_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    profile = _load_performance_feedback_profile_safe()

    if callable(apply_performance_feedback):
        try:
            out = apply_performance_feedback(out, profile)
        except Exception as e:
            out["績效回饋建議"] = f"績效回饋校正失敗：{e}"
            out["績效回饋版本"] = PERFORMANCE_FEEDBACK_VERSION
            for _c in list(GODPICK_PERFORMANCE_FEEDBACK_COLUMNS or []):
                if _c not in out.columns:
                    out[_c] = ""
    else:
        for _c in list(GODPICK_PERFORMANCE_FEEDBACK_COLUMNS or []):
            if _c not in out.columns:
                out[_c] = ""
        out["績效回饋建議"] = "績效回饋模組未載入，沿用原推薦邏輯。"
        out["績效回饋版本"] = PERFORMANCE_FEEDBACK_VERSION

    # VNext Phase 2：最小侵入串接決策引擎；不改掃描架構、不讀寫 JSON。
    if callable(apply_godpick_decision_engine):
        try:
            out = apply_godpick_decision_engine(out, profile)
        except Exception as e:
            out["建議動作"] = f"決策引擎套用失敗：{e}"
            out["決策版本"] = GODPICK_DECISION_ENGINE_VERSION
            for _c in list(GODPICK_DECISION_ENGINE_COLUMNS or []):
                if _c not in out.columns:
                    out[_c] = ""
    else:
        for _c in list(GODPICK_DECISION_ENGINE_COLUMNS or []):
            if _c not in out.columns:
                out[_c] = ""
        out["決策版本"] = GODPICK_DECISION_ENGINE_VERSION

    # >>> PHASE61_PAGE7_SIGNAL_SYNC
    # 統一補股神同步分區，畫面/Excel/8/12 共用同一套分流欄位。
    try:
        from godpick_signal_hub import add_phase61_signal_columns
        out = add_phase61_signal_columns(out)
    except Exception:
        pass
    # <<< PHASE61_PAGE7_SIGNAL_SYNC
    return out


def _render_vnext_performance_feedback_panel() -> None:
    profile = _load_performance_feedback_profile_safe()
    try:
        rows = performance_feedback_summary(profile) if callable(performance_feedback_summary) else []
    except Exception as e:
        rows = [("績效回饋", f"摘要產生失敗：{e}", "")]
    render_pro_info_card(
        "VNext Phase 4｜大盤 × 族群輪動 × 資金攻擊 × 飆股獵人",
        rows or [("績效回饋", "未取得摘要", "")],
        chips=["大盤攻擊模式", "族群輪動", "籌碼續航", "S飆股攻擊候選", "不重複計算"],
    )

def _apply_advanced_godpick_columns(df: pd.DataFrame) -> pd.DataFrame:
    """補齊進階欄位，不破壞舊欄位與既有紀錄格式。"""
    if df is None or df.empty:
        return df
    out = df.copy()
    out["買點分級"] = out.apply(_derive_buy_point_grade, axis=1)
    out["過熱風險"] = out.apply(_derive_overheat_risk, axis=1)
    out["假突破風險"] = out.apply(_derive_fake_breakout_risk, axis=1)
    out["推薦分桶"] = out.apply(_derive_recommend_bucket, axis=1)
    out["飆股起漲分數"] = pd.to_numeric(out.get("起漲前兆分數"), errors="coerce")
    out["起漲等級"] = out.apply(_derive_prelaunch_grade, axis=1)
    out["起漲摘要"] = out.apply(_derive_prelaunch_summary, axis=1)
    out["信心等級"] = out.apply(_derive_confidence_level, axis=1)
    out["買點劇本"] = out.apply(_derive_trade_script, axis=1)
    out["失效條件"] = out.apply(_derive_invalid_condition, axis=1)
    out["風險說明"] = out.apply(_derive_risk_text, axis=1)
    out["股神推論邏輯"] = out.apply(_derive_god_reasoning, axis=1)
    out["權重設定"] = _weight_text()
    if "推薦型態" not in out.columns:
        out["推薦型態"] = out.get("機會型態", "")
    out["推薦型態"] = out["推薦型態"].fillna("").astype(str).replace("", "綜合推薦")

    macro_ref = _load_latest_macro_reference()
    macro_adj = out.apply(lambda r: _macro_adjust_score(r, macro_ref), axis=1)
    out["大盤加權分"] = [x[0] for x in macro_adj]
    out["大盤風險濾網"] = [x[1] for x in macro_adj]
    out["推薦總分"] = pd.to_numeric(out["推薦總分"], errors="coerce").fillna(0) + pd.to_numeric(out["大盤加權分"], errors="coerce").fillna(0)
    out["推薦總分"] = out["推薦總分"].clip(lower=0, upper=100)

    out["大盤參考等級"] = macro_ref.get("大盤參考等級")
    out["大盤可參考分數"] = macro_ref.get("大盤可參考分數")
    out["大盤推薦權重"] = macro_ref.get("大盤推薦權重")
    out["大盤降權原因"] = macro_ref.get("大盤降權原因")
    out["大盤操作風格"] = macro_ref.get("大盤操作風格")
    out["大盤市場廣度分數"] = macro_ref.get("大盤市場廣度分數")
    out["大盤量價確認分數"] = macro_ref.get("大盤量價確認分數")
    out["大盤權值支撐分數"] = macro_ref.get("大盤權值支撐分數")
    out["大盤推薦同步分數"] = macro_ref.get("大盤推薦同步分數")
    out["大盤資料日期"] = macro_ref.get("大盤資料日期")

    tracking_df = out.apply(_build_tracking_placeholders, axis=1, result_type="expand")
    for c in ["3日追蹤預留", "5日追蹤預留", "10日追蹤預留", "20日追蹤預留"]:
        out[c] = tracking_df[c] if c in tracking_df.columns else ""

    out = _apply_god_decision_v5_columns(out)
    out = _apply_vnext_performance_feedback_columns(out)
    return out



def _build_final_god_score_row(row: Any, mode: str, market_score: float, weights: dict[str, int] | None = None) -> tuple[float, str]:
    technical_score = _safe_float(row.get("技術結構分數"), 0) or 0
    prelaunch_score = _safe_float(row.get("起漲前兆分數"), 0) or 0
    category_heat_score = _safe_float(row.get("類股熱度分數"), 0) or 0
    factor_score = _safe_float(row.get("自動因子總分"), 0) or 0
    trade_score = _safe_float(row.get("交易可行分數"), 0) or 0
    leader_advantage = _safe_float(row.get("同類股領先幅度"), 0) or 0
    pattern_score = _safe_float(row.get("型態突破分數"), 0) or 0
    burst_score = _safe_float(row.get("爆發力分數"), 0) or 0
    opportunity_score = _safe_float(row.get("機會股分數"), 0) or 0
    low_score = _safe_float(row.get("低檔位置分數"), 0) or 0
    pullback_score = _safe_float(row.get("拉回承接分數"), 0) or 0
    retest_score = _safe_float(row.get("支撐回測分數"), 0) or 0
    rebound_score = _safe_float(row.get("止跌轉強分數"), 0) or 0
    risk_score = _safe_float(row.get("風險分數"), 0) or 0

    weights = weights or _get_active_weight_map()
    total = (
        market_score * weights["市場環境"] / 100
        + technical_score * weights["技術結構"] / 100
        + prelaunch_score * weights["起漲前兆"] / 100
        + category_heat_score * weights["類股熱度"] / 100
        + factor_score * weights["自動因子"] / 100
        + trade_score * weights["交易可行"] / 100
        + pattern_score * weights["型態突破"] / 100
        + burst_score * weights["爆發力"] / 100
    )

    # 模式只做專業微調，不再硬編死權重，避免使用者調整失效。
    if mode == "飆股模式":
        total += prelaunch_score * 0.04 + burst_score * 0.04 + pattern_score * 0.03
        tag = "爆發優先 / 起漲優先"
    elif mode == "波段模式":
        total += technical_score * 0.04 + trade_score * 0.03 + risk_score * 0.03
        tag = "趨勢延續 / 波段優先"
    elif mode == "領頭羊模式":
        total += leader_advantage * 0.08 + category_heat_score * 0.04
        tag = "類股領先 / 龍頭優先"
    elif mode == "低檔轉強模式":
        total = total * 0.62 + low_score * 0.16 + rebound_score * 0.12 + opportunity_score * 0.10
        tag = "低檔轉強 / 不追高"
    elif mode == "拉回承接模式":
        total = total * 0.60 + pullback_score * 0.18 + trade_score * 0.08 + opportunity_score * 0.14
        tag = "強勢拉回 / 第二買點"
    elif mode == "回測支撐模式":
        total = total * 0.58 + retest_score * 0.20 + rebound_score * 0.08 + opportunity_score * 0.14
        tag = "突破回測 / 支撐確認"
    elif mode == "低檔拉回綜合模式":
        total = total * 0.60 + opportunity_score * 0.22 + max(low_score, pullback_score, retest_score) * 0.12 + rebound_score * 0.06
        tag = "低檔拉回 / 機會優先"
    elif mode == "保守低風險模式":
        total = total * 0.55 + opportunity_score * 0.20 + trade_score * 0.15 + max(0, 100 - risk_score) * 0.10
        tag = "低風險 / 支撐優先"
    else:
        total += (technical_score + prelaunch_score + category_heat_score + trade_score) * 0.012 + opportunity_score * 0.015
        tag = "綜合推薦"

    return _score_clip(total), tag


def _build_recommend_reason_v2(r: pd.Series) -> str:
    parts = []
    if _safe_str(r.get("市場環境")):
        parts.append(_safe_str(r.get("市場環境")))
    if _safe_float(r.get("型態突破分數"), 0) >= 78:
        parts.append(_safe_str(r.get("型態名稱")) or "型態突破")
    if _safe_float(r.get("起漲前兆分數"), 0) >= 75:
        parts.append("起漲前兆強")
    if _safe_float(r.get("交易可行分數"), 0) >= 70:
        parts.append("進出場清楚")
    if _safe_float(r.get("類股熱度分數"), 0) >= 75:
        parts.append("族群熱度高")
    if _safe_str(r.get("類股前3強")) == "是":
        parts.append("類股前3強")
    if _safe_str(r.get("是否領先同類股")) == "是":
        parts.append("領先同類股")
    if _safe_float(r.get("爆發力分數"), 0) >= 75:
        parts.append("爆發力佳")
    if _safe_float(r.get("機會股分數"), 0) >= 70:
        parts.append(_safe_str(r.get("機會型態")) or "低檔拉回機會")
    if _safe_str(r.get("機會股說明")):
        parts.append(_safe_str(r.get("機會股說明")))
    if _safe_float(r.get("風險分數"), 0) < 60:
        parts.append("風險需控管")
    text = "、".join([x for x in parts if x][:6])
    if not text:
        text = "結構偏多，列入觀察"
    entry_zone = _safe_str(r.get("建議切入區"))
    stop_loss = format_number(r.get("停損價"), 2) if pd.notna(r.get("停損價")) else "—"
    target_1 = format_number(r.get("賣出目標1"), 2) if pd.notna(r.get("賣出目標1")) else "—"
    return f"{text}｜切入區 {entry_zone or '—'}｜停損 {stop_loss}｜目標1 {target_1}"


def _avg_safe(values: list[float | None], default: float = 0.0) -> float:
    clean = [float(x) for x in values if x is not None]
    if not clean:
        return default
    return sum(clean) / len(clean)


def _fmt_seconds(sec: float) -> str:
    try:
        sec = max(0, int(sec))
    except Exception:
        sec = 0
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    if h > 0:
        return f"{h}小時 {m}分 {s}秒"
    if m > 0:
        return f"{m}分 {s}秒"
    return f"{s}秒"


def _now_text() -> str:
    return datetime.now(ZoneInfo("Asia/Taipei")).strftime("%Y-%m-%d %H:%M:%S")


def _now_date_text() -> str:
    return datetime.now(ZoneInfo("Asia/Taipei")).strftime("%Y-%m-%d")


def _now_time_text() -> str:
    return datetime.now(ZoneInfo("Asia/Taipei")).strftime("%H:%M:%S")


def _create_record_id(code: str, rec_date: str, rec_time: str, mode: str) -> str:
    raw = f"{_safe_str(code)}|{_safe_str(rec_date)}|{_safe_str(rec_time)}|{_safe_str(mode)}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def _set_status(msg: str, level: str = "info"):
    st.session_state[_k("status_msg")] = msg
    st.session_state[_k("status_type")] = level


def _save_debug_scan_summary(summary: dict[str, Any]):
    st.session_state[_k("debug_scan_summary")] = summary or {}


def _load_debug_scan_summary() -> dict[str, Any]:
    data = st.session_state.get(_k("debug_scan_summary"), {})
    return data if isinstance(data, dict) else {}


def _render_debug_scan_summary():
    data = _load_debug_scan_summary()
    if not data:
        return

    render_pro_section("推薦除錯摘要 / 掃描完整性")
    c1, c2, c3, c4, c5 = st.columns(5)
    with c1:
        st.metric("預計掃描", int(data.get("total_count", 0)))
    with c2:
        st.metric("有效K線", int(data.get("history_ok", data.get("analyzed_ok", 0))))
    with c3:
        st.metric("完整候選池", int(data.get("candidate_diagnosis_count", data.get("analyzed_ok", 0))))
    with c4:
        st.metric("作戰候選", int(data.get("action_candidate_count", data.get("passed_final", 0))))
    with c5:
        st.metric("歷史成功率", f"{float(data.get('history_success_rate_pct', 0) or 0):.1f}%")

    quality = _safe_str(data.get("scan_quality_status"))
    if quality:
        msg = f"掃描品質：{quality}｜覆蓋率 {float(data.get('scan_coverage_pct', 0) or 0):.1f}%"
        if bool(data.get("formal_recommendation_usable", False)):
            st.success(msg)
        else:
            st.error(msg + "｜本輪不可作為正式推薦依據，需重新掃描或修復資料來源。")

    # V184：把「真正資料/程式失敗」與「模型正常淘汰」分開。
    # signal/risk/prelaunch/trade filtered 是策略結果，不是程式錯誤。
    actual_failure = sum(int(data.get(k, 0) or 0) for k in ["invalid_code", "no_history", "analysis_error"])
    # V177 之後 signal/risk/prelaunch/trade 已改成 soft features；同一股票
    # 可同時命中多個軟門檻，而且仍會進 Full-Market AI，不能相加叫「淘汰」。
    soft_gate_count = int(data.get("soft_gate_survivors", 0) or 0)
    if soft_gate_count <= 0:
        soft_gate_count = max([int(data.get(k, 0) or 0) for k in [
            "signal_filtered", "risk_filtered", "prelaunch_filtered", "trade_filtered"
        ]] or [0])
    total_scan = max(1, int(data.get("total_count", 0) or 0))
    d1, d2 = st.columns(2)
    with d1:
        st.metric("真正資料/分析失敗", actual_failure, f"{actual_failure / total_scan * 100:.1f}%")
    with d2:
        st.metric("策略軟門檻提示", soft_gate_count, "仍進AI候選池；條件可重疊")
    if actual_failure > 0:
        st.info(
            f"真正需要補處理的是 {actual_failure} 檔：K線抓取失敗 {int(data.get('no_history', 0) or 0)}、"
            f"分析例外 {int(data.get('analysis_error', 0) or 0)}、代號無效 {int(data.get('invalid_code', 0) or 0)}。"
            "可用『接續上次掃描』只補失敗股，不必把正常淘汰股重新當錯誤處理。"
        )

    lines = []
    mapping = [
        ("invalid_code", "資料失敗｜代號無效"),
        ("no_history", "資料失敗｜K線取得"),
        ("analysis_error", "程式/指標失敗"),
        ("category_filtered", "前置排除｜類型"),
        ("signal_filtered", "軟門檻｜訊號不足"),
        ("risk_filtered", "軟門檻｜風險偏高"),
        ("prelaunch_filtered", "軟門檻｜起漲前兆不足"),
        ("trade_filtered", "軟門檻｜交易可行不足"),
        ("final_score_filtered", "最終分數排除"),
    ]
    for key, label in mapping:
        lines.append(f"{label}：{int(data.get(key, 0))} 檔")
    st.caption("｜".join(lines))

    # V48：推薦速度監控摘要。
    if data.get("scan_elapsed_sec") is not None or data.get("slowest_stocks"):
        with st.expander("V48 推薦速度監控", expanded=False):
            s1, s2, s3, s4 = st.columns(4)
            with s1:
                st.metric("總耗時", _fmt_seconds(_safe_float(data.get("scan_elapsed_sec"), 0) or 0))
            with s2:
                st.metric("平均每檔", _fmt_seconds(_safe_float(data.get("avg_sec_per_stock"), 0) or 0))
            with s3:
                st.metric("歷史資料成功率", f"{_safe_float(data.get('history_success_rate_pct'), 0) or 0:.2f}%")
            with s4:
                st.metric("平行工人", int(data.get("worker_count", 0) or 0))

            slowest = data.get("slowest_stocks", []) or []
            if slowest:
                st.markdown("**最慢 10 檔股票**")
                st.dataframe(pd.DataFrame(slowest), use_container_width=True, hide_index=True)

            status_summary = data.get("status_elapsed_summary", {}) or {}
            if isinstance(status_summary, dict) and status_summary:
                rows = []
                for status, payload in status_summary.items():
                    if isinstance(payload, dict):
                        rows.append({
                            "狀態": status,
                            "檔數": payload.get("count", 0),
                            "平均秒數": payload.get("avg_sec", 0),
                            "最慢秒數": payload.get("max_sec", 0),
                        })
                if rows:
                    st.markdown("**各淘汰 / 成功狀態耗時統計**")
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

            if data.get("data_source_diagnostics_available"):
                st.caption("已讀取 utils.py v48 data_source_diagnostics，可用於判斷 Yahoo / TWSE / TPEx 哪個資料源較慢或失敗。")
                diag = data.get("data_source_diagnostics", {}) or {}
                if isinstance(diag, dict):
                    # 避免整包診斷太大，畫面只顯示摘要與前幾筆。
                    st.json({k: diag.get(k) for k in list(diag.keys())[:12]}, expanded=False)
            elif data.get("data_source_diagnostics_error"):
                st.caption(f"資料源診斷讀取失敗：{data.get('data_source_diagnostics_error')}")

    history_debug = data.get("history_debug_samples", []) or []
    error_debug = data.get("error_samples", []) or []
    if history_debug or error_debug:
        with st.expander("除錯明細", expanded=False):
            if history_debug:
                st.markdown("**歷史資料抓取樣本**")
                for item in history_debug[:10]:
                    st.write(f"- {item}")
            if error_debug:
                st.markdown("**分析錯誤樣本**")
                for item in error_debug[:10]:
                    st.write(f"- {item}")


def _ensure_godpick_record_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=GODPICK_RECORD_COLUMNS)

    x = df.copy()

    if "record_id" not in x.columns and "rec_id" in x.columns:
        x["record_id"] = x["rec_id"]

    for c in GODPICK_RECORD_COLUMNS:
        if c not in x.columns:
            x[c] = None

    numeric_cols = [
        "推薦總分", "技術結構分數", "起漲前兆分數", "起漲等級", "交易可行分數", "類股熱度分數",
        "強勢族群等級", "族群資金流分數", "族群輪動狀態", "同族群強勢比例", "族群策略建議",
        "同類股領先幅度", "推薦價格", "停損價", "賣出目標1", "賣出目標2",
        "實際買進價", "實際賣出價", "實際報酬%", "最新價", "損益金額", "損益幅%", "持有天數", "大盤橋接分數"
    ]
    for c in numeric_cols:
        x[c] = pd.to_numeric(x[c], errors="coerce")

    bool_cols = ["是否領先同類股", "是否已實際買進", "是否達停損", "是否達目標1", "是否達目標2"]
    for c in bool_cols:
        x[c] = x[c].fillna(False).map(lambda v: str(v).strip().lower() in {"true", "1", "yes", "y", "是"})

    x["目前狀態"] = x["目前狀態"].fillna("觀察").replace("", "觀察")
    x["推薦日期"] = x["推薦日期"].fillna("").astype(str).replace("", _now_date_text())
    x["推薦時間"] = x["推薦時間"].fillna("").astype(str).replace("", _now_time_text())
    x["建立時間"] = x["建立時間"].fillna("").astype(str).replace("", _now_text())
    x["更新時間"] = x["更新時間"].fillna("").astype(str).replace("", _now_text())
    x["最新更新時間"] = x["最新更新時間"].fillna("").astype(str)
    x["模式績效標籤"] = x["模式績效標籤"].fillna("").astype(str)
    x["備註"] = x["備註"].fillna("").astype(str)

    need_id = x["record_id"].isna() | (x["record_id"].astype(str).str.strip() == "")
    if need_id.any():
        for idx in x[need_id].index:
            rec_date = _safe_str(x.at[idx, "推薦日期"]) or _now_date_text()
            rec_time = _safe_str(x.at[idx, "推薦時間"]) or _now_time_text()
            x.at[idx, "record_id"] = _create_record_id(
                _safe_str(x.at[idx, "股票代號"]),
                rec_date,
                rec_time,
                _safe_str(x.at[idx, "推薦模式"]),
            )

    return x[GODPICK_RECORD_COLUMNS].copy()


def _append_records_dedup_by_business_key(base_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    base_df = _ensure_godpick_record_columns(base_df)
    new_df = _ensure_godpick_record_columns(new_df)

    if new_df.empty:
        return base_df.copy()

    merged = pd.concat([base_df, new_df], ignore_index=True)
    # v25.9：推薦紀錄防呆改為「同一天 + 同股票代號 + 同推薦模式」不重複。
    # 不再把推薦時間納入 business key，避免同一天重複按匯入造成重複紀錄。
    merged["_biz_key"] = (
        merged["股票代號"].fillna("").astype(str) + "|"
        + merged["推薦日期"].fillna("").astype(str) + "|"
        + merged["推薦模式"].fillna("").astype(str)
    )
    merged["_upd"] = pd.to_datetime(merged["更新時間"], errors="coerce")
    merged = merged.sort_values(["_biz_key", "_upd"], ascending=[True, False], na_position="last")
    merged = merged.drop_duplicates(subset=["_biz_key"], keep="first")
    return _ensure_godpick_record_columns(merged.drop(columns=["_biz_key", "_upd"], errors="ignore"))



# =========================================================
# 類別推論
# =========================================================
CATEGORY_KEYWORD_RULES: list[tuple[str, list[str]]] = [
    ("晶圓代工", ["台積", "聯電", "力積電", "世界先進", "umc", "tsmc", "晶圓代工"]),
    ("IC設計", ["聯發科", "瑞昱", "聯詠", "群聯", "創意", "世芯", "智原", "敦泰", "原相", "晶心科", "矽力", "力旺", "天鈺", "義隆", "祥碩", "譜瑞", "聯陽", "瑞鼎", "義傳", "ic設計"]),
    ("封測", ["日月光", "矽品", "京元電", "頎邦", "欣銓", "矽格", "封測", "測試"]),
    ("記憶體", ["南亞科", "華邦電", "旺宏", "宇瞻", "十銓", "記憶體", "dram", "nand"]),
    ("矽晶圓", ["環球晶", "中美晶", "合晶", "嘉晶", "矽晶圓"]),
    ("半導體設備材料", ["帆宣", "漢唐", "家登", "辛耘", "中砂", "崇越", "萬潤", "均豪", "弘塑", "設備", "材料"]),
    ("IP矽智財", ["力旺", "晶心科", "智原", "創意", "世芯", "ip", "矽智財"]),
    ("AI伺服器", ["伺服器", "server", "緯穎", "廣達", "英業達", "緯創", "鴻海", "技嘉", "微星", "華碩"]),
    ("散熱", ["雙鴻", "奇鋐", "建準", "散熱", "風扇", "熱導管"]),
    ("機殼", ["勤誠", "晟銘電", "迎廣", "機殼"]),
    ("電源供應", ["台達電", "光寶科", "群電", "全漢", "康舒", "電源", "供應器"]),
    ("高速傳輸", ["高速", "傳輸", "祥碩", "譜瑞", "創惟", "威鋒", "usb4", "pcie"]),
    ("網通交換器", ["智邦", "明泰", "中磊", "智易", "啟碁", "網通", "交換器", "switch"]),
    ("光通訊", ["光通訊", "波若威", "華星光", "聯鈞", "上詮", "眾達", "聯亞", "光聖", "cpo"]),
    ("PCB載板", ["欣興", "南電", "景碩", "金像電", "健鼎", "台燿", "華通", "載板", "pcb", "銅箔基板"]),
    ("EMS代工", ["鴻海", "和碩", "廣達", "仁寶", "英業達", "緯創", "組裝"]),
    ("面板", ["友達", "群創", "彩晶", "凌巨", "面板"]),
    ("光學鏡頭", ["大立光", "玉晶光", "亞光", "今國光", "鏡頭", "光學"]),
    ("被動元件", ["國巨", "華新科", "禾伸堂", "凱美", "立隆電", "被動元件", "電容", "電阻"]),
    ("連接器", ["貿聯", "嘉澤", "信邦", "良維", "胡連", "連接器", "端子", "連接線"]),
    ("電池材料", ["康普", "美琪瑪", "立凱", "長園科", "電池", "材料", "鋰"]),
    ("金控", ["金控"]),
    ("銀行", ["銀行"]),
    ("保險", ["保險"]),
    ("證券", ["證券"]),
    ("航運", ["長榮", "陽明", "萬海", "裕民", "慧洋", "航運", "海運", "貨櫃", "散裝"]),
    ("航空觀光", ["華航", "長榮航", "星宇", "航空", "觀光", "旅遊", "飯店"]),
    ("鋼鐵", ["中鋼", "大成鋼", "東和鋼鐵", "鋼鐵", "鋼"]),
    ("塑化", ["台塑", "南亞", "台化", "台塑化", "台聚", "塑化", "化工"]),
    ("生技醫療", ["保瑞", "藥華藥", "美時", "生技", "醫療", "製藥", "藥", "醫材"]),
    ("車用電子", ["和大", "貿聯", "堤維西", "東陽", "車用", "車電", "汽車"]),
    ("綠能儲能", ["中興電", "華城", "士電", "儲能", "綠能", "太陽能", "風電"]),
    ("營建資產", ["營建", "建設", "資產"]),
    ("食品民生", ["統一", "大成", "食品", "餐飲", "飲料"]),
    ("紡織製鞋", ["儒鴻", "聚陽", "志強", "豐泰", "寶成", "紡織", "成衣", "製鞋"]),
    ("電機機械", ["上銀", "亞德客", "直得", "全球傳動", "機械", "工具機", "自動化"]),
    ("其他電子", ["電子", "電腦", "光電"]),
]

CANONICAL_CATEGORY_ALIAS = {
    "半導體": "半導體設備材料",
    "半導體設備": "半導體設備材料",
    "設備材料": "半導體設備材料",
    "半導體材料": "半導體設備材料",
    "伺服器": "AI伺服器",
    "server": "AI伺服器",
    "網通": "網通交換器",
    "交換器": "網通交換器",
    "光通訊/cpo": "光通訊",
    "載板": "PCB載板",
    "pcb": "PCB載板",
    "ems": "EMS代工",
    "鏡頭": "光學鏡頭",
    "光學": "光學鏡頭",
    "被動": "被動元件",
    "電池": "電池材料",
    "生技": "生技醫療",
    "醫療": "生技醫療",
    "車電": "車用電子",
    "綠能": "綠能儲能",
    "建材營造": "營建資產",
    "營建": "營建資產",
    "機械": "電機機械",
}

def _canonical_category(v: Any) -> str:
    text = _normalize_category(v)
    if not text:
        return ""
    key = text.lower()
    for alias, target in CANONICAL_CATEGORY_ALIAS.items():
        if key == alias.lower():
            return target
    return text

def _infer_category_from_name(name: str) -> str:
    n = _safe_str(name)
    if not n:
        return "其他"

    s = n.lower()
    for cat, keywords in CATEGORY_KEYWORD_RULES:
        for kw in keywords:
            if kw.lower() in s:
                return cat
    return "其他"

def _infer_category_from_record(name: str, raw_category: Any) -> str:
    raw_cat = _canonical_category(raw_category)
    if raw_cat:
        if raw_cat in {x[0] for x in CATEGORY_KEYWORD_RULES}:
            return raw_cat
        by_name = _infer_category_from_name(raw_cat)
        if by_name != "其他":
            return by_name
        return raw_cat
    return _infer_category_from_name(name)


# =========================================================
# GitHub / Firestore
# =========================================================
def _github_config() -> dict[str, str]:
    return {
        "token": _safe_str(st.secrets.get("GITHUB_TOKEN", "")),
        "owner": _safe_str(st.secrets.get("GITHUB_REPO_OWNER", "cheng07021028")),
        "repo": _safe_str(st.secrets.get("GITHUB_REPO_NAME", "stock-app")),
        "branch": _safe_str(st.secrets.get("GITHUB_REPO_BRANCH", "main")) or "main",
        "path": _safe_str(st.secrets.get("WATCHLIST_GITHUB_PATH", "watchlist.json")) or "watchlist.json",
    }


def _github_headers(token: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }


def _github_contents_url(owner: str, repo: str, path: str) -> str:
    return f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"


def _get_repo_watchlist_sha(cfg: dict[str, str]) -> tuple[str, str]:
    token = cfg["token"]
    if not token:
        return "", "缺少 GITHUB_TOKEN"

    try:
        resp = requests.get(
            _github_contents_url(cfg["owner"], cfg["repo"], cfg["path"]),
            headers=_github_headers(token),
            params={"ref": cfg["branch"]},
            timeout=20,
        )
        if resp.status_code == 200:
            return _safe_str(resp.json().get("sha")), ""
        if resp.status_code == 404:
            return "", ""
        return "", f"讀取 GitHub 檔案失敗：{resp.status_code} / {resp.text[:300]}"
    except Exception as e:
        return "", f"讀取 GitHub 檔案例外：{e}"


def _push_watchlist_to_github(payload: dict[str, list[dict[str, str]]]) -> tuple[bool, str]:
    cfg = _github_config()
    token = cfg["token"]
    if not token:
        return False, "未設定 GITHUB_TOKEN"

    sha, err = _get_repo_watchlist_sha(cfg)
    if err:
        return False, err

    content_text = json.dumps(payload, ensure_ascii=False, indent=2)
    encoded_content = base64.b64encode(content_text.encode("utf-8")).decode("utf-8")

    body: dict[str, Any] = {
        "message": f"update watchlist from streamlit at {_now_text()}",
        "content": encoded_content,
        "branch": cfg["branch"],
    }
    if sha:
        body["sha"] = sha

    try:
        resp = requests.put(
            _github_contents_url(cfg["owner"], cfg["repo"], cfg["path"]),
            headers=_github_headers(token),
            json=body,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            return True, f"已回寫 GitHub：{cfg['path']}"
        return False, f"GitHub API 寫入失敗：{resp.status_code} / {resp.text[:500]}"
    except Exception as e:
        return False, f"GitHub API 寫入例外：{e}"


def _firebase_config() -> dict[str, str]:
    return {
        "project_id": _safe_str(st.secrets.get("FIREBASE_PROJECT_ID", "")),
        "client_email": _safe_str(st.secrets.get("FIREBASE_CLIENT_EMAIL", "")),
        "private_key": _safe_str(st.secrets.get("FIREBASE_PRIVATE_KEY", "")),
    }


def _clean_private_key(raw_key: str) -> str:
    private_key = _safe_str(raw_key)
    private_key = private_key.replace("\\n", "\n").strip()
    if private_key.startswith("\ufeff"):
        private_key = private_key.lstrip("\ufeff")
    return private_key


def _init_firebase_app():
    if firebase_admin is None or credentials is None or firestore is None:
        raise RuntimeError("firebase-admin 未安裝或無法載入；已略過 Firestore，同步改用本機/GitHub。")
    try:
        return firebase_admin.get_app()
    except ValueError:
        pass

    cfg = _firebase_config()
    project_id = _safe_str(cfg["project_id"]).strip()
    client_email = _safe_str(cfg["client_email"]).strip()
    private_key = _clean_private_key(cfg["private_key"])

    if not project_id:
        raise ValueError("缺少 FIREBASE_PROJECT_ID")
    if not client_email:
        raise ValueError("缺少 FIREBASE_CLIENT_EMAIL")
    if not private_key:
        raise ValueError("缺少 FIREBASE_PRIVATE_KEY")
    if "BEGIN PRIVATE KEY" not in private_key or "END PRIVATE KEY" not in private_key:
        raise ValueError("FIREBASE_PRIVATE_KEY 不是有效 PEM 格式")

    cred_dict = {
        "type": "service_account",
        "project_id": project_id,
        "private_key": private_key,
        "client_email": client_email,
        "token_uri": "https://oauth2.googleapis.com/token",
    }

    cred = credentials.Certificate(cred_dict)
    return firebase_admin.initialize_app(cred, {"projectId": project_id})


def _push_watchlist_to_firestore(payload: dict[str, list[dict[str, str]]]) -> tuple[bool, str]:
    try:
        _init_firebase_app()
        db = firestore.client()
        batch = db.batch()
        now = firestore.SERVER_TIMESTAMP

        summary_ref = db.collection("system").document("watchlist_summary")
        batch.set(
            summary_ref,
            {"group_count": len(payload), "updated_at": now, "source": "streamlit_dual_write"},
            merge=True,
        )

        for group_name, items in payload.items():
            group_name = _safe_str(group_name)
            if not group_name:
                continue

            group_ref = db.collection("watchlists").document(group_name)
            batch.set(
                group_ref,
                {
                    "group_name": group_name,
                    "count": len(items),
                    "items": items,
                    "updated_at": now,
                    "source": "streamlit_dual_write",
                },
                merge=True,
            )

            new_codes = set()
            for item in items:
                code = _normalize_code(item.get("code"))
                if not code:
                    continue
                new_codes.add(code)
                stock_ref = group_ref.collection("stocks").document(code)
                batch.set(
                    stock_ref,
                    {
                        "code": code,
                        "name": _safe_str(item.get("name")) or code,
                        "market": _safe_str(item.get("market")) or "上市",
                        "category": _normalize_category(item.get("category")),
                        "group_name": group_name,
                        "updated_at": now,
                    },
                    merge=True,
                )

            existing_docs = list(group_ref.collection("stocks").stream())
            for doc in existing_docs:
                if doc.id not in new_codes:
                    batch.delete(doc.reference)

        batch.commit()
        return True, "已同步寫入 Firestore"
    except Exception as e:
        return False, f"Firestore 寫入失敗：{e}"


def _normalize_watchlist_payload(data: dict[str, list[dict[str, str]]]) -> dict[str, list[dict[str, str]]]:
    payload: dict[str, list[dict[str, str]]] = {}
    for group_name, items in data.items():
        g = _safe_str(group_name)
        if not g:
            continue
        seen = set()
        normalized_items = []
        for item in items:
            if not isinstance(item, dict):
                continue

            code = _normalize_code(item.get("code"))
            name = _safe_str(item.get("name")) or code
            market = _safe_str(item.get("market")) or "上市"
            category = _normalize_category(item.get("category"))

            if not code:
                continue
            key = (g, code)
            if key in seen:
                continue
            seen.add(key)

            row = {"code": code, "name": name, "market": market}
            if category:
                row["category"] = category
            normalized_items.append(row)

        payload[g] = sorted(normalized_items, key=lambda x: (_normalize_code(x.get("code")), _safe_str(x.get("name"))))
    return payload



def _write_watchlist_local(payload: dict[str, list[dict[str, str]]], path: str = "watchlist.json") -> tuple[bool, str]:
    """本機強制寫回 watchlist.json；GitHub / Firestore 失敗時仍保留自選股資料。"""
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        try:
            get_normalized_watchlist.clear()
        except Exception:
            pass
        return True, f"已寫入本機 {path}"
    except Exception as e:
        return False, f"本機 watchlist.json 寫入失敗：{e}"


def _force_write_watchlist_dual(data: dict[str, list[dict[str, str]]]) -> bool:
    """以本機原子寫入 + GitHub/Firestore 驗證保存自選股與空群組。"""
    payload = _normalize_watchlist_payload(data)
    if not callable(save_watchlist_permanent):
        _set_status("永久保存服務未載入，為避免假成功，本次不寫入。", "error")
        return False

    report = save_watchlist_permanent(payload)
    st.session_state["watchlist_data"] = copy.deepcopy(payload)
    st.session_state["watchlist_version"] = int(st.session_state.get("watchlist_version", 0)) + 1
    st.session_state["watchlist_last_saved_at"] = report.updated_at or _now_text()
    st.session_state["watchlist_last_saved_hash"] = report.payload_hash
    st.session_state[_k("last_dual_write_detail")] = report.messages()
    try:
        get_normalized_watchlist.clear()
    except Exception:
        pass

    remote_configured = False
    try:
        remote_configured = bool(
            (callable(durable_github_config) and durable_github_config().get("token"))
            or (callable(durable_firebase_configured) and durable_firebase_configured())
        )
    except Exception:
        remote_configured = False

    if report.permanent_ok:
        if remote_configured:
            _set_status("自選股／群組已完成本機與遠端永久保存，並通過回讀驗證。", "success")
        else:
            _set_status("自選股／群組已保存至專案固定路徑；目前未設定遠端備份，主機重建時仍可能遺失。", "warning")
        return True

    _set_status("自選股雖可能寫入部分來源，但未通過永久保存條件；請查看同步明細。", "error")
    return False



# =========================================================
# 8 頁推薦紀錄 寫入
# =========================================================
def _godpick_records_config() -> dict[str, str]:
    return {
        "token": _safe_str(st.secrets.get("GITHUB_TOKEN", "")),
        "owner": _safe_str(st.secrets.get("GITHUB_REPO_OWNER", "cheng07021028")),
        "repo": _safe_str(st.secrets.get("GITHUB_REPO_NAME", "stock-app")),
        "branch": _safe_str(st.secrets.get("GITHUB_REPO_BRANCH", "main")) or "main",
        "path": _safe_str(st.secrets.get("GODPICK_RECORDS_GITHUB_PATH", "godpick_records.json")) or "godpick_records.json",
    }


def _read_godpick_records_from_github() -> tuple[list[dict[str, Any]], str]:
    cfg = _godpick_records_config()
    token = cfg["token"]
    if not token:
        return [], "未設定 GITHUB_TOKEN"

    try:
        resp = requests.get(
            _github_contents_url(cfg["owner"], cfg["repo"], cfg["path"]),
            headers=_github_headers(token),
            params={"ref": cfg["branch"]},
            timeout=20,
        )
        if resp.status_code == 404:
            return [], ""
        if resp.status_code != 200:
            return [], f"讀取推薦紀錄失敗：{resp.status_code} / {resp.text[:300]}"

        data = resp.json()
        content = data.get("content", "")
        if not content:
            return [], ""

        decoded = base64.b64decode(content).decode("utf-8")
        payload = json.loads(decoded)
        if isinstance(payload, list):
            return payload, ""
        return [], ""
    except Exception as e:
        return [], f"讀取推薦紀錄例外：{e}"


def _get_godpick_records_sha() -> tuple[str, str]:
    cfg = _godpick_records_config()
    token = cfg["token"]
    if not token:
        return "", "缺少 GITHUB_TOKEN"

    try:
        resp = requests.get(
            _github_contents_url(cfg["owner"], cfg["repo"], cfg["path"]),
            headers=_github_headers(token),
            params={"ref": cfg["branch"]},
            timeout=20,
        )
        if resp.status_code == 200:
            return _safe_str(resp.json().get("sha")), ""
        if resp.status_code == 404:
            return "", ""
        return "", f"讀取推薦紀錄 SHA 失敗：{resp.status_code} / {resp.text[:300]}"
    except Exception as e:
        return "", f"讀取推薦紀錄 SHA 例外：{e}"


def _write_godpick_records_to_github(records: list[dict[str, Any]]) -> tuple[bool, str]:
    cfg = _godpick_records_config()
    token = cfg["token"]
    if not token:
        return False, "未設定 GITHUB_TOKEN"

    sha, err = _get_godpick_records_sha()
    if err:
        return False, err

    content_text = json.dumps(records, ensure_ascii=False, indent=2)
    encoded_content = base64.b64encode(content_text.encode("utf-8")).decode("utf-8")

    body: dict[str, Any] = {
        "message": f"update godpick records at {_now_text()}",
        "content": encoded_content,
        "branch": cfg["branch"],
    }
    if sha:
        body["sha"] = sha

    try:
        resp = requests.put(
            _github_contents_url(cfg["owner"], cfg["repo"], cfg["path"]),
            headers=_github_headers(token),
            json=body,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            return True, f"已回寫 GitHub：{cfg['path']}"
        return False, f"推薦紀錄 GitHub 寫入失敗：{resp.status_code} / {resp.text[:500]}"
    except Exception as e:
        return False, f"推薦紀錄 GitHub 寫入例外：{e}"


def _write_godpick_records_to_firestore(records: list[dict[str, Any]]) -> tuple[bool, str]:
    try:
        _init_firebase_app()
        db = firestore.client()
        batch = db.batch()
        now = firestore.SERVER_TIMESTAMP

        summary_ref = db.collection("system").document("godpick_records_summary")
        batch.set(summary_ref, {"count": len(records), "updated_at": now, "source": "streamlit_godpick_records"}, merge=True)

        records_ref = db.collection("godpick_records")
        existing_docs = list(records_ref.stream())
        existing_ids = {doc.id for doc in existing_docs}
        new_ids = set()

        for row in records:
            rec_id = _safe_str(row.get("record_id"))
            if not rec_id:
                rec_id = _create_record_id(
                    _normalize_code(row.get("股票代號")),
                    _safe_str(row.get("推薦日期")) or _now_date_text(),
                    _safe_str(row.get("推薦時間")) or _now_time_text(),
                    _safe_str(row.get("推薦模式")),
                )
                row["record_id"] = rec_id

            new_ids.add(rec_id)
            doc_ref = records_ref.document(rec_id)
            doc_data = dict(row)
            doc_data["updated_at"] = now
            batch.set(doc_ref, doc_data, merge=True)

        for old_id in existing_ids - new_ids:
            batch.delete(records_ref.document(old_id))

        batch.commit()
        return True, "已同步寫入 Firestore"
    except Exception as e:
        return False, f"推薦紀錄 Firestore 寫入失敗：{e}"




# =========================================================
# 大盤走勢串聯：讀取 0_大盤走勢.py 儲存的 macro_trend_records.json
# =========================================================
MACRO_RECORD_FILES = [
    "macro_trend_records.json",
]


def _macro_grade_weight(grade: str, score: Any) -> tuple[float, str]:
    """依大盤參考等級決定在 7_股神推薦 的自動權重，不硬篩避免漏逆勢飆股。"""
    g = _safe_str(grade)
    s = _safe_float(score, 50) or 50
    if g.startswith("A") or s >= 80:
        return 0.12, "大盤A級，作主要輔助加權"
    if g.startswith("B") or s >= 65:
        return 0.07, "大盤B級，作輔助加權"
    if g.startswith("C") or s >= 50:
        return 0.00, "大盤C級，只作風險濾網"
    return -0.08, "大盤D級，降低追價與弱勢股權重"


def _load_latest_macro_reference() -> dict[str, Any]:
    """讀取最新大盤參考結果。v33 優先 market_snapshot.json，沒有才回舊 macro_trend_records。"""
    snapshot = _read_market_snapshot_v33()
    if snapshot:
        score = _safe_float(snapshot.get("market_score"), 50) or 50
        grade = _macro_reference_grade(score)
        weight, reason = _macro_grade_weight(grade, score)
        gate = _safe_str(snapshot.get("risk_gate"))
        if gate == "normal":
            risk_filter = "可加權"
        elif gate == "selective":
            risk_filter = "只控風險"
        elif gate == "conservative":
            risk_filter = "降權防守"
        elif gate == "data_guard":
            risk_filter = "資料保護"
            weight = min(weight, -0.08)
            reason = "大盤資料不足或過舊，啟用 data_guard，保守降權但不刪剛起漲股"
        else:
            risk_filter = "中性"
        return {
            "大盤參考等級": grade,
            "大盤可參考分數": score,
            "大盤操作風格": _safe_str(snapshot.get("position_hint") or snapshot.get("market_bias") or snapshot.get("trend_comment")) or "未判定",
            "大盤推薦權重": _safe_str(snapshot.get("godpick_weight_advice")) or _macro_weight_advice_from_snapshot_v33(snapshot),
            "大盤降權原因": _safe_str(snapshot.get("trend_comment")) or reason,
            "大盤資料日期": _safe_str(snapshot.get("twse_data_date") or snapshot.get("otc_data_date") or snapshot.get("futures_data_date")),
            "大盤市場廣度分數": _safe_float(snapshot.get("market_breadth_score")),
            "大盤量價確認分數": _safe_float(snapshot.get("volume_confirm_score")),
            "大盤權值支撐分數": _safe_float(snapshot.get("large_cap_support_score")),
            "大盤推薦同步分數": score,
            "大盤風險濾網": risk_filter,
            "大盤策略模式": _safe_str(snapshot.get("risk_gate")),
            "大盤多空分數": score,
            "大盤策略建議": _safe_str(snapshot.get("position_hint")),
            "大盤風控建議": _safe_str(snapshot.get("market_risk_level")),
            "_macro_adjust_weight": weight,
        }

    base_dir = Path(__file__).resolve().parent.parent
    rows = []
    for fn in MACRO_RECORD_FILES:
        p = base_dir / fn
        if not p.exists():
            continue
        try:
            payload = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
        if isinstance(payload, list):
            data_rows = payload
        elif isinstance(payload, dict):
            if isinstance(payload.get("records"), list):
                data_rows = payload.get("records", [])
            elif isinstance(payload.get("data"), list):
                data_rows = payload.get("data", [])
            else:
                data_rows = []
        else:
            data_rows = []
        for r in data_rows:
            if isinstance(r, dict):
                rows.append(r)

    if not rows:
        return {
            "大盤參考等級": "C｜僅作風險濾網",
            "大盤可參考分數": 50.0,
            "大盤操作風格": "未讀到大盤紀錄",
            "大盤推薦權重": "0%",
            "大盤降權原因": "尚未儲存 0_大盤走勢 的投顧參考結果，7頁以中性處理",
            "大盤資料日期": "",
            "大盤市場廣度分數": None,
            "大盤量價確認分數": None,
            "大盤權值支撐分數": None,
            "大盤推薦同步分數": None,
            "大盤風險濾網": "中性",
            "_macro_adjust_weight": 0.0,
        }

    def sort_key(r):
        return (
            _safe_str(r.get("推估日期")),
            _safe_str(r.get("更新時間") or r.get("建立時間")),
            _safe_float(r.get("大盤可參考分數"), 0) or 0,
        )

    latest = sorted(rows, key=sort_key, reverse=True)[0]
    score = _safe_float(latest.get("大盤可參考分數"), 50) or 50
    grade = _safe_str(latest.get("大盤參考等級")) or _macro_reference_grade(score)
    weight, reason = _macro_grade_weight(grade, score)
    risk_filter = "中性"
    if grade.startswith("A"):
        risk_filter = "可加權"
    elif grade.startswith("B"):
        risk_filter = "輔助加權"
    elif grade.startswith("C"):
        risk_filter = "只控風險"
    elif grade.startswith("D"):
        risk_filter = "降權防守"

    return {
        "大盤參考等級": grade,
        "大盤可參考分數": score,
        "大盤操作風格": _safe_str(latest.get("今日適合操作風格")) or _safe_str(latest.get("建議動作")) or "未判定",
        "大盤推薦權重": _safe_str(latest.get("推薦加權建議")) or f"{weight*100:.0f}%",
        "大盤降權原因": _safe_str(latest.get("推薦降權原因")) or reason,
        "大盤資料日期": _safe_str(latest.get("推估日期")),
        "大盤市場廣度分數": _safe_float(latest.get("市場廣度分數")),
        "大盤量價確認分數": _safe_float(latest.get("量價確認分數")),
        "大盤權值支撐分數": _safe_float(latest.get("權值支撐分數")),
        "大盤推薦同步分數": _safe_float(latest.get("推薦同步分數")),
        "大盤風險濾網": risk_filter,
        "_macro_adjust_weight": weight,
    }


def _macro_reference_grade(score: Any) -> str:
    s = _safe_float(score, 50) or 50
    if s >= 80:
        return "A｜可作主要參考"
    if s >= 65:
        return "B｜可作輔助參考"
    if s >= 50:
        return "C｜僅作風險濾網"
    return "D｜不建議作推薦依據"


def _macro_adjust_score(row: pd.Series, macro: dict[str, Any]) -> tuple[float, str]:
    """
    大盤加權採「輔助與降權」，不硬刪股票。
    避免大盤弱但個股逆勢起漲時被漏掉。
    """
    weight = _safe_float(macro.get("_macro_adjust_weight"), 0) or 0
    macro_score = _safe_float(macro.get("大盤可參考分數"), 50) or 50
    if abs(weight) < 0.0001:
        return 0.0, _safe_str(macro.get("大盤風險濾網")) or "中性"

    rec_score = _safe_float(row.get("推薦總分"), 0) or 0
    prelaunch = _safe_float(row.get("飆股起漲分數"), row.get("起漲前兆分數")) or 0
    tech = _safe_float(row.get("技術結構分數"), 0) or 0
    risk = _safe_float(row.get("風險分數"), 50) or 50

    stock_quality = rec_score * 0.45 + prelaunch * 0.25 + tech * 0.20 + max(0, 100 - risk) * 0.10
    raw = (macro_score - 50) * weight * 0.55 + (stock_quality - 60) * weight * 0.35

    # 大盤D級時：只扣追高/弱勢，不重砍逆勢強股。
    grade = _safe_str(macro.get("大盤參考等級"))
    if grade.startswith("D"):
        if prelaunch >= 75 and tech >= 65:
            raw = max(raw, -1.5)
        elif rec_score < 70 or risk >= 65:
            raw -= 2.5
    elif grade.startswith("A"):
        if prelaunch >= 68 and tech >= 60:
            raw += 1.2

    return round(_score_clip(raw, -8, 8), 2), _safe_str(macro.get("大盤風險濾網")) or "中性"



def _normalize_godpick_record(row: dict[str, Any]) -> dict[str, Any]:
    rec_price = _safe_float(row.get("推薦價格"))
    latest_price = _safe_float(row.get("最新價"))
    stop_price = _safe_float(row.get("停損價"))
    target1 = _safe_float(row.get("賣出目標1"))
    target2 = _safe_float(row.get("賣出目標2"))

    pnl_amt = None
    pnl_pct = None
    if rec_price not in [None, 0] and latest_price is not None:
        pnl_amt = latest_price - rec_price
        pnl_pct = (pnl_amt / rec_price) * 100

    hit_stop = False
    if stop_price is not None and latest_price is not None and latest_price <= stop_price:
        hit_stop = True

    hit_target1 = False
    if target1 is not None and latest_price is not None and latest_price >= target1:
        hit_target1 = True

    hit_target2 = False
    if target2 is not None and latest_price is not None and latest_price >= target2:
        hit_target2 = True

    rec_date = _safe_str(row.get("推薦日期")) or _now_date_text()
    rec_time = _safe_str(row.get("推薦時間")) or _now_time_text()
    mode = _safe_str(row.get("推薦模式"))

    norm = {
        "record_id": _safe_str(row.get("record_id")) or _safe_str(row.get("rec_id")) or _create_record_id(
            _normalize_code(row.get("股票代號")), rec_date, rec_time, mode
        ),
        "股票代號": _normalize_code(row.get("股票代號")),
        "股票名稱": _safe_str(row.get("股票名稱")),
        "市場別": _safe_str(row.get("市場別")) or "上市",
        "類別": _normalize_category(row.get("類別")),
        "推薦模式": mode,
        "推薦等級": _safe_str(row.get("推薦等級")),
        "推薦總分": _safe_float(row.get("推薦總分")),
        "買點分級": _safe_str(row.get("買點分級")),
        "大盤參考等級": _safe_str(row.get("大盤參考等級")),
        "大盤可參考分數": _safe_float(row.get("大盤可參考分數")),
        "大盤加權分": _safe_float(row.get("大盤加權分")),
        "大盤風險濾網": _safe_str(row.get("大盤風險濾網")),
        "大盤推薦權重": _safe_str(row.get("大盤推薦權重")),
        "大盤降權原因": _safe_str(row.get("大盤降權原因")),
        "大盤操作風格": _safe_str(row.get("大盤操作風格")),
        "大盤市場廣度分數": _safe_float(row.get("大盤市場廣度分數")),
        "大盤量價確認分數": _safe_float(row.get("大盤量價確認分數")),
        "大盤權值支撐分數": _safe_float(row.get("大盤權值支撐分數")),
        "大盤推薦同步分數": _safe_float(row.get("大盤推薦同步分數")),
        "大盤資料日期": _safe_str(row.get("大盤資料日期")),
        "風險說明": _safe_str(row.get("風險說明")),
        "股神推論邏輯": _safe_str(row.get("股神推論邏輯")),
        "權重設定": _safe_str(row.get("權重設定")),
        "技術結構分數": _safe_float(row.get("技術結構分數")),
        "起漲前兆分數": _safe_float(row.get("起漲前兆分數")),
        "飆股起漲分數": _safe_float(row.get("飆股起漲分數"), row.get("起漲前兆分數")),
        "起漲摘要": _safe_str(row.get("起漲摘要")),
        "交易可行分數": _safe_float(row.get("交易可行分數")),
        "類股熱度分數": _safe_float(row.get("類股熱度分數")),
        "同類股領先幅度": _safe_float(row.get("同類股領先幅度")),
        "是否領先同類股": _safe_str(row.get("是否領先同類股")) in {"是", "True", "true", "1"},
        "推薦標籤": _safe_str(row.get("推薦標籤")),
        "推薦理由摘要": _safe_str(row.get("推薦理由摘要")),
        "推薦價格": rec_price,
        "停損價": stop_price,
        "賣出目標1": target1,
        "賣出目標2": target2,
        "推薦日期": rec_date,
        "推薦時間": rec_time,
        "建立時間": _safe_str(row.get("建立時間")) or _now_text(),
        "更新時間": _now_text(),
        "目前狀態": _safe_str(row.get("目前狀態")) or "觀察",
        "是否已實際買進": _safe_str(row.get("是否已實際買進")) in {"是", "True", "true", "1"},
        "實際買進價": _safe_float(row.get("實際買進價")),
        "實際賣出價": _safe_float(row.get("實際賣出價")),
        "實際報酬%": _safe_float(row.get("實際報酬%")),
        "最新價": latest_price,
        "最新更新時間": _safe_str(row.get("最新更新時間")),
        "損益金額": pnl_amt,
        "損益幅%": pnl_pct,
        "是否達停損": hit_stop if row.get("是否達停損") is None else (_safe_str(row.get("是否達停損")) in {"是", "True", "true", "1"}),
        "是否達目標1": hit_target1 if row.get("是否達目標1") is None else (_safe_str(row.get("是否達目標1")) in {"是", "True", "true", "1"}),
        "是否達目標2": hit_target2 if row.get("是否達目標2") is None else (_safe_str(row.get("是否達目標2")) in {"是", "True", "true", "1"}),
        "持有天數": _safe_float(row.get("持有天數")),
        "模式績效標籤": _safe_str(row.get("模式績效標籤")),
        "備註": _safe_str(row.get("備註")),
    }
    return _ensure_godpick_record_columns(pd.DataFrame([norm])).iloc[0].to_dict()


def _build_record_rows_from_rec_df(rec_df: pd.DataFrame, selected_codes: list[str]) -> list[dict[str, Any]]:
    if rec_df is None or rec_df.empty:
        return []

    work = rec_df[rec_df["股票代號"].astype(str).isin([str(x) for x in selected_codes])].copy()
    rows = []

    rec_date = _now_date_text()
    rec_time = _now_time_text()
    build_time = _now_text()

    for _, r in work.iterrows():
        code = _normalize_code(r.get("股票代號"))
        mode = _safe_str(r.get("推薦模式"))
        rows.append(
            {
                "record_id": _create_record_id(code, rec_date, rec_time, mode),
                "股票代號": code,
                "股票名稱": _safe_str(r.get("股票名稱")),
                "市場別": _safe_str(r.get("市場別")) or "上市",
                "類別": _normalize_category(r.get("類別")),
                "推薦模式": mode,
                "K線驗證標記": "已建立K線驗證資料",
                # v26.1：修正匯出/匯入紀錄時 bundle 未定義造成 NameError。
                # 這裡是由完整推薦表 rec_df 建立紀錄，因此直接使用當列 r 的價格與支撐壓力欄位。
                "推薦日價格": _safe_float(r.get("最新價") if pd.notna(r.get("最新價")) else r.get("推薦價格")),
                "推薦日支撐壓力摘要": (
                    f"近端支撐 {format_number(_safe_float(r.get('近端支撐')), 2)}｜"
                    f"主要支撐 {format_number(_safe_float(r.get('主要支撐')), 2)}｜"
                    f"近端壓力 {format_number(_safe_float(r.get('近端壓力')), 2)}｜"
                    f"停損 {format_number(_safe_float(r.get('停損參考') if pd.notna(r.get('停損參考')) else r.get('停損價')), 2)}"
                ),
                "K線查詢參數": f"stock_code={code}&source=godpick",
                "K線檢視提示": "至 3_歷史K線分析，輸入/帶入此股票，可對照推薦價、支撐、壓力、停損與後續走勢。",
                "推薦等級": _safe_str(r.get("推薦等級")),
                "推薦總分": _safe_float(r.get("推薦總分")),
                "買點分級": _safe_str(r.get("買點分級")),
                "大盤參考等級": _safe_str(r.get("大盤參考等級")),
                "大盤可參考分數": _safe_float(r.get("大盤可參考分數")),
                "大盤加權分": _safe_float(r.get("大盤加權分")),
                "大盤風險濾網": _safe_str(r.get("大盤風險濾網")),
                "大盤推薦權重": _safe_str(r.get("大盤推薦權重")),
                "大盤降權原因": _safe_str(r.get("大盤降權原因")),
                "大盤操作風格": _safe_str(r.get("大盤操作風格")),
                "大盤市場廣度分數": _safe_float(r.get("大盤市場廣度分數")),
                "大盤量價確認分數": _safe_float(r.get("大盤量價確認分數")),
                "大盤權值支撐分數": _safe_float(r.get("大盤權值支撐分數")),
                "大盤推薦同步分數": _safe_float(r.get("大盤推薦同步分數")),
                "大盤資料日期": _safe_str(r.get("大盤資料日期")),
                "大盤橋接分數": _safe_float(r.get("大盤橋接分數")),
                "大盤橋接狀態": _safe_str(r.get("大盤橋接狀態")),
                "大盤橋接加權": _safe_str(r.get("大盤橋接加權")),
                "大盤橋接風控": _safe_str(r.get("大盤橋接風控")),
                "大盤橋接策略": _safe_str(r.get("大盤橋接策略")),
                "大盤橋接更新時間": _safe_str(r.get("大盤橋接更新時間")),
                "大盤交易時段": _safe_str(r.get("大盤交易時段")),
                "大盤交易時段可用": _safe_str(r.get("大盤交易時段可用")),
                "大盤資料品質": _safe_str(r.get("大盤資料品質")),
                "大盤影響加減分": _safe_float(r.get("大盤影響加減分")),
                "大盤影響說明": _safe_str(r.get("大盤影響說明")),
                "隔夜風控分數": _safe_float(r.get("隔夜風控分數")),
                "隔夜風險等級": _safe_str(r.get("隔夜風險等級")),
                "隔夜偏向": _safe_str(r.get("隔夜偏向")),
                "隔夜解讀": _safe_str(r.get("隔夜解讀")),
                "台指夜盤漲跌": _safe_float(r.get("台指夜盤漲跌")),
                "NASDAQ漲跌%": _safe_float(r.get("NASDAQ漲跌%")),
                "S&P500漲跌%": _safe_float(r.get("S&P500漲跌%")),
                "道瓊漲跌%": _safe_float(r.get("道瓊漲跌%")),
                "費半漲跌%": _safe_float(r.get("費半漲跌%")),
                "Nasdaq期貨偏向": _safe_str(r.get("Nasdaq期貨偏向")),
                "S&P期貨偏向": _safe_str(r.get("S&P期貨偏向")),
                "匯率風險等級": _safe_str(r.get("匯率風險等級")),
                "大盤資料診斷摘要": _safe_str(r.get("大盤資料診斷摘要")),
                "風險說明": _safe_str(r.get("風險說明")),
                "股神推論邏輯": _safe_str(r.get("股神推論邏輯")),
                "權重設定": _safe_str(r.get("權重設定")),
                "推薦分桶": _safe_str(r.get("推薦分桶")),
                "起漲等級": _safe_str(r.get("起漲等級")),
                "信心等級": _safe_str(r.get("信心等級")),
                "買點劇本": _safe_str(r.get("買點劇本")),
                "失效條件": _safe_str(r.get("失效條件")),
                "假突破風險": _safe_str(r.get("假突破風險")),
                "過熱風險": _safe_str(r.get("過熱風險")),
                "3日追蹤預留": _safe_str(r.get("3日追蹤預留")),
                "5日追蹤預留": _safe_str(r.get("5日追蹤預留")),
                "10日追蹤預留": _safe_str(r.get("10日追蹤預留")),
                "20日追蹤預留": _safe_str(r.get("20日追蹤預留")),
                "技術結構分數": _safe_float(r.get("技術結構分數")),
                "起漲前兆分數": _safe_float(r.get("起漲前兆分數")),
                "飆股起漲分數": _safe_float(r.get("飆股起漲分數"), r.get("起漲前兆分數")),
                "起漲摘要": _safe_str(r.get("起漲摘要")),
                "交易可行分數": _safe_float(r.get("交易可行分數")),
                "類股熱度分數": _safe_float(r.get("類股熱度分數")),
                "同類股領先幅度": _safe_float(r.get("同類股領先幅度")),
                "是否領先同類股": _safe_str(r.get("是否領先同類股")) in {"是", "True", "true", "1"},
                "推薦標籤": "｜".join([x for x in [_safe_str(r.get("推薦標籤")), _safe_str(r.get("型態名稱")), _safe_str(r.get("爆發等級"))] if x]),
                "推薦理由摘要": _safe_str(r.get("推薦理由摘要")),
                "推薦價格": _safe_float(r.get("最新價") if pd.notna(r.get("最新價")) else r.get("推薦買點_拉回")),
                "停損價": _safe_float(r.get("停損價")),
                "賣出目標1": _safe_float(r.get("賣出目標1")),
                "賣出目標2": _safe_float(r.get("賣出目標2")),
                "推薦日期": rec_date,
                "推薦時間": rec_time,
                "建立時間": build_time,
                "更新時間": build_time,
                "目前狀態": "觀察",
                "是否已實際買進": False,
                "實際買進價": None,
                "實際賣出價": None,
                "實際報酬%": None,
                "最新價": _safe_float(r.get("最新價")),
                "最新更新時間": "",
                "損益金額": None,
                "損益幅%": None,
                "是否達停損": False,
                "是否達目標1": False,
                "是否達目標2": False,
                "持有天數": None,
                "模式績效標籤": "",
                "備註": "",
            }
        )
    return rows



# =========================================================
# 股票主檔 / 分類修正持久化
# =========================================================

# 官方產業代碼映射（TWSE / TPEX 常用）
OFFICIAL_INDUSTRY_CODE_MAP = {
    "01": "水泥工業",
    "02": "食品工業",
    "03": "塑膠工業",
    "04": "紡織纖維",
    "05": "電機機械",
    "06": "電器電纜",
    "08": "玻璃陶瓷",
    "09": "造紙工業",
    "10": "鋼鐵工業",
    "11": "橡膠工業",
    "12": "汽車工業",
    "14": "建材營造",
    "15": "航運業",
    "16": "觀光餐旅",
    "17": "金融保險",
    "18": "貿易百貨",
    "19": "綜合",
    "20": "其他",
    "21": "化學工業",
    "22": "生技醫療",
    "23": "油電燃氣",
    "24": "半導體業",
    "25": "電腦及週邊設備業",
    "26": "光電業",
    "27": "通信網路業",
    "28": "電子零組件業",
    "29": "電子通路業",
    "30": "資訊服務業",
    "31": "其他電子業",
    "32": "文化創意業",
    "33": "農業科技業",
    "34": "綠能環保",
    "35": "數位雲端",
    "36": "運動休閒",
    "37": "居家生活",
}


def _stock_master_config() -> dict[str, str]:
    return {
        "token": _safe_str(st.secrets.get("GITHUB_TOKEN", "")),
        "owner": _safe_str(st.secrets.get("GITHUB_REPO_OWNER", "cheng07021028")),
        "repo": _safe_str(st.secrets.get("GITHUB_REPO_NAME", "stock-app")),
        "branch": _safe_str(st.secrets.get("GITHUB_REPO_BRANCH", "main")) or "main",
        "master_path": _safe_str(st.secrets.get("STOCK_MASTER_GITHUB_PATH", "stock_master_cache.json")) or "stock_master_cache.json",
        "override_path": _safe_str(st.secrets.get("STOCK_CATEGORY_OVERRIDE_GITHUB_PATH", "stock_category_overrides.json")) or "stock_category_overrides.json",
    }


def _read_json_from_github(path: str) -> tuple[Any, str]:
    cfg = _stock_master_config()
    token = cfg["token"]
    if not token:
        return None, "未設定 GITHUB_TOKEN"
    try:
        resp = requests.get(
            _github_contents_url(cfg["owner"], cfg["repo"], path),
            headers=_github_headers(token),
            params={"ref": cfg["branch"]},
            timeout=20,
        )
        if resp.status_code == 404:
            return None, ""
        if resp.status_code != 200:
            return None, f"讀取 GitHub JSON 失敗：{resp.status_code} / {resp.text[:300]}"
        data = resp.json()
        content = data.get("content", "")
        if not content:
            return None, ""
        decoded = base64.b64decode(content).decode("utf-8")
        return json.loads(decoded), ""
    except Exception as e:
        return None, f"讀取 GitHub JSON 例外：{e}"


def _get_github_sha_by_path(path: str) -> tuple[str, str]:
    cfg = _stock_master_config()
    token = cfg["token"]
    if not token:
        return "", "缺少 GITHUB_TOKEN"
    try:
        resp = requests.get(
            _github_contents_url(cfg["owner"], cfg["repo"], path),
            headers=_github_headers(token),
            params={"ref": cfg["branch"]},
            timeout=20,
        )
        if resp.status_code == 200:
            return _safe_str(resp.json().get("sha")), ""
        if resp.status_code == 404:
            return "", ""
        return "", f"讀取 SHA 失敗：{resp.status_code} / {resp.text[:300]}"
    except Exception as e:
        return "", f"讀取 SHA 例外：{e}"


def _write_json_to_github(path: str, payload: Any, commit_message: str) -> tuple[bool, str]:
    cfg = _stock_master_config()
    token = cfg["token"]
    if not token:
        return False, "未設定 GITHUB_TOKEN"

    sha, err = _get_github_sha_by_path(path)
    if err:
        return False, err

    content_text = json.dumps(payload, ensure_ascii=False, indent=2)
    encoded_content = base64.b64encode(content_text.encode("utf-8")).decode("utf-8")
    body: dict[str, Any] = {
        "message": commit_message,
        "content": encoded_content,
        "branch": cfg["branch"],
    }
    if sha:
        body["sha"] = sha

    try:
        resp = requests.put(
            _github_contents_url(cfg["owner"], cfg["repo"], path),
            headers=_github_headers(token),
            json=body,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            return True, f"已回寫 GitHub：{path}"
        return False, f"GitHub 寫入失敗：{resp.status_code} / {resp.text[:500]}"
    except Exception as e:
        return False, f"GitHub 寫入例外：{e}"


def _official_industry_name(raw_value: Any) -> str:
    raw = _safe_str(raw_value)
    if not raw:
        return ""
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 1:
        digits = digits.zfill(2)
    if digits in OFFICIAL_INDUSTRY_CODE_MAP:
        return OFFICIAL_INDUSTRY_CODE_MAP[digits]
    return raw.replace("業別", "").replace("工業", "工業").strip()


def _theme_from_official(official_industry: Any, name: Any) -> str:
    official = _official_industry_name(official_industry)
    by_name = _infer_category_from_name(_safe_str(name))
    if by_name != "其他":
        return by_name
    if not official:
        return "其他_官方未知"
    mapping = {
        "水泥工業": "水泥工業",
        "食品工業": "食品民生",
        "塑膠工業": "塑化",
        "紡織纖維": "紡織製鞋",
        "電機機械": "電機機械",
        "電器電纜": "電器電纜",
        "玻璃陶瓷": "玻璃陶瓷",
        "造紙工業": "造紙工業",
        "鋼鐵工業": "鋼鐵",
        "橡膠工業": "橡膠工業",
        "汽車工業": "汽車",
        "建材營造": "營建資產",
        "航運業": "航運",
        "觀光餐旅": "航空觀光",
        "金融保險": "金融保險",
        "貿易百貨": "貿易百貨",
        "綜合": "綜合",
        "其他": "其他_主題未映射",
        "化學工業": "塑化",
        "生技醫療": "生技醫療",
        "油電燃氣": "油電燃氣",
        "半導體業": "半導體業",
        "電腦及週邊設備業": "電腦及週邊設備業",
        "光電業": "光電業",
        "通信網路業": "通信網路業",
        "電子零組件業": "電子零組件業",
        "電子通路業": "電子通路業",
        "資訊服務業": "資訊服務業",
        "其他電子業": "其他電子業",
        "文化創意業": "文化創意業",
        "農業科技業": "農業科技業",
        "綠能環保": "綠能環保",
        "數位雲端": "數位雲端",
        "運動休閒": "運動休閒",
        "居家生活": "居家生活",
    }
    return mapping.get(official, official)


def _normalize_master_columns(df: pd.DataFrame, market_label: str, code_col: str, name_col: str, industry_col: str, source_api: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    if df is None or df.empty:
        empty = pd.DataFrame(columns=["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"])
        return empty, {"rows": 0, "official_hit": 0, "raw_cols": [], "source_api": source_api}
    work = df.copy()
    for c in [code_col, name_col, industry_col]:
        if c not in work.columns:
            work[c] = ""
    work = work.rename(columns={code_col: "code", name_col: "name", industry_col: "official_industry_raw"})
    work["code"] = work["code"].map(_normalize_code)
    work["name"] = work["name"].map(_safe_str)
    work["market"] = market_label
    work["official_industry_raw_col"] = industry_col
    work["official_industry"] = work["official_industry_raw"].map(_official_industry_name)
    work["theme_category"] = work.apply(lambda r: _theme_from_official(r.get("official_industry"), r.get("name")), axis=1)
    work["category"] = work["theme_category"]
    work["source"] = f"official_{market_label}"
    work["source_api"] = source_api
    work["source_rank"] = 1
    work["待修原因"] = work["official_industry"].map(lambda x: "" if _safe_str(x) else "官方產業未抓到")
    work = work[work["code"] != ""].drop_duplicates(subset=["code"], keep="first").reset_index(drop=True)
    info = {
        "rows": len(work),
        "official_hit": int(work["official_industry"].fillna("").astype(str).str.strip().ne("").sum()),
        "raw_cols": list(df.columns),
        "source_api": source_api,
    }
    return work[["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"]].copy(), info


@st.cache_data(ttl=1800, show_spinner=False)
def _fetch_twse_master() -> tuple[pd.DataFrame, dict[str, Any]]:
    url = "https://openapi.twse.com.tw/v1/opendata/t187ap03_L"
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        payload = resp.json()
        df = pd.DataFrame(payload)
        return _normalize_master_columns(df, "上市", "公司代號", "公司簡稱", "產業別", "twse_openapi")
    except Exception:
        empty = pd.DataFrame(columns=["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"])
        return empty, {"rows": 0, "official_hit": 0, "raw_cols": [], "source_api": "twse_openapi"}


@st.cache_data(ttl=1800, show_spinner=False)
def _fetch_tpex_master(mode: str) -> tuple[pd.DataFrame, dict[str, Any]]:
    if mode == "上櫃":
        url = "https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap03_O"
    else:
        url = "https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap03_R"
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        payload = resp.json()
        df = pd.DataFrame(payload)
        return _normalize_master_columns(df, mode, "SecuritiesCompanyCode", "CompanyAbbreviation", "SecuritiesIndustryCode", f"tpex_{mode}")
    except Exception:
        empty = pd.DataFrame(columns=["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"])
        return empty, {"rows": 0, "official_hit": 0, "raw_cols": [], "source_api": f"tpex_{mode}"}


@st.cache_data(ttl=1800, show_spinner=False)
def _fetch_twse_isin_fill_map() -> dict[str, str]:
    url = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=2"
    out: dict[str, str] = {}
    try:
        tables = pd.read_html(url)
    except Exception:
        return out
    for tb in tables:
        tmp = tb.copy()
        tmp.columns = [str(c) for c in tmp.columns]
        cols = set(tmp.columns)
        if not ({"有價證券代號", "產業別"} <= cols):
            continue
        for _, r in tmp.iterrows():
            code = _normalize_code(r.get("有價證券代號"))
            industry = _official_industry_name(r.get("產業別"))
            if code and industry:
                out[code] = industry
    return out


def _build_utils_master_fallback() -> tuple[pd.DataFrame, dict[str, Any]]:
    dfs = []
    for market_arg in ["上市", "上櫃", "興櫃"]:
        try:
            df = get_all_code_name_map(market_arg)
        except Exception:
            df = pd.DataFrame()
        if df is None or df.empty:
            continue
        temp = df.copy().rename(columns={"證券代號":"code", "證券名稱":"name", "市場別":"market"})
        for c in ["code","name","market"]:
            if c not in temp.columns:
                temp[c] = ""
        temp["code"] = temp["code"].map(_normalize_code)
        temp["name"] = temp["name"].map(_safe_str)
        temp["market"] = temp["market"].map(_safe_str).replace("", market_arg)
        temp["official_industry_raw"] = ""
        temp["official_industry_raw_col"] = ""
        temp["official_industry"] = ""
        temp["theme_category"] = temp["name"].map(_infer_category_from_name).replace("其他", "其他_官方未知")
        temp["category"] = temp["theme_category"]
        temp["source"] = "utils_fallback"
        temp["source_api"] = "utils_all"
        temp["source_rank"] = 9
        temp["待修原因"] = "官方產業未抓到"
        dfs.append(temp[["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"]])
    if not dfs:
        empty = pd.DataFrame(columns=["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"])
        return empty, {"rows": 0, "official_hit": 0, "raw_cols": [], "source_api": "utils_all"}
    out = pd.concat(dfs, ignore_index=True).drop_duplicates(subset=["code"], keep="first").reset_index(drop=True)
    return out, {"rows": len(out), "official_hit": 0, "raw_cols": list(out.columns), "source_api": "utils_all"}


@st.cache_data(ttl=900, show_spinner=False)
def _load_stock_master_cache_from_repo() -> pd.DataFrame:
    cfg = _stock_master_config()
    payload, _ = _read_json_from_github(cfg["master_path"])
    cols = ["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"]
    if not isinstance(payload, list):
        return pd.DataFrame(columns=cols)
    df = pd.DataFrame(payload)
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df["code"] = df["code"].map(_normalize_code)
    df["name"] = df["name"].map(_safe_str)
    df["market"] = df["market"].map(_safe_str).replace("", "上市")
    df["official_industry"] = df["official_industry"].map(_official_industry_name)
    df["theme_category"] = df.apply(lambda r: _theme_from_official(r.get("official_industry"), r.get("name")), axis=1)
    df["category"] = df["theme_category"]
    return df[df["code"] != ""].drop_duplicates(subset=["code"], keep="first")[cols].reset_index(drop=True)


@st.cache_data(ttl=300, show_spinner=False)
def _load_stock_category_override_map() -> dict[str, dict[str, str]]:
    cfg = _stock_master_config()
    payload, _ = _read_json_from_github(cfg["override_path"])
    if not isinstance(payload, dict):
        return {}
    out = {}
    for code, item in payload.items():
        norm_code = _normalize_code(code)
        if not norm_code:
            continue
        if not isinstance(item, dict):
            item = {"category": item}
        out[norm_code] = {
            "code": norm_code,
            "name": _safe_str(item.get("name")),
            "market": _safe_str(item.get("market")),
            "category": _canonical_category(item.get("category")),
            "updated_at": _safe_str(item.get("updated_at")),
        }
    return out


def _merge_master_sources(*dfs: pd.DataFrame) -> pd.DataFrame:
    cols = ["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"]
    items = []
    for df in dfs:
        if isinstance(df, pd.DataFrame) and not df.empty:
            tmp = df.copy()
            for c in cols:
                if c not in tmp.columns:
                    tmp[c] = ""
            items.append(tmp[cols])
    if not items:
        return pd.DataFrame(columns=cols)
    merged = pd.concat(items, ignore_index=True)
    merged["source_rank"] = pd.to_numeric(merged["source_rank"], errors="coerce").fillna(999)
    merged["official_hit"] = merged["official_industry"].fillna("").astype(str).str.strip().ne("").astype(int)
    merged = merged.sort_values(["code", "official_hit", "source_rank"], ascending=[True, False, True])
    merged = merged.drop_duplicates(subset=["code"], keep="first").drop(columns=["official_hit"]).reset_index(drop=True)
    return merged


def _apply_twse_isin_fill(master_df: pd.DataFrame) -> pd.DataFrame:
    if master_df is None or master_df.empty:
        return master_df
    fill_map = _fetch_twse_isin_fill_map()
    if not fill_map:
        return master_df
    work = master_df.copy()
    mask = (work["market"].astype(str) == "上市") & (work["official_industry"].fillna("").astype(str).str.strip() == "")
    for idx in work[mask].index:
        code = _normalize_code(work.at[idx, "code"])
        fill = fill_map.get(code, "")
        if fill:
            work.at[idx, "official_industry_raw"] = fill
            work.at[idx, "official_industry_raw_col"] = "TWSE_ISIN_產業別"
            work.at[idx, "official_industry"] = fill
            work.at[idx, "theme_category"] = _theme_from_official(fill, work.at[idx, "name"])
            work.at[idx, "category"] = work.at[idx, "theme_category"]
            work.at[idx, "source"] = "twse_isin_fill"
            work.at[idx, "source_api"] = "twse_isin"
            work.at[idx, "source_rank"] = 2
            work.at[idx, "待修原因"] = ""
    return work


def _apply_master_overrides(master_df: pd.DataFrame) -> pd.DataFrame:
    if master_df is None or master_df.empty:
        master_df = pd.DataFrame(columns=["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"])
    work = master_df.copy()
    repo_df = _load_stock_master_cache_from_repo()
    work = _merge_master_sources(work, repo_df)
    override_map = _load_stock_category_override_map()
    if override_map:
        for code, item in override_map.items():
            matched = work["code"].astype(str) == str(code)
            if matched.any():
                idx = work[matched].index[0]
                if _safe_str(item.get("name")):
                    work.at[idx, "name"] = _safe_str(item.get("name"))
                if _safe_str(item.get("market")):
                    work.at[idx, "market"] = _safe_str(item.get("market"))
                if _safe_str(item.get("category")):
                    work.at[idx, "theme_category"] = _canonical_category(item.get("category"))
                    work.at[idx, "category"] = _canonical_category(item.get("category"))
                    work.at[idx, "source"] = "override"
                    work.at[idx, "source_api"] = "github_override"
                    work.at[idx, "source_rank"] = 0
                    work.at[idx, "待修原因"] = ""
    work = work[work["code"] != ""].drop_duplicates(subset=["code"], keep="first").reset_index(drop=True)
    return work


def _save_master_cache_to_repo(master_df: pd.DataFrame) -> tuple[bool, str]:
    cfg = _stock_master_config()
    cols = ["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"]
    work = master_df.copy() if isinstance(master_df, pd.DataFrame) else pd.DataFrame(columns=cols)
    for c in cols:
        if c not in work.columns:
            work[c] = ""
    work = work[work["code"].map(_normalize_code) != ""].copy()
    work["code"] = work["code"].map(_normalize_code)
    work["name"] = work["name"].map(_safe_str)
    work["market"] = work["market"].map(_safe_str)
    payload = work[cols].drop_duplicates(subset=["code"], keep="first").sort_values(["code"]).to_dict(orient="records")
    return _write_json_to_github(cfg["master_path"], payload, f"refresh stock master cache at {_now_text()}")


def _save_category_override(code: str, name: str, market: str, category: str) -> tuple[bool, str]:
    cfg = _stock_master_config()
    code = _normalize_code(code)
    if not code:
        return False, "股票代號不可空白"
    payload, _ = _read_json_from_github(cfg["override_path"])
    if not isinstance(payload, dict):
        payload = {}
    payload[code] = {
        "code": code,
        "name": _safe_str(name),
        "market": _safe_str(market) or "上市",
        "category": _canonical_category(category) or _infer_category_from_name(_safe_str(name)),
        "updated_at": _now_text(),
    }
    ok, msg = _write_json_to_github(cfg["override_path"], payload, f"update stock category override {code} at {_now_text()}")
    if ok:
        try:
            _load_stock_category_override_map.clear()
        except Exception:
            pass
    return ok, msg


def _build_master_diagnostics(twse_info=None, tpex_o_info=None, tpex_r_info=None, utils_info=None, merged=None) -> list[str]:
    twse_info = twse_info if isinstance(twse_info, dict) else {}
    tpex_o_info = tpex_o_info if isinstance(tpex_o_info, dict) else {}
    tpex_r_info = tpex_r_info if isinstance(tpex_r_info, dict) else {}
    utils_info = utils_info if isinstance(utils_info, dict) else {}
    merged_df = merged if isinstance(merged, pd.DataFrame) else pd.DataFrame()

    def _n(v, default=0):
        try:
            return int(v)
        except Exception:
            return default

    logs = []
    logs.append(f"TWSE：{_n(twse_info.get('rows'))} 筆 / 正式產業有值 {_n(twse_info.get('official_hit'))} 筆 / API: {_safe_str(twse_info.get('source_api')) or '-'}")
    if twse_info.get("raw_cols"):
        logs.append("TWSE 欄位：" + ", ".join([str(x) for x in list(twse_info.get("raw_cols", []))[:20]]))
    logs.append(f"TPEX-上櫃：{_n(tpex_o_info.get('rows'))} 筆 / 正式產業有值 {_n(tpex_o_info.get('official_hit'))} 筆 / API: {_safe_str(tpex_o_info.get('source_api')) or '-'}")
    logs.append(f"TPEX-興櫃：{_n(tpex_r_info.get('rows'))} 筆 / 正式產業有值 {_n(tpex_r_info.get('official_hit'))} 筆 / API: {_safe_str(tpex_r_info.get('source_api')) or '-'}")
    logs.append(f"utils fallback：{_n(utils_info.get('rows'))} 筆 / API: {_safe_str(utils_info.get('source_api')) or '-'}")
    if not merged_df.empty and "official_industry" in merged_df.columns:
        hit = int(merged_df["official_industry"].fillna("").astype(str).str.strip().ne("").sum())
        logs.append(f"合併後：{len(merged_df)} 筆 / 正式產業有值 {hit} 筆")
    else:
        logs.append("合併後：0 筆 / 正式產業有值 0 筆")
    return logs


def _refresh_stock_master_now() -> tuple[pd.DataFrame, list[str]]:
    try:
        _load_master_df.clear()
    except Exception:
        pass
    fresh_df = _load_master_df()
    logs = list(st.session_state.get(_k("master_diag_logs"), []))
    if fresh_df.empty:
        return fresh_df, logs + ["主檔更新失敗：官方主檔與 fallback 皆無資料"]
    ok, msg = _save_master_cache_to_repo(fresh_df)
    logs.append(msg)
    if ok:
        try:
            _load_stock_master_cache_from_repo.clear()
        except Exception:
            pass
    return fresh_df, logs


def _search_master_df(master_df: pd.DataFrame, keyword: str, market_filter: str, category_filter: str) -> pd.DataFrame:
    cols = ["code","name","market","official_industry_raw","official_industry_raw_col","official_industry","theme_category","category","source","source_api","source_rank","待修原因"]
    if master_df is None or master_df.empty:
        return pd.DataFrame(columns=cols)
    work = master_df.copy()
    kw = _safe_str(keyword)
    market_filter = _safe_str(market_filter)
    category_filter = _safe_str(category_filter)
    if market_filter and market_filter != "全部":
        work = work[work["market"].astype(str) == market_filter].copy()
    if category_filter and category_filter != "全部":
        work = work[(work["category"].astype(str) == category_filter) | (work["official_industry"].astype(str) == category_filter)].copy()
    if kw:
        work = work[
            work["code"].astype(str).str.contains(kw, case=False, na=False)
            | work["name"].astype(str).str.contains(kw, case=False, na=False)
            | work["official_industry"].astype(str).str.contains(kw, case=False, na=False)
            | work["theme_category"].astype(str).str.contains(kw, case=False, na=False)
            | work["category"].astype(str).str.contains(kw, case=False, na=False)
        ].copy()
    return work.sort_values(["market","source_rank","code"]).reset_index(drop=True)


def _render_stock_master_center(
    master_df: pd.DataFrame,
    watchlist_map: dict[str, list[dict[str, str]]],
    all_categories: list[str],
) -> pd.DataFrame:
    return master_df


@st.cache_data(ttl=1800, show_spinner=False)
def _load_master_df() -> pd.DataFrame:
    twse_df, twse_info = _fetch_twse_master()
    tpex_o_df, tpex_o_info = _fetch_tpex_master("上櫃")
    tpex_r_df, tpex_r_info = _fetch_tpex_master("興櫃")
    utils_df, utils_info = _build_utils_master_fallback()
    merged = _merge_master_sources(twse_df, tpex_o_df, tpex_r_df, utils_df)
    merged = _apply_twse_isin_fill(merged)
    merged = _apply_master_overrides(merged)
    st.session_state[_k("master_diag_logs")] = _build_master_diagnostics(twse_info, tpex_o_info, tpex_r_info, utils_info, merged)
    return merged

# =========================================================
# 主檔 / universe helpers
# =========================================================

# =========================================================
# 主檔 / universe helpers
# =========================================================
def _load_watchlist_map() -> dict[str, list[dict[str, str]]]:
    raw = st.session_state.get("watchlist_data")
    if not isinstance(raw, dict) or not raw:
        if callable(load_watchlist_permanent):
            try:
                raw, details = load_watchlist_permanent()
                st.session_state[_k("watchlist_load_detail")] = details
            except Exception as exc:
                st.session_state[_k("watchlist_load_detail")] = [f"永久來源載入失敗：{exc}"]
                raw = {}
        if not isinstance(raw, dict) or not raw:
            try:
                raw = get_normalized_watchlist()
            except Exception:
                raw = {}
        if isinstance(raw, dict):
            st.session_state["watchlist_data"] = copy.deepcopy(raw)

    result: dict[str, list[dict[str, str]]] = {}
    if isinstance(raw, dict):
        for group_name, items in raw.items():
            g = _safe_str(group_name)
            if not g:
                continue
            rows = []
            seen = set()
            if isinstance(items, list):
                for item in items:
                    if isinstance(item, dict):
                        code = _normalize_code(item.get("code"))
                        name = _safe_str(item.get("name")) or code
                        market = _safe_str(item.get("market")) or "上市"
                        category = _infer_category_from_record(name, item.get("category"))
                    else:
                        code = _normalize_code(item)
                        name = code
                        market = "上市"
                        category = ""
                    if not code or code in seen:
                        continue
                    seen.add(code)
                    rows.append({"code": code, "name": name, "market": market, "category": category, "label": f"{code} {name}"})
            result[g] = rows
    return result



@st.cache_data(ttl=300, show_spinner=False)
def _load_master_df_fallback_only() -> pd.DataFrame:
    try:
        repo_df = load_stock_master() if callable(load_stock_master) else pd.DataFrame()
    except Exception:
        repo_df = pd.DataFrame()

    if repo_df is None or repo_df.empty:
        repo_df = _load_stock_master_cache_from_repo()

    if repo_df is None or repo_df.empty:
        return pd.DataFrame(columns=["code", "name", "market", "category"])

    work = repo_df.copy()

    if "code" not in work.columns:
        work["code"] = ""
    if "name" not in work.columns:
        work["name"] = ""
    if "market" not in work.columns:
        work["market"] = "上市"
    if "category" not in work.columns:
        if "theme_category" in work.columns:
            work["category"] = work["theme_category"]
        else:
            work["category"] = ""

    work["code"] = work["code"].map(_normalize_code)
    work["name"] = work["name"].map(_safe_str)
    work["market"] = work["market"].map(_safe_str).replace("", "上市")
    work["category"] = work.apply(
        lambda r: _infer_category_from_record(r.get("name"), r.get("category")),
        axis=1,
    )

    work = _apply_master_overrides(work)

    return (
        work[work["code"] != ""]
        .drop_duplicates(subset=["code"], keep="first")
        .reset_index(drop=True)
    )


@st.cache_data(ttl=1800, show_spinner=False)
def _build_master_lookup(master_df: pd.DataFrame) -> dict[str, dict[str, str]]:
    if master_df is None or master_df.empty:
        return {}
    work = master_df.copy()
    if "code" not in work.columns:
        return {}
    out: dict[str, dict[str, str]] = {}
    for _, row in work.iterrows():
        code = _normalize_code(row.get("code"))
        if not code or code in out:
            continue
        name = _safe_str(row.get("name")) or code
        market = _safe_str(row.get("market")) or "上市"
        category = _normalize_category(row.get("category")) or _infer_category_from_record(name, row.get("category"))
        out[code] = {
            "name": name,
            "market": market,
            "category": category,
        }
    return out

def _find_name_market_category(
    code: str,
    manual_name: str,
    manual_market: str,
    manual_category: str,
    master_df_or_lookup,
) -> tuple[str, str, str]:
    code = _normalize_code(code)
    manual_name = _safe_str(manual_name)
    manual_market = _safe_str(manual_market)
    manual_category = _normalize_category(manual_category)

    if isinstance(master_df_or_lookup, dict):
        found = master_df_or_lookup.get(code, {})
        if found:
            final_name = _safe_str(found.get("name")) or manual_name or code
            final_market = _safe_str(found.get("market")) or manual_market or "上市"
            final_category = _normalize_category(found.get("category")) or manual_category or _infer_category_from_record(final_name, manual_category)
            return final_name, final_market, final_category

    if isinstance(master_df_or_lookup, pd.DataFrame) and not master_df_or_lookup.empty:
        matched = master_df_or_lookup[master_df_or_lookup["code"].astype(str) == code]
        if not matched.empty:
            row = matched.iloc[0]
            final_name = _safe_str(row.get("name")) or manual_name or code
            final_market = _safe_str(row.get("market")) or manual_market or "上市"
            final_category = _normalize_category(row.get("category")) or manual_category or _infer_category_from_record(final_name, manual_category)
            return final_name, final_market, final_category

    final_name = manual_name or code
    final_market = manual_market or "上市"
    final_category = manual_category or _infer_category_from_record(final_name, manual_category)
    return final_name, final_market, final_category

    final_name = manual_name or code
    final_market = manual_market or "上市"
    final_category = manual_category or _infer_category_from_record(final_name, manual_category)
    return final_name, final_market, final_category


def _parse_manual_codes(text: str, master_df: pd.DataFrame) -> list[dict[str, str]]:
    rows = []
    seen = set()
    raw_lines = [x.strip() for x in _safe_str(text).replace("，", "\n").replace(",", "\n").splitlines() if x.strip()]

    for raw in raw_lines:
        txt = _safe_str(raw)
        code = _normalize_code(txt)
        name = ""
        market = "上市"
        category = ""

        if not code and isinstance(master_df, pd.DataFrame) and not master_df.empty:
            matched = master_df[master_df["name"].astype(str).str.contains(txt, case=False, na=False)]
            if not matched.empty:
                row = matched.iloc[0]
                code = _normalize_code(row.get("code"))
                name = _safe_str(row.get("name"))
                market = _safe_str(row.get("market")) or "上市"
                category = _normalize_category(row.get("category"))

        if code and not name:
            name, market, category = _find_name_market_category(code, "", market, category, master_df)

        if code and code not in seen:
            seen.add(code)
            rows.append(
                {
                    "code": code,
                    "name": name or code,
                    "market": market or "上市",
                    "category": category,
                    "label": f"{code} {name or code}",
                }
            )
    return rows


def _build_universe_from_market(
    master_df: pd.DataFrame,
    market_mode: str,
    limit_count: Any,
    selected_categories: list[str],
) -> list[dict[str, str]]:
    if master_df is None or master_df.empty:
        return []

    work = master_df.copy()
    market_mode = _safe_str(market_mode)

    if market_mode == "上市":
        work = work[work["market"].astype(str) == "上市"].copy()
    elif market_mode == "上櫃":
        work = work[work["market"].astype(str) == "上櫃"].copy()
    elif market_mode == "興櫃":
        work = work[work["market"].astype(str) == "興櫃"].copy()

    clean_categories = [_normalize_category(x) for x in selected_categories if _normalize_category(x) and x != "全部"]
    if clean_categories:
        work = work[work["category"].astype(str).isin(clean_categories)].copy()

    if _safe_str(limit_count) != "全部":
        try:
            limit_n = int(limit_count)
            if limit_n > 0:
                work = work.head(limit_n).copy()
        except Exception:
            pass

    rows = []
    for _, row in work.iterrows():
        code = _normalize_code(row.get("code"))
        name = _safe_str(row.get("name")) or code
        market = _safe_str(row.get("market")) or "上市"
        category = _normalize_category(row.get("category")) or _infer_category_from_name(name)
        if code:
            rows.append(
                {
                    "code": code,
                    "name": name,
                    "market": market,
                    "category": category,
                    "label": f"{code} {name}",
                }
            )
    return rows


def _collect_all_categories(master_df: pd.DataFrame, watchlist_map: dict[str, list[dict[str, str]]]) -> list[str]:
    cats = set()

    if isinstance(master_df, pd.DataFrame) and not master_df.empty:
        for _, row in master_df.iterrows():
            name = _safe_str(row.get("name"))
            cat = _normalize_category(row.get("category")) or _infer_category_from_name(name)
            if cat:
                cats.add(cat)

    if isinstance(watchlist_map, dict):
        for _, items in watchlist_map.items():
            for item in items:
                name = _safe_str(item.get("name"))
                cat = _infer_category_from_record(name, item.get("category"))
                if cat:
                    cats.add(cat)

    return sorted(list(cats))




def _find_existing_watchlist_codes(group_name: str, codes: list[str]) -> list[str]:
    """檢查勾選股票是否已存在於自選股群組。"""
    group_name = _safe_str(group_name)
    check_codes = {_normalize_code(x) for x in codes if _normalize_code(x)}
    if not group_name or not check_codes:
        return []

    raw = st.session_state.get("watchlist_data")
    if not isinstance(raw, dict) or not raw:
        try:
            raw, details = load_watchlist_permanent() if callable(load_watchlist_permanent) else (get_normalized_watchlist(), [])
            st.session_state[_k("watchlist_load_detail")] = details
            st.session_state["watchlist_data"] = copy.deepcopy(raw)
        except Exception:
            raw = {}

    group_items = raw.get(group_name, []) if isinstance(raw, dict) else []
    exists = set()
    for item in group_items:
        if isinstance(item, dict):
            c = _normalize_code(item.get("code"))
        else:
            c = _normalize_code(item)
        if c in check_codes:
            exists.add(c)

    return sorted(exists)


def _record_business_key(row: dict[str, Any]) -> str:
    """股神推薦紀錄去重用 business key。

    必須與 ``_append_records_dedup_by_business_key`` 完全一致：同一天、
    同股票、同推薦模式只保留一筆。舊版把推薦時間納入檢查，但實際
    合併時又忽略時間，導致畫面說可匯入、保存時卻被覆蓋。
    """
    return (
        f"{_normalize_code(row.get('股票代號'))}|"
        f"{_safe_str(row.get('推薦日期'))}|"
        f"{_safe_str(row.get('推薦模式'))}"
    )


def _find_existing_godpick_record_codes(record_rows: list[dict[str, Any]]) -> tuple[list[str], list[str]]:
    """檢查永久推薦紀錄中是否已有相同 business key。"""
    if not record_rows:
        return [], []

    old_records: list[dict[str, Any]] = []
    try:
        if callable(load_records_permanent):
            old_records, details = load_records_permanent()
            st.session_state[_k("record_authority_load_detail")] = details
        else:
            old_records = _safe_json_read_local("godpick_records.json", [])
            if not isinstance(old_records, list):
                old_records = []
    except Exception as exc:
        st.session_state[_k("record_authority_load_detail")] = [f"讀取永久推薦紀錄失敗：{exc}"]
        old_records = []

    old_df = _ensure_godpick_record_columns(pd.DataFrame(old_records))
    if old_df.empty:
        return [], []

    old_keys = {_record_business_key(r.to_dict()) for _, r in old_df.iterrows()}
    dup_codes: list[str] = []
    dup_keys: list[str] = []
    for row in record_rows:
        key = _record_business_key(row)
        if key in old_keys:
            code = _normalize_code(row.get("股票代號"))
            if code:
                dup_codes.append(code)
            dup_keys.append(key)
    return sorted(set(dup_codes)), sorted(set(dup_keys))


def _append_stock_to_watchlist(group_name: str, code: str, name: str, market: str, category: str):
    group_name = _safe_str(group_name)
    code = _normalize_code(code)
    name = _safe_str(name) or code
    market = _safe_str(market) or "上市"
    category = _canonical_category(category) or _infer_category_from_record(name, category)

    if not group_name:
        return False, "群組不可空白"
    if not code:
        return False, "股票代號不可空白"

    raw = st.session_state.get("watchlist_data")
    if not isinstance(raw, dict) or not raw:
        try:
            raw, details = load_watchlist_permanent() if callable(load_watchlist_permanent) else (get_normalized_watchlist(), [])
            st.session_state[_k("watchlist_load_detail")] = details
            st.session_state["watchlist_data"] = copy.deepcopy(raw)
        except Exception:
            raw = {}

    if group_name not in raw or not isinstance(raw[group_name], list):
        raw[group_name] = []

    for item in raw[group_name]:
        if isinstance(item, dict) and _normalize_code(item.get("code")) == code:
            return False, f"{code} 已存在於 {group_name}"

    row = {"code": code, "name": name, "market": market}
    if category:
        row["category"] = category

    raw[group_name].append(row)
    ok = _force_write_watchlist_dual(raw)
    if ok:
        return True, f"已加入 {group_name}：{code} {name}"
    return False, _safe_str(st.session_state.get(_k("status_msg"), "寫入失敗"))


def _append_multiple_stocks_to_watchlist(group_name: str, rows: list[dict[str, str]]) -> tuple[int, list[str]]:
    group_name = _safe_str(group_name)
    if not group_name:
        return 0, ["請先選擇群組。"]

    raw = st.session_state.get("watchlist_data")
    if not isinstance(raw, dict) or not raw:
        try:
            raw, details = load_watchlist_permanent() if callable(load_watchlist_permanent) else (get_normalized_watchlist(), [])
            st.session_state[_k("watchlist_load_detail")] = details
            st.session_state["watchlist_data"] = copy.deepcopy(raw)
        except Exception:
            raw = {}

    if group_name not in raw or not isinstance(raw[group_name], list):
        raw[group_name] = []

    existing_codes = {_normalize_code(x.get("code")) for x in raw[group_name] if isinstance(x, dict)}
    added = 0
    messages = []

    for row in rows:
        code = _normalize_code(row.get("code"))
        name = _safe_str(row.get("name")) or code
        market = _safe_str(row.get("market")) or "上市"
        category = _normalize_category(row.get("category")) or _infer_category_from_name(name)

        if not code:
            continue

        if code in existing_codes:
            messages.append(f"{code} 已存在於 {group_name}")
            continue

        item = {"code": code, "name": name, "market": market}
        if category:
            item["category"] = category

        raw[group_name].append(item)
        existing_codes.add(code)
        added += 1
        messages.append(f"已加入 {group_name}：{code} {name}")

    if added > 0:
        ok = _force_write_watchlist_dual(raw)
        if not ok:
            return 0, [_safe_str(st.session_state.get(_k("status_msg"), "GitHub / Firestore 寫入失敗"))]

    return added, messages


def _create_watchlist_group(group_name: str) -> tuple[bool, str]:
    group_name = _safe_str(group_name)
    if not group_name:
        return False, "群組名稱不可空白"

    raw = st.session_state.get("watchlist_data")
    if not isinstance(raw, dict) or raw is None:
        try:
            raw, details = load_watchlist_permanent() if callable(load_watchlist_permanent) else (get_normalized_watchlist(), [])
            st.session_state[_k("watchlist_load_detail")] = details
            st.session_state["watchlist_data"] = copy.deepcopy(raw)
        except Exception:
            raw = {}

    if not isinstance(raw, dict):
        raw = {}

    if group_name in raw:
        return False, f"群組已存在：{group_name}"

    raw[group_name] = []
    ok = _force_write_watchlist_dual(raw)
    if ok:
        return True, f"已新增群組：{group_name}"
    return False, _safe_str(st.session_state.get(_k("status_msg"), "新增群組失敗"))



def _show_import_result_notice(title: str, added_count: int, selected_count: int, messages: list[str], module_name: str):
    """v26.5：匯入成功 / 重複防呆 / 失敗提示統一顯示。"""
    duplicate_msgs = []
    fail_msgs = []
    success_msgs = []

    for msg in messages or []:
        s = _safe_str(msg)
        if not s:
            continue
        if any(k in s for k in ["已存在", "已在", "略過", "重複", "防呆"]):
            duplicate_msgs.append(s)
        elif any(k in s for k in ["失敗", "例外", "錯誤", "未設定"]):
            fail_msgs.append(s)
        else:
            success_msgs.append(s)

    duplicate_count = max(selected_count - int(added_count or 0), 0)

    if added_count > 0 and duplicate_count > 0:
        st.success(f"{title}：成功新增 {added_count} 筆；另有 {duplicate_count} 筆疑似重複或未寫入，請看明細。")
    elif added_count > 0:
        st.success(f"{title}：成功新增 {added_count} 筆到 {module_name}。")
    elif duplicate_count > 0 or duplicate_msgs:
        st.warning(f"{title}：沒有新增資料，可能已存在；防呆已阻擋重複匯入。")
    else:
        st.warning(f"{title}：沒有新增資料，請查看寫入明細。")

    with st.expander(f"{title}｜寫入明細", expanded=True):
        st.write(f"- 勾選筆數：{selected_count}")
        st.write(f"- 新增筆數：{added_count}")
        if duplicate_count > 0:
            st.write(f"- 可能重複 / 略過筆數：{duplicate_count}")
        if success_msgs:
            st.write("#### 成功 / 同步訊息")
            for msg in success_msgs:
                st.write(f"- {msg}")
        if duplicate_msgs:
            st.write("#### 防呆略過")
            for msg in duplicate_msgs:
                st.write(f"- {msg}")
        if fail_msgs:
            st.write("#### 失敗 / 異常")
            for msg in fail_msgs:
                st.write(f"- {msg}")



def _append_godpick_records(record_rows: list[dict[str, Any]], force_duplicate: bool = False) -> tuple[int, list[str]]:
    """將 07 股神推薦結果寫入唯一權威紀錄檔。

    V174：一般寫入採鎖定式增量 upsert，直接重讀當下最新
    ``godpick_records.json`` 後合併，不再先下載整份 GitHub/Firestore，
    也不再以頁面舊 DataFrame 整檔回寫。這可避免跨頁、多人或
    Streamlit rerun 將新紀錄回退成舊日期。
    """
    if not record_rows:
        return 0, ["沒有可寫入的推薦紀錄。"]

    try:
        normalized_rows = [_normalize_godpick_record(x) for x in record_rows if isinstance(x, dict)]
        new_df = _ensure_godpick_record_columns(pd.DataFrame(normalized_rows))
        if new_df.empty:
            return 0, ["匯入資料正規化後沒有有效推薦紀錄。"]

        if not force_duplicate and callable(upsert_records_authority_fast):
            report, stats = upsert_records_authority_fast(
                new_df.to_dict(orient="records"),
                reason="07 股神推薦完成自動紀錄／手動寫入",
            )
            authority_detail = ""
            if callable(records_authority_status):
                try:
                    authority = records_authority_status()
                    authority_detail = (
                        f"權威檔：{authority.get('path')}｜{authority.get('count', 0)}筆｜"
                        f"最新推薦日期 {authority.get('latest_recommendation_date') or '未取得'}｜"
                        f"state {'有效' if authority.get('valid') else '待修復'}"
                    )
                except Exception as authority_exc:
                    authority_detail = f"權威檔驗證失敗：{authority_exc}"
            save_details = report.messages()
            details = [
                f"權威增量保存：{'成功' if report.permanent_ok else '失敗'}｜新增 {stats.get('added', 0)}｜更新 {stats.get('updated', 0)}",
                authority_detail,
                *save_details,
            ]
            st.session_state[_k("last_record_write_detail")] = [x for x in details if x]
            changed_count = int(stats.get("changed", 0) or 0)
            if bool(report.permanent_ok):
                if changed_count == 0:
                    return 0, ["相同日期、股票與推薦模式已存在；權威資料未重複新增。", authority_detail, *save_details]
                return changed_count, [
                    f"已直接寫入唯一權威檔：新增 {stats.get('added', 0)} 筆、更新 {stats.get('updated', 0)} 筆。",
                    authority_detail,
                    *save_details,
                ]
            return 0, ["推薦紀錄未寫入權威檔；本輪不得顯示保存成功。", authority_detail, *save_details]

        # 明確允許重複時才走相容完整保存流程。
        if callable(load_records_permanent):
            old_records, load_details = load_records_permanent()
        else:
            old_records = _safe_json_read_local("godpick_records.json", [])
            if not isinstance(old_records, list):
                old_records = []
            load_details = ["永久保存服務未載入，使用本機相容模式。"]
        old_df = _ensure_godpick_record_columns(pd.DataFrame(old_records))
        duplicated = new_df.copy()
        now_tag = str(int(time.time() * 1000))
        for idx in duplicated.index:
            raw = (
                f"{_safe_str(duplicated.at[idx, '股票代號'])}|"
                f"{_safe_str(duplicated.at[idx, '推薦日期'])}|"
                f"{_safe_str(duplicated.at[idx, '推薦時間'])}|"
                f"{_safe_str(duplicated.at[idx, '推薦模式'])}|duplicate|{now_tag}|{idx}"
            )
            duplicated.at[idx, "record_id"] = hashlib.md5(raw.encode("utf-8")).hexdigest()
            note = _safe_str(duplicated.at[idx, "備註"])
            duplicated.at[idx, "備註"] = (note + "；" if note else "") + "使用者確認重複紀錄"
            duplicated.at[idx, "更新時間"] = _now_text()
        merged_df = _ensure_godpick_record_columns(pd.concat([old_df, duplicated], ignore_index=True))
        if callable(save_records_sync_fast):
            report = save_records_sync_fast(merged_df.to_dict(orient="records"), reason="07 使用者確認重複紀錄")
            details = list(load_details) + report.messages()
            st.session_state[_k("last_record_write_detail")] = details
            if report.permanent_ok:
                return len(duplicated), [f"已依使用者確認新增重複紀錄 {len(duplicated)} 筆。", *details]
            return 0, ["重複紀錄未通過永久保存驗證。", *details]
        # V191-H3 data-integrity rule: never bypass the authority service with a
        # direct full-file write.  A transient import/service failure must fail
        # closed rather than risk replacing 1,800+ cumulative records with a
        # partial/stale in-page DataFrame.
        msg_local = "V191-H3防歸零：推薦紀錄永久服務未載入，已拒絕直接覆寫 godpick_records.json；請修復服務後再重試。"
        st.session_state[_k("last_record_write_detail")] = [msg_local]
        return 0, [msg_local]
    except Exception as exc:
        st.session_state[_k("last_record_write_detail")] = [f"永久推薦紀錄例外：{exc}"]
        return 0, [f"寫入股神推薦紀錄失敗：{exc}"]

def _normalize_recommend_list_payload(payload) -> list[dict[str, Any]]:
    """v26：支援 10_推薦清單 的 list / dict 格式。"""
    if isinstance(payload, dict):
        if isinstance(payload.get("recommendations"), list):
            payload = payload.get("recommendations")
        elif isinstance(payload.get("records"), list):
            payload = payload.get("records")
        elif isinstance(payload.get("data"), list):
            payload = payload.get("data")
        else:
            payload = []
    if not isinstance(payload, list):
        return []
    return [dict(x) for x in payload if isinstance(x, dict)]


def _recommend_list_business_key(row: dict[str, Any]) -> str:
    """v26：推薦清單防呆，同一天 + 同股票 + 同推薦模式不重複。"""
    return (
        f"{_normalize_code(row.get('股票代號'))}|"
        f"{_safe_str(row.get('推薦日期'))}|"
        f"{_safe_str(row.get('推薦模式'))}"
    )


def _build_recommend_list_rows_from_rec_df(rec_df: pd.DataFrame, selected_codes: list[str]) -> list[dict[str, Any]]:
    if rec_df is None or rec_df.empty:
        return []
    codes = {_normalize_code(x) for x in selected_codes if _normalize_code(x)}
    if not codes or "股票代號" not in rec_df.columns:
        return []

    rec_date = _now_date_text()
    rec_time = _now_time_text()
    build_time = _now_text()
    work = rec_df[rec_df["股票代號"].astype(str).map(lambda x: _normalize_code(x) in codes)].copy()

    rows: list[dict[str, Any]] = []
    for _, r in work.iterrows():
        row = dict(r)
        code = _normalize_code(row.get("股票代號"))
        mode = _safe_str(row.get("推薦模式"))
        row["股票代號"] = code
        row["股票名稱"] = _safe_str(row.get("股票名稱"))
        row["市場別"] = _safe_str(row.get("市場別")) or "上市"
        row["類別"] = _normalize_category(row.get("類別"))
        row["推薦日期"] = _safe_str(row.get("推薦日期")) or rec_date
        row["推薦時間"] = _safe_str(row.get("推薦時間")) or rec_time
        row["推薦模式"] = mode
        row["資料來源"] = _safe_str(row.get("資料來源")) or "7_股神推薦_完整推薦表"
        row["狀態"] = _safe_str(row.get("狀態")) or "觀察中"
        row["建立時間"] = _safe_str(row.get("建立時間")) or build_time
        row["更新時間"] = build_time
        if not _safe_str(row.get("record_id")):
            row["record_id"] = _create_record_id(code, row["推薦日期"], row["推薦時間"], mode)
        rows.append(row)
    return rows


def _append_recommend_list_from_full_table(rec_df: pd.DataFrame, selected_codes: list[str]) -> tuple[int, list[str]]:
    """
    v26：從完整推薦表勾選資料寫入 10_推薦清單。
    防呆：同一天 + 股票代號 + 推薦模式 不重複。
    """
    new_rows = _build_recommend_list_rows_from_rec_df(rec_df, selected_codes)
    if not new_rows:
        return 0, ["沒有可寫入推薦清單的資料。"]

    github_payload, github_msg = _read_json_from_github_path(GODPICK_LIST_FILE, [])
    local_payload = _safe_json_read_local(GODPICK_LIST_FILE, [])

    old_rows = []
    old_rows.extend(_normalize_recommend_list_payload(github_payload))
    old_rows.extend(_normalize_recommend_list_payload(local_payload))

    merged_map: dict[str, dict[str, Any]] = {}
    for row in old_rows:
        key = _recommend_list_business_key(row)
        if key.strip("|"):
            merged_map[key] = row

    added = 0
    messages: list[str] = []
    for row in new_rows:
        key = _recommend_list_business_key(row)
        code = _normalize_code(row.get("股票代號"))
        name = _safe_str(row.get("股票名稱"))
        if not key.strip("|") or not code:
            messages.append(f"{code or '空代號'} 資料不完整，未寫入推薦清單")
            continue
        if key in merged_map:
            messages.append(f"{code} {name} 今日同推薦模式已在 10_推薦清單，略過")
            continue
        merged_map[key] = row
        added += 1
        messages.append(f"{code} {name} 已加入 10_推薦清單")

    merged_rows = list(merged_map.values())

    local_ok, local_msg = _safe_json_write_local(GODPICK_LIST_FILE, merged_rows)
    github_ok, github_msg2 = _write_json_to_github_path(GODPICK_LIST_FILE, merged_rows)

    messages.append(local_msg)
    messages.append(github_msg2 if github_ok else f"GitHub 同步略過/失敗：{github_msg2 or github_msg}")

    return added, messages



# =========================================================
# 歷史資料 / 指標
# =========================================================
def _prepare_history_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    temp = df.copy()
    if "日期" not in temp.columns:
        possible_date = [c for c in temp.columns if str(c).lower() in {"date", "日期"}]
        if possible_date:
            temp = temp.rename(columns={possible_date[0]: "日期"})
        else:
            return pd.DataFrame()

    temp["日期"] = pd.to_datetime(temp["日期"], errors="coerce")
    temp = temp.dropna(subset=["日期"]).sort_values("日期").drop_duplicates("日期", keep="last").reset_index(drop=True)

    rename_map = {}
    for c in temp.columns:
        cs = str(c).lower()
        if cs == "open":
            rename_map[c] = "開盤價"
        elif cs == "high":
            rename_map[c] = "最高價"
        elif cs == "low":
            rename_map[c] = "最低價"
        elif cs == "close":
            rename_map[c] = "收盤價"
        elif cs == "volume":
            rename_map[c] = "成交股數"
    temp = temp.rename(columns=rename_map)

    for col in ["成交股數", "成交金額", "開盤價", "最高價", "最低價", "收盤價", "成交筆數"]:
        if col in temp.columns:
            temp[col] = pd.to_numeric(temp[col], errors="coerce")

    if "收盤價" not in temp.columns:
        return pd.DataFrame()

    temp = temp.dropna(subset=["收盤價"]).copy()
    # 供應商偶爾會回傳 0 元或重複日期的占位列；這些列不可作為最新價。
    temp = temp[pd.to_numeric(temp["收盤價"], errors="coerce") > 0].copy()
    if temp.empty:
        return pd.DataFrame()

    close = temp["收盤價"]
    high = temp["最高價"] if "最高價" in temp.columns else close
    low = temp["最低價"] if "最低價" in temp.columns else close
    vol = pd.to_numeric(temp["成交股數"], errors="coerce") if "成交股數" in temp.columns else pd.Series(index=temp.index, dtype=float)

    for n in [5, 10, 20, 60, 120, 240]:
        temp[f"MA{n}"] = close.rolling(n).mean()

    low_9 = low.rolling(9).min()
    high_9 = high.rolling(9).max()
    rsv = (close - low_9) / (high_9 - low_9).replace(0, pd.NA) * 100
    temp["K"] = rsv.ewm(alpha=1 / 3, adjust=False).mean()
    temp["D"] = temp["K"].ewm(alpha=1 / 3, adjust=False).mean()
    temp["J"] = 3 * temp["K"] - 2 * temp["D"]

    ema12 = close.ewm(span=12, adjust=False).mean()
    ema26 = close.ewm(span=26, adjust=False).mean()
    temp["DIF"] = ema12 - ema26
    temp["DEA"] = temp["DIF"].ewm(span=9, adjust=False).mean()
    temp["MACD_HIST"] = temp["DIF"] - temp["DEA"]

    prev_close = close.shift(1)
    tr1 = high - low
    tr2 = (high - prev_close).abs()
    tr3 = (low - prev_close).abs()
    tr = pd.concat([tr1, tr2, tr3], axis=1).max(axis=1)
    temp["ATR14"] = tr.rolling(14).mean()

    temp["VOL5"] = vol.rolling(5).mean()
    temp["VOL20"] = vol.rolling(20).mean()
    temp["RET5"] = close.pct_change(5) * 100
    temp["RET20"] = close.pct_change(20) * 100
    temp["RET60"] = close.pct_change(60) * 100
    temp["RET120"] = close.pct_change(120) * 100
    temp["UP_DAY"] = (close > close.shift(1)).astype(float)
    temp["MA20_SLOPE"] = temp["MA20"].diff(3)
    temp["MA60_SLOPE"] = temp["MA60"].diff(3)

    return temp


def _get_history_smart(stock_no: str, stock_name: str, market_type: str, start_date: date, end_date: date) -> tuple[pd.DataFrame, str, dict[str, Any]]:
    primary = _safe_str(market_type) or "上市"
    attempt_summary: list[dict[str, Any]] = []

    # V34 加速重點：只呼叫 get_history_data 一次。
    # utils.get_history_data 內部已經包含：
    # 1. 本機 disk cache
    # 2. Yahoo 高速日線
    # 3. TWSE / TPEx 官方 fallback
    # 4. 上市 / 上櫃市場別候選
    # 原本外層又再重跑上市/上櫃/興櫃，會讓同一檔股票重複下載 2~4 輪，
    # 是畫面卡在 1/100、平均每檔 1 分鐘以上的主要原因之一。
    try:
        df = get_history_data(
            stock_no=stock_no,
            stock_name=stock_name,
            market_type=primary,
            start_date=start_date,
            end_date=end_date,
        )
    except TypeError:
        try:
            df = get_history_data(
                stock_no=stock_no,
                stock_name=stock_name,
                market_type=primary,
                start_dt=start_date,
                end_dt=end_date,
            )
        except Exception as e1:
            try:
                df = get_history_data(code=stock_no, start_date=start_date, end_date=end_date)
            except Exception as e2:
                attempt_summary.append({
                    "market_type": primary,
                    "rows": 0,
                    "source": "history_fetch_exception",
                    "error": f"{e1} / fallback: {e2}",
                })
                df = pd.DataFrame()
    except Exception as e:
        attempt_summary.append({
            "market_type": primary,
            "rows": 0,
            "source": "history_fetch_exception",
            "error": str(e),
        })
        df = pd.DataFrame()

    prepared_df = _prepare_history_df(df)
    if not prepared_df.empty:
        history_debug = {
            "ok": True,
            "stock_no": stock_no,
            "stock_name": stock_name,
            "used_market": primary,
            "attempts": attempt_summary + [{
                "market_type": primary,
                "rows": int(len(prepared_df)),
                "source": "history_fetch_ok_v34_single_call",
                "error": "",
            }],
            "rows": len(prepared_df),
        }
        return prepared_df, primary, history_debug

    attempt_summary.append({
        "market_type": primary,
        "rows": 0,
        "source": "history_fetch_empty_v34_single_call",
        "error": "",
    })

    # V180：禁止外層把完整歷史資料管線再跑一輪。
    # get_history_data 本身已包含 HTTP Retry、Yahoo query1/query2、TWSE/TPEx 當月補尾與完整月份 fallback。
    # 舊版在所有來源都失敗後又 sleep + 再呼叫 get_history_data，等於慢股/斷網時把最壞耗時放大近 2 倍。
    # no_history 仍屬 retryable，使用者可按「接續上次掃描」只補失敗股，不會因加速而永久漏股。
    if V180_DISABLE_DUPLICATE_HISTORY_RETRY:
        attempt_summary.append({
            "market_type": primary,
            "rows": 0,
            "source": "v180_no_duplicate_full_pipeline_retry",
            "error": "完整資料源已跑過；留待斷點續掃補抓",
        })
    else:
        try:
            delay_seed = sum(ord(ch) for ch in str(stock_no)) % 7
            time.sleep(0.12 + delay_seed * 0.025)
            retry_df = get_history_data(
                stock_no=stock_no,
                stock_name=stock_name,
                market_type=primary,
                start_date=start_date,
                end_date=end_date,
            )
            retry_prepared = _prepare_history_df(retry_df)
            if not retry_prepared.empty:
                attempt_summary.append({
                    "market_type": primary,
                    "rows": int(len(retry_prepared)),
                    "source": "history_fetch_ok_compat_retry",
                    "error": "",
                })
                return retry_prepared, primary, {
                    "ok": True, "stock_no": stock_no, "stock_name": stock_name,
                    "used_market": primary, "attempts": attempt_summary, "rows": len(retry_prepared),
                }
        except Exception as retry_error:
            attempt_summary.append({
                "market_type": primary, "rows": 0,
                "source": "history_fetch_retry_exception_compat", "error": str(retry_error),
            })

    # V35：掃描時失敗股票不再立即跑慢速 debug。
    # 原本 no_history 會再跑 get_history_data_debug，等於失敗股票又多打一輪官方資料源，
    # 在大量掃描時會造成 1/100 卡很久。若需要詳細診斷，可把 HISTORY_DEBUG_ON_FAIL 改 True。
    debug_attempts = attempt_summary.copy()
    if HISTORY_DEBUG_ON_FAIL and callable(get_history_data_debug):
        try:
            debug_info = get_history_data_debug(
                stock_no=stock_no,
                stock_name=stock_name,
                market_type=primary,
                start_date=start_date,
                end_date=end_date,
            )
            debug_attempts.append({
                "market_type": _safe_str(debug_info.get("market_type")) or primary,
                "rows": int(debug_info.get("rows", 0) or 0),
                "source": _safe_str(debug_info.get("source")) or "history_debug_v35_optional",
                "error": _safe_str(debug_info.get("error")),
            })
        except Exception as e:
            debug_attempts.append({
                "market_type": primary,
                "rows": 0,
                "source": "history_debug_exception",
                "error": str(e),
            })

    return pd.DataFrame(), primary, {
        "ok": False,
        "stock_no": stock_no,
        "stock_name": stock_name,
        "used_market": primary,
        "attempts": debug_attempts,
        "rows": 0,
    }


# =========================================================
# 計分
# =========================================================
def _build_prelaunch_scores(df: pd.DataFrame, signal_snapshot: dict, sr_snapshot: dict, radar: dict) -> dict[str, Any]:
    if df is None or df.empty:
        return {
            "起漲前兆分數": 0.0,
            "均線轉強分": 0.0,
            "量能啟動分": 0.0,
            "突破準備分": 0.0,
            "動能翻多分": 0.0,
            "支撐防守分": 0.0,
        }

    last = df.iloc[-1]
    prev = df.iloc[-2] if len(df) >= 2 else last

    close_now = _safe_float(last.get("收盤價"))
    ma5 = _safe_float(last.get("MA5"))
    ma10 = _safe_float(last.get("MA10"))
    ma20 = _safe_float(last.get("MA20"))
    ma60 = _safe_float(last.get("MA60"))
    ma20_slope = _safe_float(last.get("MA20_SLOPE"), 0) or 0
    vol5 = _safe_float(last.get("VOL5"))
    vol20 = _safe_float(last.get("VOL20"))
    ret5 = _safe_float(last.get("RET5"), 0) or 0
    k_now = _safe_float(last.get("K"))
    d_now = _safe_float(last.get("D"))
    k_prev = _safe_float(prev.get("K"))
    d_prev = _safe_float(prev.get("D"))
    hist_now = _safe_float(last.get("MACD_HIST"))
    hist_prev = _safe_float(prev.get("MACD_HIST"))
    res20 = _safe_float(sr_snapshot.get("res_20"))
    sup20 = _safe_float(sr_snapshot.get("sup_20"))

    trend_score = 0.0
    if close_now is not None and ma20 is not None and close_now >= ma20:
        trend_score += 25
    if close_now is not None and ma60 is not None and close_now >= ma60:
        trend_score += 18
    if ma5 is not None and ma10 is not None and ma5 >= ma10:
        trend_score += 18
    if ma20_slope > 0:
        trend_score += 22
    trend_score = _score_clip(trend_score)

    volume_score = 0.0
    if vol5 not in [None, 0] and vol20 not in [None, 0]:
        ratio = vol5 / vol20
        if ratio >= 1.8:
            volume_score = 85
        elif ratio >= 1.4:
            volume_score = 72
        elif ratio >= 1.1:
            volume_score = 60
        elif ratio >= 0.9:
            volume_score = 45
        else:
            volume_score = 25

    breakout_score = 0.0
    if close_now is not None and res20 not in [None, 0]:
        dist = ((res20 - close_now) / res20) * 100
        if 0 <= dist <= 2:
            breakout_score = 90
        elif 2 < dist <= 5:
            breakout_score = 72
        elif 5 < dist <= 8:
            breakout_score = 55
        elif dist < 0:
            breakout_score = 60
        else:
            breakout_score = 30

    momentum_score = 0.0
    if k_prev is not None and d_prev is not None and k_now is not None and d_now is not None:
        if k_prev <= d_prev and k_now > d_now:
            momentum_score += 45
        elif k_now > d_now:
            momentum_score += 28
    if hist_now is not None:
        if hist_prev is not None and hist_prev <= 0 < hist_now:
            momentum_score += 35
        elif hist_now > 0:
            momentum_score += 20
    radar_m = _safe_float(radar.get("momentum"), 50) or 50
    momentum_score += radar_m * 0.2
    momentum_score = _score_clip(momentum_score)

    support_score = 0.0
    if close_now is not None and sup20 not in [None, 0]:
        dist_sup = ((close_now - sup20) / sup20) * 100
        if 0 <= dist_sup <= 2:
            support_score = 85
        elif 2 < dist_sup <= 5:
            support_score = 70
        elif 5 < dist_sup <= 8:
            support_score = 55
        elif dist_sup < 0:
            support_score = 20
        else:
            support_score = 40

    if ret5 > 12:
        breakout_score -= 15
    if ret5 > 20:
        breakout_score -= 25

    total = _avg_safe([trend_score, volume_score, breakout_score, momentum_score, support_score], 0)
    return {
        "起漲前兆分數": _score_clip(total),
        "均線轉強分": _score_clip(trend_score),
        "量能啟動分": _score_clip(volume_score),
        "突破準備分": _score_clip(breakout_score),
        "動能翻多分": _score_clip(momentum_score),
        "支撐防守分": _score_clip(support_score),
    }


def _build_risk_filter(df: pd.DataFrame, signal_snapshot: dict, sr_snapshot: dict, strictness: str) -> dict[str, Any]:
    if df is None or df.empty:
        return {"是否通過風險過濾": False, "風險分數": 0.0, "淘汰原因": "無歷史資料"}

    last = df.iloc[-1]
    close_now = _safe_float(last.get("收盤價"))
    ma20 = _safe_float(last.get("MA20"))
    ma60 = _safe_float(last.get("MA60"))
    atr14 = _safe_float(last.get("ATR14"))
    vol20 = _safe_float(last.get("VOL20"))
    ret20 = _safe_float(last.get("RET20"), 0) or 0
    pressure_dist = None
    res20 = _safe_float(sr_snapshot.get("res_20"))
    if close_now not in [None] and res20 not in [None, 0]:
        pressure_dist = ((res20 - close_now) / res20) * 100

    rules = {
        "寬鬆": {"min_days": 60, "min_vol20": 300000, "max_atr_pct": 11.0, "max_ret20": 35.0},
        "標準": {"min_days": 90, "min_vol20": 800000, "max_atr_pct": 8.5, "max_ret20": 28.0},
        "嚴格": {"min_days": 120, "min_vol20": 1200000, "max_atr_pct": 6.5, "max_ret20": 22.0},
    }
    cfg = rules.get(_safe_str(strictness), rules["標準"])

    reasons = []
    risk_score = 100.0

    if len(df) < cfg["min_days"]:
        reasons.append(f"歷史資料不足{cfg['min_days']}天")
        risk_score -= 30
    if vol20 not in [None] and vol20 < cfg["min_vol20"]:
        reasons.append("量能不足")
        risk_score -= 22
    if close_now not in [None] and atr14 not in [None]:
        atr_pct = atr14 / close_now * 100 if close_now != 0 else 999
        if atr_pct > cfg["max_atr_pct"]:
            reasons.append("波動過大")
            risk_score -= 18
    if close_now not in [None] and ma20 not in [None] and ma60 not in [None]:
        if close_now < ma20 and close_now < ma60:
            reasons.append("中期結構偏弱")
            risk_score -= 20
    if ret20 > cfg["max_ret20"]:
        reasons.append("近20日漲幅過大")
        risk_score -= 16

    if pressure_dist is not None and pressure_dist < 0:
        risk_score -= 4
    elif pressure_dist is not None and pressure_dist > 10:
        risk_score -= 8

    signal_score = _safe_float(signal_snapshot.get("score"), 0) or 0
    risk_score += max(min(signal_score * 1.8, 12), -12)
    risk_score = _score_clip(risk_score)

    passed = len(reasons) == 0 or risk_score >= 55
    return {
        "是否通過風險過濾": passed,
        "風險分數": risk_score,
        "淘汰原因": "；".join(reasons) if reasons else "",
    }


def _build_trade_feasibility(df: pd.DataFrame, sr_snapshot: dict, signal_snapshot: dict) -> dict[str, Any]:
    if df is None or df.empty:
        return {
            "交易可行分數": 0.0,
            "追價風險分數": 0.0,
            "拉回買點分數": 0.0,
            "突破買點分數": 0.0,
            "風險報酬評級": "—",
        }

    last = df.iloc[-1]
    close_now = _safe_float(last.get("收盤價"), 0) or 0
    atr14 = _safe_float(last.get("ATR14"), 0) or max(close_now * 0.03, 1.0)
    ma20 = _safe_float(last.get("MA20"))
    res20 = _safe_float(sr_snapshot.get("res_20"))
    sup20 = _safe_float(sr_snapshot.get("sup_20"))

    pullback_buy = ma20 if ma20 is not None else (sup20 if sup20 is not None else close_now)
    breakout_buy = res20 if res20 is not None else close_now
    stop_price = sup20 if sup20 is not None else max(close_now - atr14, 0)
    target_1 = res20 if res20 is not None and res20 > close_now else close_now + atr14 * 1.5
    target_2 = target_1 + atr14 * 1.2

    def _rr(entry: float, stop: float, target: float) -> float:
        risk = entry - stop
        reward = target - entry
        if risk <= 0:
            return 0.0
        return reward / risk

    rr_pullback = _rr(pullback_buy, stop_price, target_1) if pullback_buy and stop_price is not None and target_1 else 0.0
    rr_breakout = _rr(breakout_buy, stop_price, target_2) if breakout_buy and stop_price is not None and target_2 else 0.0

    pullback_score = 25 + min(rr_pullback * 28, 45)
    breakout_score = 25 + min(rr_breakout * 22, 40)

    chase_risk = 0.0
    if ma20 not in [None, 0] and close_now not in [None]:
        bias = ((close_now - ma20) / ma20) * 100
        if bias >= 12:
            chase_risk = 88
        elif bias >= 8:
            chase_risk = 72
        elif bias >= 5:
            chase_risk = 58
        else:
            chase_risk = 35

    signal_score = _safe_float(signal_snapshot.get("score"), 0) or 0
    feasibility = _avg_safe(
        [_score_clip(pullback_score), _score_clip(breakout_score), _score_clip(100 - chase_risk), 50 + signal_score * 5],
        0,
    )

    if feasibility >= 80:
        rr_grade = "A"
    elif feasibility >= 68:
        rr_grade = "B"
    elif feasibility >= 55:
        rr_grade = "C"
    else:
        rr_grade = "D"

    return {
        "交易可行分數": _score_clip(feasibility),
        "追價風險分數": _score_clip(chase_risk),
        "拉回買點分數": _score_clip(pullback_score),
        "突破買點分數": _score_clip(breakout_score),
        "風險報酬評級": rr_grade,
    }


def _build_mode_score(
    mode: str,
    technical_score: float,
    prelaunch_score: float,
    category_heat_score: float,
    factor_score: float,
    trade_score: float,
    leader_advantage: float,
) -> tuple[float, str]:
    mode = _safe_str(mode)

    if mode == "飆股模式":
        total = prelaunch_score * 0.35 + technical_score * 0.25 + category_heat_score * 0.20 + factor_score * 0.10 + trade_score * 0.10
        tag = "突破前夜 / 起漲優先"
    elif mode == "波段模式":
        total = technical_score * 0.30 + category_heat_score * 0.25 + factor_score * 0.20 + trade_score * 0.15 + prelaunch_score * 0.10
        tag = "趨勢延續 / 波段優先"
    elif mode == "領頭羊模式":
        total = leader_advantage * 0.30 + category_heat_score * 0.25 + technical_score * 0.20 + prelaunch_score * 0.15 + factor_score * 0.10
        tag = "類股領先 / 龍頭優先"
    else:
        total = technical_score * 0.30 + prelaunch_score * 0.20 + category_heat_score * 0.20 + factor_score * 0.15 + trade_score * 0.15
        tag = "綜合推薦"

    return _score_clip(total), tag


def _build_auto_factor_scores(df: pd.DataFrame, signal_snapshot: dict, sr_snapshot: dict, radar: dict) -> dict[str, Any]:
    last = df.iloc[-1]
    close_now = _safe_float(last.get("收盤價"))
    ma20 = _safe_float(last.get("MA20"))
    ma60 = _safe_float(last.get("MA60"))
    ma120 = _safe_float(last.get("MA120"))
    atr14 = _safe_float(last.get("ATR14"))
    vol5 = _safe_float(last.get("VOL5"))
    vol20 = _safe_float(last.get("VOL20"))
    ret20 = _safe_float(last.get("RET20"))
    ret60 = _safe_float(last.get("RET60"))

    signal_score = _safe_float(signal_snapshot.get("score"), 0) or 0
    radar_trend = _safe_float(radar.get("trend"), 50) or 50
    radar_momentum = _safe_float(radar.get("momentum"), 50) or 50
    radar_volume = _safe_float(radar.get("volume"), 50) or 50
    radar_structure = _safe_float(radar.get("structure"), 50) or 50
    sup20 = _safe_float(sr_snapshot.get("sup_20"))

    eps_proxy = 50.0
    if close_now not in [None, 0]:
        trend_bonus = 0.0
        if ma120 is not None and close_now > ma120:
            trend_bonus += 18
        if ma60 is not None and close_now > ma60:
            trend_bonus += 12
        if ma20 is not None and close_now > ma20:
            trend_bonus += 8

        vol_penalty = 0.0
        if atr14 is not None:
            atr_pct = atr14 / close_now * 100
            if atr_pct <= 2.5:
                vol_penalty = 0
            elif atr_pct <= 5:
                vol_penalty = 6
            else:
                vol_penalty = 12

        eps_proxy = _score_clip(30 + trend_bonus + radar_structure * 0.25 + radar_trend * 0.20 - vol_penalty)

    revenue_proxy = _score_clip(25 + (_safe_float(ret20, 0) or 0) * 0.9 + (_safe_float(ret60, 0) or 0) * 0.35 + radar_momentum * 0.30 + radar_volume * 0.20)
    profit_proxy = _score_clip(30 + signal_score * 6 + radar_trend * 0.28 + radar_structure * 0.22 + (_safe_float(ret60, 0) or 0) * 0.35)

    lock_proxy = 45.0
    if close_now not in [None, 0]:
        vol_ratio = None
        if vol5 not in [None, 0] and vol20 not in [None, 0]:
            vol_ratio = vol5 / vol20

        atr_pct = None
        if atr14 is not None:
            atr_pct = atr14 / close_now * 100

        lock_bonus = 0.0
        if ma20 is not None and close_now >= ma20:
            lock_bonus += 12
        if sup20 is not None and close_now >= sup20:
            lock_bonus += 10
        if vol_ratio is not None:
            if 0.7 <= vol_ratio <= 1.15:
                lock_bonus += 12
            elif vol_ratio < 0.7:
                lock_bonus += 8
        if atr_pct is not None:
            if atr_pct <= 2.5:
                lock_bonus += 14
            elif atr_pct <= 4:
                lock_bonus += 8

        lock_proxy = _score_clip(20 + lock_bonus + radar_structure * 0.24)

    recent = df.tail(5).copy()
    up_days_5 = int(recent["UP_DAY"].sum()) if "UP_DAY" in recent.columns else 0
    inst_proxy = _score_clip(20 + up_days_5 * 10 + signal_score * 5 + radar_momentum * 0.25 + radar_volume * 0.20)

    factor_summary = (
        f"EPS代理 {format_number(eps_proxy,1)} / "
        f"營收動能代理 {format_number(revenue_proxy,1)} / "
        f"獲利代理 {format_number(profit_proxy,1)} / "
        f"大戶鎖碼代理 {format_number(lock_proxy,1)} / "
        f"法人連買代理 {format_number(inst_proxy,1)}"
    )

    return {
        "auto_factor_total": _avg_safe([eps_proxy, revenue_proxy, profit_proxy, lock_proxy, inst_proxy], 0),
        "eps_proxy": eps_proxy,
        "revenue_proxy": revenue_proxy,
        "profit_proxy": profit_proxy,
        "lock_proxy": lock_proxy,
        "inst_proxy": inst_proxy,
        "factor_summary": factor_summary,
    }


def _build_trade_plan(df: pd.DataFrame, sr_snapshot: dict, signal_snapshot: dict) -> dict[str, Any]:
    last = df.iloc[-1]
    close_now = _safe_float(last.get("收盤價"), 0) or 0
    atr14 = _safe_float(last.get("ATR14"), 0) or max(close_now * 0.03, 1.0)
    ma20 = _safe_float(last.get("MA20"))
    res20 = _safe_float(sr_snapshot.get("res_20"))
    sup20 = _safe_float(sr_snapshot.get("sup_20"))
    res60 = _safe_float(sr_snapshot.get("res_60"))
    score = _safe_float(signal_snapshot.get("score"), 0) or 0

    breakout_buy = res20 if res20 is not None else close_now
    pullback_buy = ma20 if ma20 is not None else (sup20 if sup20 is not None else close_now)
    stop_price = sup20 if sup20 is not None else max(close_now - atr14, 0)
    sell_target_1 = res20 if res20 is not None and res20 > close_now else close_now + atr14 * 1.5
    sell_target_2 = res60 if res60 is not None and res60 > sell_target_1 else sell_target_1 + atr14 * 1.2

    if score >= 4:
        launch_tag = "強勢起漲候選"
    elif score >= 2:
        launch_tag = "偏多轉強候選"
    elif score <= -2:
        launch_tag = "不建議追價"
    else:
        launch_tag = "等待表態"

    def _rr(entry: float, stop: float, target: float) -> str:
        risk = entry - stop
        reward = target - entry
        if risk <= 0:
            return "—"
        return f"1 : {reward / risk:.2f}"

    rr1 = _rr(pullback_buy, stop_price, sell_target_1) if pullback_buy and stop_price is not None and sell_target_1 else "—"
    rr2 = _rr(breakout_buy, stop_price, sell_target_2) if breakout_buy and stop_price is not None and sell_target_2 else "—"

    return {
        "launch_tag": launch_tag,
        "breakout_buy": breakout_buy,
        "pullback_buy": pullback_buy,
        "stop_price": stop_price,
        "sell_target_1": sell_target_1,
        "sell_target_2": sell_target_2,
        "rr1": rr1,
        "rr2": rr2,
    }



def _annotate_kline_freshness(frame: pd.DataFrame) -> pd.DataFrame:
    """以本輪全市場共同最新交易日驗證每檔 K 線日期。

    這種做法不需要硬編台灣休市日：春節、颱風停市或週末時，全市場的共同
    最新交易日本身就是正確基準；個別股票若仍停在更早日期，才標記為落後。
    """
    if frame is None or frame.empty:
        return frame
    out = frame.copy()
    if "K線最後交易日" not in out.columns:
        out["K線最後交易日"] = ""
    dates = pd.to_datetime(out["K線最後交易日"], errors="coerce").dt.normalize()
    stock_latest = dates.dropna().max() if dates.notna().any() else pd.NaT

    # 全市場股票若剛好都命中同一份舊快取，只用 max(K線日期) 仍會把舊資料
    # 誤判成最新。再與 01 大盤走勢已保存的市場資料日期交叉驗證。
    macro_latest = pd.NaT
    try:
        macro_ref = _load_latest_macro_reference()
        macro_text = _safe_str(macro_ref.get("大盤資料日期")) if isinstance(macro_ref, dict) else ""
        macro_latest = pd.to_datetime(macro_text, errors="coerce")
        if not pd.isna(macro_latest):
            macro_latest = macro_latest.normalize()
    except Exception:
        macro_latest = pd.NaT

    reference_dates = [d for d in [stock_latest, macro_latest] if not pd.isna(d)]
    market_latest = max(reference_dates) if reference_dates else pd.NaT
    reference_source = "K線全市場共同日期"
    if not pd.isna(macro_latest) and (pd.isna(stock_latest) or macro_latest > stock_latest):
        reference_source = "01大盤資料日期"

    lags: list[int] = []
    states: list[str] = []
    for dt in dates:
        if pd.isna(dt) or pd.isna(market_latest):
            lags.append(999)
            states.append("未知｜無法驗證K線日期")
            continue
        lag = max(len(pd.bdate_range(dt + pd.Timedelta(days=1), market_latest)), 0)
        lags.append(int(lag))
        states.append("最新交易日" if lag == 0 else f"落後{lag}個交易日")
    out["K線落後交易日"] = lags
    out["K線資料新鮮度"] = states
    out["本輪市場最新交易日"] = market_latest.strftime("%Y-%m-%d") if not pd.isna(market_latest) else ""
    out["K線日期驗證基準"] = reference_source if not pd.isna(market_latest) else "無法驗證"
    return out

def _analyze_stock_bundle(stock_no: str, stock_name: str, market_type: str, start_dt: date, end_dt: date, risk_strictness: str) -> dict[str, Any]:
    try:
        hist_df, used_market, history_debug = _get_history_smart(
            stock_no=stock_no,
            stock_name=stock_name,
            market_type=market_type,
            start_date=start_dt,
            end_date=end_dt,
        )
        if hist_df.empty:
            return {
                "ok": False,
                "error_stage": "history",
                "error_message": "抓不到歷史資料",
                "used_market": used_market,
                "history_debug": history_debug,
            }

        signal_snapshot = compute_signal_snapshot(hist_df)
        sr_snapshot = compute_support_resistance_snapshot(hist_df)
        radar = _ensure_radar_dict(compute_radar_scores(hist_df))
        auto_factor = _build_auto_factor_scores(hist_df, signal_snapshot, sr_snapshot, radar)
        trade_plan = _build_trade_plan(hist_df, sr_snapshot, signal_snapshot)
        pattern_info = _build_pattern_breakout_scores(hist_df, sr_snapshot, signal_snapshot)
        burst_info = _build_burst_scores(hist_df)
        opportunity_info = _build_opportunity_scores(hist_df, sr_snapshot, signal_snapshot, radar)
        prelaunch = _build_prelaunch_scores(hist_df, signal_snapshot, sr_snapshot, radar)
        risk_filter = _build_risk_filter(hist_df, signal_snapshot, sr_snapshot, risk_strictness)
        trade_feasibility = _build_trade_feasibility(hist_df, sr_snapshot, signal_snapshot)
        entry_decision = _build_entry_decision_scores(hist_df, sr_snapshot, opportunity_info, trade_plan, trade_feasibility)

        last = hist_df.iloc[-1]
        first = hist_df.iloc[0]
        last_kline_date = ""
        try:
            last_kline_date = pd.to_datetime(last.get("日期"), errors="coerce").strftime("%Y-%m-%d")
        except Exception:
            last_kline_date = ""

        close_now = _safe_float(last.get("收盤價"))
        close_first = _safe_float(first.get("收盤價"))

        def _history_numeric_series(names: list[str]) -> pd.Series:
            for col in names:
                if col not in hist_df.columns:
                    continue
                values = pd.to_numeric(hist_df[col], errors="coerce").replace([np.inf, -np.inf], np.nan)
                if values.notna().any():
                    return values.astype(float)
            return pd.Series(dtype="float64")

        # Providers use different aliases. Missing volume must not be silently
        # treated as zero liquidity, which previously excluded most candidates.
        volume_series = _history_numeric_series(["成交股數", "成交量", "Volume", "volume", "總量", "成交量(股)"])
        amount_series = _history_numeric_series(["成交金額", "成交額", "成交值", "Amount", "amount", "成交金額(元)"])
        positive_volume = volume_series.where(volume_series > 0).dropna() if not volume_series.empty else pd.Series(dtype="float64")
        positive_amount = amount_series.where(amount_series > 0).dropna() if not amount_series.empty else pd.Series(dtype="float64")
        # 週末、休市日或供應商最後一列常會是 0；使用最近一筆有效交易日，
        # 不可把 20 日均成交額正常的股票誤判成零成交、低流動性。
        volume_last = _safe_float(positive_volume.iloc[-1]) if not positive_volume.empty else None
        volume_5 = _safe_float(last.get("VOL5"))
        volume_20 = _safe_float(last.get("VOL20"))
        if volume_5 in [None, 0] and not positive_volume.empty:
            volume_5 = _safe_float(positive_volume.tail(5).mean())
        if volume_20 in [None, 0] and not positive_volume.empty:
            volume_20 = _safe_float(positive_volume.tail(20).mean())
        # 「均量比」必須反映最新一日點火量，而不是 5 日均量 / 20 日均量。
        # 後者反應太慢，會讓剛放量突破或漲停的股票在前置篩選階段被漏掉。
        volume_5_20_ratio = None
        if volume_5 not in [None, 0] and volume_20 not in [None, 0]:
            volume_5_20_ratio = volume_5 / volume_20
        daily_volume_ratio = None
        if not positive_volume.empty:
            latest_valid_volume = _safe_float(positive_volume.iloc[-1], 0) or 0
            prior20_volume = positive_volume.iloc[:-1].tail(20)
            prior20_mean = _safe_float(prior20_volume.mean(), 0) if not prior20_volume.empty else 0
            if latest_valid_volume > 0 and prior20_mean not in [None, 0]:
                daily_volume_ratio = latest_valid_volume / prior20_mean
        volume_ratio = daily_volume_ratio if daily_volume_ratio not in [None, 0] else volume_5_20_ratio

        turnover_m = None
        avg20_turnover_m = None
        liquidity_source = "缺少成交額/成交量"
        # Direct turnover has the most reliable unit. Only fall back to
        # price × shares when no direct turnover field is available.
        if not positive_amount.empty:
            turnover_m = (_safe_float(positive_amount.iloc[-1], 0) or 0) / 1_000_000
            avg20_turnover_m = (_safe_float(positive_amount.tail(20).mean(), 0) or 0) / 1_000_000
            liquidity_source = "最近有效交易日K線成交金額"
        elif close_now not in [None, 0] and volume_last not in [None, 0]:
            turnover_m = close_now * volume_last / 1_000_000
            if volume_20 not in [None, 0]:
                avg20_turnover_m = close_now * volume_20 / 1_000_000
            liquidity_source = "收盤價×成交股數回推"
        liquidity_status = "有效" if any((_safe_float(v, 0) or 0) > 0 for v in [turnover_m, avg20_turnover_m, volume_last, volume_20]) else "缺失"
        ma20_now = _safe_float(last.get("MA20"))
        ma60_now = _safe_float(last.get("MA60"))
        close_vs_ma20_pct = None
        close_vs_ma60_pct = None
        if close_now not in [None, 0] and ma20_now not in [None, 0]:
            close_vs_ma20_pct = (close_now - ma20_now) / ma20_now * 100
        if close_now not in [None, 0] and ma60_now not in [None, 0]:
            close_vs_ma60_pct = (close_now - ma60_now) / ma60_now * 100
        period_pct = None
        if close_now is not None and close_first not in [None, 0]:
            period_pct = ((close_now / close_first) - 1) * 100

        # 資金流實戰層需要明確知道近期漲幅，不再只看長區間漲跌。
        ret5_now = _safe_float(last.get("RET5"), 0) or 0
        ret20_now = _safe_float(last.get("RET20"), 0) or 0
        ret60_now = _safe_float(last.get("RET60"), 0) or 0

        # Phase 8.8：補上「當日點火 / 漲停續強」專用量價欄位。
        # 過去只看 5/20 日均線與靜態 RR，會把剛出現價格發現、放量突破的
        # 強勢股在 limit-up / explosive engines 執行前就刪掉。
        close_series = _history_numeric_series(["收盤價", "Close", "close"])
        open_series = _history_numeric_series(["開盤價", "Open", "open"])
        high_series = _history_numeric_series(["最高價", "High", "high"])
        low_series = _history_numeric_series(["最低價", "Low", "low"])
        valid_close = close_series.dropna()
        prior_close = _safe_float(valid_close.iloc[-2], 0) if len(valid_close) >= 2 else 0
        open_now = _safe_float(open_series.dropna().iloc[-1], 0) if not open_series.dropna().empty else (_safe_float(last.get("開盤價"), 0) or close_now or 0)
        high_now = _safe_float(high_series.dropna().iloc[-1], 0) if not high_series.dropna().empty else (_safe_float(last.get("最高價"), 0) or close_now or 0)
        low_now = _safe_float(low_series.dropna().iloc[-1], 0) if not low_series.dropna().empty else (_safe_float(last.get("最低價"), 0) or close_now or 0)
        day_gain_pct = ((close_now / prior_close) - 1.0) * 100.0 if close_now not in [None, 0] and prior_close > 0 else 0.0
        opening_gap_pct = ((open_now / prior_close) - 1.0) * 100.0 if open_now > 0 and prior_close > 0 else 0.0
        day_range = max(0.0, (high_now or 0) - (low_now or 0))
        close_location_pct = ((close_now - low_now) / day_range * 100.0) if close_now not in [None, 0] and day_range > 0 else 50.0
        upper_shadow_pct = ((high_now - max(open_now, close_now)) / day_range * 100.0) if close_now not in [None, 0] and day_range > 0 else 0.0
        prior20_high = 0.0
        valid_high = high_series.dropna()
        if len(valid_high) >= 2:
            prior20_high = _safe_float(valid_high.iloc[:-1].tail(20).max(), 0) or 0
        breakout20_pct = ((close_now / prior20_high) - 1.0) * 100.0 if close_now not in [None, 0] and prior20_high > 0 else 0.0
        distance_to_20d_high_pct = max(0.0, ((prior20_high - close_now) / prior20_high) * 100.0) if close_now not in [None, 0] and prior20_high > 0 else 0.0
        consecutive_up_days = 0
        if len(valid_close) >= 2:
            vals = valid_close.tail(8).tolist()
            for i in range(len(vals) - 1, 0, -1):
                if vals[i] > vals[i - 1]:
                    consecutive_up_days += 1
                else:
                    break
        strong_close_flag = bool(close_location_pct >= 75 and day_gain_pct >= 2.0 and upper_shadow_pct <= 35)

        res20 = _safe_float(sr_snapshot.get("res_20"))
        sup20 = _safe_float(sr_snapshot.get("sup_20"))
        pressure_dist = None
        support_dist = None
        if close_now is not None and res20 not in [None, 0]:
            pressure_dist = ((res20 - close_now) / res20) * 100
        if close_now is not None and sup20 not in [None, 0]:
            support_dist = ((close_now - sup20) / sup20) * 100

        radar_avg = _avg_safe(
            [
                _safe_float(radar.get("trend")),
                _safe_float(radar.get("momentum")),
                _safe_float(radar.get("volume")),
                _safe_float(radar.get("position")),
                _safe_float(radar.get("structure")),
            ],
            50.0,
        )

        technical_score = _score_clip(
            (radar_avg * 0.55)
            + ((_safe_float(signal_snapshot.get("score"), 0) or 0) * 7.5)
            + ((_safe_float(period_pct, 0) or 0) * 0.18)
        )

        return {
            "ok": True,
            "used_market": used_market,
            "history_debug": history_debug,
            "signal_snapshot": signal_snapshot,
            "sr_snapshot": sr_snapshot,
            "radar": radar,
            "auto_factor": auto_factor,
            "trade_plan": trade_plan,
            "pattern_info": pattern_info,
            "burst_info": burst_info,
            "opportunity_info": opportunity_info,
            "prelaunch": prelaunch,
            "risk_filter": risk_filter,
            "trade_feasibility": trade_feasibility,
            "entry_decision": entry_decision,
            "close_now": close_now,
            "history_last_date": last_kline_date,
            "period_pct": period_pct,
            "ret5": ret5_now,
            "ret20": ret20_now,
            "ret60": ret60_now,
            "turnover_m": turnover_m,
            "avg20_turnover_m": avg20_turnover_m,
            "pressure_dist": pressure_dist,
            "support_dist": support_dist,
            "radar_avg": radar_avg,
            "technical_score": technical_score,
            "volume_last": volume_last,
            "volume_5": volume_5,
            "volume_20": volume_20,
            "volume_ratio": volume_ratio,
            "daily_volume_ratio": daily_volume_ratio,
            "volume_5_20_ratio": volume_5_20_ratio,
            "day_gain_pct": day_gain_pct,
            "opening_gap_pct": opening_gap_pct,
            "close_location_pct": close_location_pct,
            "upper_shadow_pct": upper_shadow_pct,
            "breakout20_pct": breakout20_pct,
            "distance_to_20d_high_pct": distance_to_20d_high_pct,
            "consecutive_up_days": consecutive_up_days,
            "strong_close_flag": strong_close_flag,
            "prior20_high": prior20_high,
            "liquidity_status": liquidity_status,
            "liquidity_source": liquidity_source,
            "close_vs_ma20_pct": close_vs_ma20_pct,
            "close_vs_ma60_pct": close_vs_ma60_pct,
        }
    except Exception as e:
        return {
            "ok": False,
            "error_stage": "analysis",
            "error_message": str(e),
            "used_market": _safe_str(market_type) or "未知",
            "history_debug": {},
        }


def _bundle_momentum_rescue_profile(bundle: dict[str, Any], market_type: str = "") -> dict[str, Any]:
    """前置篩選的強勢動能救援層。

    目的不是把漲停股直接宣稱為可買，而是避免它們在訊號 / RR / 過熱等
    傳統波段條件前置階段就被永久刪除。只要量價、流動性與收盤結構成立，
    先保留到後段飆股、漲停回放、主流資金與正式風控引擎，再決定是否列入
    R1-M 強勢動能條件雷達。
    """
    market = _safe_str(market_type or bundle.get("used_market")).replace(" ", "")
    day_gain = _safe_float(bundle.get("day_gain_pct"), 0) or 0
    close_loc = _safe_float(bundle.get("close_location_pct"), 50) or 50
    day_vol = _safe_float(bundle.get("daily_volume_ratio"), 0) or 0
    breakout20 = _safe_float(bundle.get("breakout20_pct"), -99) or -99
    upper_shadow = _safe_float(bundle.get("upper_shadow_pct"), 0) or 0
    turnover = max(_safe_float(bundle.get("turnover_m"), 0) or 0, _safe_float(bundle.get("avg20_turnover_m"), 0) or 0)
    ret5 = _safe_float(bundle.get("ret5"), 0) or 0
    close_ma20 = _safe_float(bundle.get("close_vs_ma20_pct"), 0) or 0

    gain_score = 98 if day_gain >= 9.0 else 92 if day_gain >= 6.5 else 82 if day_gain >= 4.0 else 68 if day_gain >= 2.0 else 35
    volume_score = 100 if day_vol >= 2.0 else 88 if day_vol >= 1.5 else 76 if day_vol >= 1.2 else 62 if day_vol >= 1.0 else 35
    breakout_score = 96 if breakout20 >= 0 else 84 if breakout20 >= -1.0 else 68 if breakout20 >= -3.0 else 35
    liquidity_score = 100 if turnover >= 2000 else 90 if turnover >= 800 else 78 if turnover >= 300 else 65 if turnover >= 150 else 48 if turnover >= 100 else 15
    score = gain_score * 0.25 + close_loc * 0.22 + volume_score * 0.20 + breakout_score * 0.18 + liquidity_score * 0.15
    if upper_shadow > 45:
        score -= 18
    elif upper_shadow > 35:
        score -= 8
    if ret5 > 35 or close_ma20 > 28:
        score -= 15
    score = round(max(0.0, min(100.0, score)), 1)

    reasons: list[str] = []
    if day_gain >= 9.0:
        reasons.append("接近漲停/漲停型強勢收盤")
    elif day_gain >= 4.0:
        reasons.append("單日價格動能明顯")
    if close_loc >= 80:
        reasons.append("收盤位於當日高檔")
    if day_vol >= 1.2:
        reasons.append(f"當日量比{day_vol:.2f}倍")
    if breakout20 >= 0:
        reasons.append("收盤突破前20日高點")
    elif breakout20 >= -1.0:
        reasons.append("接近前20日高點")
    if turnover >= 300:
        reasons.append("成交額具可交易性")

    eligible = bool(
        market not in {"興櫃", "Emerging"}
        and turnover >= 100
        and 2.5 <= day_gain <= 10.3
        and close_loc >= 68
        and (day_vol >= 1.15 or turnover >= 500)
        and (breakout20 >= -1.5 or day_gain >= 6.0)
        and upper_shadow <= 42
        and ret5 <= 38
        and close_ma20 <= 30
        and score >= 68
    )
    kind = "LIMITUP_MOMENTUM" if eligible and day_gain >= 8.5 else "BREAKOUT_MOMENTUM" if eligible else ""
    return {
        "eligible": eligible,
        "kind": kind,
        "score": score,
        "reason": "、".join(reasons) if reasons else "未達強勢動能救援條件",
    }


def _bundle_prebreakout_rescue_profile(bundle: dict[str, Any], market_type: str = "") -> dict[str, Any]:
    """前置篩選的強勢前兆救援層。

    與已經發動的 R1-M 不同，本層找的是「尚未大漲，但主流/族群、技術前兆、
    交易可行性與流動性已同步」的候選。它只負責讓股票通過前置資料管線，
    後段仍須由正式引擎判定為 R1-P；絕不因此直接升級成正式推薦。
    """
    market = _safe_str(market_type or bundle.get("used_market")).replace(" ", "")
    auto_factor = bundle.get("auto_factor", {}) or {}
    prelaunch = bundle.get("prelaunch", {}) or {}
    trade = bundle.get("trade_feasibility", {}) or {}
    opportunity = bundle.get("opportunity_info", {}) or {}
    signal = bundle.get("signal_snapshot", {}) or {}

    turnover = max(
        _safe_float(bundle.get("turnover_m"), 0) or 0,
        _safe_float(bundle.get("avg20_turnover_m"), 0) or 0,
    )
    technical = _safe_float(bundle.get("technical_score"), 0) or 0
    prelaunch_score = _safe_float(prelaunch.get("起漲前兆分數"), 0) or 0
    trade_score = _safe_float(trade.get("交易可行分數"), 0) or 0
    signal_score = _safe_float(signal.get("score"), 0) or 0
    opportunity_score = _safe_float(opportunity.get("機會股分數"), 0) or 0
    auto_total = _safe_float(auto_factor.get("auto_factor_total"), 0) or 0
    ret5 = _safe_float(bundle.get("ret5"), 0) or 0
    ret20 = _safe_float(bundle.get("ret20"), 0) or 0
    volume_ratio = max(
        _safe_float(bundle.get("daily_volume_ratio"), 0) or 0,
        _safe_float(bundle.get("volume_5_20_ratio"), 0) or 0,
        _safe_float(bundle.get("volume_ratio"), 0) or 0,
    )
    pressure_dist = _safe_float(bundle.get("pressure_dist"), 99) or 99
    support_dist = _safe_float(bundle.get("support_dist"), 99) or 99

    liquidity_score = 100 if turnover >= 2000 else 92 if turnover >= 800 else 82 if turnover >= 300 else 72 if turnover >= 150 else 58 if turnover >= 100 else 20
    proximity_score = 92 if 0 <= pressure_dist <= 3.0 else 78 if pressure_dist <= 5.0 else 62 if support_dist <= 3.5 else 35
    volume_score = 92 if volume_ratio >= 1.5 else 82 if volume_ratio >= 1.2 else 70 if volume_ratio >= 1.0 else 55 if volume_ratio >= 0.8 else 30
    score = (
        technical * 0.19 + prelaunch_score * 0.22 + trade_score * 0.12
        + signal_score * 0.10 + opportunity_score * 0.10 + auto_total * 0.08
        + liquidity_score * 0.09 + proximity_score * 0.06 + volume_score * 0.04
    )
    if ret5 > 22 or ret20 > 90:
        score -= 10
    if ret5 < -14 or ret20 < -25:
        score -= 10
    score = round(max(0.0, min(100.0, score)), 1)

    reasons: list[str] = []
    if prelaunch_score >= 62:
        reasons.append(f"起漲前兆{prelaunch_score:.0f}")
    if technical >= 62:
        reasons.append(f"技術結構{technical:.0f}")
    if trade_score >= 55:
        reasons.append(f"交易可行{trade_score:.0f}")
    if turnover >= 150:
        reasons.append("成交額具可交易性")
    if pressure_dist <= 5:
        reasons.append("接近壓力/突破觸發區")
    if volume_ratio >= 1.0:
        reasons.append(f"量能比{volume_ratio:.2f}")

    eligible = bool(
        market not in {"興櫃", "Emerging"}
        and turnover >= 100
        and score >= 60
        and max(prelaunch_score, technical, opportunity_score) >= 58
        and trade_score >= 42
        and signal_score >= 30
        and -14 <= ret5 <= 26
        and -25 <= ret20 <= 100
        and (pressure_dist <= 6.0 or support_dist <= 4.0 or volume_ratio >= 1.0)
    )
    return {
        "eligible": eligible,
        "kind": "PREBREAKOUT_LEADER" if eligible else "",
        "score": score,
        "reason": "、".join(reasons) if reasons else "未達強勢前兆救援條件",
    }


def _analyze_one_stock_for_recommend(
    item: dict[str, str],
    master_lookup: dict[str, dict[str, str]],
    start_dt: date,
    end_dt: date,
    min_signal_score: float,
    clean_categories: list[str],
    mode: str,
    risk_strictness: str,
    min_prelaunch_score: float,
    min_trade_score: float,
):
    code = _normalize_code(item.get("code"))
    manual_name = _safe_str(item.get("name"))
    manual_market = _safe_str(item.get("market"))
    manual_category = _normalize_category(item.get("category"))

    if not code:
        return {"status": "invalid_code", "code": "", "message": "股票代號空白"}

    stock_name, market_type, category = _find_name_market_category(code, manual_name, manual_market, manual_category, master_lookup)

    if clean_categories and category not in clean_categories:
        return {"status": "category_filtered", "code": code, "message": f"類型不符合：{category}"}

    bundle = _analyze_stock_bundle(
        stock_no=code,
        stock_name=stock_name,
        market_type=market_type,
        start_dt=start_dt,
        end_dt=end_dt,
        risk_strictness=risk_strictness,
    )
    if not bundle or not bundle.get("ok", False):
        history_debug = bundle.get("history_debug", {}) if isinstance(bundle, dict) else {}
        if isinstance(bundle, dict) and bundle.get("error_stage") == "analysis":
            return {
                "status": "analysis_error",
                "code": code,
                "message": _safe_str(bundle.get("error_message")) or "分析錯誤",
                "history_debug": history_debug,
            }
        return {
            "status": "no_history",
            "code": code,
            "message": "抓不到歷史資料",
            "history_debug": history_debug,
        }

    signal_score = _safe_float(bundle["signal_snapshot"].get("score"), 0) or 0
    opportunity_info = bundle.get("opportunity_info", {}) or {}
    momentum_rescue = _bundle_momentum_rescue_profile(bundle, market_type)
    prebreak_rescue = _bundle_prebreakout_rescue_profile(bundle, market_type)
    rescue_eligible = bool(momentum_rescue.get("eligible") or prebreak_rescue.get("eligible"))
    rescue_profile = momentum_rescue if momentum_rescue.get("eligible") else prebreak_rescue
    opportunity_score = _safe_float(opportunity_info.get("機會股分數"), 0) or 0
    opportunity_core = max(
        _safe_float(opportunity_info.get("低檔位置分數"), 0) or 0,
        _safe_float(opportunity_info.get("拉回承接分數"), 0) or 0,
        _safe_float(opportunity_info.get("支撐回測分數"), 0) or 0,
        _safe_float(opportunity_info.get("止跌轉強分數"), 0) or 0,
    )
    opportunity_mode = _is_opportunity_mode(mode)
    risk_pass = bool(bundle["risk_filter"].get("是否通過風險過濾", False))
    opportunity_chase = _safe_float(opportunity_info.get("追高風險分_機會"), 50) or 50
    prelaunch_score = _safe_float(bundle["prelaunch"].get("起漲前兆分數"), 0) or 0
    trade_score = _safe_float(bundle["trade_feasibility"].get("交易可行分數"), 0) or 0

    # V177：舊訊號／傳統風控／起漲／交易門檻不再在 AI 之前 return。
    # 只要 K 線分析成功，就保留到完整全市場 AI 母體；舊門檻改成診斷特徵。
    # 真正硬風控仍由後段 K 線新鮮度、低流動性、異常價格、LOCKDOWN 與正式治理層把關。
    if callable(evaluate_legacy_soft_gates):
        soft_gate = evaluate_legacy_soft_gates(
            signal_score=signal_score, min_signal_score=min_signal_score,
            risk_pass=risk_pass, risk_reason=_safe_str(bundle["risk_filter"].get("淘汰原因")) or "傳統風控未通過",
            prelaunch_score=prelaunch_score, min_prelaunch_score=min_prelaunch_score,
            trade_score=trade_score, min_trade_score=min_trade_score,
            opportunity_mode=opportunity_mode, opportunity_score=opportunity_score,
            opportunity_core=opportunity_core, opportunity_chase=opportunity_chase,
            rescue_eligible=rescue_eligible,
        )
    else:
        # Fail-open only for discovery visibility; downstream hard controls remain unchanged.
        soft_gate = {"soft_statuses": [], "soft_stages": [], "soft_reasons": [], "rescued_stages": [], "soft_count": 0, "soft_state": "PASS｜相容模式"}
    rescued_stages = list(soft_gate.get("rescued_stages", []))
    soft_filter_stages = list(soft_gate.get("soft_stages", []))
    soft_filter_reasons = list(soft_gate.get("soft_reasons", []))
    soft_filter_statuses = list(soft_gate.get("soft_statuses", []))

    auto_factor_total = _safe_float(bundle["auto_factor"].get("auto_factor_total"), 0) or 0
    technical_score = _safe_float(bundle.get("technical_score"), 0) or 0

    base_composite = _score_clip(technical_score * 0.40 + auto_factor_total * 0.32 + prelaunch_score * 0.18 + trade_score * 0.10)

    return {
        "status": "ok",
        "row": {
            "股票代號": code,
            "股票名稱": stock_name,
            "市場別": bundle["used_market"],
            "類別": category or _infer_category_from_record(stock_name, category),
            "最新價": bundle["close_now"],
            "K線最後交易日": _safe_str(bundle.get("history_last_date")),
            "K線落後交易日": 999,
            "K線資料新鮮度": "待全市場比對",
            "區間漲跌幅%": bundle["period_pct"],
            "訊號分數": signal_score,
            "雷達均分": bundle["radar_avg"],
            "技術結構分數": technical_score,
            "起漲前兆分數": prelaunch_score,
            "交易可行分數": trade_score,
            "追價風險分數": _safe_float(bundle["trade_feasibility"].get("追價風險分數"), 0) or 0,
            "拉回買點分數": _safe_float(bundle["trade_feasibility"].get("拉回買點分數"), 0) or 0,
            "突破買點分數": _safe_float(bundle["trade_feasibility"].get("突破買點分數"), 0) or 0,
            "風險報酬評級": _safe_str(bundle["trade_feasibility"].get("風險報酬評級")),
            "自動因子總分": auto_factor_total,
            "EPS代理分數": bundle["auto_factor"]["eps_proxy"],
            "營收動能代理分數": bundle["auto_factor"]["revenue_proxy"],
            "獲利代理分數": bundle["auto_factor"]["profit_proxy"],
            "大戶鎖碼代理分數": bundle["auto_factor"]["lock_proxy"],
            "法人連買代理分數": bundle["auto_factor"]["inst_proxy"],
            "20日壓力距離%": bundle["pressure_dist"],
            "20日支撐距離%": bundle["support_dist"],
            "個股原始總分": base_composite,
            "市場環境分數": None,
            "市場環境": "",
            "型態名稱": _safe_str(bundle["pattern_info"].get("型態名稱")),
            "型態突破分數": _safe_float(bundle["pattern_info"].get("型態突破分數"), 0) or 0,
            "突破風險": _safe_str(bundle["pattern_info"].get("突破風險")),
            "爆發力分數": _safe_float(bundle["burst_info"].get("爆發力分數"), 0) or 0,
            "爆發等級": _safe_str(bundle["burst_info"].get("爆發等級")),
            "推薦型態": _safe_str(opportunity_info.get("推薦型態")),
            "機會型態": _safe_str(opportunity_info.get("機會型態")),
            "低檔位置分數": _safe_float(opportunity_info.get("低檔位置分數"), 0) or 0,
            "拉回承接分數": _safe_float(opportunity_info.get("拉回承接分數"), 0) or 0,
            "支撐回測分數": _safe_float(opportunity_info.get("支撐回測分數"), 0) or 0,
            "止跌轉強分數": _safe_float(opportunity_info.get("止跌轉強分數"), 0) or 0,
            "機會股分數": _safe_float(opportunity_info.get("機會股分數"), 0) or 0,
            "機會股說明": _safe_str(opportunity_info.get("機會股說明")),
            "進場時機": _safe_str(bundle.get("entry_decision", {}).get("進場時機")),
            "進場時機分數": _safe_float(bundle.get("entry_decision", {}).get("進場時機分數"), 0) or 0,
            "建議動作": _safe_str(bundle.get("entry_decision", {}).get("建議動作")),
            "等待條件": _safe_str(bundle.get("entry_decision", {}).get("等待條件")),
            "近端支撐": bundle.get("entry_decision", {}).get("近端支撐"),
            "主要支撐": bundle.get("entry_decision", {}).get("主要支撐"),
            "近端壓力": bundle.get("entry_decision", {}).get("近端壓力"),
            "突破確認價": bundle.get("entry_decision", {}).get("突破確認價"),
            "停損參考": bundle.get("entry_decision", {}).get("停損參考"),
            "操作區間": _safe_str(bundle.get("entry_decision", {}).get("操作區間")),
            "風險報酬比_決策": bundle.get("entry_decision", {}).get("風險報酬比_決策"),
            "追高風險分數_決策": _safe_float(bundle.get("entry_decision", {}).get("追高風險分數_決策"), 0) or 0,
            "追高風險等級": _safe_str(bundle.get("entry_decision", {}).get("追高風險等級")),
            "是否建議追價": _safe_str(bundle.get("entry_decision", {}).get("是否建議追價")),
            "風險扣分原因": _safe_str(bundle.get("entry_decision", {}).get("風險扣分原因")),
            "決策說明": _safe_str(bundle.get("entry_decision", {}).get("決策說明")),
            "建議切入區": _build_entry_zone_text(bundle["trade_plan"]["pullback_buy"], bundle["trade_plan"]["breakout_buy"]),
            "起漲判斷": bundle["trade_plan"]["launch_tag"],
            "推薦買點_突破": bundle["trade_plan"]["breakout_buy"],
            "推薦買點_拉回": bundle["trade_plan"]["pullback_buy"],
            "停損價": bundle["trade_plan"]["stop_price"],
            "賣出目標1": bundle["trade_plan"]["sell_target_1"],
            "賣出目標2": bundle["trade_plan"]["sell_target_2"],
            "風險報酬_拉回": bundle["trade_plan"]["rr1"],
            "風險報酬_突破": bundle["trade_plan"]["rr2"],
            "自動因子摘要": bundle["auto_factor"]["factor_summary"],
            "雷達摘要": _safe_str(bundle["radar"].get("summary")) or "—",
            "風險分數": _safe_float(bundle["risk_filter"].get("風險分數"), 0) or 0,
            "淘汰原因": _safe_str(bundle["risk_filter"].get("淘汰原因")),
            "均線轉強分": _safe_float(bundle["prelaunch"].get("均線轉強分"), 0) or 0,
            "量能啟動分": _safe_float(bundle["prelaunch"].get("量能啟動分"), 0) or 0,
            "突破準備分": _safe_float(bundle["prelaunch"].get("突破準備分"), 0) or 0,
            "動能翻多分": _safe_float(bundle["prelaunch"].get("動能翻多分"), 0) or 0,
            "支撐防守分": _safe_float(bundle["prelaunch"].get("支撐防守分"), 0) or 0,
            "最新成交量": _safe_float(bundle.get("volume_last"), 0) or 0,
            "5日均量": _safe_float(bundle.get("volume_5"), 0) or 0,
            "20日均量": _safe_float(bundle.get("volume_20"), 0) or 0,
            "最新成交量_張": round((_safe_float(bundle.get("volume_last"), 0) or 0) / 1000, 1),
            "20日均量_張": round((_safe_float(bundle.get("volume_20"), 0) or 0) / 1000, 1),
            "成交額百萬": _safe_float(bundle.get("turnover_m"), 0) or 0,
            "20日均成交額百萬": _safe_float(bundle.get("avg20_turnover_m"), 0) or 0,
            "流動性資料狀態": _safe_str(bundle.get("liquidity_status")) or "缺失",
            "流動性資料來源": _safe_str(bundle.get("liquidity_source")) or "缺少成交額/成交量",
            "均量比": _safe_float(bundle.get("volume_ratio"), 0) or 0,
            "當日量比": _safe_float(bundle.get("daily_volume_ratio"), 0) or 0,
            "5日20日量比": _safe_float(bundle.get("volume_5_20_ratio"), 0) or 0,
            "今日漲幅%": _safe_float(bundle.get("day_gain_pct"), 0) or 0,
            "開盤跳空%": _safe_float(bundle.get("opening_gap_pct"), 0) or 0,
            "當日收盤位置%": _safe_float(bundle.get("close_location_pct"), 50) or 50,
            "突破20日高點%": _safe_float(bundle.get("breakout20_pct"), 0) or 0,
            "距20日高點%": _safe_float(bundle.get("distance_to_20d_high_pct"), 0) or 0,
            "上影線比例%": _safe_float(bundle.get("upper_shadow_pct"), 0) or 0,
            "連續上漲天數": int(_safe_float(bundle.get("consecutive_up_days"), 0) or 0),
            "強勢收盤旗標": "是" if bool(bundle.get("strong_close_flag")) else "否",
            "盤後動能救援分": _safe_float(momentum_rescue.get("score"), 0) or 0,
            "盤前強勢前兆分": _safe_float(prebreak_rescue.get("score"), 0) or 0,
            "前置保留類型": (rescue_profile.get("kind", "") if rescued_stages else ("FULL-MARKET-AI" if soft_filter_stages else "")),
            "前置保留原因": (f"前置救援通過：{rescue_profile.get('reason', '')}；原本會被過濾：{'、'.join(rescued_stages)}" if rescued_stages else ("V177全市場AI保留：" + "、".join(soft_filter_reasons) if soft_filter_reasons else "")),
            "前置軟篩選狀態": _safe_str(soft_gate.get("soft_state")),
            "前置軟篩選階段": "、".join(soft_filter_stages),
            "前置軟篩選原因": "｜".join(soft_filter_reasons),
            "前置軟篩選數": int(soft_gate.get("soft_count", 0) or 0),
            "AI發現母體版本": FULL_MARKET_DISCOVERY_VERSION,
            "近5日漲幅%": _safe_float(bundle.get("ret5"), 0) or 0,
            "近20日漲幅%": _safe_float(bundle.get("ret20"), 0) or 0,
            "近60日漲幅%": _safe_float(bundle.get("ret60"), 0) or 0,
            "收盤距MA20%": _safe_float(bundle.get("close_vs_ma20_pct"), 0) or 0,
            "收盤距MA60%": _safe_float(bundle.get("close_vs_ma60_pct"), 0) or 0,
            "推薦模式": mode,
        },
        "history_ok": True,
        "soft_filter_statuses": soft_filter_statuses,
        "history_debug": bundle.get("history_debug", {}),
    }


def _sector_flow_grade(score: Any) -> str:
    score = _safe_float(score, 0) or 0
    if score >= 85:
        return "S級資金主流"
    if score >= 75:
        return "A級強勢輪動"
    if score >= 65:
        return "B級轉強族群"
    if score >= 55:
        return "C級觀察族群"
    return "弱勢/資金不足"


def _sector_rotation_state(row: pd.Series) -> str:
    flow = _safe_float(row.get("族群資金流分數"), 0) or 0
    heat = _safe_float(row.get("類股熱度分數"), 0) or 0
    accel = _safe_float(row.get("類股加速度"), 0) or 0
    avg_ret = _safe_float(row.get("類股平均漲幅"), 0) or 0
    strong_ratio = _safe_float(row.get("同族群強勢比例"), 0) or 0
    if heat >= 78 and accel >= 76 and strong_ratio >= 35:
        return "主流加速"
    if flow >= 72 and avg_ret <= 6 and accel >= 65:
        return "低位吸金"
    if heat >= 72 and accel < 60:
        return "高檔鈍化"
    if flow >= 65 and strong_ratio >= 25:
        return "輪動轉強"
    if flow < 55:
        return "資金退潮"
    return "中性輪動"


def _sector_strategy_text(row: pd.Series) -> str:
    state = _safe_str(row.get("族群輪動狀態"))
    grade = _safe_str(row.get("強勢族群等級"))
    if "主流加速" in state:
        return "主流族群，優先找拉回承接與回測支撐，不盲目追高。"
    if "低位吸金" in state:
        return "族群尚未大漲但資金轉入，優先找低檔轉強與剛起漲。"
    if "輪動轉強" in state:
        return "族群開始輪動，先挑類股內前段班並控管停損。"
    if "高檔鈍化" in state:
        return "族群熱但加速度下降，避免追價，等拉回或量縮整理。"
    if "退潮" in state or "弱勢" in grade:
        return "族群資金不足，只保留個股訊號很強且風險低者。"
    return "族群中性，個股條件需明確優於同類股。"


def _sector_flow_summary(row: pd.Series) -> str:
    return (
        f"{_safe_str(row.get('強勢族群等級'))}｜{_safe_str(row.get('族群輪動狀態'))}｜"
        f"強勢比例{format_number(row.get('同族群強勢比例'),1)}%｜"
        f"量能{format_number(row.get('同族群平均量能分'),1)}｜"
        f"密度{format_number(row.get('同族群推薦密度'),1)}%"
    )


def _compute_category_strength(base_df: pd.DataFrame) -> pd.DataFrame:
    if base_df is None or base_df.empty:
        return pd.DataFrame(columns=[
            "類別", "類股平均總分", "類股平均訊號", "類股平均漲幅", "類股熱度分數",
            "族群資金流分數", "強勢族群等級", "族群輪動狀態", "同族群強勢比例",
            "同族群推薦密度", "同族群平均量能分", "族群策略建議", "族群資金流說明"
        ])

    work = base_df.copy()

    def _numeric_col(name: str) -> pd.Series:
        if name not in work.columns:
            return pd.Series([np.nan] * len(work), index=work.index, dtype="float64")
        return pd.to_numeric(work[name], errors="coerce")

    def _fill_alias(target: str, aliases: list[str], scale: float = 1.0) -> None:
        base = _numeric_col(target)
        for alias in aliases:
            alt = _numeric_col(alias) * scale
            base = base.where(base.notna() & (base.abs() > 1e-9), alt)
        work[target] = base.fillna(0.0)

    # 類股榜過去在欄位經過決策引擎改名後直接補 0，導致所有族群都顯示
    # 「弱勢 / 資金退潮」。這裡改用同義欄位回補，且只有真缺值才補。
    _fill_alias("個股原始總分", ["推薦總分", "候選強度分", "股神實戰總分", "可操作分"])
    _fill_alias("訊號分數", [])
    if (work["訊號分數"].abs() <= 1e-9).all():
        work["訊號分數"] = _numeric_col("買進分數").fillna(_numeric_col("起漲前兆分數")).fillna(0) / 10.0
    _fill_alias("區間漲跌幅%", ["近5日漲幅%", "今日漲幅%"])
    _fill_alias("雷達均分", ["爆發雷達分", "隔日爆發分", "飆股攻擊分", "主流領漲回補分", "漲停回放分"])
    _fill_alias("自動因子總分", ["Alpha選股潛力分", "股神實戰總分"])
    _fill_alias("起漲前兆分數", ["盤後動能救援分", "隔日爆發分"])
    _fill_alias("交易可行分數", ["進場可執行分", "Entry進場買點分"])
    _fill_alias("型態突破分數", ["飆股攻擊分", "突破準備分"])
    _fill_alias("爆發力分數", ["爆發雷達分", "隔日爆發分", "盤後動能救援分"])
    _fill_alias("量能啟動分", [])
    if (work["量能啟動分"].abs() <= 1e-9).all():
        vr = _numeric_col("當日量比").fillna(_numeric_col("均量比")).fillna(0)
        work["量能啟動分"] = np.select(
            [vr >= 2.0, vr >= 1.5, vr >= 1.2, vr >= 1.0],
            [100.0, 88.0, 75.0, 60.0],
            default=35.0,
        )

    work["_sector_strong_flag"] = (
        (work["個股原始總分"].fillna(0) >= 70)
        | (work["起漲前兆分數"].fillna(0) >= 72)
        | ((work["型態突破分數"].fillna(0) >= 70) & (work["量能啟動分"].fillna(0) >= 62))
    ).astype(float)
    work["_sector_candidate_flag"] = (
        (work["個股原始總分"].fillna(0) >= 62)
        | (work["起漲前兆分數"].fillna(0) >= 65)
        | (work["交易可行分數"].fillna(0) >= 68)
    ).astype(float)

    grp = (
        work.groupby("類別", dropna=False)
        .agg(
            股票數=("股票代號", "count"),
            類股平均總分=("個股原始總分", "mean"),
            類股平均訊號=("訊號分數", "mean"),
            類股平均漲幅=("區間漲跌幅%", "mean"),
            類股平均雷達=("雷達均分", "mean"),
            類股平均自動因子=("自動因子總分", "mean"),
            類股平均起漲前兆=("起漲前兆分數", "mean"),
            類股平均交易可行=("交易可行分數", "mean"),
            類股平均型態突破=("型態突破分數", "mean"),
            類股平均爆發力=("爆發力分數", "mean"),
            同族群平均量能分=("量能啟動分", "mean"),
            同族群強勢比例=("_sector_strong_flag", "mean"),
            同族群推薦密度=("_sector_candidate_flag", "mean"),
        )
        .reset_index()
    )

    grp["同族群強勢比例"] = (grp["同族群強勢比例"].fillna(0) * 100).clip(0, 100)
    grp["同族群推薦密度"] = (grp["同族群推薦密度"].fillna(0) * 100).clip(0, 100)

    # V177：小樣本類股不能因 1/1 或 2/2 就被視為 100% 廣度。
    # 以全市場先驗做 empirical-Bayes 收縮，樣本越大越接近原始比例。
    if callable(apply_sector_bayesian_shrinkage):
        try:
            global_strong_pct = float(work["_sector_strong_flag"].fillna(0).mean() * 100.0)
            global_candidate_pct = float(work["_sector_candidate_flag"].fillna(0).mean() * 100.0)
            global_volume_score = float(pd.to_numeric(work["量能啟動分"], errors="coerce").fillna(0).mean())
            grp = apply_sector_bayesian_shrinkage(
                grp,
                global_strong_pct=global_strong_pct,
                global_candidate_pct=global_candidate_pct,
                global_volume_score=global_volume_score,
            )
        except Exception:
            grp["同族群強勢比例_原始"] = grp["同族群強勢比例"]
            grp["同族群推薦密度_原始"] = grp["同族群推薦密度"]
            grp["同族群平均量能分_原始"] = grp["同族群平均量能分"]
            grp["族群樣本可信度"] = 0.0
            grp["族群樣本校正說明"] = "樣本校正模組失敗，沿用原值"
            grp["族群樣本校正版本"] = SECTOR_SHRINKAGE_VERSION
    else:
        grp["同族群強勢比例_原始"] = grp["同族群強勢比例"]
        grp["同族群推薦密度_原始"] = grp["同族群推薦密度"]
        grp["同族群平均量能分_原始"] = grp["同族群平均量能分"]
        grp["族群樣本可信度"] = 0.0
        grp["族群樣本校正說明"] = "樣本校正模組未載入"
        grp["族群樣本校正版本"] = SECTOR_SHRINKAGE_VERSION

    grp["類股熱度分數"] = (
        grp["類股平均總分"] * 0.28
        + grp["類股平均訊號"] * 5.5
        + grp["類股平均漲幅"].fillna(0) * 0.32
        + grp["類股平均雷達"] * 0.16
        + grp["類股平均自動因子"] * 0.12
        + grp["類股平均起漲前兆"] * 0.12
    ).apply(lambda x: _score_clip(x))

    grp["類股加速度"] = (
        grp["類股平均起漲前兆"] * 0.45
        + grp["類股平均交易可行"] * 0.20
        + grp["類股平均訊號"] * 4.0
        + grp["類股平均漲幅"].fillna(0) * 0.18
    ).apply(lambda x: _score_clip(x))

    grp["族群資金流分數"] = (
        grp["同族群強勢比例"].fillna(0) * 0.30
        + grp["同族群推薦密度"].fillna(0) * 0.18
        + grp["同族群平均量能分"].fillna(0) * 0.22
        + grp["類股加速度"].fillna(0) * 0.18
        + grp["類股平均爆發力"].fillna(0) * 0.07
        + grp["類股平均型態突破"].fillna(0) * 0.05
    ).apply(lambda x: _score_clip(x))

    grp["強勢族群等級"] = grp["族群資金流分數"].apply(_sector_flow_grade)
    grp["族群輪動狀態"] = grp.apply(_sector_rotation_state, axis=1)
    grp["族群策略建議"] = grp.apply(_sector_strategy_text, axis=1)
    grp["族群資金流說明"] = grp.apply(_sector_flow_summary, axis=1)

    grp = grp.sort_values(["族群資金流分數", "類股熱度分數", "類股平均總分"], ascending=[False, False, False]).reset_index(drop=True)
    grp["類股熱度排名"] = range(1, len(grp) + 1)
    return grp

def _build_hot_stock_candidates(base_df: pd.DataFrame, final_df: pd.DataFrame, min_total_score: float) -> pd.DataFrame:
    if base_df is None or base_df.empty:
        return pd.DataFrame()

    final_codes = set()
    if isinstance(final_df, pd.DataFrame) and not final_df.empty and "股票代號" in final_df.columns:
        final_codes = set(final_df["股票代號"].astype(str).tolist())

    work = base_df.copy()
    work = work[~work["股票代號"].astype(str).isin(final_codes)].copy()
    if work.empty:
        return pd.DataFrame()

    # Phase 6：市場領漲回補雷達。
    # 針對 6/12 類型的「隔夜催化 + 主流族群全面攻擊」漏網股，
    # 先保留到領漲回補分頁，不讓 RR/停損距離在盤前就把它刪掉。
    if "主流領漲回補分" in work.columns or "領漲回補角色" in work.columns:
        leader_score = pd.to_numeric(work.get("主流領漲回補分", 0), errors="coerce").fillna(0)
        leader_theme = pd.to_numeric(work.get("漲停族群相似度", 0), errors="coerce").fillna(0)
        leader_role = work.get("領漲回補角色", pd.Series([""] * len(work), index=work.index)).astype(str)
        leader_bucket = work.get("領漲回補分區", pd.Series([""] * len(work), index=work.index)).astype(str)
        amount_m = pd.to_numeric(work.get("成交額百萬", 0), errors="coerce").fillna(0)
        leader_mask = (
            (leader_score >= 70)
            & (leader_theme >= 62)
            & (amount_m >= 80)
            & leader_role.str.contains(r"L\+｜領漲回補雷達|L｜主流強勢回補|T｜題材轉強追蹤", na=False)
            & ~leader_role.str.contains("N｜非領漲回補", na=False)
            & ~leader_bucket.str.contains("低流動性排除", na=False)
        )
        leader_df = work[leader_mask].copy()
        if not leader_df.empty:
            if "補抓原因" not in leader_df.columns:
                leader_df["補抓原因"] = ""
            leader_df["補抓原因"] = leader_df.apply(
                lambda r: _safe_str(r.get("補抓原因")) or f"Phase6市場領漲回補：{_safe_str(r.get('領漲回補角色'))}｜{_safe_str(r.get('錯失原因診斷'))}",
                axis=1,
            )
            sort_cols = [c for c in ["主流領漲回補分", "市場領漲相似分", "漲停族群相似度", "爆發雷達分", "族群攻擊強度", "成交額百萬"] if c in leader_df.columns]
            if sort_cols:
                leader_df = leader_df.sort_values(sort_cols, ascending=[False] * len(sort_cols))
            return leader_df.reset_index(drop=True).head(30)

    # Phase 5：先用獨立飆股雷達做「漏網回補」。
    # 這裡不要求 Entry/Risk 完全通過，避免可能隔日點火的股票被穩健風控提前刪掉；
    # 但仍保留角色/風險欄位，讓它只出現在飆股雷達或高風險觀察，不混成主推薦。
    if "爆發雷達分" in work.columns or "飆股雷達角色" in work.columns:
        radar_score = pd.to_numeric(work.get("爆發雷達分", 0), errors="coerce").fillna(0)
        radar_role = work.get("飆股雷達角色", pd.Series([""] * len(work), index=work.index)).astype(str)
        radar_bucket = work.get("飆股雷達分區", pd.Series([""] * len(work), index=work.index)).astype(str)
        radar_mask = (
            (radar_score >= 62)
            & radar_role.str.contains(r"S\+｜漲停雷達|S｜飆股攻擊候選|B\+｜盤中點火追蹤|R｜高風險爆發觀察", na=False)
            & ~radar_role.str.contains("X｜假強排除", na=False)
            & ~radar_bucket.str.contains("假強排除", na=False)
        )
        radar_df = work[radar_mask].copy()
        if not radar_df.empty:
            if "補抓原因" not in radar_df.columns:
                radar_df["補抓原因"] = ""
            radar_df["補抓原因"] = radar_df.apply(
                lambda r: _safe_str(r.get("補抓原因")) or f"Phase5飆股雷達漏網回補：{_safe_str(r.get('飆股雷達角色'))}｜{_safe_str(r.get('飆股雷達原因'))}",
                axis=1,
            )
            sort_cols = [c for c in ["爆發雷達分", "隔日爆發分", "局部題材火種分", "飆股攻擊分", "族群攻擊強度", "主流資金分", "成交額百萬"] if c in radar_df.columns]
            if sort_cols:
                radar_df = radar_df.sort_values(sort_cols, ascending=[False] * len(sort_cols))
            return radar_df.reset_index(drop=True).head(30)

    score_floor = max(float(min_total_score) - 10.0, 45.0)
    hot_mask = (
        (pd.to_numeric(work.get("推薦總分"), errors="coerce").fillna(0) >= score_floor)
        & (pd.to_numeric(work.get("起漲前兆分數"), errors="coerce").fillna(0) >= 70)
        & (pd.to_numeric(work.get("交易可行分數"), errors="coerce").fillna(0) >= 60)
        & (pd.to_numeric(work.get("類股熱度分數"), errors="coerce").fillna(0) >= 68)
        & (pd.to_numeric(work.get("訊號分數"), errors="coerce").fillna(-999) >= 0)
        & (work.get("起漲判斷", "").astype(str).isin(["強勢起漲候選", "偏多轉強候選"]))
    )
    hot_df = work[hot_mask].copy()
    if hot_df.empty:
        return pd.DataFrame()

    hot_df["補抓原因"] = hot_df.apply(
        lambda r: "、".join([
            x for x in [
                "起漲前兆強" if _safe_float(r.get("起漲前兆分數"), 0) >= 75 else "",
                "交易可行佳" if _safe_float(r.get("交易可行分數"), 0) >= 68 else "",
                "類股熱度高" if _safe_float(r.get("類股熱度分數"), 0) >= 72 else "",
                "類股前3強" if _safe_str(r.get("類股前3強")) == "是" else "",
                "領先同類股" if _safe_str(r.get("是否領先同類股")) == "是" else "",
            ] if x
        ]) or "接近主名單門檻但具起漲結構" ,
        axis=1,
    )
    hot_df = hot_df.sort_values(
        ["型態突破分數", "爆發力分數", "起漲前兆分數", "類股熱度分數", "交易可行分數", "推薦總分", "訊號分數"],
        ascending=[False, False, False, False, False, False, False],
    ).reset_index(drop=True)
    return hot_df



# =========================================================
# V34 高速掃描優化：不做預篩、不改評分、不漏股票
# =========================================================
def _v22_json_safe(obj: Any):
    try:
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, pd.Timestamp):
            return obj.isoformat()
        if pd.isna(obj):
            return None
    except Exception:
        pass
    if isinstance(obj, dict):
        return {str(k): _v22_json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_v22_json_safe(v) for v in obj]
    try:
        import numpy as np
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, (np.bool_,)):
            return bool(obj)
    except Exception:
        pass
    return obj


def _v22_scan_signature(
    universe_items: list[dict[str, str]],
    start_dt: date,
    end_dt: date,
    min_total_score: float,
    min_signal_score: float,
    selected_categories: list[str],
    mode: str,
    risk_strictness: str,
    min_prelaunch_score: float,
    min_trade_score: float,
) -> str:
    codes = [str(x.get("code", "")).strip() for x in universe_items if str(x.get("code", "")).strip()]
    raw = {
        "codes": codes,
        "start_dt": str(start_dt),
        "end_dt": str(end_dt),
        "min_total_score": float(min_total_score),
        "min_signal_score": float(min_signal_score),
        "selected_categories": sorted([str(x) for x in selected_categories]),
        "mode": str(mode),
        "risk_strictness": str(risk_strictness),
        "min_prelaunch_score": float(min_prelaunch_score),
        "min_trade_score": float(min_trade_score),
        "weights": GODPICK_ACTIVE_SCORE_WEIGHTS,
        "macro_bridge": _read_macro_mode_bridge(),
        "version": "v27.3",
    }
    text = json.dumps(raw, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()


def _v22_checkpoint_path() -> Path:
    return Path(GODPICK_SCAN_CHECKPOINT_FILE)


def _v22_load_checkpoint(signature: str) -> dict[str, Any]:
    path = _v22_checkpoint_path()
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            payload = json.load(f)
        if not isinstance(payload, dict):
            return {}
        if str(payload.get("signature", "")) != str(signature):
            return {}
        results = payload.get("processed_results", [])
        if not isinstance(results, list):
            payload["processed_results"] = []
        return payload
    except Exception:
        return {}


def _v22_save_checkpoint(signature: str, processed_results: list[dict[str, Any]], total_count: int, finished: bool = False) -> None:
    try:
        path = _v22_checkpoint_path()
        payload = {
            "version": "v22_godpick_fast_cache_resume",
            "signature": signature,
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished": bool(finished),
            "total_count": int(total_count),
            "processed_count": int(len(processed_results)),
            "processed_results": _v22_json_safe(processed_results),
        }
        tmp = path.with_suffix(".tmp")
        with tmp.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2, default=str)
        tmp.replace(path)
    except Exception:
        pass


def _v22_clear_checkpoint() -> tuple[bool, str]:
    try:
        path = _v22_checkpoint_path()
        if path.exists():
            path.unlink()
            return True, "已清除斷點續掃檔。"
        return True, "目前沒有斷點續掃檔。"
    except Exception as e:
        return False, f"清除斷點續掃檔失敗：{e}"


def _v22_checkpoint_status() -> dict[str, Any]:
    path = _v22_checkpoint_path()
    if not path.exists():
        return {"exists": False, "path": str(path), "processed_count": 0, "total_count": 0, "updated_at": ""}
    try:
        with path.open("r", encoding="utf-8") as f:
            payload = json.load(f)
        return {
            "exists": True,
            "path": str(path),
            "processed_count": int(payload.get("processed_count", len(payload.get("processed_results", []) or [])) or 0),
            "total_count": int(payload.get("total_count", 0) or 0),
            "updated_at": str(payload.get("updated_at", "")),
            "finished": bool(payload.get("finished", False)),
        }
    except Exception as e:
        return {"exists": True, "path": str(path), "processed_count": 0, "total_count": 0, "updated_at": "", "error": str(e)}



# =========================================================
# V118 實戰品質防呆：避免冷門低量、無趨勢股票排到前面
# =========================================================
def _apply_v118_liquidity_trend_guard(df: pd.DataFrame | None) -> pd.DataFrame:
    """降低冷門低量 / 無趨勢股票排序與分數。

    不刪股票，避免漏掉潛伏股；但低量與無趨勢個股會被降分、降排序，
    並在表格中標示量能狀態 / 趨勢狀態 / 實戰品質提醒。
    """
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()

    out = df.copy()

    def n(col: str, default: float = 0.0) -> pd.Series:
        if col in out.columns:
            return pd.to_numeric(out[col], errors="coerce").fillna(default)
        return pd.Series([default] * len(out), index=out.index, dtype="float64")

    volume_score = n("量能啟動分", 0)
    volume_ratio = n("均量比", 0)
    vol20 = n("20日均量", 0)
    tech = n("技術結構分數", 0)
    trend = n("均線轉強分", 0)
    momentum = n("動能翻多分", 0)
    prelaunch = n("起漲前兆分數", 0)
    period_pct = n("區間漲跌幅%", 0)
    ma20_pct = n("收盤距MA20%", 0)
    ma60_pct = n("收盤距MA60%", 0)
    official_complete = n("官方資料完整度", 0)
    inst_score = n("法人籌碼分數", 50)
    official_score = n("官方因子總分", 50)

    very_low_liquidity = ((vol20 > 0) & (vol20 < 300000)) | ((volume_score < 35) & (volume_ratio < 0.75))
    low_liquidity = very_low_liquidity | ((volume_score < 45) & (volume_ratio < 0.90))

    weak_trend = (
        (tech < 55)
        & (trend < 52)
        & (momentum < 52)
        & (prelaunch < 58)
        & (period_pct <= 0)
        & (ma20_pct <= 0)
    )
    no_uptrend = weak_trend | ((ma20_pct < -3) & (ma60_pct < -3) & (period_pct <= 3))

    official_buffer = ((official_complete >= 60) & (official_score >= 76) & (inst_score >= 60))

    penalty = pd.Series([0.0] * len(out), index=out.index, dtype="float64")
    penalty += very_low_liquidity.astype(float) * 14
    penalty += (low_liquidity & ~very_low_liquidity).astype(float) * 8
    penalty += no_uptrend.astype(float) * 12
    penalty += (low_liquidity & no_uptrend).astype(float) * 8
    penalty -= official_buffer.astype(float) * 4
    penalty = penalty.clip(lower=0, upper=28)

    quality = (100 - penalty - low_liquidity.astype(float) * 8 - no_uptrend.astype(float) * 10).clip(lower=0, upper=100)

    out["實戰品質分"] = quality.round(1)
    out["量能狀態"] = [
        "極低量/冷門" if bool(vl) else ("量能不足" if bool(ll) else "量能可接受")
        for vl, ll in zip(very_low_liquidity.tolist(), low_liquidity.tolist())
    ]
    out["趨勢狀態"] = ["無明確上升趨勢" if bool(nt) else "趨勢可接受" for nt in no_uptrend.tolist()]
    out["實戰降分"] = penalty.round(1)

    reasons = []
    for i in out.index:
        r = []
        if bool(very_low_liquidity.loc[i]):
            r.append("20日均量偏低或量能明顯不足")
        elif bool(low_liquidity.loc[i]):
            r.append("量能未確認")
        if bool(no_uptrend.loc[i]):
            r.append("尚未形成上升趨勢")
        if bool(official_buffer.loc[i]) and r:
            r.append("官方因子佳但僅緩衝，不取代量價確認")
        reasons.append("；".join(r) if r else "OK")
    out["實戰品質提醒"] = reasons

    for col, ratio in [("推薦總分", 1.00), ("夜間股神總分", 0.90), ("隔日進場分數", 1.00), ("隔日實戰排序分", 1.15), ("波段潛力分數", 0.70)]:
        if col in out.columns:
            out[col] = (pd.to_numeric(out[col], errors="coerce").fillna(0) - penalty * ratio).clip(lower=0, upper=100).round(2)

    if "推薦理由摘要" in out.columns:
        base = out["推薦理由摘要"].astype(str)
        out["推薦理由摘要"] = [b if reason == "OK" else (b + "｜" + reason if b and b != "nan" else reason) for b, reason in zip(base.tolist(), reasons)]
    if "夜間風險提醒" in out.columns:
        base = out["夜間風險提醒"].astype(str)
        out["夜間風險提醒"] = [b if reason == "OK" else (b + "；" + reason if b and b != "nan" else reason) for b, reason in zip(base.tolist(), reasons)]

    return out




# =========================================================
# V139：動態資金流 + 主升起漲嚴選
# 目的：
# - 不再使用固定熱門題材白名單。
# - 以當前掃描結果、族群資金流、量能、趨勢、成交金額與大盤狀態，動態判斷主流族群。
# - 嚴禁極低量/冷門股進主要推薦；但保留少數「有量、有趨勢、接近主升」的隱藏飆股候選。
# =========================================================
def _apply_v139_dynamic_hot_money_breakout_rules(df: pd.DataFrame | None) -> pd.DataFrame:
    """主流資金 + 近期強勢 + 流動性嚴選。

    這一層不再把「小量低基期」當成主要推薦。主推薦必須同時具備：
    1. 當前族群有資金流；
    2. 個股成交量 / 成交金額足夠；
    3. 近 5 / 20 日有明確強勢或突破延續；
    4. 交易可行與隔日進場分數不能太低。

    低量股不刪除，避免完全漏掉潛伏股；但只放到觀察區，不進主要推薦。
    """
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame()
    if df.empty:
        return df.copy()

    out = df.copy()

    def n(col: str, default: float = 0.0) -> pd.Series:
        if col in out.columns:
            return pd.to_numeric(out[col], errors="coerce").fillna(default).astype("float64")
        return pd.Series([default] * len(out), index=out.index, dtype="float64")

    def s(col: str, default: str = "") -> pd.Series:
        if col in out.columns:
            return out[col].fillna(default).astype(str)
        return pd.Series([default] * len(out), index=out.index, dtype="object")

    def _norm_group(v: Any) -> str:
        text = str(v).strip()
        if not text or text.lower() in {"nan", "none", "null", "--"}:
            return "未分類"
        return text

    if "族群名稱" in out.columns:
        group = s("族群名稱").map(_norm_group)
    elif "主題類別" in out.columns:
        group = s("主題類別").map(_norm_group)
    elif "類別" in out.columns:
        group = s("類別").map(_norm_group)
    elif "產業" in out.columns:
        group = s("產業").map(_norm_group)
    else:
        group = pd.Series(["未分類"] * len(out), index=out.index, dtype="object")

    out["族群名稱"] = group

    price = n("最新價", 0)
    latest_vol = n("最新成交量", 0)
    avg20_vol = n("20日均量", 0)
    avg5_vol = n("5日均量", 0)
    latest_lot = n("最新成交量_張", 0)
    avg20_lot = n("20日均量_張", 0)
    vol_ratio = n("均量比", 0)
    volume_score = n("量能啟動分", 0)
    trend_score = n("均線轉強分", 0)
    momentum_score = n("動能翻多分", 0)
    breakout_score = n("突破準備分", 0)
    support_score = n("支撐防守分", 0)
    tech_score = n("技術結構分數", 0)
    prelaunch_score = n("起漲前兆分數", 0)
    trade_score = n("交易可行分數", 0)
    entry_score = n("隔日進場分數", 0)
    quality_score = n("實戰品質分", 50)
    group_money_score = n("族群資金流分數", 50)
    category_heat = n("類股熱度分數", 50)
    period_pct = n("區間漲跌幅%", 0)
    ret5 = n("近5日漲幅%", 0)
    ret20 = n("近20日漲幅%", 0)
    ma20_pct = n("收盤距MA20%", 0)
    ma60_pct = n("收盤距MA60%", 0)
    total_score = n("推薦總分", 0)
    night_score = n("夜間股神總分", total_score)

    # 成交額：優先用既有欄位；否則以 最新價 x 成交量 估算百萬元。
    if "成交額百萬" in out.columns:
        turnover_m = n("成交額百萬", 0)
    elif "V133成交額百萬" in out.columns:
        turnover_m = n("V133成交額百萬", 0)
    else:
        turnover_m = (price * latest_vol / 1_000_000).replace([np.inf, -np.inf], 0).fillna(0)

    if "20日均成交額百萬" in out.columns:
        avg20_turnover_m = n("20日均成交額百萬", 0)
    elif "V133二十日均成交額百萬" in out.columns:
        avg20_turnover_m = n("V133二十日均成交額百萬", 0)
    else:
        avg20_turnover_m = (price * avg20_vol / 1_000_000).replace([np.inf, -np.inf], 0).fillna(0)

    # 若舊資料沒有張數欄，從成交股數換算。
    latest_lot = latest_lot.where(latest_lot > 0, latest_vol / 1000)
    avg20_lot = avg20_lot.where(avg20_lot > 0, avg20_vol / 1000)

    out["最新成交量_張"] = latest_lot.round(1)
    out["20日均量_張"] = avg20_lot.round(1)
    out["成交額百萬"] = turnover_m.round(1)
    out["20日均成交額百萬"] = avg20_turnover_m.round(1)

    # V148：四層實戰分流。
    # 核心觀念：主推薦不等於「分數高」，而是必須同時通過
    # 主流資金、成交量/成交金額、近期強勢與買點品質。
    # 低量股保留在觀察區，避免完全漏掉剛起漲；但不再進主要推薦。
    recent_momentum = (
        (ret5 >= 3.0)
        | (ret20 >= 7.0)
        | ((ret5 >= 1.5) & (ret20 >= 4.5) & (vol_ratio >= 1.25))
        | ((breakout_score >= 66) & (volume_score >= 62) & (vol_ratio >= 1.20))
    )
    strong_recent_momentum = (
        (ret5 >= 5.0)
        | (ret20 >= 11.0)
        | ((ret5 >= 3.0) & (ret20 >= 8.0) & (vol_ratio >= 1.30))
        | ((period_pct >= 12.0) & (ret5 >= 2.0))
    )
    # 「剛起漲」可以不要求已大漲，但必須有放量、突破/起漲分數與不是低量冷門。
    early_breakout_momentum = (
        ((ret5 >= 1.0) | (ret20 >= 3.0) | (breakout_score >= 62) | (prelaunch_score >= 65))
        & ((vol_ratio >= 1.25) | (volume_score >= 64))
        & (ret20 <= 22)
    )

    # 流動性門檻改用「張數 + 成交金額」雙條件，避免幾百張冷門股因百分比漂亮而進主推薦。
    # 一般可交易：可列觀察；主流作戰：才能進主推薦。
    liquidity_ok = (
        ((turnover_m >= 100) & (latest_lot >= 1500))
        | ((avg20_turnover_m >= 70) & (avg20_lot >= 1500) & (vol_ratio >= 1.08))
        | ((latest_lot >= 3500) & (price >= 15) & (vol_ratio >= 1.05))
    )
    strong_liquidity = (
        ((turnover_m >= 220) & (latest_lot >= 2500))
        | ((avg20_turnover_m >= 150) & (avg20_lot >= 3000) & (vol_ratio >= 1.10))
        | ((latest_lot >= 7000) & (turnover_m >= 130))
    )
    hot_liquidity = (
        ((turnover_m >= 160) & (latest_lot >= 2500))
        | ((latest_lot >= 5000) & (vol_ratio >= 1.15))
        | ((avg20_turnover_m >= 120) & (avg20_lot >= 2500))
    )
    cold_stock = (
        (latest_lot < 1000)
        | (turnover_m < 60)
        | ((avg20_turnover_m < 45) & (avg20_lot < 1200) & (vol_ratio < 1.45))
        | s("量能狀態").str.contains("極低量|冷門|量能不足", na=False)
    )
    # 主推薦專用門檻，比一般 liquidity_ok 更硬。
    main_liquidity_gate = (
        ((turnover_m >= 180) & (latest_lot >= 3000))
        | ((turnover_m >= 280) & (latest_lot >= 1800))
        | ((avg20_turnover_m >= 160) & (avg20_lot >= 3000) & (vol_ratio >= 1.10))
    )

    trend_ok = (
        (tech_score >= 55)
        | (trend_score >= 55)
        | (momentum_score >= 58)
        | (prelaunch_score >= 58)
        | ((ma20_pct >= -1.0) & (ret20 >= 0))
    )
    strong_trend = (
        ((trend_score >= 65) | (momentum_score >= 65) | (breakout_score >= 65) | (prelaunch_score >= 68))
        & (ma20_pct >= -1.5)
        & (ret20 >= -1)
    )
    hidden_breakout = (
        liquidity_ok
        & trend_ok
        & recent_momentum
        & ((vol_ratio >= 1.25) | (volume_score >= 62) | (turnover_m >= 100))
        & ((breakout_score >= 58) | (prelaunch_score >= 60) | (momentum_score >= 60) | (support_score >= 60))
        & (ret20 <= 28)
    )

    def _turnover_score(v: float) -> float:
        try:
            return float(max(0, min(100, (v / 260.0) * 100)))
        except Exception:
            return 0.0

    def _lot_score(v: float) -> float:
        try:
            return float(max(0, min(100, (v / 8000.0) * 100)))
        except Exception:
            return 0.0

    out["人氣量能分"] = (
        turnover_m.map(_turnover_score) * 0.35
        + avg20_turnover_m.map(_turnover_score) * 0.20
        + latest_lot.map(_lot_score) * 0.25
        + (vol_ratio.clip(lower=0, upper=3) / 3 * 100) * 0.20
    ).clip(0, 100).round(1)

    tmp = pd.DataFrame({
        "_group": group,
        "_turnover": turnover_m.clip(lower=0),
        "_avg20_turnover": avg20_turnover_m.clip(lower=0),
        "_latest_lot": latest_lot.clip(lower=0),
        "_avg20_lot": avg20_lot.clip(lower=0),
        "_vol_ratio": vol_ratio.clip(lower=0),
        "_vol_score": volume_score.clip(lower=0, upper=100),
        "_trend": trend_score.clip(lower=0, upper=100),
        "_momentum": momentum_score.clip(lower=0, upper=100),
        "_breakout": breakout_score.clip(lower=0, upper=100),
        "_prelaunch": prelaunch_score.clip(lower=0, upper=100),
        "_trade": trade_score.clip(lower=0, upper=100),
        "_entry": entry_score.clip(lower=0, upper=100),
        "_quality": quality_score.clip(lower=0, upper=100),
        "_group_money": group_money_score.clip(lower=0, upper=100),
        "_category_heat": category_heat.clip(lower=0, upper=100),
        "_ret5": ret5.clip(lower=-20, upper=35),
        "_ret20": ret20.clip(lower=-30, upper=60),
        "_period": period_pct.clip(lower=-30, upper=80),
        "_liquidity_ok": liquidity_ok.astype(float),
        "_strong_liquidity": strong_liquidity.astype(float),
        "_hot_liquidity": hot_liquidity.astype(float),
        "_trend_ok": trend_ok.astype(float),
        "_strong_trend": strong_trend.astype(float),
        "_recent": recent_momentum.astype(float),
        "_strong_recent": strong_recent_momentum.astype(float),
        "_hidden": hidden_breakout.astype(float),
    }, index=out.index)

    group_stats = tmp.groupby("_group", dropna=False).agg(
        樣本數=("_group", "size"),
        平均成交額=("_turnover", "mean"),
        平均20日成交額=("_avg20_turnover", "mean"),
        平均成交張數=("_latest_lot", "mean"),
        平均20日張數=("_avg20_lot", "mean"),
        平均均量比=("_vol_ratio", "mean"),
        平均量能分=("_vol_score", "mean"),
        平均趨勢分=("_trend", "mean"),
        平均動能分=("_momentum", "mean"),
        平均突破分=("_breakout", "mean"),
        平均起漲分=("_prelaunch", "mean"),
        平均交易分=("_trade", "mean"),
        平均隔日分=("_entry", "mean"),
        平均品質分=("_quality", "mean"),
        平均族群資金=("_group_money", "mean"),
        平均類股熱度=("_category_heat", "mean"),
        平均5日漲幅=("_ret5", "mean"),
        平均20日漲幅=("_ret20", "mean"),
        平均區間漲跌=("_period", "mean"),
        流動性合格率=("_liquidity_ok", "mean"),
        強流動率=("_strong_liquidity", "mean"),
        熱門流動率=("_hot_liquidity", "mean"),
        趨勢合格率=("_trend_ok", "mean"),
        強趨勢率=("_strong_trend", "mean"),
        近期強勢率=("_recent", "mean"),
        強近期率=("_strong_recent", "mean"),
        隱藏起漲率=("_hidden", "mean"),
    )

    group_stats["族群流動性分數_calc"] = (
        group_stats["平均成交額"].map(_turnover_score) * 0.30
        + group_stats["平均20日成交額"].map(_turnover_score) * 0.20
        + group_stats["平均成交張數"].map(_lot_score) * 0.22
        + group_stats["平均20日張數"].map(_lot_score) * 0.12
        + (group_stats["平均均量比"].clip(lower=0, upper=3) / 3 * 100) * 0.08
        + group_stats["流動性合格率"].clip(0, 1) * 100 * 0.08
    ).clip(0, 100)

    group_stats["族群資金流分數_calc"] = (
        group_stats["族群流動性分數_calc"] * 0.24
        + group_stats["平均族群資金"] * 0.12
        + group_stats["平均類股熱度"] * 0.08
        + group_stats["強流動率"].clip(0, 1) * 100 * 0.12
        + group_stats["熱門流動率"].clip(0, 1) * 100 * 0.08
        + group_stats["近期強勢率"].clip(0, 1) * 100 * 0.14
        + group_stats["強近期率"].clip(0, 1) * 100 * 0.08
        + group_stats["強趨勢率"].clip(0, 1) * 100 * 0.08
        + group_stats["平均突破分"] * 0.03
        + group_stats["平均起漲分"] * 0.03
    ).clip(0, 100).round(1)

    group_stats = group_stats.sort_values(["族群資金流分數_calc", "族群流動性分數_calc", "樣本數"], ascending=[False, False, False])
    group_stats["族群熱度排名_calc"] = range(1, len(group_stats) + 1)

    market_mode = s("大盤橋接狀態").replace("", np.nan).fillna(s("大盤策略模式")).astype(str)
    bearish = market_mode.str.contains("空|偏空|防守", na=False)
    market_text = "空頭防守" if bool(bearish.any()) else ("偏多進攻" if market_mode.str.contains("多|進攻", na=False).any() else "震盪選股")
    hot_threshold = 78 if market_text == "空頭防守" else (72 if market_text == "震盪選股" else 68)

    hot_groups = set(group_stats[
        (
            (group_stats["族群資金流分數_calc"] >= hot_threshold)
            & (group_stats["族群流動性分數_calc"] >= (66 if market_text != "空頭防守" else 72))
            & (group_stats["近期強勢率"] >= 0.22)
            & (
                (group_stats["樣本數"] >= 2)
                | ((group_stats["樣本數"] == 1) & (group_stats["平均成交額"] >= 250) & (group_stats["強近期率"] >= 1))
            )
        )
    ].index.astype(str).tolist())

    group_score_map = group_stats["族群資金流分數_calc"].to_dict()
    group_liq_map = group_stats["族群流動性分數_calc"].to_dict()
    group_rank_map = group_stats["族群熱度排名_calc"].to_dict()
    group_size_map = group_stats["樣本數"].to_dict()

    dynamic_group_score = group.map(lambda g: float(group_score_map.get(g, 0))).astype("float64")
    dynamic_group_liq = group.map(lambda g: float(group_liq_map.get(g, 0))).astype("float64")
    dynamic_group_rank = group.map(lambda g: int(group_rank_map.get(g, 999))).astype("int64")
    dynamic_group_size = group.map(lambda g: int(group_size_map.get(g, 0))).astype("int64")
    dynamic_hot_group = group.map(lambda g: "是" if str(g) in hot_groups else "否")

    individual_hot_flow = (
        (strong_liquidity & strong_recent_momentum & (out["人氣量能分"] >= 68))
        | ((turnover_m >= 350) & (latest_lot >= 4000) & strong_recent_momentum)
        | ((turnover_m >= 500) & (ret5 >= 3.0) & (ret20 >= 8.0))
    )
    hot_flow_pass = dynamic_hot_group.eq("是") | individual_hot_flow

    out["資金流熱門族群"] = ["是" if bool(x) else "否" for x in hot_flow_pass.tolist()]
    out["族群熱度排名"] = dynamic_group_rank
    out["族群資金流分數"] = dynamic_group_score.round(1)
    out["族群流動性分數"] = dynamic_group_liq.round(1)
    out["族群樣本數"] = dynamic_group_size
    out["族群判斷依據"] = [
        f"{g}｜排名{int(r)}｜資金{float(gs):.1f}｜流動性{float(gl):.1f}｜樣本{int(sz)}｜{market_text}"
        for g, r, gs, gl, sz in zip(group.tolist(), dynamic_group_rank.tolist(), dynamic_group_score.tolist(), dynamic_group_liq.tolist(), dynamic_group_size.tolist())
    ]
    out["大盤趨勢模式"] = market_text
    out["近期強勢狀態"] = [
        "強勢主升" if bool(sr) else ("近期轉強" if bool(rm) else "未見近期強勢")
        for sr, rm in zip(strong_recent_momentum.tolist(), recent_momentum.tolist())
    ]
    out["主升量價檢查"] = [
        "OK" if (bool(lo) and bool(rm) and not bool(cs)) else "不列主推薦：" + "、".join([x for x in [
            "成交量不足" if not bool(lo) else "",
            "近期漲幅/突破不足" if not bool(rm) else "",
            "低量冷門" if bool(cs) else "",
        ] if x])
        for lo, rm, cs in zip(liquidity_ok.tolist(), recent_momentum.tolist(), cold_stock.tolist())
    ]

    practical_score = (
        dynamic_group_score * 0.18
        + dynamic_group_liq * 0.14
        + out["人氣量能分"] * 0.16
        + strong_recent_momentum.astype(float) * 100 * 0.12
        + recent_momentum.astype(float) * 100 * 0.06
        + volume_score.clip(0, 100) * 0.08
        + trend_score.clip(0, 100) * 0.08
        + momentum_score.clip(0, 100) * 0.08
        + breakout_score.clip(0, 100) * 0.04
        + prelaunch_score.clip(0, 100) * 0.03
        + trade_score.clip(0, 100) * 0.03
    ).clip(0, 100).round(1)

    out["股神輸出排序"] = practical_score
    out["候補排序分"] = (
        practical_score
        + hidden_breakout.astype(float) * 6
        + strong_liquidity.astype(float) * 5
        + strong_recent_momentum.astype(float) * 6
        - cold_stock.astype(float) * 25
        - (~hot_flow_pass).astype(float) * 12
        - (~recent_momentum).astype(float) * 12
        - bearish.astype(float) * 3
    ).clip(0, 100).round(1)

    # V151：主推薦硬門檻。
    # 推薦總分只代表「候選強度」；能不能進主推薦，必須再通過買進分數、
    # 風險報酬比、隔日進場、交易可行與追價風險檢查。
    # 這層只調整分流與排序，不刪除候選股，避免漏掉可觀察標的。
    buy_score = n("買進分數", 0)
    buy_alt = n("實戰買點分數", 0)
    buy_score = buy_score.where(buy_score > 0, buy_alt)
    buy_score = buy_score.where(buy_score > 0, entry_score)

    rr_score = n("風險報酬比", 0)
    rr_alt = n("風險報酬比_決策", 0)
    rr_score = rr_score.where(rr_score > 0, rr_alt)

    chase_score = n("追價風險分", 0)
    chase_alt = n("追高風險分數_決策", 0)
    chase_score = chase_score.where(chase_score > 0, chase_alt)
    chase_level = s("追高風險等級")
    no_buy_reason = s("高分禁買原因")
    direct_buy_text = s("是否可直接買進")
    chase_advice = s("是否建議追價")

    high_chase_risk = (
        (chase_score >= 70)
        | chase_level.str.contains("高|過熱|不追", na=False)
        | no_buy_reason.str.strip().ne("")
        | direct_buy_text.str.contains("否|不適合|等拉回", na=False)
        | chase_advice.str.contains("否|不追|禁止", na=False)
    )
    strict_buy_gate = buy_score >= 65
    strict_rr_gate = rr_score >= 1.5
    strict_entry_gate = entry_score >= 60
    strict_trade_gate = trade_score >= 60

    main_pass = (
        hot_flow_pass
        & main_liquidity_gate
        & strong_liquidity
        & hot_liquidity
        & strong_trend
        & strong_recent_momentum
        & ((dynamic_group_score >= (70 if market_text != "空頭防守" else 76)) | individual_hot_flow)
        & ((dynamic_group_liq >= (68 if market_text != "空頭防守" else 74)) | individual_hot_flow)
        & (trade_score >= 64)
        & (entry_score >= (58 if market_text != "空頭防守" else 66))
        & (quality_score >= 74)
        & ~cold_stock
    )
    hard_main_gate = (
        main_pass
        & strict_buy_gate
        & strict_rr_gate
        & strict_entry_gate
        & strict_trade_gate
        & ~high_chase_risk
    )
    waiting_pullback = main_pass & ~hard_main_gate

    early_watch = (
        hot_flow_pass
        & liquidity_ok
        & early_breakout_momentum
        & hidden_breakout
        & trend_ok
        & (trade_score >= 56)
        & (entry_score >= (40 if market_text != "空頭防守" else 50))
        & (quality_score >= 66)
        & ~cold_stock
    )
    defensive_watch = (
        liquidity_ok
        & strong_liquidity
        & (quality_score >= 72)
        & (trade_score >= 60)
        & (trend_ok | strong_trend)
        & ~cold_stock
        & ~hard_main_gate
        & ~early_watch
    )
    hq_watch = (
        liquidity_ok
        & trend_ok
        & (recent_momentum | early_breakout_momentum)
        & (practical_score >= 58)
        & ~cold_stock
        & ~hard_main_gate
        & ~early_watch
        & ~defensive_watch
    )

    level = []
    candidate_grade = []
    main_show = []
    display_zone = []
    reason_list = []
    advice_list = []
    for i in out.index:
        reasons = []
        if not bool(hot_flow_pass.loc[i]):
            reasons.append("非目前資金流熱門族群/個股資金未達門檻")
        if bool(cold_stock.loc[i]):
            reasons.append("低量/冷門，封鎖主要推薦")
        if not bool(liquidity_ok.loc[i]):
            reasons.append("成交額或成交張數不足")
        if not bool(recent_momentum.loc[i]):
            reasons.append("近期未形成主升/突破動能")
        if not bool(trend_ok.loc[i]):
            reasons.append("趨勢尚未確認")
        if not bool(main_liquidity_gate.loc[i]):
            reasons.append("未達主推薦成交量/成交金額門檻")
        if not bool(strong_recent_momentum.loc[i]):
            reasons.append("未達主推薦近期強勢門檻")
        if (float(dynamic_group_score.loc[i]) < (70 if market_text != "空頭防守" else 76)) and not bool(individual_hot_flow.loc[i]):
            reasons.append("族群資金流未達主推薦門檻")
        if float(buy_score.loc[i]) < 65:
            reasons.append(f"買進分數不足({float(buy_score.loc[i]):.1f}<65)")
        if float(rr_score.loc[i]) < 1.5:
            reasons.append(f"風險報酬比不足({float(rr_score.loc[i]):.2f}<1.5)")
        if float(entry_score.loc[i]) < 60:
            reasons.append(f"隔日進場分數不足({float(entry_score.loc[i]):.1f}<60)")
        if float(trade_score.loc[i]) < 60:
            reasons.append(f"交易可行分數不足({float(trade_score.loc[i]):.1f}<60)")
        if bool(high_chase_risk.loc[i]):
            reasons.append("追價/過熱風險過高，等待拉回或盤中確認")
        if market_text == "空頭防守" and float(entry_score.loc[i]) < 66:
            reasons.append("大盤偏空，進場條件加嚴")

        if bool(hard_main_gate.loc[i]):
            lv, cg, show, zone = "主流強勢作戰股", "主推薦", "是", "今日主推薦"
            advice = "通過主推薦硬門檻：主流資金＋高流動性＋近期強勢＋買點/RR合格；仍依突破/回測價執行，不開高追價。"
        elif bool(waiting_pullback.loc[i]):
            lv, cg, show, zone = "高分等待拉回股", "等待拉回", "否", "等待拉回候選"
            advice = "候選強度高，但買進分數、R/R或追價風險未通過主推薦硬門檻；只列觀察，等拉回/盤中確認。"
        elif bool(early_watch.loc[i]):
            lv, cg, show, zone = "剛起漲觀察股", "觀察", "否", "剛起漲觀察"
            advice = "有放量與起漲訊號，但尚未通過主推薦流動性/強勢門檻；等放量突破或回測承接確認。"
        elif bool(defensive_watch.loc[i]):
            lv, cg, show, zone = "穩健防守觀察股", "觀察", "否", "穩健防守觀察"
            advice = "流動性與品質尚可，但短線主升動能不足；只適合追蹤，不列主推薦。"
        elif bool(hq_watch.loc[i]):
            lv, cg, show, zone = "高品質觀察股", "觀察", "否", "高品質觀察"
            advice = "具備部分條件但尚未達主要推薦門檻。"
        else:
            lv, cg, show, zone = "排除/低量觀察股", "排除觀察", "否", "低量/弱勢排除" if bool(cold_stock.loc[i]) else "觀察等待"
            advice = "未達主流資金、近期強勢或成交量門檻；不列主要推薦。"

        level.append(lv)
        candidate_grade.append(cg)
        main_show.append(show)
        display_zone.append(zone)
        if show == "是":
            reason_list.append("OK｜通過主推薦硬門檻")
        else:
            reason_list.append("不列主推薦：" + ("；".join(reasons) if reasons else "未通過主推薦硬門檻"))
        advice_list.append(advice)

    out["股神推薦層級"] = level
    out["候補等級"] = candidate_grade
    out["是否主要顯示"] = main_show
    out["主表篩選"] = main_show
    out["流動性等級"] = ["高流動性" if bool(sl) else ("可交易" if bool(lo) else "低流動性") for sl, lo in zip(strong_liquidity.tolist(), liquidity_ok.tolist())]
    out["限制原因"] = reason_list
    out["股神實戰建議"] = advice_list
    out["顯示分區"] = display_zone
    out["主升起漲候選"] = ["是" if bool(x) else "否" for x in hidden_breakout.tolist()]
    out["主表篩選說明"] = [
        "主流強勢作戰股：主要顯示" if show == "是" else "不列主推薦：保留於觀察/排除分區"
        for show in main_show
    ]
    out["是否可直接買進"] = [
        "可小量試單｜仍需盤中確認" if show == "是" else "否｜僅列觀察/等待拉回"
        for show in main_show
    ]
    out["推薦用途"] = [
        "今日主推薦｜主流強勢作戰" if show == "是" else ("高分追蹤候選｜等拉回不追高" if lv == "高分等待拉回股" else "觀察候選｜等待訊號補強")
        for show, lv in zip(main_show, level)
    ]

    tier_rank = {"主流強勢作戰股": 1, "高分等待拉回股": 2, "剛起漲觀察股": 3, "穩健防守觀察股": 4, "高品質觀察股": 5, "排除/低量觀察股": 6, "觀察等待": 7}
    out["_hot_money_tier_rank"] = [tier_rank.get(x, 9) for x in level]

    # v147 hot-money safety: some legacy recommendation rows do not carry every
    # ranking column.  sort_values requires all keys to exist, so we backfill the
    # ranking columns from the local calculation series before sorting.  This keeps
    # the liquidity/hot-money upgrade without crashing on older saved data or
    # lighter scan modes.
    sort_defaults = {
        "候補排序分": out.get("候補排序分", practical_score),
        "股神輸出排序": out.get("股神輸出排序", practical_score),
        "人氣量能分": out.get("人氣量能分", pd.Series([0] * len(out), index=out.index)),
        "買進分數": buy_score,
        "風險報酬比": rr_score,
        "近20日漲幅%": ret20,
        "隔日進場分數": entry_score,
        "交易可行分數": trade_score,
        "推薦總分": total_score,
    }
    for _col, _default in sort_defaults.items():
        if _col not in out.columns:
            if isinstance(_default, pd.Series):
                out[_col] = _default.reindex(out.index).fillna(0)
            else:
                out[_col] = _default
        out[_col] = pd.to_numeric(out[_col], errors="coerce").fillna(0)

    _sort_cols = ["_hot_money_tier_rank", "買進分數", "風險報酬比", "隔日進場分數", "交易可行分數", "候補排序分", "股神輸出排序", "人氣量能分", "近20日漲幅%", "推薦總分"]
    out = out.sort_values(
        _sort_cols,
        ascending=[True, False, False, False, False, False, False, False, False, False],
    ).drop(columns=["_hot_money_tier_rank"], errors="ignore").reset_index(drop=True)

    return out


def _build_recommend_df(
    universe_items: list[dict[str, str]],
    master_df: pd.DataFrame,
    start_dt: date,
    end_dt: date,
    min_total_score: float,
    min_signal_score: float,
    selected_categories: list[str],
    mode: str,
    risk_strictness: str,
    min_prelaunch_score: float,
    min_trade_score: float,
    resume_scan: bool = False,
    reuse_finished_checkpoint: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    clean_categories = [_normalize_category(x) for x in selected_categories if _normalize_category(x) and x != "全部"]
    if not universe_items:
        _save_debug_scan_summary({})
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    total_count = len(universe_items)
    worker_count = min(SCAN_MAX_WORKERS, max(6, total_count // 80 if total_count >= 240 else 6))
    master_lookup = _build_master_lookup(master_df)

    progress_wrap = st.container()
    progress_bar = progress_wrap.progress(0, text="準備開始推薦...")
    progress_text = progress_wrap.empty()

    start_ts = time.time()
    done_count = 0
    base_rows = []
    debug_summary = {
        "total_count": total_count,
        # history_ok：成功取得並完成 K 線/指標分析；即使之後被訊號、風控、起漲或交易門檻篩掉，仍屬有效K線。
        # analyzed_ok：通過推薦前置篩選並進入完整候選池。兩者不可混用。
        "history_ok": 0,
        "analyzed_ok": 0,
        "passed_final": 0,
        "invalid_code": 0,
        "category_filtered": 0,
        "no_history": 0,
        "analysis_error": 0,
        "signal_filtered": 0,
        "risk_filtered": 0,
        "prelaunch_filtered": 0,
        "trade_filtered": 0,
        "soft_gate_survivors": 0,
        "full_market_ai_pool": 0,
        "ai_pool_share_of_kline_pct": 0.0,
        "full_market_discovery_version": FULL_MARKET_DISCOVERY_VERSION,
        "final_score_filtered": 0,
        "history_debug_samples": [],
        "error_samples": [],
        "worker_count": worker_count,
        "checkpoint_retryable_count": 0,
        "speed_version": "v48_3_price_date_integrity_and_entry_readiness",
        "scan_started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "scan_elapsed_sec": 0.0,
        "avg_sec_per_stock": 0.0,
        "history_success_rate_pct": 0.0,
        "scan_speed_samples": [],
        "slowest_stocks": [],
        "status_elapsed_summary": {},
        "data_source_diagnostics_available": False,
        "v180_checkpoint_reused": False,
        "v180_checkpoint_reused_count": 0,
        "v180_postscan_nonblocking": True,
    }

    scan_signature = _v22_scan_signature(
        universe_items,
        start_dt,
        end_dt,
        min_total_score,
        min_signal_score,
        selected_categories,
        mode,
        risk_strictness,
        min_prelaunch_score,
        min_trade_score,
    )
    st.session_state[_k("active_scan_signature_v181")] = scan_signature
    processed_results: list[dict[str, Any]] = []
    processed_codes: set[str] = set()

    def _consume_scan_result(result: dict[str, Any], from_checkpoint: bool = False) -> None:
        if not isinstance(result, dict):
            debug_summary["analysis_error"] += 1
            debug_summary["error_samples"].append("未知錯誤：future.result 非 dict")
            return
        status = _safe_str(result.get("status")) or "analysis_error"
        code = _safe_str(result.get("code"))
        if not code and isinstance(result.get("row"), dict):
            code = _safe_str(result.get("row", {}).get("股票代號"))

        # 只有真正完成或確定性排除才可視為斷點已完成。
        # no_history / analysis_error 必須在「斷點續掃」時重新補抓，不能永久被舊失敗結果跳過。
        if code and status not in _RETRYABLE_SCAN_STATUSES:
            processed_codes.add(code)

        history_valid = bool(result.get("history_ok", False)) or status in _KLINE_VALID_STATUSES
        if history_valid:
            debug_summary["history_ok"] = int(debug_summary.get("history_ok", 0)) + 1

        if status == "ok":
            row = result.get("row")
            if isinstance(row, dict):
                base_rows.append(row)
                debug_summary["analyzed_ok"] += 1
                softs = result.get("soft_filter_statuses", []) or []
                if softs:
                    debug_summary["soft_gate_survivors"] = int(debug_summary.get("soft_gate_survivors", 0)) + 1
                    for soft_status in softs:
                        if soft_status in {"signal_filtered", "risk_filtered", "prelaunch_filtered", "trade_filtered"}:
                            debug_summary[soft_status] = int(debug_summary.get(soft_status, 0)) + 1
        else:
            debug_summary[status] = int(debug_summary.get(status, 0)) + 1
            msg = _safe_str(result.get("message"))
            if status == "no_history":
                hdbg = result.get("history_debug", {}) or {}
                attempt_lines = []
                for att in hdbg.get("attempts", [])[:3]:
                    market = _safe_str(att.get("market_type")) or "未知市場"
                    rows = att.get("rows", 0)
                    err = _safe_str(att.get("error"))
                    source = _safe_str(att.get("source"))
                    attempt_lines.append(f"{market} rows={rows} source={source} err={err}")
                debug_summary["history_debug_samples"].append(f"{code}：{msg}｜" + " / ".join(attempt_lines))
            elif status == "analysis_error":
                debug_summary["error_samples"].append(f"{code}：{msg}")

    # V180：相同交易日 / 相同條件的「開始推薦」可直接重用已完成斷點，
    # 避免使用者只是重按開始就重新下載 1000~2000 檔 K 線。
    # 「重新推薦」不走此路徑，仍會強制抓新資料。盤中 TTL 僅 5 分鐘；收盤後/週末較長。
    if reuse_finished_checkpoint and not resume_scan:
        checkpoint_payload = _v22_load_checkpoint(scan_signature)
        reusable = False
        if isinstance(checkpoint_payload, dict) and checkpoint_payload.get("finished"):
            checkpoint_results = checkpoint_payload.get("processed_results", [])
            try:
                path = _v22_checkpoint_path()
                age_sec = max(time.time() - path.stat().st_mtime, 0.0) if path.exists() else 10**9
            except Exception:
                age_sec = 10**9
            try:
                now_tw = datetime.now(ZoneInfo("Asia/Taipei"))
                during_market = now_tw.weekday() < 5 and (9, 0) <= (now_tw.hour, now_tw.minute) <= (13, 40)
                ttl_sec = 300 if during_market else (18 * 3600 if now_tw.weekday() < 5 else 36 * 3600)
            except Exception:
                ttl_sec = 300
            reusable = bool(
                isinstance(checkpoint_results, list)
                and len(checkpoint_results) >= total_count
                and age_sec <= ttl_sec
            )
            if reusable:
                latest_by_code: dict[str, dict[str, Any]] = {}
                for old_result in checkpoint_results:
                    if not isinstance(old_result, dict):
                        continue
                    old_code = _normalize_code(old_result.get("code"))
                    if not old_code and isinstance(old_result.get("row"), dict):
                        old_code = _normalize_code(old_result.get("row", {}).get("股票代號"))
                    if old_code:
                        latest_by_code[old_code] = old_result
                # 只重用確定性成功/排除結果；no_history / analysis_error / future_exception
                # 仍要重新補抓，避免把舊失敗當成有效快取，也避免除錯統計重複計數。
                reused_results: list[dict[str, Any]] = []
                retryable_reuse_count = 0
                for old_result in latest_by_code.values():
                    old_status = _safe_str(old_result.get("status")) or "analysis_error"
                    if old_status in _RETRYABLE_SCAN_STATUSES:
                        retryable_reuse_count += 1
                        continue
                    reused_results.append(old_result)
                    _consume_scan_result(old_result, from_checkpoint=True)
                processed_results.extend(reused_results)
                debug_summary["v180_checkpoint_reused"] = True
                debug_summary["v180_checkpoint_reused_count"] = len(reused_results)
                debug_summary["checkpoint_retryable_count"] = retryable_reuse_count
                progress_text.caption(
                    f"V180 快取命中：沿用同條件有效結果 {len(reused_results)}/{total_count} 檔；"
                    f"舊失敗待補抓 {retryable_reuse_count} 檔。若要強制更新全部行情，請按『重新推薦』。"
                )

    if resume_scan:
        checkpoint_payload = _v22_load_checkpoint(scan_signature)
        checkpoint_results = checkpoint_payload.get("processed_results", []) if isinstance(checkpoint_payload, dict) else []
        if isinstance(checkpoint_results, list) and checkpoint_results:
            # 同一股票只保留斷點內最後一筆結果；舊的 no_history / analysis_error 不載入完成集合，
            # 讓本次續掃真正只補抓失敗股票，而不是顯示「已完成」後永遠不重試。
            latest_by_code: dict[str, dict[str, Any]] = {}
            no_code_results: list[dict[str, Any]] = []
            for old_result in checkpoint_results:
                if not isinstance(old_result, dict):
                    continue
                old_code = _normalize_code(old_result.get("code"))
                if not old_code and isinstance(old_result.get("row"), dict):
                    old_code = _normalize_code(old_result.get("row", {}).get("股票代號"))
                if old_code:
                    latest_by_code[old_code] = old_result
                else:
                    no_code_results.append(old_result)

            retryable_count = 0
            retained_results: list[dict[str, Any]] = []
            for old_result in list(latest_by_code.values()) + no_code_results:
                old_status = _safe_str(old_result.get("status")) or "analysis_error"
                if old_status in _RETRYABLE_SCAN_STATUSES:
                    retryable_count += 1
                    continue
                retained_results.append(old_result)
                _consume_scan_result(old_result, from_checkpoint=True)

            processed_results.extend(retained_results)
            debug_summary["checkpoint_retryable_count"] = retryable_count
            progress_text.caption(
                f"已載入斷點有效結果：{len(retained_results)} / {total_count} 檔；"
                f"本次將重新補抓 {retryable_count} 檔舊失敗股票及其他未完成股票。"
            )

    pending_items = []
    for item in universe_items:
        c = _normalize_code(item.get("code"))
        if c and c in processed_codes:
            continue
        pending_items.append(item)

    done_count = len(processed_results)

    if pending_items:
        with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix="godpick_scan") as executor:
            # V48：推薦速度監控。future 需要保留股票代號、名稱與送出時間，才能找出真正慢在哪幾檔。
            futures: dict[Any, dict[str, Any]] = {}
            for item in pending_items:
                submit_ts = time.time()
                future = executor.submit(
                    _analyze_one_stock_for_recommend,
                    item,
                    master_lookup,
                    start_dt,
                    end_dt,
                    min_signal_score,
                    clean_categories,
                    mode,
                    risk_strictness,
                    min_prelaunch_score,
                    min_trade_score,
                )
                futures[future] = {
                    "code": _normalize_code(item.get("code")),
                    "name": _safe_str(item.get("name")),
                    "market": _safe_str(item.get("market")),
                    "submitted_at": submit_ts,
                }

            speed_samples: list[dict[str, Any]] = []
            status_elapsed_bucket: dict[str, list[float]] = {}

            for future in as_completed(futures):
                done_count += 1
                future_meta = futures.get(future, {}) or {}
                elapsed_one = max(time.time() - float(future_meta.get("submitted_at", time.time())), 0.0)
                try:
                    result = future.result()
                    if not isinstance(result, dict):
                        result = {"status": "analysis_error", "code": future_meta.get("code", ""), "message": "future.result 非 dict"}

                    # V48：把單檔耗時寫入結果，不影響推薦分數，只用於診斷。
                    result["scan_elapsed_sec"] = round(elapsed_one, 3)
                    result["scan_code"] = _normalize_code(result.get("code") or future_meta.get("code"))
                    result["scan_name"] = _safe_str(future_meta.get("name"))
                    result["scan_market"] = _safe_str(future_meta.get("market"))
                    result_status = _safe_str(result.get("status")) or "unknown"
                    status_elapsed_bucket.setdefault(result_status, []).append(elapsed_one)
                    speed_samples.append({
                        "股票代號": result.get("scan_code") or _normalize_code(future_meta.get("code")),
                        "股票名稱": result.get("scan_name") or _safe_str(future_meta.get("name")),
                        "市場別": result.get("scan_market") or _safe_str(future_meta.get("market")),
                        "狀態": result_status,
                        "耗時秒": round(elapsed_one, 3),
                        "訊息": _safe_str(result.get("message"))[:120],
                    })

                    processed_results.append(result)
                    _consume_scan_result(result)

                    if done_count % max(1, V22_CHECKPOINT_EVERY) == 0 or done_count == total_count:
                        _v22_save_checkpoint(scan_signature, processed_results, total_count, finished=False)
                except Exception as e:
                    debug_summary["analysis_error"] += 1
                    debug_summary["error_samples"].append(f"future.result 例外：{e}")
                    status_elapsed_bucket.setdefault("future_exception", []).append(elapsed_one)
                    speed_samples.append({
                        "股票代號": _normalize_code(future_meta.get("code")),
                        "股票名稱": _safe_str(future_meta.get("name")),
                        "市場別": _safe_str(future_meta.get("market")),
                        "狀態": "future_exception",
                        "耗時秒": round(elapsed_one, 3),
                        "訊息": str(e)[:120],
                    })
                should_update_progress = (
                    done_count == 1
                    or done_count == total_count
                    or done_count % max(1, PROGRESS_UPDATE_EVERY) == 0
                    or done_count / total_count >= 0.98
                )
                if should_update_progress:
                    elapsed = time.time() - start_ts
                    avg_per_stock = elapsed / done_count if done_count > 0 else 0
                    remain_count = max(total_count - done_count, 0)
                    eta_sec = avg_per_stock * remain_count
                    ratio = done_count / total_count if total_count > 0 else 0

                    progress_bar.progress(min(max(ratio * 0.85, 0.0), 0.85), text=f"K線與特徵掃描中... {done_count}/{total_count} ({ratio*100:.1f}%)")
                    progress_text.caption(
                        f"已完成 {done_count}/{total_count}｜"
                        f"已花時間：{_fmt_seconds(elapsed)}｜"
                        f"預估剩餘：{_fmt_seconds(eta_sec)}｜"
                        f"平均每檔：約 {_fmt_seconds(avg_per_stock)}｜平行工人：{worker_count}｜V48速度監控掃描"
                    )
    else:
        progress_text.caption(f"斷點資料已涵蓋全部 {total_count} 檔，直接整理結果。")

    _v22_save_checkpoint(scan_signature, processed_results, total_count, finished=True)

    progress_bar.progress(0.86, text=f"K線掃描完成 {total_count} 檔｜正在整理市場、類股與最終決策...")
    postscan_start_ts = time.time()
    total_elapsed = time.time() - start_ts

    # V48：整理速度監控資訊。只做顯示與診斷，不影響推薦結果。
    try:
        all_speed_samples = locals().get("speed_samples", []) or []
        all_status_elapsed_bucket = locals().get("status_elapsed_bucket", {}) or {}
        slowest = sorted(all_speed_samples, key=lambda x: _safe_float(x.get("耗時秒"), 0) or 0, reverse=True)[:10]
        status_summary = {}
        for k, vals in all_status_elapsed_bucket.items():
            vals = [float(v) for v in vals if v is not None]
            if not vals:
                continue
            status_summary[k] = {
                "count": len(vals),
                "avg_sec": round(sum(vals) / len(vals), 3),
                "max_sec": round(max(vals), 3),
            }
        debug_summary["scan_elapsed_sec"] = round(total_elapsed, 3)
        debug_summary["avg_sec_per_stock"] = round(total_elapsed / max(done_count, 1), 3)
        debug_summary["scan_speed_samples"] = all_speed_samples[:200]
        debug_summary["slowest_stocks"] = slowest
        debug_summary["status_elapsed_summary"] = status_summary
        debug_summary["history_success_rate_pct"] = round((debug_summary.get("history_ok", 0) / max(total_count, 1)) * 100, 2)
        debug_summary["candidate_success_rate_pct"] = round((debug_summary.get("analyzed_ok", 0) / max(total_count, 1)) * 100, 2)
        debug_summary["scan_finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # v48 utils.py 若有提供資料源診斷，推薦頁會在除錯摘要顯示。
        try:
            from utils import get_data_source_diagnostics  # type: ignore
            diag = get_data_source_diagnostics()
            if isinstance(diag, dict):
                debug_summary["data_source_diagnostics_available"] = True
                debug_summary["data_source_diagnostics"] = diag
        except Exception as diag_err:
            debug_summary["data_source_diagnostics_available"] = False
            debug_summary["data_source_diagnostics_error"] = str(diag_err)
    except Exception as speed_err:
        debug_summary["speed_monitor_error"] = str(speed_err)

    progress_text.caption(
        f"K線掃描完成｜掃描耗時：{_fmt_seconds(total_elapsed)}｜"
        f"平均每檔：{_fmt_seconds(debug_summary.get('avg_sec_per_stock', 0))}｜"
        f"歷史資料成功率：{debug_summary.get('history_success_rate_pct', 0)}%"
    )

    base_df = pd.DataFrame(base_rows)
    if base_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    debug_summary["full_market_ai_pool"] = int(len(base_df))
    debug_summary["ai_pool_share_of_kline_pct"] = round(100.0 * len(base_df) / max(int(debug_summary.get("history_ok", 0)), 1), 2)

    # V177：此處的 base_df 現在是「所有 K 線分析成功股票」的 AI 母體，
    # 不再只是通過舊訊號/風控/起漲/交易門檻後的殘餘候選。
    # 正式評分前先以本輪市場共同最新日期驗證每檔行情，避免不同日期價格混算。
    base_df = _annotate_kline_freshness(base_df)
    try:
        debug_summary["market_latest_kline_date"] = _safe_str(base_df.get("本輪市場最新交易日", pd.Series([""])).iloc[0])
        debug_summary["stale_kline_count"] = int(pd.to_numeric(base_df.get("K線落後交易日", 999), errors="coerce").fillna(999).gt(0).sum())
    except Exception:
        pass

    category_strength_df = _compute_category_strength(base_df)
    if category_strength_df.empty:
        base_df["類股平均總分"] = None
        base_df["類股平均訊號"] = None
        base_df["類股平均漲幅"] = None
        base_df["類股熱度分數"] = None
        base_df["類股熱度排名"] = None
        base_df["類股加速度"] = None
        base_df["類股平均型態突破"] = None
        base_df["類股平均爆發力"] = None
        base_df["族群資金流分數"] = None
        base_df["強勢族群等級"] = ""
        base_df["族群輪動狀態"] = ""
        base_df["同族群強勢比例"] = None
        base_df["同族群推薦密度"] = None
        base_df["同族群平均量能分"] = None
        base_df["同族群強勢比例_原始"] = None
        base_df["同族群推薦密度_原始"] = None
        base_df["同族群平均量能分_原始"] = None
        base_df["族群樣本可信度"] = None
        base_df["族群樣本校正說明"] = ""
        base_df["族群樣本校正版本"] = SECTOR_SHRINKAGE_VERSION
        base_df["族群策略建議"] = ""
        base_df["族群資金流說明"] = ""
    else:
        base_df = base_df.merge(
            category_strength_df[
                ["類別", "類股平均總分", "類股平均訊號", "類股平均漲幅", "類股熱度分數", "類股熱度排名", "類股加速度", "類股平均型態突破", "類股平均爆發力", "族群資金流分數", "強勢族群等級", "族群輪動狀態", "同族群強勢比例", "同族群推薦密度", "同族群平均量能分", "同族群強勢比例_原始", "同族群推薦密度_原始", "同族群平均量能分_原始", "族群樣本可信度", "族群樣本校正說明", "族群樣本校正版本", "族群策略建議", "族群資金流說明"]
            ],
            on="類別",
            how="left",
        )

    market_info = _build_market_environment(base_df)
    base_df["市場環境分數"] = market_info.get("score", 50)
    base_df["市場環境"] = market_info.get("label", "中性")

    base_df["同類股領先幅度"] = (base_df["個股原始總分"] - base_df["類股平均總分"].fillna(0)).apply(lambda x: _score_clip(50 + x))
    base_df["是否領先同類股"] = (base_df["個股原始總分"] >= base_df["類股平均總分"].fillna(0)).map({True: "是", False: "否"})
    base_df["類股內排名"] = base_df.groupby("類別")["個股原始總分"].rank(method="dense", ascending=False).astype(int)
    base_df["類股前3強"] = base_df["類股內排名"].apply(lambda x: "是" if pd.notna(x) and int(x) <= 3 else "否")

    _v181_weight_map = _get_active_weight_map()
    _v181_market_score = _safe_float(market_info.get("score"), 50) or 50
    mode_scores = [
        _build_final_god_score_row(
            row=r,
            mode=_safe_str(mode),
            market_score=_v181_market_score,
            weights=_v181_weight_map,
        )
        for r in base_df.to_dict(orient="records")
    ]
    base_df["推薦總分"] = [x[0] for x in mode_scores]
    base_df["推薦標籤"] = [x[1] for x in mode_scores]

    # V13：類股資金流只做加權與排序，不做硬篩選，避免漏掉股票。
    if "族群資金流分數" in base_df.columns:
        flow_score = pd.to_numeric(base_df["族群資金流分數"], errors="coerce").fillna(50)
        sector_bonus = ((flow_score - 60) / 10).clip(lower=-2.0, upper=4.0)
        base_df["族群資金流加權"] = sector_bonus.round(2)
        base_df["推薦總分"] = (pd.to_numeric(base_df["推薦總分"], errors="coerce").fillna(0) + sector_bonus).clip(lower=0, upper=100)
    else:
        base_df["族群資金流加權"] = 0.0

    def _recommend(score: float) -> str:
        if score >= 90:
            return "股神級"
        if score >= 84:
            return "強烈關注"
        if score >= 72:
            return "優先觀察"
        if score >= 60:
            return "可列追蹤"
        return "觀察"

    base_df["推薦等級"] = base_df["推薦總分"].apply(_recommend)

    # V59：加入上漲機率估計。此為條件機率，不等於保證上漲。
    try:
        upside_estimates = [_estimate_upside_probability_row(r) for r in base_df.to_dict(orient="records")]
        base_df["上漲機率估計%"] = [x.get("上漲機率估計%") for x in upside_estimates]
        base_df["上漲機率等級"] = [x.get("上漲機率等級") for x in upside_estimates]
        base_df["上漲機率信心"] = [x.get("上漲機率信心") for x in upside_estimates]
        base_df["上漲機率說明"] = [x.get("上漲機率說明") for x in upside_estimates]
        base_df["上漲機率因子明細"] = [x.get("上漲機率因子明細") for x in upside_estimates]
    except Exception as prob_err:
        base_df["上漲機率估計%"] = None
        base_df["上漲機率等級"] = "無法估計"
        base_df["上漲機率信心"] = "低"
        base_df["上漲機率說明"] = f"上漲機率估計失敗：{prob_err}"
        base_df["上漲機率因子明細"] = "{}"

    # V90：夜間 23:00 後使用的隔日進場股神層。
    # 不覆蓋原本推薦總分；新增夜間股神總分、隔日進場分數、預估進場點、法人/基本面資料完整度。
    try:
        if callable(enrich_night_strategy):
            base_df = enrich_night_strategy(base_df)
        else:
            base_df["資料完整度"] = "夜間策略模組未載入"
    except Exception as night_err:
        base_df["資料完整度"] = "夜間策略計算失敗"
        base_df["夜間風險提醒"] = f"夜間隔日策略未套用：{night_err}"

    # V118：實戰品質防呆。降低冷門低量 / 無上升趨勢股票的分數與排序，
    # 避免官方因子或型態分數把沒有量、沒有趨勢的冷門股推到前面。
    try:
        base_df = _apply_v118_liquidity_trend_guard(base_df)
    except Exception as quality_err:
        base_df["實戰品質提醒"] = f"V118實戰品質檢查失敗：{quality_err}"


    def _reason_builder(r):
        reason_parts = []
        if _safe_float(r.get("均線轉強分"), 0) >= 70:
            reason_parts.append("均線結構轉強")
        if _safe_float(r.get("量能啟動分"), 0) >= 65:
            reason_parts.append("量能明顯放大")
        if _safe_float(r.get("突破準備分"), 0) >= 70:
            reason_parts.append("接近壓力突破位")
        if _safe_float(r.get("動能翻多分"), 0) >= 65:
            reason_parts.append("動能翻多")
        if _safe_float(r.get("支撐防守分"), 0) >= 65:
            reason_parts.append("支撐防守佳")
        if _safe_str(r.get("是否領先同類股")) == "是":
            reason_parts.append("領先同類股")
        if _safe_str(r.get("類股前3強")) == "是":
            reason_parts.append("類股前3強")
        if _safe_float(r.get("類股熱度分數"), 0) >= 75:
            reason_parts.append("所屬類股熱度高")
        if _safe_float(r.get("族群資金流分數"), 0) >= 75:
            reason_parts.append("族群資金流強")
        if _safe_str(r.get("族群輪動狀態")) in ["低位吸金", "輪動轉強", "主流加速"]:
            reason_parts.append(_safe_str(r.get("族群輪動狀態")))
        if _safe_float(r.get("交易可行分數"), 0) >= 70:
            reason_parts.append("風險報酬佳")
        if not reason_parts:
            reason_parts.append("結構偏多，列入觀察")
        return "、".join(reason_parts[:6])

    base_df["推薦理由摘要"] = [_build_recommend_reason_v2(r) for r in base_df.to_dict(orient="records")]

    for c in ["3日績效%", "5日績效%", "10日績效%", "20日績效%"]:
        if c not in base_df.columns:
            base_df[c] = pd.NA

    # V139：先依「動態資金流熱門族群 + 有量 + 有趨勢 + 主升起漲」重新分流，
    # 再輸出主要推薦，避免冷門股或固定題材白名單干擾。
    try:
        base_df = _apply_v139_dynamic_hot_money_breakout_rules(base_df)
    except Exception as v139_err:
        base_df["股神實戰建議"] = f"動態資金流檢查失敗：{v139_err}"

    # VNext：套用歷史績效回饋校正；其中決策引擎已包含正式推薦引擎與 Phase105/108 AI overlay。
    progress_bar.progress(0.90, text="市場與類股整理完成｜正在執行最終股神決策引擎...")
    _v181_decision_started = time.time()
    base_df = _apply_vnext_performance_feedback_columns(base_df)
    debug_summary["v181_decision_engine_sec"] = round(time.time() - _v181_decision_started, 3)

    # V181：決策引擎內部的正式推薦引擎已呼叫每日學習 overlay；舊版又在 Page7
    # 立即重算一次 1700 檔，造成「掃描100%後卡很久」。只有 AI 欄位確實不存在時才 fallback。
    _v181_ai_ready = {"AI綜合決策分", "AI模型版本", "AI發現母體"}.issubset(base_df.columns)
    if callable(apply_daily_learning_overlay) and not _v181_ai_ready:
        try:
            base_df = apply_daily_learning_overlay(base_df)
            debug_summary["v181_learning_overlay_fallback"] = True
        except Exception as learning_error:
            debug_summary["phase105_learning_error"] = str(learning_error)
    else:
        debug_summary["v181_learning_overlay_fallback"] = False

    # V183：SuperAI 必須等官方因子合併與掃描品質報告完成後才計算。
    # 此處只記錄起點，避免在資料治理尚未完成時提前給出 READY。
    _v183_super_started = 0.0

    base_score = pd.to_numeric(base_df.get("推薦總分", 0), errors="coerce").fillna(0)
    practical_score = pd.to_numeric(base_df.get("股神實戰總分", base_score), errors="coerce").fillna(base_score)
    main_mask = base_df.get("是否主要顯示", pd.Series(["否"] * len(base_df), index=base_df.index)).astype(str).eq("是")
    role_text = base_df.get("推薦角色", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    filter_state = base_df.get("實戰過濾狀態", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    hard_veto_text = base_df.get("硬否決原因", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    true_veto_text = base_df.get("真禁買原因", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    breakout_status = base_df.get("突破確認狀態", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    blocked_decision_mask = (
        role_text.str.contains("過熱禁買", na=False)
        | filter_state.str.contains("BLOCK", na=False)
        | true_veto_text.str.strip().ne("")
    )
    # Phase 3：有硬否決原因但被決策引擎改判 B 的標的，不再被 hard_veto_text 直接排除；
    # 這類屬於「假陰性修正 / 等突破確認」，應出現在候選名單而不是消失。
    feedback_main_mask = role_text.str.contains("股神主推薦", na=False) & (practical_score >= 84)
    early_potential_mask = role_text.str.contains("早期潛伏", na=False) & (practical_score >= max(60, min(68, float(min_total_score))))
    confirm_mask = role_text.str.contains("等突破確認", na=False) & (practical_score >= 60)
    breakout_wait_mask = breakout_status.str.contains("WAIT", na=False) & (practical_score >= 58)
    radar_role_text = base_df.get("飆股雷達角色", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    radar_bucket_text = base_df.get("飆股雷達分區", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    radar_score = pd.to_numeric(base_df.get("爆發雷達分", 0), errors="coerce").fillna(0)
    amount_m = pd.to_numeric(base_df.get("成交額百萬", 0), errors="coerce").fillna(0)
    leader_role_text = base_df.get("領漲回補角色", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    leader_bucket_text = base_df.get("領漲回補分區", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    leader_score = pd.to_numeric(base_df.get("主流領漲回補分", 0), errors="coerce").fillna(0)
    leader_theme = pd.to_numeric(base_df.get("漲停族群相似度", 0), errors="coerce").fillna(0)
    # Phase 6：領漲回補候選是「檢討 6/12 漏網強勢股」的第三條路，
    # 不混成 A 主推薦，但要出現在匯出雷達，避免只剩弱勢觀察。
    leader_replay_mask = (
        (leader_score >= 70)
        & (leader_theme >= 62)
        & (amount_m >= 80)
        & leader_role_text.str.contains(r"L\+｜領漲回補雷達|L｜主流強勢回補|T｜題材轉強追蹤", na=False)
        & ~leader_role_text.str.contains("N｜非領漲回補", na=False)
        & ~leader_bucket_text.str.contains("低流動性排除", na=False)
    )
    # Phase 5：穩健推薦與飆股雷達雙引擎分流。
    # 飆股雷達候選不因 Entry/Risk/RR 先被刪掉，但它只進雷達/高風險分頁，不當成無腦買進清單。
    explosive_radar_mask = (
        (
            radar_role_text.str.contains(r"S\+｜漲停雷達|S｜飆股攻擊候選|B\+｜盤中點火追蹤", na=False)
            & (radar_score >= 66)
            & (amount_m >= 80)
        )
        | (
            radar_role_text.str.contains("R｜高風險爆發觀察", na=False)
            & (radar_score >= 70)
            & (amount_m >= 120)
        )
    )
    explosive_radar_mask = explosive_radar_mask & ~radar_role_text.str.contains("X｜假強排除", na=False) & ~radar_bucket_text.str.contains("假強排除", na=False)
    allowed_decision_mask = ~blocked_decision_mask & ~role_text.str.contains("弱勢觀察", na=False)
    ai_decision_score = pd.to_numeric(base_df.get("AI綜合決策分", 0), errors="coerce").fillna(0)
    ai_recall_score = pd.to_numeric(base_df.get("AI召回分", 0), errors="coerce").fillna(0)
    ai_qualification = base_df.get("AI推薦資格", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    ai_retain_flag = base_df.get("AI召回保留旗標", pd.Series(["否"] * len(base_df), index=base_df.index)).astype(str).eq("是")
    ai_tactical_text = base_df.get("AI過熱型態", pd.Series([""] * len(base_df), index=base_df.index)).astype(str)
    ai_absolute_block = (
        true_veto_text.str.lower().str.contains("lockdown|全面禁買|低流動|興櫃|行情落後|k線落後|資料待更新", regex=True, na=False)
        | pd.to_numeric(base_df.get("K線落後交易日", 999), errors="coerce").fillna(999).ne(0)
        | ai_qualification.str.contains("LOCKDOWN", na=False)
    )
    # Phase107：一般AI候選仍需通過既有操作許可；跨市場新強股只保證進入診斷／雷達，
    # 不直接繞過正式推薦風控。健康主升整理不得再因舊版「過熱」字樣被整檔刪除。
    ai_recall_mask = (ai_decision_score >= 66) & (ai_recall_score >= 68) & allowed_decision_mask & ~ai_qualification.str.contains("LOCKDOWN", na=False)
    ai_emerging_retain_mask = ai_retain_flag & ~ai_absolute_block & ~ai_tactical_text.eq("爆量噴出末升")
    ai_blowoff_radar_mask = ai_retain_flag & ~ai_absolute_block & ai_tactical_text.eq("爆量噴出末升")
    stable_final_mask = (base_score >= min_total_score) & allowed_decision_mask & (main_mask | feedback_main_mask | early_potential_mask | confirm_mask | breakout_wait_mask)
    final_df = base_df[stable_final_mask | explosive_radar_mask | leader_replay_mask | ai_recall_mask | ai_emerging_retain_mask | ai_blowoff_radar_mask].copy()

    # 若沒有主要推薦，不用冷門股硬湊；只保留少數高品質觀察作為輔助參考。
    if final_df.empty:
        # 不再用低量股硬湊名單。若主推薦為空，只補少量「有成交量 + 近期轉強」的觀察股。
        watch_mask = (
            base_df.get("股神推薦層級", pd.Series([""] * len(base_df), index=base_df.index)).astype(str).isin(["剛起漲觀察股", "高品質觀察股"])
            & (pd.to_numeric(base_df.get("股神輸出排序", 0), errors="coerce").fillna(0) >= 58)
            & (pd.to_numeric(base_df.get("成交額百萬", 0), errors="coerce").fillna(0) >= 80)
            & (pd.to_numeric(base_df.get("最新成交量_張", 0), errors="coerce").fillna(0) >= 1200)
            & (base_df.get("近期強勢狀態", pd.Series([""] * len(base_df), index=base_df.index)).astype(str).isin(["近期轉強", "強勢主升"]))
            & ~base_df.get("量能狀態", pd.Series([""] * len(base_df), index=base_df.index)).astype(str).str.contains("極低量|冷門|量能不足", na=False)
        )
        _watch_sort_cols = ["股神實戰總分", "選股潛力分", "股神輸出排序", "人氣量能分", "近20日漲幅%", "隔日進場分數", "交易可行分數", "推薦總分"]
        _active_watch_sort_cols = [c for c in _watch_sort_cols if c in base_df.columns]
        final_df = base_df[watch_mask].copy()
        if _active_watch_sort_cols:
            final_df = final_df.sort_values(
                _active_watch_sort_cols,
                ascending=[False] * len(_active_watch_sort_cols),
            )
        final_df = final_df.head(10).copy()

    debug_summary["final_score_filtered"] = max(len(base_df) - len(final_df), 0)
    debug_summary["passed_final"] = len(final_df)
    _save_debug_scan_summary(debug_summary)

    sort_cols = ["AI綜合決策分", "AI召回分", "AI Alpha品質分", "AI Timing時機分", "AI Risk風控分", "AI Continuation延續分", "股神實戰總分", "選股潛力分", "進場買點分", "風控安全分", "績效校正分", "股神輸出排序", "候補排序分", "人氣量能分", "成交額百萬", "最新成交量_張", "近20日漲幅%", "近5日漲幅%", "族群資金流分數", "族群流動性分數", "隔日進場分數", "交易可行分數", "實戰品質分", "夜間股神總分", "推薦總分", "區間漲跌幅%"]
    active_sort_cols = [c for c in sort_cols if c in final_df.columns]
    if active_sort_cols:
        final_df = final_df.sort_values(
            active_sort_cols,
            ascending=[False] * len(active_sort_cols),
        ).reset_index(drop=True)
    else:
        final_df = final_df.reset_index(drop=True)

    if "勾選" not in final_df.columns:
        final_df.insert(0, "勾選", False)

    hot_pick_df = _build_hot_stock_candidates(base_df, final_df, min_total_score)

    # Phase 8.2：完整候選診斷池與正式作戰名單分離。
    # base_df 保留所有成功分析股票；final_df 只保留最終作戰候選。
    try:
        if callable(canonicalize_final_partition):
            final_df = canonicalize_final_partition(final_df)
            governed_candidate_df = canonicalize_final_partition(base_df)
            # Build the primary action/reference list from the complete governed
            # candidate pool.  The old score/Top-N display filter must not erase
            # A/A-/R1 rows; when none exist, the helper returns a small, clearly
            # labelled conditional-reference list instead of a blank page.
            final_df = _operational_recommendation_rows(governed_candidate_df, refresh_decision=False)
        else:
            governed_candidate_df = base_df.copy()
            final_df = _operational_recommendation_rows(final_df, refresh_decision=False)

        # V189：禁止在 SuperAI/V188 交易品質治理完成前保存 decision frame。
        # V188 的 Alpha/Trade/RR/作戰優先分尚未產生時，這裡只能是中間母體，
        # 不得寫入供 UI/Excel/輪動共用的最終快取。
        progress_bar.progress(0.97, text="正式決策完成｜正在套用官方因子、SuperAI與V188交易品質治理...")
        # Phase104：掃描品質必須在官方因子合併後計算。舊版先建立治理報告，
        # 後續才把 official_factors_cache 併入顯示名單，造成快取明明有 886 筆
        # 完整度>=60，掃描報告卻永遠看到預設 0。
        governed_candidate_df = _apply_official_factor_cache_v109(governed_candidate_df)
        scan_report = (
            build_scan_quality_report(
                debug_summary,
                universe_size=total_count,
                candidate_count=len(governed_candidate_df),
                final_count=len(final_df),
                candidate_frame=governed_candidate_df,
            )
            if callable(build_scan_quality_report)
            else {}
        )
        # V183：先把掃描品質/正式推薦許可寫入完整母體，再由 SuperAI 評分。
        # 因此 8/10 類似『最新可信0%＋正式推薦暫停』的情境只能輸出 WAIT-DATA，
        # 不會出現掃描報告說暫停、SuperAI卻說READY的矛盾。
        if callable(apply_scan_quality_to_frame):
            governed_candidate_df = apply_scan_quality_to_frame(governed_candidate_df, scan_report)
            final_df = apply_scan_quality_to_frame(final_df, scan_report)
        _v183_super_started = time.time()
        if callable(apply_super_ai_engine):
            try:
                governed_candidate_df = apply_super_ai_engine(governed_candidate_df)
                debug_summary["v183_super_ai_rows"] = int(len(governed_candidate_df))
                debug_summary["v183_super_ai_sec"] = round(time.time() - _v183_super_started, 3)
                if isinstance(final_df, pd.DataFrame) and not final_df.empty and "股票代號" in final_df.columns:
                    _super_cols = [c for c in getattr(__import__("godpick_super_ai_engine"), "SUPER_AI_COLUMNS", []) if c in governed_candidate_df.columns]
                    # V188 trade-quality is demotion-only, but demotion must flow back
                    # to the actual action list as well as the full candidate diagnosis.
                    _v188_authority_cols = [c for c in [
                        "正式推薦分區", "操作許可", "是否正式推薦", "正式推薦資格",
                        "下週是否可直接買", "正式推薦動作",
                    ] if c in governed_candidate_df.columns]
                    _merge_cols = list(dict.fromkeys([*_super_cols, *_v188_authority_cols]))
                    if _merge_cols:
                        _super_map = governed_candidate_df[["股票代號", *_merge_cols]].drop_duplicates("股票代號", keep="first")
                        final_df = final_df.drop(columns=[c for c in _merge_cols if c in final_df.columns], errors="ignore").merge(_super_map, on="股票代號", how="left")
            except Exception as super_ai_error:
                debug_summary["v183_super_ai_error"] = str(super_ai_error)
                debug_summary["v183_super_ai_sec"] = round(time.time() - _v183_super_started, 3)
        else:
            debug_summary["v183_super_ai_error"] = "Super AI engine unavailable"

        # V189：只有「官方因子 + 掃描治理 + SuperAI + V188交易品質」全部完成後，
        # 才能建立真正的單次決策快取。這修正 V188 欄位在主排名全部變 0 / - / -- 的根因。
        _v188_cache_diag = {}
        if callable(inspect_v188_decision_frame):
            try:
                _v188_cache_diag = inspect_v188_decision_frame(governed_candidate_df)
            except Exception as _v188_diag_err:
                _v188_cache_diag = {"complete": False, "reason": f"V188快取檢查失敗：{_v188_diag_err}"}
        else:
            _v188_cache_diag = {
                "complete": bool("V188股神作戰優先分" in governed_candidate_df.columns and "SuperAI Alpha分" in governed_candidate_df.columns and "SuperAI Trade分" in governed_candidate_df.columns),
                "reason": "fallback-column-check",
            }

        if bool(_v188_cache_diag.get("complete")):
            governed_candidate_df["V181最終決策已完成"] = "是"
            governed_candidate_df["V181最終決策版本"] = PAGE07_SPEED_FIX_VERSION
            governed_candidate_df["V189_V188最終快取完整"] = "是"
            governed_candidate_df["V189快取守門版本"] = V189_CACHE_GUARD_VERSION
            st.session_state[_k("decision_frame_store_v181")] = governed_candidate_df.copy()
            st.session_state[_k("decision_frame_scan_signature_v181")] = scan_signature
            st.session_state[_k("v188_cache_integrity_v189")] = dict(_v188_cache_diag)
            debug_summary["v181_decision_cache_rows"] = int(len(governed_candidate_df))
            debug_summary["v189_v188_final_cache_complete"] = True
            debug_summary["v189_v188_cache_reason"] = _safe_str(_v188_cache_diag.get("reason"))
        else:
            # 不讓本輪半成品冒充 V188 完成品；相同 scan signature 的舊快取也清除。
            st.session_state.pop(_k("decision_frame_store_v181"), None)
            st.session_state.pop(_k("decision_frame_scan_signature_v181"), None)
            st.session_state[_k("v188_cache_integrity_v189")] = dict(_v188_cache_diag)
            debug_summary["v189_v188_final_cache_complete"] = False
            debug_summary["v189_v188_cache_reason"] = _safe_str(_v188_cache_diag.get("reason"))

        candidate_diagnosis_df = (
            build_candidate_diagnosis(governed_candidate_df)
            if callable(build_candidate_diagnosis)
            else governed_candidate_df.copy()
        )
        st.session_state[_k("candidate_diagnosis_store")] = candidate_diagnosis_df.copy()
        st.session_state[_k("scan_quality_report")] = dict(scan_report)
        debug_summary.update({
            "candidate_diagnosis_count": len(candidate_diagnosis_df),
            "action_candidate_count": len(final_df),
            "scan_quality_status": scan_report.get("掃描品質狀態", ""),
            "scan_coverage_pct": scan_report.get("掃描覆蓋率%", 0),
            "history_success_rate_pct": scan_report.get("歷史資料成功率%", debug_summary.get("history_success_rate_pct", 0)),
            "formal_recommendation_usable": scan_report.get("正式推薦可用", False),
        })
        _save_debug_scan_summary(debug_summary)
    except Exception as governance_err:
        debug_summary["execution_governance_error"] = str(governance_err)
        _save_debug_scan_summary(debug_summary)

    _v181_postscan_sec = time.time() - postscan_start_ts
    debug_summary["v181_postscan_final_pipeline_sec"] = round(_v181_postscan_sec, 3)
    debug_summary["v181_total_with_final_sec"] = round(time.time() - start_ts, 3)
    _save_debug_scan_summary(debug_summary)
    progress_bar.progress(1.0, text=f"股神推薦完成｜{total_count} 檔掃描與最終決策均已完成")
    progress_text.caption(
        f"完成｜K線掃描 {_fmt_seconds(total_elapsed)}｜最終結果運算 {_fmt_seconds(_v181_postscan_sec)}｜"
        f"總耗時 {_fmt_seconds(time.time() - start_ts)}｜V189 V188完整決策快取已建立"
    )
    return final_df, category_strength_df, hot_pick_df



def _extract_checked_codes_from_editor_state(editor_key: str, source_df: pd.DataFrame, state_key: str | None = None) -> list[str]:
    """
    v40：穩定版勾選狀態。
    修正 st.data_editor 勾選後，因為 rerun / 表格重繪 / 上方元件異動，
    checkbox 瞬間跳回未勾選的問題。
    """
    if source_df is None or source_df.empty or "股票代號" not in source_df.columns:
        return []

    persist_key = state_key or f"{editor_key}__checked_codes"
    previous = {_normalize_code(x) for x in st.session_state.get(persist_key, []) if _normalize_code(x)}

    base_df = source_df.reset_index(drop=True).copy()
    checked_set: set[str] = set(previous)
    visible_order: list[str] = []

    def _is_true(v: Any) -> bool:
        if isinstance(v, bool):
            return bool(v)
        return str(v).strip().lower() in {"true", "1", "yes", "y", "是", "勾選", "checked"}

    # 回傳表格中的 True 只補入；False 不直接清空，避免 rerun 誤清。
    for _, row in base_df.iterrows():
        code = _normalize_code(row.get("股票代號"))
        if not code:
            continue
        visible_order.append(code)
        if _is_true(row.get("勾選", False)):
            checked_set.add(code)

    # edited_rows 是使用者本次明確改動。True 加入，False 移除。
    raw_state = st.session_state.get(editor_key, {})
    edited_rows = raw_state.get("edited_rows", {}) if isinstance(raw_state, dict) else {}
    if isinstance(edited_rows, dict):
        for raw_idx, changes in edited_rows.items():
            try:
                idx = int(raw_idx)
            except Exception:
                continue
            if idx < 0 or idx >= len(base_df):
                continue
            if not isinstance(changes, dict) or "勾選" not in changes:
                continue
            code = _normalize_code(base_df.iloc[idx].get("股票代號"))
            if not code:
                continue
            if _is_true(changes.get("勾選")):
                checked_set.add(code)
            else:
                checked_set.discard(code)

    visible_set = set(visible_order)
    final_codes = [c for c in visible_order if c in checked_set and c in visible_set]
    st.session_state[persist_key] = final_codes
    return final_codes



def _stable_checkbox_editor_on_change(editor_key: str, code_map_key: str, persist_key: str) -> None:
    """
    v44：data_editor checkbox callback。
    Streamlit data_editor 在 rerun 時，畫面會先用舊 dataframe 重建，容易讓勾選欄位看起來跳回未勾選。
    這個 callback 在 rerun 前先把 edited_rows / edited_cells 寫入穩定的 session_state 清單，
    下一輪重建 dataframe 時即可直接用該清單回填勾選欄位。
    """
    try:
        code_map = [_normalize_code(x) for x in st.session_state.get(code_map_key, []) if _normalize_code(x)]
        if not code_map:
            return

        selected = {_normalize_code(x) for x in st.session_state.get(persist_key, []) if _normalize_code(x)}
        raw_state = st.session_state.get(editor_key, {})
        if not isinstance(raw_state, dict):
            return

        def _is_true(v: Any) -> bool:
            if isinstance(v, bool):
                return bool(v)
            return str(v).strip().lower() in {"true", "1", "yes", "y", "是", "勾選", "checked"}

        def _apply_row(idx: Any, val: Any) -> None:
            try:
                i = int(idx)
            except Exception:
                return
            if i < 0 or i >= len(code_map):
                return
            code = _normalize_code(code_map[i])
            if not code:
                return
            if _is_true(val):
                selected.add(code)
            else:
                selected.discard(code)

        # 新版 Streamlit data_editor state：edited_rows = {row_index: {col: value}}
        edited_rows = raw_state.get("edited_rows", {})
        if isinstance(edited_rows, dict):
            for raw_idx, changes in edited_rows.items():
                if isinstance(changes, dict) and "勾選" in changes:
                    _apply_row(raw_idx, changes.get("勾選"))

        # 舊版 / 部分環境：edited_cells = {"row:column": value}
        edited_cells = raw_state.get("edited_cells", {})
        if isinstance(edited_cells, dict):
            for raw_key, val in edited_cells.items():
                key_text = str(raw_key)
                if ":勾選" in key_text or key_text.endswith(":0"):
                    _apply_row(key_text.split(":", 1)[0], val)

        # 只保留目前表格看得到的股票，避免舊結果殘留；依畫面順序排序。
        visible = set(code_map)
        st.session_state[persist_key] = [c for c in code_map if c in selected and c in visible]
    except Exception:
        # callback 不能讓主頁掛掉
        return

def _format_percent_value(x: Any, digits: int = 2) -> str:
    """V93：百分比欄位安全格式化。

    舊版快取/匯出資料有時已經是字串，例如 "12.30%"、"--"、"資料不足"。
    直接用 f"{x:,.2f}%" 會在 Streamlit Cloud 造成 ValueError。
    這裡統一轉數字；無法轉換時保留原文字，避免整頁中斷。
    """
    try:
        if x is None or pd.isna(x):
            return ""
    except Exception:
        if x is None:
            return ""
    try:
        if isinstance(x, str):
            s = x.strip()
            if not s or s in {"—", "-", "--", "nan", "NaN", "None", "資料不足"}:
                return "" if s in {"nan", "NaN", "None"} else s
            s_num = s.replace("%", "").replace(",", "")
            return f"{float(s_num):,.{digits}f}%"
        return f"{float(x):,.{digits}f}%"
    except Exception:
        return str(x)


def _format_df(df: pd.DataFrame) -> pd.DataFrame:
    show = df.copy()
    price_cols = ["最新價", "推薦買點_突破", "推薦買點_拉回", "近端支撐", "主要支撐", "近端壓力", "突破確認價", "突破確認價_隔日", "回測承接價", "停損參考", "停損價", "停損價_隔日", "第一壓力價", "賣出目標1", "賣出目標2", "PER本益比", "估算EPS"]
    pct_cols = ["區間漲跌幅%", "20日壓力距離%", "20日支撐距離%", "類股平均漲幅", "法人買超占量比%", "3日績效%", "5日績效%", "10日績效%", "20日績效%"]
    score_cols = [
        "訊號分數", "雷達均分", "技術結構分數", "起漲前兆分數", "飆股起漲分數", "大盤可參考分數", "大盤加權分", "大盤市場廣度分數", "大盤量價確認分數", "大盤權值支撐分數", "大盤推薦同步分數", "建議部位%", "建議倉位%", "第一筆進場%", "最大風險%", "風險報酬比", "追價風險分", "停損距離%", "目標報酬%", "交易可行分數",
        "追價風險分數", "拉回買點分數", "突破買點分數",
        "低檔位置分數", "拉回承接分數", "支撐回測分數", "止跌轉強分數", "機會股分數",
        "進場時機分數", "近端支撐", "主要支撐", "近端壓力", "突破確認價", "停損參考", "風險報酬比_決策", "追高風險分數_決策",
        "自動因子總分", "EPS代理分數", "營收動能代理分數", "獲利代理分數",
        "大戶鎖碼代理分數", "法人連買代理分數",
        "個股原始總分", "市場環境分數", "型態突破分數", "爆發力分數", "類股平均總分", "類股平均訊號", "類股熱度分數",
        "類股加速度", "族群資金流分數", "族群資金流加權", "同族群強勢比例", "同族群推薦密度", "同族群平均量能分", "同類股領先幅度", "推薦總分", "風險分數",
        "均線轉強分", "量能啟動分", "突破準備分", "動能翻多分", "支撐防守分",
        "夜間股神總分", "隔日實戰排序分", "隔日進場分數", "波段潛力分數", "技術趨勢分數", "量價動能分數",
        "法人籌碼分數", "大戶鎖碼分數", "基本面成長分數", "營收成長分數", "EPS成長分數", "估值風險分數",
        "外資近1日買賣超", "投信近1日買賣超", "自營商近1日買賣超", "三大法人近1日合計"
    ]

    for c in price_cols:
        if c in show.columns:
            show[c] = show[c].apply(lambda x: format_number(x, 2) if pd.notna(x) else "")
    for c in pct_cols:
        if c in show.columns:
            show[c] = show[c].apply(lambda x: _format_percent_value(x, 2))
    for c in score_cols:
        if c in show.columns:
            show[c] = show[c].apply(lambda x: format_number(x, 1) if pd.notna(x) else "")

    return show


def _save_recommend_result_to_state(rec_df: pd.DataFrame, category_strength_df: pd.DataFrame, hot_pick_df: pd.DataFrame) -> bool:
    """Save only a non-empty usable/reference result.

    A failed scan or a zero-row final filter must not erase the last valid
    recommendation JSON/session result.  The new scan diagnostics remain in
    session_state, while the last non-empty list stays available for reference.
    """
    if rec_df is None or not isinstance(rec_df, pd.DataFrame) or rec_df.empty:
        candidate_df = st.session_state.get(_k("candidate_diagnosis_store"))
        if isinstance(candidate_df, pd.DataFrame) and not candidate_df.empty:
            st.session_state[_k("empty_scan_preserved_previous")] = False
            st.session_state[_k("empty_scan_notice")] = "本輪正式/A-為0，但已保存本輪完整候選快照，不再沿用舊日期。"
            st.session_state[_k("rec_df_store")] = pd.DataFrame()
            st.session_state[_k("category_strength_store")] = category_strength_df.copy() if isinstance(category_strength_df, pd.DataFrame) else pd.DataFrame()
            st.session_state[_k("hot_pick_store")] = hot_pick_df.copy() if isinstance(hot_pick_df, pd.DataFrame) else pd.DataFrame()
            st.session_state[_k("result_saved_at")] = _now_text()
            save_ok, save_msgs = _save_latest_recommendation_pack(pd.DataFrame(), category_strength_df, hot_pick_df)
            st.session_state[_k("latest_pack_permanent_ok")] = bool(save_ok)
            if not save_ok:
                st.session_state[_k("latest_pack_permanent_error")] = "本輪候選已在本機完成，但最新推薦永久錨點未通過遠端驗證；系統不會把它宣稱為永久保存成功。"
            else:
                st.session_state[_k("latest_pack_permanent_error")] = ""
            return bool(save_ok)
        st.session_state[_k("empty_scan_preserved_previous")] = True
        st.session_state[_k("empty_scan_notice")] = "本輪沒有任何候選資料，未覆蓋上一輪快照。"
        return False
    st.session_state[_k("empty_scan_preserved_previous")] = False
    st.session_state[_k("empty_scan_notice")] = ""
    st.session_state[_k("rec_df_store")] = rec_df.copy()
    st.session_state[_k("category_strength_store")] = category_strength_df.copy()
    st.session_state[_k("hot_pick_store")] = hot_pick_df.copy()
    st.session_state[_k("result_saved_at")] = _now_text()
    save_ok, save_msgs = _save_latest_recommendation_pack(rec_df, category_strength_df, hot_pick_df)
    st.session_state[_k("latest_pack_permanent_ok")] = bool(save_ok)
    if not save_ok:
        st.session_state[_k("latest_pack_permanent_error")] = "本輪推薦已在本機完成，但最新推薦永久錨點未通過遠端驗證；請查看『本輪推薦永久保存明細』。"
    else:
        st.session_state[_k("latest_pack_permanent_error")] = ""
    return bool(save_ok)


def _load_recommend_result_from_state() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rec_df = st.session_state.get(_k("rec_df_store"))
    cat_df = st.session_state.get(_k("category_strength_store"))
    hot_df = st.session_state.get(_k("hot_pick_store"))

    if isinstance(rec_df, pd.DataFrame) and isinstance(cat_df, pd.DataFrame) and not rec_df.empty:
        rec_df = _ensure_v92_night_compat_df(rec_df, source="session_rec_df")
        if not isinstance(hot_df, pd.DataFrame):
            hot_df = pd.DataFrame()
        else:
            hot_df = _ensure_v92_night_compat_df(hot_df, source="session_hot_pick")
        # 寫回 session，避免同一輪頁面重繪反覆補欄。
        st.session_state[_k("rec_df_store")] = rec_df.copy()
        st.session_state[_k("hot_pick_store")] = hot_df.copy()
        candidate_df = st.session_state.get(_k("candidate_diagnosis_store"))
        if not isinstance(candidate_df, pd.DataFrame) or candidate_df.empty:
            try:
                st.session_state[_k("candidate_diagnosis_store")] = (
                    build_candidate_diagnosis(rec_df) if callable(build_candidate_diagnosis) else rec_df.copy()
                )
            except Exception:
                pass
        return rec_df.copy(), cat_df.copy(), hot_df.copy()

    rec_df, cat_df, hot_df, saved_at = _load_latest_recommendation_pack()
    if isinstance(rec_df, pd.DataFrame) and not rec_df.empty:
        rec_df = _ensure_v92_night_compat_df(rec_df, source="loaded_rec_df")
        if isinstance(hot_df, pd.DataFrame) and not hot_df.empty:
            hot_df = _ensure_v92_night_compat_df(hot_df, source="loaded_hot_pick")
        st.session_state[_k("rec_df_store")] = rec_df.copy()
        st.session_state[_k("category_strength_store")] = cat_df.copy()
        st.session_state[_k("hot_pick_store")] = hot_df.copy()
        st.session_state[_k("result_saved_at")] = saved_at or _now_text()
        return rec_df.copy(), cat_df.copy(), hot_df.copy()

    return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()


# =========================================================
# Excel 匯出
# =========================================================
@st.cache_data(ttl=300, show_spinner=False)

def _get_full_table_default_cols() -> list[str]:
    return [
        "股票代號", "股票名稱", "市場別", "類別", "類股內排名", "類股前3強",
        "推薦模式", "推薦型態", "機會型態", "推薦等級", "推薦總分", "實戰品質分", "量能狀態", "趨勢狀態", "實戰降分", "夜間股神總分", "隔日實戰排序分", "隔日進場分數", "波段潛力分數",
        "進場型態_隔日", "隔日建議動作", "預估進場點", "回測承接價", "突破確認價_隔日", "停損價_隔日", "第一壓力價", "觀察週期",
        "法人籌碼分數", "大戶鎖碼分數", "基本面成長分數", "營收成長分數", "EPS成長分數", "估值風險分數", "PER本益比", "資料完整度",
        "夜間股神建議", "隔日作戰策略", "夜間風險提醒",
        "上漲機率估計%", "上漲機率等級", "上漲機率信心", "買點分級", "買點狀態", "進場型態", "高分禁買旗標", "高分禁買原因", "實戰買點分數", "實戰操作建議", "支撐距離%", "壓力空間%", "近5日漲幅%", "長上影風險", 
        "機會股分數", "低檔位置分數", "拉回承接分數", "支撐回測分數", "止跌轉強分數",
        "信心等級", "推薦分桶", "市場環境分數",
        "型態名稱", "型態突破分數", "爆發等級", "爆發力分數",
        "技術結構分數", "起漲前兆分數", "起漲等級", "交易可行分數", "類股熱度分數",
        "強勢族群等級", "族群資金流分數", "族群輪動狀態", "同族群強勢比例", "族群策略建議",
        "同類股領先幅度", "是否領先同類股", "建議切入區", "最新價",
        "推薦買點_拉回", "推薦買點_突破", "停損價", "賣出目標1", "賣出目標2",
        "推薦標籤", "機會股說明", "股神推論邏輯", "風險說明", "推薦理由摘要",
        "3日績效%", "5日績效%", "10日績效%", "20日績效%",
    ]


def _get_full_table_order_for_export(rec_df: pd.DataFrame) -> list[str]:
    """
    讓 Excel 的「完整推薦表」與畫面上的「完整推薦表」欄位順序完全一致。
    會吃使用者在完整推薦表欄位管理中套用並永久記錄的順序。
    """
    if rec_df is None or rec_df.empty:
        return []
    available_cols = list(rec_df.columns)
    default_cols = _get_full_table_default_cols()
    saved_order = _load_persistent_column_order("full_table")
    full_order = _normalize_column_order(saved_order if saved_order else default_cols, available_cols, default_cols)
    return [c for c in full_order if c in rec_df.columns]


def _export_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    """Excel 匯出專用：只取存在欄位；如果全缺，回傳原表，避免分頁空白。"""
    if df is None or df.empty:
        return pd.DataFrame()
    use_cols = [c for c in cols if c in df.columns]
    if not use_cols:
        return df.copy()
    return df[use_cols].copy()


def _safe_sort_export_df(df: pd.DataFrame, sort_cols: list[str], ascending: list[bool] | None = None) -> pd.DataFrame:
    """Excel 匯出專用：排序欄位不存在時自動跳過，避免三個榜單空白或錯誤。"""
    if df is None or df.empty:
        return pd.DataFrame()
    work = df.copy()
    real_cols = [c for c in sort_cols if c in work.columns]
    if not real_cols:
        return work.reset_index(drop=True)
    real_asc = []
    if ascending is None:
        real_asc = [False] * len(real_cols)
    else:
        for c in real_cols:
            idx = sort_cols.index(c)
            real_asc.append(ascending[idx] if idx < len(ascending) else False)
    return work.sort_values(real_cols, ascending=real_asc).reset_index(drop=True)




def _phase41_apply_week_battle_columns(rec_export: pd.DataFrame) -> pd.DataFrame:
    """Phase 4.1：集中產生下週作戰分區，避免匯出與畫面各自重複判斷。

    注意：完整推薦表不是買進清單。本函式只做分流欄位，不改原本分數、不寫 JSON。
    """
    if rec_export is None or not isinstance(rec_export, pd.DataFrame) or rec_export.empty:
        return pd.DataFrame()
    work = rec_export.copy()

    def _text_col(col: str) -> pd.Series:
        if col in work.columns:
            return work[col].fillna("").astype(str)
        return pd.Series([""] * len(work), index=work.index, dtype="object")

    role = _text_col("推薦角色")
    hunter_role = _text_col("飆股獵人角色")
    state = _text_col("實戰過濾狀態")
    level = _text_col("股神推薦層級") + "｜" + _text_col("顯示分區") + "｜" + _text_col("推薦用途") + "｜" + _text_col("限制原因")

    # Phase 4.1 修正：不能再因為「真禁買原因 / 硬否決原因」有文字就整列丟排除。
    # B / C+ / C- 本來就會保留風控限制文字，真正排除只看角色 D 或 BLOCK 狀態。
    true_block = role.str.contains("D｜過熱禁買|過熱禁買", na=False) | state.str.contains("BLOCK", na=False)
    role_attack = role.str.contains("S｜飆股攻擊候選|飆股攻擊候選", na=False)
    role_main = role.str.contains("A｜股神主推薦|股神主推薦", na=False)
    role_breakout = role.str.contains("B｜等突破確認|等突破確認", na=False)
    role_early = role.str.contains("C\\+｜早期潛伏|早期潛伏", na=False) | state.str.contains("EARLY", na=False)
    role_weak = role.str.contains("C-｜弱勢觀察|弱勢觀察", na=False) | state.str.contains("WATCH", na=False)
    hunter_breakout = hunter_role.str.contains("B\\+｜盤中突破可追|盤中突破可追|B｜等突破確認|等突破確認", na=False)

    # 以決策角色為主，獵人角色只補沒有明確 B/C 角色的股票，避免 C+/C- 被錯誤升級成可追。
    attack = role_attack & ~true_block
    main = role_main & ~true_block
    early = role_early & ~true_block & ~attack & ~main
    weak = role_weak & ~true_block & ~attack & ~main & ~early
    breakout = (
        role_breakout
        | state.str.contains("WAIT", na=False)
        | _text_col("突破確認狀態").str.contains("WAIT", na=False)
        | (hunter_breakout & ~role_early & ~role_weak)
    ) & ~true_block & ~attack & ~main & ~early & ~weak
    exclude = true_block | ((~attack) & (~main) & (~breakout) & (~early) & (~weak) & level.str.contains("排除|禁買|成交額或成交張數不足|低量", na=False))
    weak = weak & ~exclude

    bucket = pd.Series("觀察名單｜等待條件補強", index=work.index, dtype="object")
    action = pd.Series("只觀察，不主動買進", index=work.index, dtype="object")
    desc = pd.Series("條件尚未完整，等待買點、量能、族群或風控改善。", index=work.index, dtype="object")

    bucket.loc[attack | main] = "下週可進攻名單"
    action.loc[attack] = "盤中觸發後可小量進攻"
    action.loc[main] = "可依條件分批進攻"
    desc.loc[attack] = "S 飆股攻擊候選：不是開盤直接買，需站上觸發價且量能確認。"
    desc.loc[main] = "A 股神主推薦：買點、風控與分數通過，仍需分批與嚴守失效條件。"

    bucket.loc[breakout] = "盤中突破追蹤名單"
    action.loc[breakout] = "突破前不買，站上觸發價再評估"
    desc.loc[breakout] = "B / B+：有題材或族群/資金訊號，但買點未確認；只做盤中突破追蹤。"

    bucket.loc[early] = "早期潛伏觀察名單"
    action.loc[early] = "最多小量潛伏，不追高"
    desc.loc[early] = "C+：早期轉強或低檔潛伏，尚未達主攻條件，需等量價續強。"

    bucket.loc[weak] = "弱勢觀察清單"
    action.loc[weak] = "不買，僅追蹤是否轉強"
    desc.loc[weak] = "C-：訊號不足或風控尚未修復，僅保留觀察。"

    bucket.loc[exclude] = "禁止買進／排除名單"
    action.loc[exclude] = "禁止新倉"
    desc.loc[exclude] = "D / BLOCK：過熱、風控失衡或條件不足，未解除前不進場。"

    work["下週作戰分區"] = bucket
    work["下週作戰說明"] = desc
    work["下週操作動作"] = action
    work["下週是否可進攻"] = pd.Series("否", index=work.index, dtype="object")
    work.loc[attack | main, "下週是否可進攻"] = "是｜但仍需觸發/分批"
    work.loc[breakout, "下週是否可進攻"] = "突破確認後才可"
    work.loc[early, "下週是否可進攻"] = "僅小量潛伏"
    work["下週作戰版本"] = "phase4_2_mainstream_week_plan_20260605"

    # Phase 4.2：主流資金分層。完整推薦表可保留所有候選，但 Excel/畫面分頁必須把冷門股隔離。
    def _num_col(col: str, default: float = 0.0) -> pd.Series:
        if col in work.columns:
            return pd.to_numeric(work[col], errors="coerce").fillna(default).astype(float)
        return pd.Series([default] * len(work), index=work.index, dtype="float64")

    main_score = _num_col("主流資金分", 50)
    amount_m = _num_col("成交額百萬", 0)
    cold_text = _text_col("冷門股警示") + "｜" + _text_col("主流股判定") + "｜" + _text_col("主流資金角色")
    cold = cold_text.str.contains("冷門|低流動性", na=False)
    severe_cold = cold_text.str.contains("低流動性排除|冷門禁追", na=False) | (amount_m < 50)

    main_bucket = pd.Series("弱勢觀察", index=work.index, dtype="object")
    main_action = pd.Series("僅觀察", index=work.index, dtype="object")
    main_desc = pd.Series("尚未通過主流資金與買點條件。", index=work.index, dtype="object")

    main_bucket.loc[attack | main] = "主流攻擊候選"
    main_action.loc[attack | main] = "可依盤中觸發/風控條件進攻"
    main_desc.loc[attack | main] = "主流資金與買點較完整，仍需觸發價、量能與停損條件。"

    main_bucket.loc[breakout] = "主流突破追蹤"
    main_action.loc[breakout] = "突破前不買，放量站上觸發價再評估"
    main_desc.loc[breakout] = "具主流/族群訊號但買點未確認，列入盤中突破追蹤。"

    main_bucket.loc[early] = "早期潛伏觀察"
    main_action.loc[early] = "最多小量觀察，不追高"
    main_desc.loc[early] = "早期轉強但尚未成為主流攻擊，等待量價續強。"

    main_bucket.loc[weak] = "弱勢觀察"
    main_action.loc[weak] = "不買，只追蹤是否轉強"
    main_desc.loc[weak] = "訊號不足或風控尚未修復。"

    main_bucket.loc[exclude] = "禁止買進排除"
    main_action.loc[exclude] = "禁止新倉"
    main_desc.loc[exclude] = "過熱、禁買或風控失衡。"

    main_bucket.loc[cold & ~severe_cold] = "冷門潛伏觀察"
    main_action.loc[cold & ~severe_cold] = "冷門股隔離；不可追高，只能觀察"
    main_desc.loc[cold & ~severe_cold] = "低成交額/低量股，放量訊號容易失真，不列主流突破名單。"

    main_bucket.loc[severe_cold] = "低流動性排除"
    main_action.loc[severe_cold] = "低流動性排除，禁止追高"
    main_desc.loc[severe_cold] = "成交額過低，容易滑價、假突破與無法出場。"

    work["主流作戰分區"] = main_bucket
    work["主流作戰說明"] = main_desc + "｜主流資金分" + main_score.round(1).astype(str) + "｜成交額" + amount_m.round(1).astype(str) + "百萬"
    work["主流操作動作"] = main_action
    return work


def _phase41_split_week_battle_views(rec_export: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Phase 4.2 Excel 分頁：主流攻擊 / 主流突破 / 潛伏 / 冷門隔離 / 低流動性排除。"""
    if rec_export is None or not isinstance(rec_export, pd.DataFrame) or rec_export.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    work = _phase41_apply_week_battle_columns(rec_export)

    def _text_col(col: str) -> pd.Series:
        if col in work.columns:
            return work[col].fillna("").astype(str)
        return pd.Series([""] * len(work), index=work.index, dtype="object")

    bucket = _text_col("主流作戰分區")
    main_attack_mask = bucket.eq("主流攻擊候選")
    main_breakout_mask = bucket.eq("主流突破追蹤")
    early_mask = bucket.eq("早期潛伏觀察")
    cold_mask = bucket.eq("冷門潛伏觀察")
    weak_mask = bucket.eq("弱勢觀察")
    exclude_mask = bucket.isin(["低流動性排除", "禁止買進排除"])

    attack_main_df = _safe_sort_export_df(
        work.loc[main_attack_mask].copy(),
        ["主流資金分", "飆股攻擊分", "股神實戰總分", "Entry進場買點分", "Risk風控安全分", "隔日大漲機率分", "風險報酬比"],
        [False, False, False, False, False, False, False],
    )
    breakout_df = _safe_sort_export_df(
        work.loc[main_breakout_mask].copy(),
        ["主流資金分", "飆股攻擊分", "隔日大漲機率分", "族群攻擊強度", "資金攻擊有效分", "盤中轉強觸發價"],
        [False, False, False, False, False, True],
    )
    early_df = _safe_sort_export_df(
        work.loc[early_mask].copy(),
        ["主流資金分", "股神實戰總分", "Entry進場買點分", "Risk風控安全分", "風險報酬比", "候選強度分"],
        [False, False, False, False, False, False],
    )
    cold_df = _safe_sort_export_df(
        work.loc[cold_mask].copy(),
        ["成交額百萬", "主流資金分", "飆股攻擊分", "股神實戰總分", "族群攻擊強度"],
        [False, False, False, False, False],
    )
    weak_df = _safe_sort_export_df(
        work.loc[weak_mask].copy(),
        ["主流資金分", "候選強度分", "股神實戰總分", "Entry進場買點分", "Risk風控安全分"],
        [False, False, False, False, False],
    )
    exclude_df = _safe_sort_export_df(
        work.loc[exclude_mask].copy(),
        ["主流資金分", "成交額百萬", "最新成交量_張", "候選強度分", "推薦總分"],
        [False, False, False, False, False],
    )
    return attack_main_df, breakout_df, early_df, cold_df, weak_df, exclude_df, work


# 相容舊函式名稱：避免其他舊流程仍呼叫 v151 split 時失效。
def _v151_split_main_observe_views(rec_export: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    attack_main_df, breakout_df, early_df, cold_df, weak_df, exclude_df, _ = _phase41_split_week_battle_views(rec_export)
    # 舊版第二個回傳值是今日主推薦；Phase 4.2 將主流攻擊合併成可進攻名單。
    return attack_main_df, attack_main_df, early_df, breakout_df, pd.concat([cold_df, weak_df, exclude_df], ignore_index=True)

def _build_export_views(rec_df: pd.DataFrame, category_strength_df: pd.DataFrame, top_n: int, full_order: list[str] | None = None):
    if rec_df is None or rec_df.empty:
        empty = pd.DataFrame()
        return empty, empty, empty, empty

    if full_order is None:
        full_order = _get_full_table_order_for_export(rec_df)

    # Excel「完整推薦表」必須和畫面上的完整推薦表欄位一致。
    rec_export = rec_df[[c for c in full_order if c in rec_df.columns]].copy() if full_order else rec_df.copy()
    try:
        if callable(prune_empty_recommendation_columns):
            rec_export = prune_empty_recommendation_columns(rec_export)
    except Exception:
        pass

    # v70 修正：類股強度榜不能只依賴 session_state。
    # 如果快取或舊資料沒有 category_strength_df，就直接用本次推薦結果重算，避免 Excel 分頁空白。
    if isinstance(category_strength_df, pd.DataFrame) and not category_strength_df.empty:
        cat_export = category_strength_df.copy()
    else:
        try:
            cat_export = _compute_category_strength(rec_df.copy())
        except Exception:
            cat_export = pd.DataFrame()

    if isinstance(cat_export, pd.DataFrame) and not cat_export.empty:
        cat_sort_cols = ["族群資金流分數", "類股熱度分數", "類股平均總分", "股票數"]
        cat_export = _safe_sort_export_df(cat_export, cat_sort_cols, [False, False, False, False]).head(top_n).copy()
    else:
        # v72：即使 _compute_category_strength 失敗，也直接從推薦表用最基本欄位重建類股榜，避免分頁空白。
        try:
            base = rec_df.copy()
            cat_col = "類別" if "類別" in base.columns else None
            score_col = "推薦總分" if "推薦總分" in base.columns else None
            if cat_col:
                if score_col:
                    base[score_col] = pd.to_numeric(base[score_col], errors="coerce")
                    cat_export = (
                        base.groupby(cat_col, dropna=False)
                        .agg(股票數=(cat_col, "size"), 類股平均總分=(score_col, "mean"), 類股最高分=(score_col, "max"))
                        .reset_index()
                        .rename(columns={cat_col: "類別"})
                        .sort_values(["類股平均總分", "股票數"], ascending=[False, False])
                        .head(top_n)
                    )
                else:
                    cat_export = (
                        base.groupby(cat_col, dropna=False)
                        .size()
                        .reset_index(name="股票數")
                        .rename(columns={cat_col: "類別"})
                        .sort_values("股票數", ascending=False)
                        .head(top_n)
                    )
        except Exception:
            cat_export = pd.DataFrame()

    # v70 修正：同類股領先榜欄位兼容。
    # 舊推薦紀錄可能缺少部分欄位，過去直接取欄會導致空白/失敗。
    leader_df = _safe_sort_export_df(
        rec_df,
        ["是否領先同類股", "推薦總分", "類股熱度分數", "同類股領先幅度", "類股內排名"],
        [False, False, False, False, True],
    )
    leader_cols = [
        "股票代號", "股票名稱", "市場別", "類別", "類股內排名", "類股前3強",
        "是否領先同類股", "同類股領先幅度", "市場環境分數", "型態名稱", "型態突破分數",
        "爆發力分數", "飆股起漲分數", "起漲前兆分數", "起漲等級", "起漲摘要",
        "個股原始總分", "類股平均總分", "類股熱度分數", "族群資金流分數",
        "強勢族群等級", "推薦總分", "上漲機率估計%", "買點分級", "推薦理由摘要",
    ]
    leader_export = _export_cols(leader_df, leader_cols).head(top_n).copy()

    # v70 修正：自動因子榜欄位兼容。
    # 若 EPS/營收/獲利代理欄位不存在，至少用自動因子總分、推薦總分、技術分數排序並匯出。
    factor_rank = _safe_sort_export_df(
        rec_df,
        ["自動因子總分", "EPS代理分數", "營收動能代理分數", "獲利代理分數", "推薦總分", "起漲前兆分數"],
        [False, False, False, False, False, False],
    )
    factor_cols = [
        "股票代號", "股票名稱", "市場別", "類別", "市場環境分數", "型態名稱", "型態突破分數",
        "爆發等級", "爆發力分數", "自動因子總分", "EPS代理分數", "營收動能代理分數",
        "獲利代理分數", "大戶鎖碼代理分數", "法人連買代理分數", "起漲前兆分數",
        "交易可行分數", "推薦總分", "上漲機率估計%", "自動因子摘要", "推薦理由摘要",
    ]
    factor_export = _export_cols(factor_rank, factor_cols).head(top_n).copy()

    return rec_export, cat_export, leader_export, factor_export


def _excel_safe_value(v):
    """Excel 匯出專用：把 pandas/numpy 型別轉成 openpyxl 可寫入型別。"""
    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass
    try:
        if hasattr(v, "item"):
            v = v.item()
    except Exception:
        pass
    if isinstance(v, (list, tuple, set, dict)):
        try:
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return v


def _excel_safe_df(df: pd.DataFrame, fallback_title: str = "無資料") -> pd.DataFrame:
    """
    v72：Excel 匯出防空白。
    1) 移除重複欄名，避免 Arrow / Excel 顯示異常。
    2) MultiIndex 欄位轉字串。
    3) 空表仍寫入提示列，避免 Excel 分頁完全空白造成誤判。
    """
    if df is None or not isinstance(df, pd.DataFrame):
        return pd.DataFrame({"狀態": [fallback_title]})
    work = df.copy()
    try:
        work.columns = [" / ".join(map(str, c)) if isinstance(c, tuple) else str(c) for c in work.columns]
    except Exception:
        work.columns = [str(c) for c in work.columns]
    try:
        work = work.loc[:, ~pd.Index(work.columns).duplicated()].copy()
    except Exception:
        pass
    if work.empty:
        return pd.DataFrame({"狀態": [fallback_title]})
    real_cols = [c for c in work.columns if str(c).strip() and str(c) != "勾選"]
    if not real_cols:
        return pd.DataFrame({"狀態": [fallback_title]})
    return work.reset_index(drop=True)


def _write_df_to_ws(wb, sheet_name: str, df: pd.DataFrame, fallback_title: str):
    """v72：不用 pandas ExcelWriter，改用 openpyxl 逐格寫入，避免下載後整個分頁空白。"""
    safe_name = str(sheet_name)[:31]
    ws = wb.create_sheet(title=safe_name)
    work = _excel_safe_df(df, fallback_title=fallback_title)
    headers = [str(c) for c in work.columns]
    ws.append(headers)
    try:
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
    except Exception:
        pass
    for _, row in work.iterrows():
        ws.append([_excel_safe_value(row.get(c, "")) for c in work.columns])
    try:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if safe_name == "股神推薦總排名":
            from openpyxl.styles import PatternFill, Font
            from openpyxl.formatting.rule import ColorScaleRule
            ws.sheet_properties.tabColor = "00A6A6"
            header_map = {str(cell.value): cell.column for cell in ws[1]}
            score_col = header_map.get("股神推薦優先分")
            rank_col = header_map.get("股神推薦總排名")
            if score_col and ws.max_row >= 2:
                letter = ws.cell(row=1, column=score_col).column_letter
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    ColorScaleRule(start_type="num", start_value=50, start_color="FECACA", mid_type="num", mid_value=70, mid_color="FEF3C7", end_type="num", end_value=90, end_color="BBF7D0"),
                )
            if rank_col:
                for row_idx in range(2, min(ws.max_row, 4) + 1):
                    ws.cell(row=row_idx, column=rank_col).font = Font(bold=True)
                    ws.cell(row=row_idx, column=rank_col).fill = PatternFill("solid", fgColor=("FFD966" if row_idx == 2 else "D9EAF7"))
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                cell_val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)
    except Exception:
        pass
    return ws




def _phase6_split_market_leader_replay_views(rec_export: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Phase 6 Excel 分頁：領漲回補雷達 / 題材轉強追蹤。

    專門檢討 6/12 類型的漏網股：記憶體、半導體、PCB、被動元件、面板/太陽能等主流領漲族群。
    """
    if rec_export is None or not isinstance(rec_export, pd.DataFrame) or rec_export.empty:
        return pd.DataFrame(), pd.DataFrame()
    work = rec_export.copy()

    def _txt(col: str) -> pd.Series:
        if col in work.columns:
            return work[col].fillna("").astype(str)
        return pd.Series([""] * len(work), index=work.index, dtype="object")

    role = _txt("領漲回補角色")
    bucket = _txt("領漲回補分區")
    score = pd.to_numeric(work.get("主流領漲回補分", 0), errors="coerce").fillna(0)
    theme = pd.to_numeric(work.get("漲停族群相似度", 0), errors="coerce").fillna(0)
    amount_m = pd.to_numeric(work.get("成交額百萬", 0), errors="coerce").fillna(0)

    leader_mask = (
        (bucket.eq("領漲回補雷達") | role.str.contains(r"L\+｜領漲回補雷達|L｜主流強勢回補", na=False))
        & (score >= 72)
        & (theme >= 62)
        & (amount_m >= 80)
    )
    theme_mask = (
        (bucket.eq("題材轉強追蹤") | role.str.contains("T｜題材轉強追蹤", na=False))
        & (score >= 64)
        & (theme >= 58)
        & (amount_m >= 50)
        & ~leader_mask
    )
    sort_cols = ["主流領漲回補分", "市場領漲相似分", "漲停族群相似度", "爆發雷達分", "隔日爆發分", "族群攻擊強度", "成交額百萬"]
    leader_df = _safe_sort_export_df(work.loc[leader_mask].copy(), sort_cols, [False] * len(sort_cols))
    theme_df = _safe_sort_export_df(work.loc[theme_mask].copy(), sort_cols, [False] * len(sort_cols))
    return leader_df, theme_df



def _phase62_split_miss_replay_views(rec_export: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Phase 6.2 Excel 分頁：漲停漏選回放 / 漏選原因診斷 / 已覆蓋雷達。

    這是回放檢討，不等同買進清單。目的是找出「明明像強勢/漲停股，卻被風控、流動性或族群判斷提前降級」的股票。
    """
    if rec_export is None or not isinstance(rec_export, pd.DataFrame) or rec_export.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    work = rec_export.copy()
    try:
        if "回放校正版本" not in work.columns or work.get("回放校正版本", pd.Series(dtype="object")).astype(str).str.strip().eq("").all():
            from godpick_miss_replay_engine import apply_godpick_miss_replay_engine
            work = apply_godpick_miss_replay_engine(work)
    except Exception:
        pass

    def _txt(col: str) -> pd.Series:
        if col not in work.columns:
            return pd.Series([""] * len(work), index=work.index, dtype="object")
        return work[col].fillna("").astype(str).str.strip()

    role = _txt("回放校正角色")
    bucket = _txt("回放校正分區")
    score = pd.to_numeric(work.get("漲停回放分", 0), errors="coerce").fillna(0)
    risk = pd.to_numeric(work.get("強勢股漏選風險分", 0), errors="coerce").fillna(0)

    miss_mask = bucket.eq("漏選回放校正") | role.str.contains(r"M\+｜漲停漏選回放", na=False) | ((score >= 82) & (risk >= 78))
    diag_mask = bucket.eq("漏選原因診斷") | role.str.contains(r"M｜強勢漏選追蹤", na=False) | ((score >= 72) & (risk >= 70))
    covered_mask = bucket.eq("已覆蓋雷達") | role.str.contains(r"K｜已納入雷達", na=False)

    sort_cols = ["漲停回放分", "強勢股漏選風險分", "主流領漲回補分", "爆發雷達分", "族群攻擊強度", "成交額百萬"]
    miss_df = _safe_sort_export_df(work.loc[miss_mask].copy(), sort_cols, [False] * len(sort_cols))
    diag_df = _safe_sort_export_df(work.loc[diag_mask & ~miss_mask].copy(), sort_cols, [False] * len(sort_cols))
    covered_df = _safe_sort_export_df(work.loc[covered_mask & ~miss_mask & ~diag_mask].copy(), sort_cols, [False] * len(sort_cols))
    return miss_df, diag_df, covered_df

def _phase5_split_explosive_radar_views(rec_export: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Phase 5 Excel 分頁：飆股雷達 / 高風險爆發觀察 / 假強排除。

    這是獨立於穩健推薦的第二條路，不把 R 類股票混進主流攻擊候選。
    """
    if rec_export is None or not isinstance(rec_export, pd.DataFrame) or rec_export.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    if any(c in rec_export.columns for c in ["下週作戰版本", "主流作戰分區", "飆股雷達分區", "領漲回補分區"]):
        work = rec_export.copy()
    else:
        try:
            work = _phase41_apply_week_battle_columns(rec_export)
        except Exception:
            work = rec_export.copy()

    def _txt(col: str) -> pd.Series:
        if col in work.columns:
            return work[col].fillna("").astype(str)
        return pd.Series([""] * len(work), index=work.index, dtype="object")

    role = _txt("飆股雷達角色")
    bucket = _txt("飆股雷達分區")
    score = pd.to_numeric(work.get("爆發雷達分", 0), errors="coerce").fillna(0)

    radar_mask = (
        (bucket.eq("飆股雷達") | role.str.contains(r"S\+｜漲停雷達|S｜飆股攻擊候選|B\+｜盤中點火追蹤", na=False))
        & (score >= 62)
    )
    risk_mask = bucket.eq("高風險爆發觀察") | role.str.contains("R｜高風險爆發觀察", na=False)
    fake_mask = bucket.eq("假強排除") | role.str.contains("X｜假強排除", na=False)

    radar_df = _safe_sort_export_df(
        work.loc[radar_mask].copy(),
        ["爆發雷達分", "隔日爆發分", "局部題材火種分", "漏網回補分", "飆股攻擊分", "族群攻擊強度", "主流資金分", "成交額百萬"],
        [False, False, False, False, False, False, False, False],
    )
    risk_df = _safe_sort_export_df(
        work.loc[risk_mask & ~radar_mask].copy(),
        ["爆發雷達分", "隔日爆發分", "局部題材火種分", "漏網回補分", "族群攻擊強度", "成交額百萬"],
        [False, False, False, False, False, False],
    )
    fake_df = _safe_sort_export_df(
        work.loc[fake_mask].copy(),
        ["爆發雷達分", "隔日爆發分", "成交額百萬"],
        [False, False, False],
    )
    return radar_df, risk_df, fake_df


def _phase63_split_formal_recommendation_views(rec_export: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Phase 6.3 Excel 分頁：正式推薦 / 盤中雷達 / 高風險觀察 / 不可直接買 / 正式排除。

    目的：避免完整推薦表、D 禁買、弱勢觀察、高風險雷達混在一起，被誤解為下週可買清單。
    不重算選股，只讀 godpick_formal_recommendation_engine 產生的共用欄位。
    """
    if rec_export is None or not isinstance(rec_export, pd.DataFrame) or rec_export.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    work = rec_export.copy()
    if "正式推薦分區" not in work.columns:
        try:
            from godpick_formal_recommendation_engine import apply_formal_recommendation_engine
            work = apply_formal_recommendation_engine(work)
        except Exception:
            work["正式推薦分區"] = "不可直接買觀察"
            work["正式推薦資格"] = "WATCH｜觀察不買"
            work["下週是否可直接買"] = "不可"
            work["可操作分"] = 0
            work["正式推薦排序分"] = 0
            work["正式推薦動作"] = "只觀察，不主動買進。"

    def _txt(col: str) -> pd.Series:
        if col in work.columns:
            return work[col].fillna("").astype(str)
        return pd.Series([""] * len(work), index=work.index, dtype="object")

    bucket = _txt("正式推薦分區")
    formal_mask = bucket.eq("正式下週主推薦")
    a_minus_mask = bucket.eq("A-｜準主推薦小量試單")
    intraday_mask = bucket.eq("盤中雷達追蹤")
    risk_mask = bucket.eq("高風險雷達觀察")
    watch_mask = bucket.isin(["不可直接買觀察", "早期潛伏觀察"])
    exclude_mask = bucket.eq("正式排除清單")

    sort_cols = ["正式推薦排序分", "可操作分", "Entry進場買點分", "Risk風控安全分", "風險報酬比", "主流資金分", "成交額百萬"]
    radar_sort = ["正式推薦排序分", "爆發雷達分", "隔日爆發分", "主流領漲回補分", "漲停回放分", "族群攻擊強度", "成交額百萬"]
    formal_df = _safe_sort_export_df(work.loc[formal_mask].copy(), sort_cols, [False] * len(sort_cols))
    a_minus_df = _safe_sort_export_df(work.loc[a_minus_mask].copy(), sort_cols, [False] * len(sort_cols))
    intraday_df = _safe_sort_export_df(work.loc[intraday_mask].copy(), radar_sort, [False] * len(radar_sort))
    risk_df = _safe_sort_export_df(work.loc[risk_mask].copy(), radar_sort, [False] * len(radar_sort))
    watch_df = _safe_sort_export_df(work.loc[watch_mask].copy(), ["可操作分", "正式推薦排序分", "股神實戰總分", "主流資金分"], [False, False, False, False])
    exclude_df = _safe_sort_export_df(work.loc[exclude_mask].copy(), ["正式推薦排序分", "可操作分", "追價風險分", "成交額百萬"], [False, False, False, False])
    return formal_df, a_minus_df, intraday_df, risk_df, watch_df, exclude_df


def _phase71_split_intraday_radar_layers(intraday_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Phase 7.1：盤中雷達分層。

    39 檔雷達資料可以保留，但人工盯盤主清單必須收斂。
    - 盤中核心雷達：R1，約 10~12 檔，放第一眼主表。
    - 盤中備援雷達：R2，保留輪動機會。
    - 盤中低優先觀察：R3，資料保留但不作為主盯盤清單。
    """
    if intraday_df is None or not isinstance(intraday_df, pd.DataFrame) or intraday_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    work = intraday_df.copy()
    if "盤中雷達優先級" not in work.columns or "盤中盯盤順序" not in work.columns:
        try:
            from godpick_formal_recommendation_engine import apply_formal_recommendation_engine
            work = apply_formal_recommendation_engine(work)
            work = work.loc[work.get("正式推薦分區", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str).eq("盤中雷達追蹤")].copy()
        except Exception:
            work["盤中雷達優先級"] = "R2｜備援雷達"
            work["盤中盯盤順序"] = range(1, len(work) + 1)
            work["盤中雷達分層"] = "盤中備援雷達"
            work["盤中雷達分層說明"] = "備援輪動名單。"
    def _txt(col: str) -> pd.Series:
        if col in work.columns:
            return work[col].fillna("").astype(str)
        return pd.Series([""] * len(work), index=work.index, dtype="object")
    pri = _txt("盤中雷達優先級")
    sort_cols = ["盤中盯盤順序", "正式推薦排序分", "可操作分", "爆發雷達分", "主流資金分", "成交額百萬"]
    asc = [True, False, False, False, False, False]
    core = _safe_sort_export_df(work.loc[pri.str.startswith("R1")].copy(), sort_cols, asc)
    backup = _safe_sort_export_df(work.loc[pri.str.startswith("R2")].copy(), sort_cols, asc)
    low = _safe_sort_export_df(work.loc[pri.str.startswith("R3")].copy(), sort_cols, asc)
    full = _safe_sort_export_df(work.copy(), sort_cols, asc)
    return core, backup, low, full


def _phase70_build_battle_dashboard(
    formal_df: pd.DataFrame,
    a_minus_df: pd.DataFrame,
    intraday_df: pd.DataFrame,
    risk_df: pd.DataFrame,
    exclude_df: pd.DataFrame,
) -> pd.DataFrame:
    """Phase 8：真正可操作的股神作戰總表。

    第一張作戰表只允許：正式主推薦、A- 準主推薦、R1 盤中核心雷達。
    高風險觀察與正式排除只能留在診斷分頁，絕不可混入主推薦總表。
    ``risk_df`` / ``exclude_df`` 保留參數相容性，但故意不併入主表。
    """
    parts: list[pd.DataFrame] = []
    for zone, df in [
        ("1｜正式主推薦", formal_df),
        ("2｜A-準主推薦", a_minus_df),
        ("3｜盤中核心雷達", intraday_df),
    ]:
        if isinstance(df, pd.DataFrame) and not df.empty:
            tmp = df.copy()
            if "股神作戰區" not in tmp.columns:
                tmp["股神作戰區"] = zone
            parts.append(tmp)
    if not parts:
        return pd.DataFrame()
    work = pd.concat(parts, ignore_index=True, sort=False)
    if "股神作戰優先序" not in work.columns:
        zone_rank = {
            "正式下週主推薦": 10,
            "A-｜準主推薦小量試單": 20,
            "盤中雷達追蹤": 30,
        }
        work["股神作戰優先序"] = work.get("正式推薦分區", pd.Series([""] * len(work))).map(zone_rank).fillna(50)
    cols = [
        "最終操作結論", "是否正式推薦", "操作許可", "正式推薦等級", "候選性質",
        "推薦可信度分", "建議倉位上限%", "風控否決旗標", "決策一致性",
        "股神作戰區", "主要依據工作表", "股票代號", "股票名稱", "類別", "產業",
        "正式推薦分區", "正式推薦資格", "下週是否可直接買", "正式推薦動作",
        "股神推薦總排名", "股神推薦優先分", "股神推薦等級", "股神推薦用途", "股神推薦分數說明",
        "可操作分", "正式推薦排序分", "推薦總分", "買進分數", "Entry進場買點分", "Risk風控安全分", "風險報酬比",
        "主要進場路徑", "主要進場參考價", "回測承接參考價", "突破確認參考價", "守價回測參考價", "守價回測距離%",
        "推薦升級判定路徑", "路徑風險報酬比", "風報比計算口徑", "正式與A近門檻說明",
        "隔日耗竭風險分", "隔日耗竭風險等級", "隔日可執行優先分", "進場績效計算口徑",
        "強勢動能分", "強勢動能判定", "強勢前兆分", "強勢前兆判定", "紅燈逆勢反轉分", "紅燈逆勢反轉判定",
        "大盤風控層級", "大盤條件覆寫", "逆勢操作限制", "官方因子資料日期", "官方因子落後交易日", "官方因子新鮮度", "大盤資料日期", "大盤資料落後交易日", "大盤資料新鮮度", "大盤與K線對齊狀態", "股神資料總新鮮度", "股神資料警示", "紅燈反轉首觸禁買", "主流強勢替代進場", "大盤原始橋接狀態",
        "今日漲幅%", "當日量比", "當日收盤位置%", "突破20日高點%", "上影線比例%", "強勢前兆進場條件", "強勢前兆風控", "紅燈逆勢反轉分", "紅燈逆勢反轉判定", "大盤風控層級", "大盤條件覆寫", "逆勢操作限制", "盤前強勢前兆分", "前置保留類型", "前置保留原因",
        "主流主升優先分", "主流主升判定", "主流主升操作限制",
        "主流資金分", "族群輪動分", "族群攻擊強度", "族群廣度分", "族群成交額分", "族群主升確認",
        "爆發雷達分", "隔日爆發分", "漲停回放分", "強勢股漏選風險分",
        "實戰觸發價", "觸發後守價", "盤中觸發確認條件", "開盤跳空處理",
        "盤中雷達優先級", "盤中盯盤順序", "盤中雷達分層", "盤中雷達分層說明", "核心雷達品質檢查", "核心雷達降級原因",
        "盤中雷達動作", "正式推薦排除原因", "股神作戰提示",
    ]
    use = [c for c in cols if c in work.columns]
    if use:
        work = work[use + [c for c in work.columns if c not in use]].copy()
    return _safe_sort_export_df(
        work,
        ["股神作戰優先序", "主流主升優先分", "正式推薦排序分", "可操作分", "推薦可信度分", "爆發雷達分", "成交額百萬"],
        [True, False, False, False, False, False, False],
    )


def _phase80_build_recommendation_summary(
    formal_df: pd.DataFrame,
    a_minus_df: pd.DataFrame,
    intraday_core_df: pd.DataFrame,
    risk_df: pd.DataFrame,
    exclude_df: pd.DataFrame,
    total_candidates: int,
    scan_report: dict[str, Any] | None = None,
) -> pd.DataFrame:
    formal_n = len(formal_df) if isinstance(formal_df, pd.DataFrame) else 0
    a_minus_n = len(a_minus_df) if isinstance(a_minus_df, pd.DataFrame) else 0
    if isinstance(a_minus_df, pd.DataFrame) and not a_minus_df.empty:
        _a_permit = a_minus_df.get("操作許可", pd.Series([""] * len(a_minus_df), index=a_minus_df.index)).fillna("").astype(str)
        _a_cap = pd.to_numeric(a_minus_df.get("建議倉位上限%", 0), errors="coerce").fillna(0)
        a_blocked_n = int((_a_permit.str.contains("禁止|等大盤|封鎖", regex=True, na=False) | _a_cap.le(0)).sum())
    else:
        a_blocked_n = 0
    a_actionable_n = max(0, a_minus_n - a_blocked_n)
    intraday_n = len(intraday_core_df) if isinstance(intraday_core_df, pd.DataFrame) else 0
    risk_n = len(risk_df) if isinstance(risk_df, pd.DataFrame) else 0
    exclude_n = len(exclude_df) if isinstance(exclude_df, pd.DataFrame) else 0
    report = scan_report if isinstance(scan_report, dict) else {}
    quality = _safe_str(report.get("掃描品質狀態")) or "未知｜舊資料未記錄掃描完整性"
    usable = bool(report.get("正式推薦可用", False))

    limited = usable and _safe_str(report.get("掃描品質等級")) in {"limited", "warning"}
    scope = _safe_str(report.get("推薦適用範圍"))
    prefix = f"{scope}：" if limited and scope else ""
    if not usable:
        status = quality or "資料品質不足｜禁止正式推薦"
        conclusion = (
            "掃描或資料品質未達正式操作標準。系統已禁止輸出正式買進結論；"
            "資料缺失不會再被當成低流動性，請先補齊成交額/成交量或修復K線來源。"
        )
    elif formal_n > 0:
        conclusion = f"{prefix}本輪有 {formal_n} 檔正式推薦；只依進場區、觸發價與停損分批操作。"
        status = "限定資料池｜有正式推薦" if limited else "有正式推薦"
    elif a_minus_n > 0:
        if a_blocked_n == a_minus_n:
            conclusion = (
                f"{prefix}本輪有 {a_minus_n} 檔個股達 A-／正式候選資格，但目前全數受大盤紅燈封鎖；"
                "建議倉位0%，等待大盤解除風控後再做盤中觸發確認。"
            )
            status = "A-資格候選｜大盤封鎖"
        else:
            conclusion = (
                f"{prefix}本輪有 {a_minus_n} 檔 A- 準主推薦，其中 {a_actionable_n} 檔可等待盤中觸發，"
                f"{a_blocked_n} 檔受大盤風控封鎖。"
            )
            status = "限定資料池｜僅準主推薦" if limited else "無直接買進｜僅準主推薦"
    elif intraday_n > 0:
        conclusion = f"{prefix}本輪沒有正式推薦；保留 {intraday_n} 檔盤中核心雷達，未觸發前不可買。"
        status = "限定資料池｜只看盤中雷達" if limited else "無正式推薦｜只看盤中雷達"
    else:
        conclusion = f"{prefix}沒有股票同時通過買點、風控與風險報酬門檻；系統選擇空手等待。"
        status = "限定資料池｜空手等待" if limited else "完整掃描｜空手等待"

    return pd.DataFrame([{
        "本輪結論": status,
        "掃描品質狀態": quality,
        "正式推薦可用": "是" if usable else "否",
        "預計掃描數": int(report.get("預計掃描數", 0) or 0),
        "成功分析數": int(report.get("成功分析數", 0) or 0),
        "掃描覆蓋率%": float(report.get("掃描覆蓋率%", 0) or 0),
        "歷史資料成功率%": float(report.get("歷史資料成功率%", 0) or 0),
        "有效K線資料率%": float(report.get("有效K線資料率%", report.get("歷史資料成功率%", 0)) or 0),
        "流動性資料覆蓋率%": float(report.get("流動性資料覆蓋率%", 0) or 0),
        "官方因子覆蓋率%": float(report.get("官方因子覆蓋率%", 0) or 0),
        "官方紀錄匹配率%": float(report.get("官方紀錄匹配率%", 0) or 0),
        "官方有效因子覆蓋率%": float(report.get("官方有效因子覆蓋率%", report.get("官方因子覆蓋率%", 0)) or 0),
        "官方最新可信覆蓋率%": float(report.get("官方最新可信覆蓋率%", 0) or 0),
        "官方來源可信覆蓋率%": float(report.get("官方來源可信覆蓋率%", 0) or 0),
        "官方日期T-1內覆蓋率%": float(report.get("官方日期T-1內覆蓋率%", 0) or 0),
        "官方同日對齊覆蓋率%": float(report.get("官方同日對齊覆蓋率%", 0) or 0),
        "官方落後1日覆蓋率%": float(report.get("官方落後1日覆蓋率%", 0) or 0),
        "推薦適用範圍": scope,
        "倉位折減係數": float(report.get("倉位折減係數", 0) or 0),
        "完整候選診斷數": int(report.get("完整候選診斷數", total_candidates) or total_candidates or 0),
        "正式推薦檔數": formal_n,
        "A-準主推薦檔數": a_minus_n,
        "A-可操作檔數": a_actionable_n,
        "A-大盤封鎖檔數": a_blocked_n,
        "盤中核心雷達檔數": intraday_n,
        "高風險觀察檔數": risk_n,
        "正式排除檔數": exclude_n,
        "候選診斷總數": int(total_candidates or 0),
        "操作說明": conclusion,
        "掃描品質說明": _safe_str(report.get("掃描品質說明")),
        "核心紀律": "採四路徑：波段型看 Entry/Risk/RR；R1-M 看已發動量價；R1-P 看主流資金、族群與起漲前兆；R1-RB 看紅燈後恐慌反彈領漲。所有 R1 都是條件雷達，不可開盤盲目追價。",
        "版本": "phase9_5_persistent_records_panic_rebound_20260721",
    }])


def _phase80_allowed_codes(rec_df: pd.DataFrame, selected_codes: list[str], target: str) -> tuple[list[str], list[str]]:
    """防止正式排除/高風險觀察被寫成推薦紀錄或推薦清單。"""
    selected = [_normalize_code(x) for x in selected_codes if _normalize_code(x)]
    scan_report = st.session_state.get(_k("scan_quality_report"), {})
    formal_scan_ok = bool(isinstance(scan_report, dict) and scan_report.get("正式推薦可用", False))
    if target != "record" and not formal_scan_ok:
        # 推薦清單仍需完整掃描；推薦紀錄可保留個股資料合格的R1研究雷達。
        return [], selected
    if rec_df is None or not isinstance(rec_df, pd.DataFrame) or rec_df.empty or not selected:
        return [], selected
    work = rec_df[rec_df["股票代號"].astype(str).map(_normalize_code).isin(set(selected))].copy()
    if "正式推薦分區" not in work.columns:
        try:
            from godpick_formal_recommendation_engine import apply_formal_recommendation_engine
            work = apply_formal_recommendation_engine(work)
        except Exception:
            return [], selected
    bucket = work.get("正式推薦分區", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str)
    radar_pri = work.get("盤中雷達優先級", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str)
    if target == "record":
        # V159：推薦紀錄是模型績效資料庫，正式/A- 與 R1（含 R1-M 強勢動能、R1-P 強勢前兆）都必須留下；
        # 以「正式推薦分區 / 盤中雷達優先級」區分，不把雷達冒充正式買進推薦。
        formal_mask = bucket.isin(["正式下週主推薦", "A-｜準主推薦小量試單"])
        radar_mask = bucket.eq("盤中雷達追蹤") & radar_pri.str.startswith("R1")
        allowed_mask = (formal_mask if formal_scan_ok else pd.Series([False] * len(work), index=work.index)) | radar_mask
        if callable(assess_individual_sample_quality):
            quality_mask = work.apply(lambda r: bool(assess_individual_sample_quality(r)[0]), axis=1)
            allowed_mask &= quality_mask
    elif target == "list":
        allowed_mask = bucket.isin(["正式下週主推薦", "A-｜準主推薦小量試單"]) | (bucket.eq("盤中雷達追蹤") & radar_pri.str.startswith("R1"))
    else:
        allowed_mask = pd.Series([True] * len(work), index=work.index)
    allowed = work.loc[allowed_mask, "股票代號"].astype(str).map(_normalize_code).tolist()
    allowed = list(dict.fromkeys([x for x in allowed if x]))
    rejected = [x for x in selected if x not in set(allowed)]
    return allowed, rejected



# =========================================================
# V159 推薦紀錄完整保存 + 每輪自動記錄
# =========================================================
def _v159_record_level_from_row(row: pd.Series | dict[str, Any]) -> str:
    bucket = _safe_str(row.get("正式推薦分區"))
    radar = _safe_str(row.get("盤中雷達優先級"))
    if bucket == "正式下週主推薦":
        return "正式主推薦"
    if bucket == "A-｜準主推薦小量試單":
        return "A-準主推薦"
    if bucket == "盤中雷達追蹤" and radar.startswith("R1-RB"):
        return "R1-RB恐慌反彈領漲雷達"
    if bucket == "盤中雷達追蹤" and radar.startswith("R1-M"):
        return "R1-M強勢動能雷達"
    if bucket == "盤中雷達追蹤" and radar.startswith("R1-P"):
        return "R1-P強勢前兆雷達"
    if bucket == "盤中雷達追蹤" and radar.startswith("R1"):
        return "R1核心雷達"
    return bucket or radar or "未分層"


def _normalize_godpick_record(row: dict[str, Any]) -> dict[str, Any]:
    """V159：保留推薦列全部共用欄位，再補齊紀錄控制欄位。

    舊版由固定小字典重建資料，會把正式推薦分區、R1-M、進場可執行性、
    強勢動能證據與 K 線日期全部丟掉；推薦紀錄因此無法還原當天決策。
    """
    raw = dict(row or {})
    rec_price = _safe_float(raw.get("推薦價格"), _safe_float(raw.get("最新價")))
    latest_price = _safe_float(raw.get("最新價"), rec_price)
    stop_price = _safe_float(raw.get("停損價"), _safe_float(raw.get("實戰停損參考"), _safe_float(raw.get("停損參考"))))
    target1 = _safe_float(raw.get("賣出目標1"), _safe_float(raw.get("第一壓力價")))
    target2 = _safe_float(raw.get("賣出目標2"))
    rec_date = _safe_str(raw.get("推薦日期")) or _now_date_text()
    rec_time = _safe_str(raw.get("推薦時間")) or _now_time_text()
    mode = _safe_str(raw.get("推薦模式")) or "股神推薦"

    pnl_amt = None
    pnl_pct = None
    if rec_price not in [None, 0] and latest_price is not None:
        pnl_amt = latest_price - rec_price
        pnl_pct = pnl_amt / rec_price * 100

    raw.update({
        "record_id": _safe_str(raw.get("record_id")) or _safe_str(raw.get("rec_id")) or _create_record_id(
            _normalize_code(raw.get("股票代號")), rec_date, rec_time, mode
        ),
        "股票代號": _normalize_code(raw.get("股票代號")),
        "股票名稱": _safe_str(raw.get("股票名稱")),
        "市場別": _safe_str(raw.get("市場別")) or "上市",
        "類別": _normalize_category(raw.get("類別")),
        "推薦模式": mode,
        "推薦日期": rec_date,
        "推薦時間": rec_time,
        "建立時間": _safe_str(raw.get("建立時間")) or _now_text(),
        "更新時間": _now_text(),
        "目前狀態": _safe_str(raw.get("目前狀態")) or ("雷達觀察" if "R1" in _safe_str(raw.get("紀錄層級")) else "觀察"),
        "推薦價格": rec_price,
        "最新價": latest_price,
        "停損價": stop_price,
        "賣出目標1": target1,
        "賣出目標2": target2,
        "損益金額": pnl_amt if raw.get("損益金額") in [None, ""] else raw.get("損益金額"),
        "損益幅%": pnl_pct if raw.get("損益幅%") in [None, ""] else raw.get("損益幅%"),
        "是否已實際買進": _safe_str(raw.get("是否已實際買進")) in {"是", "True", "true", "1"},
        "是否達停損": bool(stop_price is not None and latest_price is not None and latest_price <= stop_price),
        "是否達目標1": bool(target1 is not None and latest_price is not None and latest_price >= target1),
        "是否達目標2": bool(target2 is not None and latest_price is not None and latest_price >= target2),
    })
    return _ensure_godpick_record_columns(pd.DataFrame([raw])).iloc[0].to_dict()


def _build_record_rows_from_rec_df(rec_df: pd.DataFrame, selected_codes: list[str]) -> list[dict[str, Any]]:
    """V159：直接保存完整推薦列，不再用舊版固定欄位字典截斷新引擎證據。"""
    if rec_df is None or not isinstance(rec_df, pd.DataFrame) or rec_df.empty or "股票代號" not in rec_df.columns:
        return []
    codes = {_normalize_code(x) for x in (selected_codes or []) if _normalize_code(x)}
    if not codes:
        return []
    work = rec_df[rec_df["股票代號"].astype(str).map(_normalize_code).isin(codes)].copy()
    rec_date = _now_date_text()
    rec_time = _now_time_text()
    build_time = _now_text()
    rows: list[dict[str, Any]] = []
    for _, r in work.iterrows():
        raw = r.to_dict()
        code = _normalize_code(raw.get("股票代號"))
        if not code:
            continue
        level = _v159_record_level_from_row(raw)
        latest = _safe_float(raw.get("最新價"), _safe_float(raw.get("推薦價格")))
        stop = _safe_float(raw.get("實戰停損參考"), _safe_float(raw.get("停損參考"), _safe_float(raw.get("停損價"))))
        raw.update({
            "record_id": _create_record_id(code, rec_date, rec_time, _safe_str(raw.get("推薦模式")) or "股神推薦"),
            "股票代號": code,
            "股票名稱": _safe_str(raw.get("股票名稱")),
            "市場別": _safe_str(raw.get("市場別")) or "上市",
            "類別": _normalize_category(raw.get("類別")),
            "推薦模式": _safe_str(raw.get("推薦模式")) or "股神推薦",
            "推薦日期": rec_date,
            "推薦時間": rec_time,
            "建立時間": build_time,
            "更新時間": build_time,
            "最新更新時間": build_time,
            "紀錄來源": _safe_str(raw.get("紀錄來源")) or "07_股神推薦",
            "自動記錄": _safe_str(raw.get("自動記錄")) or "否",
            "紀錄層級": level,
            "本輪推薦版本": _safe_str(raw.get("正式推薦版本")) or _safe_str(raw.get("決策版本")) or "V159",
            "目前狀態": "雷達觀察" if level.startswith("R1") else "新推薦",
            "推薦價格": latest,
            "推薦日價格": latest,
            "最新價": latest,
            "停損價": stop,
            "K線驗證標記": "已建立K線驗證資料",
            "K線查詢參數": f"stock_code={code}&source=godpick",
            "K線檢視提示": "至 3_歷史K線分析，對照推薦價、支撐、壓力、停損與後續走勢。",
            "是否已實際買進": False,
            "損益金額": None,
            "損益幅%": None,
            "是否達停損": False,
            "是否達目標1": False,
            "是否達目標2": False,
            "備註": _safe_str(raw.get("備註")),
        })
        rows.append(_normalize_godpick_record(raw))
    return rows


def _v159_auto_record_actionable_recommendations(source_df: pd.DataFrame, *, background_write: bool = False) -> tuple[int, list[str]]:
    """保存正式/A-/R1/R1-M/R1-P；整體掃描不足時仍保留個股資料合格的雷達樣本。

    正式/A- 仍需整體掃描達正式可用；R1/R1-M 是研究型雷達，只要該檔個股
    K線、價格與成交資料完整即可保存，避免全域覆蓋率讓校正資料整批歸零。
    """
    action, formal_scan_ok, partition_notes = _v191_actionable_tracking_frame(source_df)
    if action.empty:
        return 0, list(partition_notes or ["本輪沒有正式/A-/R1可自動記錄資料。"] )

    quality_notes: dict[str, tuple[str, str]] = {}
    for _, row in action.iterrows():
        code = _normalize_code(row.get("股票代號"))
        if callable(assess_individual_sample_quality):
            try:
                eligible, reason, confidence = assess_individual_sample_quality(row)
            except Exception as exc:
                eligible, reason, confidence = True, f"品質判定例外沿用：{exc}", "中"
        else:
            eligible, reason, confidence = True, "未載入個股品質服務，沿用既有判定", "中"
        quality_notes[code] = (reason, confidence)

    action["紀錄來源"] = "07_股神推薦｜推薦完成自動記錄"
    action["自動記錄"] = "是"
    _exec_ctx_v191 = st.session_state.get(_k("recommend_execution_context_v191"), {})
    if not isinstance(_exec_ctx_v191, dict):
        _exec_ctx_v191 = {}
    action["推薦執行來源"] = _safe_str(_exec_ctx_v191.get("owner")) or "07_股神推薦"
    action["推薦觸發方式"] = _safe_str(_exec_ctx_v191.get("trigger")) or "手動操作"
    action["紀錄層級"] = action.apply(_v159_record_level_from_row, axis=1)

    def _sample_meta(row: pd.Series) -> pd.Series:
        level = _safe_str(row.get("紀錄層級"))
        code = _normalize_code(row.get("股票代號"))
        _, confidence = quality_notes.get(code, ("", "中"))
        if level == "正式主推薦":
            sample_type, weight, formal_perf = "A｜正式交易樣本", 1.00, "是"
        elif level == "A-準主推薦":
            sample_type, weight, formal_perf = "A-｜準主推薦樣本", 0.90, "是"
        elif level.startswith("R1-RB"):
            sample_type, weight, formal_perf = "B｜R1-RB恐慌反彈領漲雷達", 0.72, "否"
        elif level.startswith("R1-M"):
            sample_type, weight, formal_perf = "B｜R1-M強勢動能雷達", 0.75, "否"
        elif level.startswith("R1-P"):
            sample_type, weight, formal_perf = "B｜R1-P強勢前兆雷達", 0.70, "否"
        else:
            sample_type, weight, formal_perf = "B｜R1核心雷達", 0.75, "否"
        return pd.Series({
            "校正樣本類型": sample_type,
            "校正樣本用途": "正式推薦績效與權重校正" if formal_perf == "是" else "雷達觸發、動能延續與失效條件校正",
            "校正樣本權重": weight,
            "是否納入正式推薦績效": formal_perf,
            "是否納入權重校正": "是",
            "個股資料品質": "可追蹤",
            "樣本可信度": confidence,
            "校正樣本建立版本": CALIBRATION_SAMPLE_VERSION,
        })

    meta_df = action.apply(_sample_meta, axis=1)
    for col in meta_df.columns:
        action[col] = meta_df[col]
    codes = action["股票代號"].astype(str).map(_normalize_code).tolist()
    rows = _build_record_rows_from_rec_df(action, codes)
    if background_write and rows and callable(upsert_records_authority_fast):
        try:
            _page07_record_authority_executor_v181().submit(_v181_background_record_upsert, copy.deepcopy(rows))
            st.session_state[_k("v181_record_authority_scheduled_at")] = _now_text()
            messages = [
                f"V181：{len(rows)} 筆正式/A-/R1紀錄已排程背景權威寫入；畫面不再等待 GitHub/Firestore 權威恢復。",
                "寫入仍使用 code＋推薦日期＋推薦模式 business key 防重；背景完成後第8頁會自動讀取最新 authority。",
            ]
            if not formal_scan_ok:
                messages.insert(0, "整體掃描未達正式可用：本輪僅保存個股資料合格的R1/R1-M研究雷達，不宣稱正式推薦。")
            return len(rows), messages
        except Exception as exc:
            # 排程失敗才退回同步，不能因此遺失正式推薦紀錄。
            fallback_note = f"V181背景排程失敗，改同步權威寫入：{exc}"
        else:
            fallback_note = ""
    else:
        fallback_note = ""
    added, messages = _append_godpick_records(rows, force_duplicate=False)
    messages = [*[f"H7行動分區｜{x}" for x in (partition_notes or [])], *messages]
    if fallback_note:
        messages = [fallback_note, *messages]
    if not formal_scan_ok:
        messages = ["整體掃描未達正式可用：本輪僅保存個股資料合格的R1/R1-M研究雷達，不宣稱正式推薦。", *messages]
    return added, messages



def _phase90_build_master_recommendation_rank(source_df: pd.DataFrame, top_n: int = 30) -> pd.DataFrame:
    """建立唯一第一優先的「股神推薦總排名」。

    保留原本正式/A-/R1/R2/觀察分頁，但使用者不必在多張表間自行比較。
    V188 起第一排序採 ``V188股神作戰優先分``（Alpha × Trade × RR × 資料證據 ×
    追價治理 × 類股集中度）；舊 ``股神推薦優先分`` 保留作 Alpha/候選強度參考。
    正式排除永不進榜；好股票如果現在不是好買點，Trade Grade 會明確降級。
    """
    if source_df is None or not isinstance(source_df, pd.DataFrame) or source_df.empty:
        return pd.DataFrame()
    try:
        work = canonicalize_final_partition(source_df) if callable(canonicalize_final_partition) else source_df.copy()
    except Exception:
        work = source_df.copy()
    try:
        from godpick_formal_recommendation_engine import apply_formal_recommendation_engine
        required = {"股神推薦優先分", "股神推薦總排名", "股神推薦等級", "股神推薦用途"}
        if not required.issubset(set(work.columns)):
            work = apply_formal_recommendation_engine(work)
    except Exception:
        pass
    if work.empty or "股票代號" not in work.columns:
        return pd.DataFrame()

    _v188_rank_diag = {}
    if callable(inspect_v188_decision_frame):
        try:
            _v188_rank_diag = inspect_v188_decision_frame(work)
        except Exception as _rank_diag_err:
            _v188_rank_diag = {"complete": False, "reason": f"V188排名檢查失敗：{_rank_diag_err}"}
    else:
        _v188_rank_diag = {"complete": all(c in work.columns for c in ["V188股神作戰優先分", "SuperAI Alpha分", "SuperAI Trade分", "V188交易許可"]), "reason": "fallback-column-check"}
    if not bool(_v188_rank_diag.get("complete")):
        st.session_state[_k("v188_rank_block_reason_v189")] = _safe_str(_v188_rank_diag.get("reason")) or "V188交易品質資料尚未完成"
        return pd.DataFrame()
    st.session_state[_k("v188_rank_block_reason_v189")] = ""

    bucket = work.get("正式推薦分區", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str)
    _rank_score_col = "V188股神作戰優先分"
    score = pd.to_numeric(work.get(_rank_score_col, 0), errors="coerce").fillna(0.0)
    freshness = work.get("K線資料新鮮度", pd.Series([""] * len(work), index=work.index)).fillna("").astype(str)
    # 正式排除不進主排名；資料過期只留在診斷表，不應占用第一眼名單。
    keep = ~bucket.eq("正式排除清單")
    keep &= score.ge(50.0)
    keep &= ~freshness.str.contains("過期|落後|待更新", regex=True, na=False)
    rank = work.loc[keep].copy()
    if rank.empty:
        return pd.DataFrame()

    sort_cols = [
        "V188股神作戰優先分", "SuperAI Trade分", "SuperAI Alpha分", "股神推薦優先分", "今日訊號新鮮分", "主流主升優先分", "隔日可執行優先分", "實戰操作品質分", "進場可執行分", "強勢動能分",
        "強勢前兆分", "主流資金分", "族群攻擊強度", "流動性參考成交額百萬",
    ]
    for col in sort_cols:
        if col not in rank.columns:
            rank[col] = 0.0
        rank[col] = pd.to_numeric(rank[col], errors="coerce").fillna(0.0)
    rank = rank.sort_values(sort_cols, ascending=[False] * len(sort_cols), kind="mergesort")
    rank["股票代號"] = rank["股票代號"].astype(str).map(_normalize_code)
    rank = rank.loc[rank["股票代號"].ne("")].drop_duplicates(subset=["股票代號"], keep="first")
    rank = rank.head(max(1, int(top_n or 30))).copy().reset_index(drop=True)
    rank["股神推薦總排名"] = range(1, len(rank) + 1)

    cols = [
        "股神推薦總排名", "V188股神作戰優先分", "SuperAI Alpha分", "SuperAI Alpha等級", "SuperAI Trade分", "SuperAI Trade等級", "SuperAI最終作戰等級",
        "V188交易許可", "V188正式推薦資格", "SuperAI執行風報比", "SuperAI風報比來源", "V188RR治理", "V188T+1追價治理",
        "V188個股資料證據", "V188市場對齊治理", "V188類股集中治理", "V188類股集中扣分", "SuperAI校準後隔日上漲機率%",
        "股神推薦優先分", "股神推薦等級", "股神推薦用途",
        "今日訊號新鮮分", "近5次入榜次數", "連續入榜次數", "重複推薦校正分",
        "推薦輪動狀態", "今日新進榜", "前次推薦排名", "本次分數變化", "重複推薦說明",
        "主流主升優先分", "主流主升判定", "主流主升操作限制",
        "股票代號", "股票名稱", "市場別", "類別", "產業",
        "最終操作結論", "操作許可", "是否正式推薦", "正式推薦分區", "盤中雷達優先級", "核心雷達品質檢查", "核心雷達降級原因",
        "推薦總分", "候選強度分", "實戰操作品質分", "進場可執行分", "買進分數",
        "Entry進場買點分", "Risk風控安全分", "實戰風險報酬比", "風險報酬比", "追價風險分",
        "主要進場路徑", "主要進場參考價", "回測承接參考價", "突破確認參考價", "守價回測參考價", "守價回測距離%",
        "推薦升級判定路徑", "路徑風險報酬比", "風報比計算口徑", "正式與A近門檻說明",
        "隔日耗竭風險分", "隔日耗竭風險等級", "隔日可執行優先分", "進場績效計算口徑",
        "強勢動能分", "強勢動能判定", "強勢前兆分", "強勢前兆判定",
        "紅燈逆勢反轉分", "紅燈逆勢反轉判定", "大盤風控層級", "大盤條件覆寫", "逆勢操作限制",
        "官方因子資料日期", "官方因子落後交易日", "官方因子新鮮度", "大盤資料日期", "大盤資料落後交易日", "大盤資料新鮮度", "大盤與K線對齊狀態", "股神資料總新鮮度", "股神資料警示", "紅燈反轉首觸禁買", "主流強勢替代進場", "大盤原始橋接狀態",
        "主流資金分", "族群輪動分", "族群攻擊強度", "族群廣度分", "族群成交額分", "族群主升確認",
        "今日漲幅%", "當日量比", "當日收盤位置%",
        "最新價", "預估進場點", "實戰觸發價", "觸發後守價", "守價回測參考價", "守價回測距離%", "實戰停損參考", "第一壓力價",
        "建議倉位上限%", "正式推薦動作", "盤中觸發確認條件", "失效條件",
        "股神推薦分數說明", "正式推薦排除原因",
    ]
    use = [c for c in cols if c in rank.columns]
    return rank[use + [c for c in rank.columns if c not in use]].copy()


def _phase90_navigation_table() -> pd.DataFrame:
    return pd.DataFrame([
        {"優先序": 1, "活頁/表格": "股神推薦總排名", "真正用途": "唯一第一優先；V188 依 Alpha×Trade 交易品質排序，舊股神分數只保留作股票強度參考", "是否買進清單": "否，仍須看V188交易許可與觸發條件"},
        {"優先序": 2, "活頁/表格": "正式下週主推薦", "真正用途": "已通過買點、風控與風報比，可依條件分批操作", "是否買進清單": "是，仍須盤中確認"},
        {"優先序": 3, "活頁/表格": "A-準主推薦小量試單", "真正用途": "接近正式門檻，只能觸發且守價後小量試單", "是否買進清單": "條件式"},
        {"優先序": 4, "活頁/表格": "強勢動能核心雷達", "真正用途": "已發動強勢股；只等回測守住或再突破放量", "是否買進清單": "不是，禁止開盤盲追"},
        {"優先序": 5, "活頁/表格": "強勢前兆核心雷達", "真正用途": "尚未發動但主流、族群與前兆較完整的股票", "是否買進清單": "不是，等待觸發"},
        {"優先序": 6, "活頁/表格": "資料待更新雷達", "真正用途": "模型找到候選，但個股K線不是最新；更新資料並重新推薦前不得操作", "是否買進清單": "禁止買進"},
        {"優先序": 7, "活頁/表格": "候選診斷總表", "真正用途": "模型檢討、漏選原因與所有候選證據", "是否買進清單": "絕對不是"},
    ])

def _phase92_render_zero_formal_diagnostics(source_df: pd.DataFrame) -> None:
    """正式與 A- 同時為 0 時，顯示真正的門檻阻擋與最接近升級候選。

    這張診斷不會硬湊推薦數量；它用來辨識是資料、大盤、流動性、
    買點、風報比或風控哪一層把候選全部擋下，避免使用者只看到 0。
    """
    if source_df is None or not isinstance(source_df, pd.DataFrame) or source_df.empty:
        return
    try:
        work = canonicalize_final_partition(source_df) if callable(canonicalize_final_partition) else source_df.copy()
    except Exception:
        work = source_df.copy()
    try:
        required = {"推薦升級判定路徑", "路徑風險報酬比", "正式與A近門檻說明"}
        if not required.issubset(set(work.columns)):
            from godpick_formal_recommendation_engine import apply_formal_recommendation_engine
            work = apply_formal_recommendation_engine(work)
    except Exception:
        pass
    if work.empty:
        return

    reason_cols = [c for c in [
        "正式與A近門檻說明", "進場阻擋原因", "正式推薦排除原因", "真禁買原因",
        "不可直接買原因", "隔日風險標記", "前置保留原因",
    ] if c in work.columns]
    if reason_cols:
        reason_text = work[reason_cols].fillna("").astype(str).agg("｜".join, axis=1)
    else:
        reason_text = pd.Series([""] * len(work), index=work.index)

    blocker_rules = [
        ("資料/K線未更新", r"K線|資料待更新|資料缺失|非最新"),
        ("大盤風控禁止", r"大盤|紅燈|全面防守|禁止進攻"),
        ("流動性不足", r"低流動性|冷門|成交額.*不足|流動性資料"),
        ("離買點過遠", r"距可執行買點|距買點|觸發距離"),
        ("風險報酬不足", r"路徑RR|風報比|上方空間"),
        ("停損距離過大", r"停損距離|停損過遠"),
        ("Entry買點不足", r"Entry"),
        ("Risk風控不足", r"Risk"),
        ("追價/過熱", r"追價|過熱|禁買|耗竭"),
    ]
    blocker_rows = []
    for label, pattern in blocker_rules:
        count = int(reason_text.str.contains(pattern, regex=True, na=False).sum())
        if count > 0:
            blocker_rows.append({"主要阻擋層": label, "候選檔數": count})
    if blocker_rows:
        blocker_df = pd.DataFrame(blocker_rows).sort_values("候選檔數", ascending=False, kind="mergesort")
    else:
        blocker_df = pd.DataFrame([{"主要阻擋層": "尚無結構化阻擋原因", "候選檔數": len(work)}])

    st.warning(
        "正式推薦與 A- 同時為 0，不代表市場沒有強勢股；代表目前沒有股票通過對應的可執行路徑。"
        "下方會列出阻擋層與最接近升級候選，系統不會為了湊數把弱股包裝成正式推薦。"
    )
    with st.expander("為什麼正式／A- 都是 0？｜門檻阻擋與近門檻候選", expanded=True):
        st.dataframe(blocker_df, use_container_width=True, hide_index=True)

        score_cols = [c for c in [
            "股神推薦優先分", "隔日可執行優先分", "實戰操作品質分", "進場可執行分",
            "Entry進場買點分", "Risk風控安全分", "路徑風險報酬比", "主流資金分", "族群攻擊強度",
        ] if c in work.columns]
        for col in score_cols:
            work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0.0)
        if score_cols:
            work = work.sort_values(score_cols, ascending=[False] * len(score_cols), kind="mergesort")
        if "股票代號" in work.columns:
            work["股票代號"] = work["股票代號"].astype(str).map(_normalize_code)
            work = work.loc[work["股票代號"].ne("")].drop_duplicates("股票代號", keep="first")
        near_cols = [c for c in [
            "股票代號", "股票名稱", "股神推薦優先分", "正式推薦分區", "推薦升級判定路徑",
            "進場可執行判定", "進場可執行分", "路徑風險報酬比", "風報比計算口徑",
            "距最近可執行買點%", "Entry進場買點分", "Risk風控安全分", "追價風險分",
            "流動性參考成交額百萬", "正式與A近門檻說明",
        ] if c in work.columns]
        if near_cols:
            st.caption("以下只是最接近升級的診斷候選；是否可買仍以操作許可、觸發價與守價條件為準。")
            st.dataframe(_format_df(work.head(12)[near_cols]), use_container_width=True, hide_index=True)



_PHASE93_MARKET_CONTEXT_COLUMNS = [
    "大盤風險燈號", "大盤橋接風控", "大盤策略模式", "大盤策略建議",
    "大盤風控建議", "今日大盤結論", "大盤橋接狀態", "大盤橋接分數",
    "大盤多空分數", "大盤資料品質", "大盤資料日期", "大盤資料落後交易日", "大盤資料新鮮度",
    "大盤橋接更新時間", "大盤原始橋接狀態", "大盤交易時段", "大盤交易時段可用",
    "隔日大盤方向", "隔日大盤分數", "隔日大盤信心", "隔日大盤預測加減分",
    "隔日建議總部位上限%", "隔日大盤預測理由",
]


def _phase93_single_source_decision_frame(
    rec_df: pd.DataFrame | None,
    candidate_df: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """以同一份完整候選池重建最終分區，供畫面、排名、摘要與 Excel 共用。

    V181：完整掃描已在 _build_recommend_df 內完成一次正式決策。若本輪快取存在且
    scan signature 一致，直接重用完整 decision frame；禁止畫面/快照/自動紀錄再次
    對 1000~2000 檔執行 apply_formal_recommendation_engine。

    舊版主排名使用 candidate_diagnosis_store，但該診斷表已裁掉大盤欄位；
    排名函式再次套用正式推薦引擎時，便把紅燈市場誤當成中性，產生 A-；
    摘要與作戰表卻使用保留大盤欄位的 rec_df，因此統計仍為 0。
    Phase 9.3 先把目前大盤情境補回完整候選池，再只執行一次最終引擎，
    之後所有 UI/Excel 分頁一律從同一份 decision frame 分流。
    """
    _cached_decision = st.session_state.get(_k("decision_frame_store_v181"))
    _cached_sig = _safe_str(st.session_state.get(_k("decision_frame_scan_signature_v181")))
    _active_sig = _safe_str(st.session_state.get(_k("active_scan_signature_v181")))
    if (
        isinstance(_cached_decision, pd.DataFrame)
        and not _cached_decision.empty
        and _cached_sig
        and _cached_sig == _active_sig
        and "正式推薦分區" in _cached_decision.columns
    ):
        _cache_ok = False
        _cache_diag = {}
        if callable(inspect_v188_decision_frame):
            try:
                _cache_diag = inspect_v188_decision_frame(_cached_decision)
                _cache_ok = bool(_cache_diag.get("complete"))
            except Exception as _cache_check_err:
                _cache_diag = {"complete": False, "reason": f"V188快取檢查失敗：{_cache_check_err}"}
        else:
            _cache_ok = all(c in _cached_decision.columns for c in ["V188股神作戰優先分", "SuperAI Alpha分", "SuperAI Trade分", "V188交易許可"])

        if not _cache_ok and callable(repair_v188_decision_frame) and callable(apply_super_ai_engine):
            try:
                _repaired, _repair_diag = repair_v188_decision_frame(
                    _cached_decision,
                    super_ai_callable=apply_super_ai_engine,
                    official_factor_callable=_apply_official_factor_cache_v109,
                    scan_quality_callable=apply_scan_quality_to_frame if callable(apply_scan_quality_to_frame) else None,
                    scan_report=st.session_state.get(_k("scan_quality_report"), {}),
                    canonicalize_callable=canonicalize_final_partition if callable(canonicalize_final_partition) else None,
                )
                if bool(_repair_diag.get("complete")):
                    _repaired["V181最終決策已完成"] = "是"
                    _repaired["V181最終決策版本"] = PAGE07_SPEED_FIX_VERSION
                    _repaired["V189_V188最終快取完整"] = "是"
                    _repaired["V189快取守門版本"] = V189_CACHE_GUARD_VERSION
                    st.session_state[_k("decision_frame_store_v181")] = _repaired.copy()
                    st.session_state[_k("decision_frame_scan_signature_v181")] = _cached_sig
                    st.session_state[_k("v188_cache_integrity_v189")] = dict(_repair_diag)
                    st.session_state[_k("v189_v188_cache_repairs")] = int(st.session_state.get(_k("v189_v188_cache_repairs"), 0) or 0) + 1
                    _cached_decision = _repaired
                    _cache_ok = True
                    _cache_diag = dict(_repair_diag)
            except Exception as _repair_err:
                _cache_diag = {"complete": False, "reason": f"V188快取自動補算失敗：{_repair_err}"}

        if _cache_ok:
            st.session_state[_k("v181_decision_cache_hits")] = int(st.session_state.get(_k("v181_decision_cache_hits"), 0) or 0) + 1
            st.session_state[_k("v188_rank_block_reason_v189")] = ""
            return _cached_decision.copy().reset_index(drop=True)

        # V189：不能把半成品快取當 V188 完成品。讓流程走 fallback 重建；若仍失敗，
        # 主排名會明確封鎖，不再用舊股神分數補 0 冒充 V188。
        st.session_state[_k("v188_rank_block_reason_v189")] = _safe_str(_cache_diag.get("reason")) or "V188快取欄位不完整"

    if isinstance(candidate_df, pd.DataFrame) and not candidate_df.empty:
        source = candidate_df.copy()
    elif isinstance(rec_df, pd.DataFrame) and not rec_df.empty:
        source = rec_df.copy()
    else:
        return pd.DataFrame()

    context = rec_df.copy() if isinstance(rec_df, pd.DataFrame) and not rec_df.empty else pd.DataFrame()

    # 全市場大盤欄位屬於同一輪的全域情境。候選診斷表若已裁欄，
    # 由原始推薦結果第一個有效值回填，避免重新分類時失去市場風控。
    if not context.empty:
        for col in _PHASE93_MARKET_CONTEXT_COLUMNS:
            if col not in context.columns:
                continue
            values = context[col].dropna()
            values = values.loc[values.astype(str).str.strip().ne("")]
            if values.empty:
                continue
            current_missing = col not in source.columns
            if not current_missing:
                current_text = source[col].fillna("").astype(str).str.strip()
                current_missing = bool(current_text.eq("").all())
            if current_missing:
                source[col] = values.iloc[0]

    # 若原始結果也沒有大盤欄位，再讀 01 大盤走勢的永久橋接檔補齊。
    market_evidence = False
    for col in ["大盤風險燈號", "大盤橋接風控", "大盤策略模式", "大盤橋接狀態"]:
        if col in source.columns and source[col].fillna("").astype(str).str.strip().ne("").any():
            market_evidence = True
            break
    if not market_evidence:
        try:
            bridge = _read_macro_mode_bridge()
            if isinstance(bridge, dict) and bridge:
                source = _apply_macro_bridge_columns(source, bridge, enabled=True)
        except Exception:
            pass

    try:
        from godpick_formal_recommendation_engine import apply_formal_recommendation_engine
        st.session_state[_k("v181_decision_cache_misses")] = int(st.session_state.get(_k("v181_decision_cache_misses"), 0) or 0) + 1
        source = apply_formal_recommendation_engine(source)
    except Exception:
        # 若正式引擎暫時載入失敗，仍使用既有欄位，不讓整頁崩潰。
        pass
    try:
        source = canonicalize_final_partition(source) if callable(canonicalize_final_partition) else source
    except Exception:
        pass

    # V189 fallback：舊 Session / Reboot 後若只有 pre-V188 候選，也必須先補齊
    # 官方因子、掃描治理與 SuperAI/V188，再交給主排名。
    if callable(repair_v188_decision_frame) and callable(apply_super_ai_engine):
        try:
            source, _fallback_diag = repair_v188_decision_frame(
                source,
                super_ai_callable=apply_super_ai_engine,
                official_factor_callable=_apply_official_factor_cache_v109,
                scan_quality_callable=apply_scan_quality_to_frame if callable(apply_scan_quality_to_frame) else None,
                scan_report=st.session_state.get(_k("scan_quality_report"), {}),
                canonicalize_callable=canonicalize_final_partition if callable(canonicalize_final_partition) else None,
            )
            st.session_state[_k("v188_cache_integrity_v189")] = dict(_fallback_diag)
            if bool(_fallback_diag.get("complete")) and _active_sig:
                source["V181最終決策已完成"] = "是"
                source["V181最終決策版本"] = PAGE07_SPEED_FIX_VERSION
                source["V189_V188最終快取完整"] = "是"
                source["V189快取守門版本"] = V189_CACHE_GUARD_VERSION
                st.session_state[_k("decision_frame_store_v181")] = source.copy()
                st.session_state[_k("decision_frame_scan_signature_v181")] = _active_sig
                st.session_state[_k("v188_rank_block_reason_v189")] = ""
            elif not bool(_fallback_diag.get("complete")):
                st.session_state[_k("v188_rank_block_reason_v189")] = _safe_str(_fallback_diag.get("reason")) or "V188 fallback補算未完成"
        except Exception as _fallback_v188_err:
            st.session_state[_k("v188_rank_block_reason_v189")] = f"V188 fallback補算失敗：{_fallback_v188_err}"
    return source.reset_index(drop=True)

def _phase80_render_actionable_panel(rec_df: pd.DataFrame) -> None:
    if rec_df is None or not isinstance(rec_df, pd.DataFrame) or rec_df.empty:
        return
    scan_report = st.session_state.get(_k("scan_quality_report"), {})
    candidate_df = st.session_state.get(_k("candidate_diagnosis_store"))
    decision_source = _phase93_single_source_decision_frame(rec_df, candidate_df)
    if decision_source.empty:
        return
    formal, a_minus, intraday, risk, watch, excluded = _phase63_split_formal_recommendation_views(decision_source)
    live_intraday, stale_intraday = _phase94_split_intraday_by_kline_freshness(intraday)
    core, _, _, _ = _phase71_split_intraday_radar_layers(live_intraday)
    candidate_n = len(decision_source)
    summary = _phase80_build_recommendation_summary(
        formal, a_minus, core, risk, excluded, candidate_n, scan_report=scan_report
    )

    rank_source = decision_source
    master_rank = _phase90_build_master_recommendation_rank(rank_source, top_n=20)
    render_pro_section("股神推薦總排名｜真正第一優先")
    st.caption("V188 第一順位改看『V188股神作戰優先分』與 Alpha×Trade：好股票不等於現在是好買點。RR、T+1追價、逐股資料證據、大盤對齊與類股集中都會壓低 Trade Grade；舊『股神推薦優先分/等級』保留作股票強度參考，不能單獨當買進許可。")
    if callable(rotation_diagnostics):
        try:
            rotation_info = rotation_diagnostics(decision_source)
        except Exception:
            rotation_info = {"available": False}
        if rotation_info.get("available"):
            if rotation_info.get("warning"):
                st.warning(
                    f"推薦黏著警示：前10名有 {rotation_info.get('top10_repeat3_count', 0)} 檔近5次至少入榜3次，"
                    f"其中 {rotation_info.get('top10_sticky_without_signal', 0)} 檔缺少今日新訊號。"
                    "系統已套用輪動校正；請優先看『推薦輪動狀態』與『今日訊號新鮮分』。"
                )
            else:
                st.info(
                    f"推薦輪動檢查：前10名新進榜 {rotation_info.get('top10_new_count', 0)} 檔，"
                    f"平均今日訊號新鮮分 {rotation_info.get('top10_average_signal_freshness', 0):.1f}。"
                )
    if isinstance(master_rank, pd.DataFrame) and not master_rank.empty:
        top_row = master_rank.iloc[0]
        render_pro_kpi_row([
            {"label": "第一名", "value": f"{_safe_str(top_row.get('股票代號'))} {_safe_str(top_row.get('股票名稱'))}", "delta": _safe_str(top_row.get("SuperAI最終作戰等級")) or _safe_str(top_row.get("股神推薦用途"))},
            {"label": "V188作戰優先分", "value": format_number(top_row.get("V188股神作戰優先分", top_row.get("股神推薦優先分")), 1), "delta": _safe_str(top_row.get("V188交易許可"))},
            {"label": "Alpha / Trade", "value": f"{_safe_str(top_row.get('SuperAI Alpha等級')) or '-'} / {_safe_str(top_row.get('SuperAI Trade等級')) or '-'}", "delta": f"RR {format_number(top_row.get('SuperAI執行風報比'),2)}"},
            {"label": "操作許可", "value": _safe_str(top_row.get("V188交易許可")) or _safe_str(top_row.get("操作許可")) or _safe_str(top_row.get("下週是否可直接買")), "delta": "Prediction ≠ Permission"},
            {"label": "前20名", "value": str(len(master_rank)), "delta": "依交易品質排序"},
        ])
        master_cols = [c for c in [
            "股神推薦總排名", "V188股神作戰優先分", "SuperAI Alpha分", "SuperAI Alpha等級", "SuperAI Trade分", "SuperAI Trade等級", "SuperAI最終作戰等級",
            "V188交易許可", "V188正式推薦資格", "SuperAI執行風報比", "SuperAI風報比來源", "V188RR治理", "V188T+1追價治理",
            "V188個股資料證據", "V188市場對齊治理", "V188類股集中治理", "SuperAI校準後隔日上漲機率%",
            "股神推薦優先分", "股神推薦等級", "股神推薦用途",
            "今日訊號新鮮分", "近5次入榜次數", "連續入榜次數", "重複推薦校正分",
            "推薦輪動狀態", "今日新進榜", "前次推薦排名", "本次分數變化",
            "主流主升優先分", "主流主升判定", "主流主升操作限制",
            "推薦資格路徑", "資料受限A-", "A-建議單檔上限%", "推薦漏斗階段",
            "股票代號", "股票名稱", "類別", "最終操作結論", "操作許可",
            "推薦總分", "隔日可執行優先分", "實戰操作品質分", "Entry進場買點分", "Risk風控安全分",
            "主要進場路徑", "主要進場參考價", "推薦升級判定路徑", "路徑風險報酬比", "風報比計算口徑", "正式與A近門檻說明",
            "隔日耗竭風險分", "隔日耗竭風險等級", "強勢動能分", "強勢前兆分",
            "主流資金分", "族群輪動分", "族群攻擊強度", "族群廣度分", "族群成交額分", "族群主升確認",
            "大盤資料日期", "大盤資料落後交易日", "大盤資料新鮮度",
            "最新價", "實戰觸發價", "觸發後守價", "實戰停損參考", "正式推薦動作",
        ] if c in master_rank.columns]
        st.dataframe(_format_df(master_rank[master_cols]), use_container_width=True, hide_index=True)
    else:
        _v188_block_reason = _safe_str(st.session_state.get(_k("v188_rank_block_reason_v189")))
        if _v188_block_reason:
            st.error(
                "V189 已封鎖不完整的 V188 排名：" + _v188_block_reason +
                "。系統不會再以舊『股神推薦優先分』補 0 冒充 V188；請重新推薦，或讓本頁自動補算完成後再看排名。"
            )
        elif isinstance(stale_intraday, pd.DataFrame) and not stale_intraday.empty:
            st.error(
                f"不是完全沒有候選：模型找到 {len(stale_intraday)} 檔盤中雷達，但其個股K線不是最新交易日，"
                "因此已封鎖排名與操作。請先更新個股K線後重新推薦；Excel 會列在『資料待更新雷達』。"
            )
        else:
            st.info("目前沒有資料新鮮且分數達 50 分的推薦/觀察候選；請先完成最新行情掃描。")

    render_pro_section("分區作戰明細｜正式／A-／核心雷達")
    # 推薦漏斗：明確分辨「市場沒有合格股票」與「資料來源把資格鎖死」。
    funnel_bucket = decision_source.get("正式推薦分區", pd.Series([""] * len(decision_source), index=decision_source.index)).fillna("").astype(str)
    fresh_series = decision_source.get("K線資料新鮮度", pd.Series([""] * len(decision_source), index=decision_source.index)).fillna("").astype(str)
    official_series = decision_source.get("官方因子新鮮度", pd.Series([""] * len(decision_source), index=decision_source.index)).fillna("").astype(str)
    data_limited_series = decision_source.get("資料受限A-", pd.Series([""] * len(decision_source), index=decision_source.index)).fillna("").astype(str)
    funnel_fresh = int(fresh_series.str.contains("最新", na=False).sum())
    funnel_official = int(official_series.str.contains("最新|對齊|READY|已驗證T-1|降級可用", regex=True, na=False).sum())
    funnel_a = int(funnel_bucket.eq("A-｜準主推薦小量試單").sum())
    funnel_formal = int(funnel_bucket.eq("正式下週主推薦").sum())
    funnel_dq = int(data_limited_series.eq("是").sum())
    st.caption(
        f"推薦通過漏斗：候選 {len(decision_source)} → 最新K線 {funnel_fresh} → 官方因子對齊 {funnel_official} "
        f"→ A- {funnel_a}（其中資料受限 {funnel_dq}）→ 正式推薦 {funnel_formal}。"
        "資料受限A-每日最多1檔、單檔最多1%，不等同正式推薦。"
    )
    row = summary.iloc[0]
    render_pro_kpi_row([
        {"label": "本輪結論", "value": _safe_str(row.get("本輪結論")), "delta": ""},
        {"label": "預計掃描", "value": str(int(row.get("預計掃描數", 0))), "delta": "檔"},
        {"label": "成功分析", "value": str(int(row.get("成功分析數", 0))), "delta": f"{float(row.get('有效K線資料率%', 0)):.1f}%"},
        {"label": "流動性覆蓋", "value": f"{float(row.get('流動性資料覆蓋率%', 0)):.1f}%", "delta": _safe_str(row.get("推薦適用範圍"))},
        {"label": "官方有效覆蓋", "value": f"{float(row.get('官方有效因子覆蓋率%', row.get('官方因子覆蓋率%', 0))):.1f}%", "delta": f"最新可信 {float(row.get('官方最新可信覆蓋率%', 0)):.1f}%｜來源可信 {float(row.get('官方來源可信覆蓋率%', 0)):.1f}%｜T-1內 {float(row.get('官方日期T-1內覆蓋率%', 0)):.1f}%"},
        {"label": "正式推薦", "value": str(int(row.get("正式推薦檔數", 0))), "delta": "檔"},
        {"label": "A-準主推薦", "value": str(int(row.get("A-準主推薦檔數", 0))), "delta": f"可操作{int(row.get('A-可操作檔數', 0))}／封鎖{int(row.get('A-大盤封鎖檔數', 0))}"},
    ])

    scan_usable = bool(scan_report.get("正式推薦可用", False)) if isinstance(scan_report, dict) else False
    if isinstance(stale_intraday, pd.DataFrame) and not stale_intraday.empty:
        st.error(
            f"資料新鮮度封鎖：{len(stale_intraday)} 檔雷達候選使用落後K線，已禁止列入當日作戰表。"
            "這些股票不是消失，而是被移到『資料待更新雷達』等待重掃。"
        )
    if not scan_usable:
        st.warning(_safe_str(row.get("操作說明")) or "目前為條件式參考清單；下單前請更新最新價並重新推薦。")
        if _safe_str(row.get("掃描品質說明")):
            st.caption(_safe_str(row.get("掃描品質說明")))
    else:
        if len(formal) > 0:
            st.success(_safe_str(row.get("操作說明")))
        else:
            st.warning(_safe_str(row.get("操作說明")))
    if scan_usable and len(formal) == 0 and len(a_minus) == 0:
        _phase92_render_zero_formal_diagnostics(rank_source)
    battle = _phase70_build_battle_dashboard(formal, a_minus, core, risk, excluded)
    if callable(build_action_table):
        try:
            battle = build_action_table(battle, include_intraday=True)
        except Exception:
            pass
    st.caption(_safe_str(row.get("核心紀律")))

    if isinstance(battle, pd.DataFrame) and not battle.empty:
        show_cols = [c for c in [
            "最終操作結論", "股票代號", "股票名稱", "類別", "是否正式推薦", "操作許可",
            "股神推薦總排名", "股神推薦優先分", "股神推薦等級", "股神推薦用途",
            "主流主升優先分", "主流主升判定", "主流主升操作限制",
            "正式推薦等級", "正式推薦判定來源", "實戰操作品質分", "推薦可信度分", "模型隔日上漲機率%", "模型預測信心分", "模型預測等級", "模型下行風險%", "崩跌後反彈過熱", "推薦資格路徑", "資料受限A-", "A-建議單檔上限%", "推薦漏斗階段", "候選強度分", "建議倉位上限%",
            "Entry進場買點分", "Risk風控安全分", "實戰風險報酬比", "風險報酬比", "追價風險分",
            "主要進場路徑", "主要進場參考價", "回測承接參考價", "突破確認參考價", "守價回測參考價", "守價回測距離%",
            "推薦升級判定路徑", "路徑風險報酬比", "風報比計算口徑", "正式與A近門檻說明",
            "隔日耗竭風險分", "隔日耗竭風險等級", "隔日可執行優先分", "進場績效計算口徑", "流動性參考成交額百萬",
            "強勢動能分", "強勢動能判定", "強勢前兆分", "強勢前兆判定", "紅燈逆勢反轉分", "紅燈逆勢反轉判定",
            "大盤風控層級", "大盤條件覆寫", "逆勢操作限制", "大盤資料日期", "大盤資料落後交易日", "大盤資料新鮮度",
            "主流資金分", "族群輪動分", "族群攻擊強度", "族群廣度分", "族群成交額分", "族群主升確認",
            "今日漲幅%", "當日量比", "當日收盤位置%", "動能進場條件", "動能風險控制", "強勢前兆進場條件", "強勢前兆風控",
            "最新價", "預估進場點", "實戰觸發價", "觸發後守價", "守價回測參考價", "守價回測距離%", "實戰停損參考", "實戰停損距離%", "實戰壓力空間%", "停損參考", "第一壓力價",
            "正式推薦動作", "失效條件",
        ] if c in battle.columns]
        st.dataframe(_format_df(battle[show_cols]), use_container_width=True, hide_index=True)
    elif scan_usable:
        st.info("完整掃描已完成，但本輪沒有可列入作戰表的正式推薦、A-準主推薦或盤中核心雷達；空手也是正式決策。")


def _phase82_compact_operational_view(df: pd.DataFrame, purpose: str) -> pd.DataFrame:
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    work = canonicalize_final_partition(df) if callable(canonicalize_final_partition) else df.copy()
    work["分頁用途"] = purpose
    common = [
        "分頁用途", "最終操作結論", "股票代號", "股票名稱", "市場別", "類別", "產業",
        "正式推薦分區", "是否正式推薦", "操作許可", "正式推薦等級", "正式推薦判定來源", "候選性質",
        "盤中雷達優先級", "盤中盯盤順序", "盤中雷達分層", "核心雷達品質檢查", "核心雷達降級原因",
        "股神推薦總排名", "股神推薦優先分", "股神推薦等級", "股神推薦用途", "股神推薦分數說明",
        "主流主升優先分", "主流主升判定", "主流主升操作限制",
        "候選強度分", "股神實戰總分", "可操作分", "實戰操作品質分", "推薦可信度分", "模型隔日上漲機率%", "模型預測信心分", "模型預測等級", "模型預估超額報酬%", "模型下行風險%", "模型預測限制", "崩跌後反彈過熱",
        "資料完整度評分", "買進分數", "Entry進場買點分", "Risk風控安全分",
        "主要進場路徑", "主要進場參考價", "回測承接參考價", "突破確認參考價", "守價回測參考價", "守價回測距離%", "隔日耗竭風險分", "隔日耗竭風險等級", "隔日可執行優先分", "進場績效計算口徑",
        "風險報酬比", "追價風險分", "停損距離%", "壓力空間%", "近5日漲幅%", "近20日漲幅%",
        "強勢動能分", "強勢動能判定", "強勢前兆分", "強勢前兆判定", "紅燈逆勢反轉分", "紅燈逆勢反轉判定", "大盤風控層級", "大盤條件覆寫", "逆勢操作限制", "官方因子資料日期", "官方因子落後交易日", "官方因子新鮮度", "大盤資料日期", "大盤資料落後交易日", "大盤資料新鮮度", "大盤與K線對齊狀態", "股神資料總新鮮度", "股神資料警示", "紅燈反轉首觸禁買", "主流強勢替代進場", "大盤原始橋接狀態", "今日漲幅%", "開盤跳空%", "當日量比", "當日收盤位置%", "突破20日高點%", "上影線比例%", "動能進場條件", "動能風險控制", "強勢前兆進場條件", "強勢前兆風控", "紅燈逆勢反轉分", "紅燈逆勢反轉判定", "大盤風控層級", "大盤條件覆寫", "逆勢操作限制", "盤前強勢前兆分", "前置保留類型", "前置保留原因", "前置軟篩選狀態", "前置軟篩選階段", "前置軟篩選原因", "前置軟篩選數", "AI發現母體版本",
        "主流資金分", "族群輪動分", "族群攻擊強度", "族群廣度分", "族群成交額分", "族群主升確認", "成交額百萬", "20日均成交額百萬", "流動性參考成交額百萬", "流動性等級", "流動性資料狀態", "流動性資料來源",
        "最新價", "預估進場點", "實戰觸發價", "觸發後守價", "停損參考", "第一壓力價",
        "建議倉位上限%", "正式推薦動作", "正式推薦排除原因", "失效條件", "開盤跳空處理",
        "引擎輔助訊號", "分區互斥檢查", "掃描品質狀態", "正式推薦可用",
    ]
    cols = [c for c in common if c in work.columns]
    return work[cols].copy().reset_index(drop=True)




def _phase94_split_intraday_by_kline_freshness(intraday_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """將盤中雷達分成「可用最新 K 線」與「資料待更新」。

    正式推薦可用與研究雷達可顯示是兩個不同層級：官方因子不足時可以
    顯示研究雷達，但 K 線落後的候選不得冒充當日盤中雷達。
    """
    if intraday_df is None or not isinstance(intraday_df, pd.DataFrame) or intraday_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    work = intraday_df.copy()
    freshness = work.get(
        "K線資料新鮮度",
        pd.Series([""] * len(work), index=work.index),
    ).fillna("").astype(str)
    fresh_mask = freshness.eq("最新交易日")
    return work.loc[fresh_mask].copy(), work.loc[~fresh_mask].copy()


def _phase94_mark_research_only(frame: pd.DataFrame, reason: str) -> pd.DataFrame:
    """官方因子不足時保留研究雷達，但明確禁止當成買進清單。"""
    if frame is None or not isinstance(frame, pd.DataFrame) or frame.empty:
        return pd.DataFrame()
    out = frame.copy()
    out["是否正式推薦"] = "否"
    out["操作許可"] = "僅供研究｜不可直接買進"
    out["最終操作結論"] = "RESEARCH｜研究雷達：資料補齊後重新評分"
    out["正式推薦動作"] = "先補齊官方因子，再重新推薦；未完成前禁止下單。"
    warning = out.get("股神資料警示", pd.Series([""] * len(out), index=out.index)).fillna("").astype(str)
    out["股神資料警示"] = warning.where(warning.ne(""), reason)
    return out


def _phase94_build_stale_radar_export(stale_df: pd.DataFrame) -> pd.DataFrame:
    """輸出被新鮮度閘門擋下的雷達，避免 Excel 看起來像完全沒有資料。"""
    if stale_df is None or not isinstance(stale_df, pd.DataFrame) or stale_df.empty:
        return pd.DataFrame()
    out = _phase82_compact_operational_view(
        stale_df,
        "資料待更新雷達｜只供診斷，更新 K 線後重新推薦",
    )
    if out.empty:
        return out
    out["是否正式推薦"] = "否"
    out["操作許可"] = "資料待更新｜禁止買進"
    out["最終操作結論"] = "DATA-WAIT｜K線落後：更新後重新評分"
    out["正式推薦動作"] = "先更新個股 K 線，再重新推薦；目前不得依舊價操作。"
    warning = out.get("股神資料警示", pd.Series([""] * len(out), index=out.index)).fillna("").astype(str)
    fallback = out.get("K線資料新鮮度", pd.Series(["K線日期未驗證"] * len(out), index=out.index)).fillna("K線日期未驗證").astype(str)
    out["股神資料警示"] = warning.where(warning.ne(""), fallback)
    return out


def _build_excel_bytes(
    rec_export: pd.DataFrame,
    cat_export: pd.DataFrame,
    leader_export: pd.DataFrame,
    factor_export: pd.DataFrame,
    candidate_diagnosis_export: pd.DataFrame | None = None,
    scan_report: dict[str, Any] | None = None,
) -> bytes:
    """Phase 8.3 practical workbook with evidence-aware liquidity recovery.

    Operational lists are mutually exclusive.  Raw multi-engine signals are
    consolidated into one diagnostic sheet instead of duplicating the same stock
    across contradictory tabs.  The full analysed pool is exported separately
    from the filtered action list.
    """
    from openpyxl import Workbook

    output = io.BytesIO()
    wb = Workbook()
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    raw_candidate_source = candidate_diagnosis_export if isinstance(candidate_diagnosis_export, pd.DataFrame) and not candidate_diagnosis_export.empty else None
    governed = _phase93_single_source_decision_frame(rec_export, raw_candidate_source)
    candidate_source = governed.copy()
    report = scan_report if isinstance(scan_report, dict) else st.session_state.get(_k("scan_quality_report"), {})
    if not isinstance(report, dict):
        report = {}

    formal_df, a_minus_df, intraday_df, risk_df, watch_df, exclude_df = _phase63_split_formal_recommendation_views(governed)
    live_intraday_df, stale_intraday_df = _phase94_split_intraday_by_kline_freshness(intraday_df)
    core_df, backup_df, low_df, intraday_full_df = _phase71_split_intraday_radar_layers(live_intraday_df)
    formal_usable = bool(report.get("正式推薦可用", False))

    battle_df = _phase70_build_battle_dashboard(formal_df, a_minus_df, core_df, risk_df, exclude_df) if formal_usable else pd.DataFrame()
    if callable(build_action_table):
        try:
            battle_df = build_action_table(battle_df, include_intraday=True)
        except Exception:
            pass

    formal_compact = _phase82_compact_operational_view(formal_df if formal_usable else pd.DataFrame(), "正式主推薦｜可依進場條件分批操作")
    a_minus_compact = _phase82_compact_operational_view(a_minus_df if formal_usable else pd.DataFrame(), "A-準主推薦｜觸發且守價後小量試單")
    core_compact = _phase82_compact_operational_view(core_df, "R1盤中核心雷達｜未觸發前不可買")
    if not formal_usable:
        core_compact = _phase94_mark_research_only(core_compact, "官方因子不足｜只供研究雷達")
    if not core_compact.empty and "盤中雷達優先級" in core_compact.columns:
        _core_priority = core_compact["盤中雷達優先級"].fillna("").astype(str)
        momentum_core_compact = core_compact.loc[_core_priority.str.startswith("R1-M")].copy()
        prebreak_core_compact = core_compact.loc[_core_priority.str.startswith("R1-P")].copy()
    else:
        momentum_core_compact = pd.DataFrame()
        prebreak_core_compact = pd.DataFrame()
    backup_compact = _phase82_compact_operational_view(backup_df, "R2盤中備援雷達｜只做輪動備援")
    low_compact = _phase82_compact_operational_view(low_df, "R3低優先｜不列主要盯盤")
    intraday_full_compact = _phase82_compact_operational_view(intraday_full_df, "盤中雷達完整診斷｜非預先買進清單")
    if not formal_usable:
        backup_compact = _phase94_mark_research_only(backup_compact, "官方因子不足｜只供研究雷達")
        low_compact = _phase94_mark_research_only(low_compact, "官方因子不足｜只供研究雷達")
        intraday_full_compact = _phase94_mark_research_only(intraday_full_compact, "官方因子不足｜只供研究雷達")
    stale_radar_compact = _phase94_build_stale_radar_export(stale_intraday_df)
    if not intraday_full_compact.empty:
        _momentum_status = intraday_full_compact.get("強勢動能判定", pd.Series([""] * len(intraday_full_compact), index=intraday_full_compact.index)).fillna("").astype(str)
        _prebreak_status = intraday_full_compact.get("強勢前兆判定", pd.Series([""] * len(intraday_full_compact), index=intraday_full_compact.index)).fillna("").astype(str)
        momentum_full_compact = intraday_full_compact.loc[~_momentum_status.str.startswith("BLOCK") & _momentum_status.ne("")].copy()
        prebreak_full_compact = intraday_full_compact.loc[~_prebreak_status.str.startswith("BLOCK") & _prebreak_status.ne("")].copy()
    else:
        momentum_full_compact = pd.DataFrame()
        prebreak_full_compact = pd.DataFrame()
    risk_compact = _phase82_compact_operational_view(risk_df, "高風險觀察｜禁止追價")
    watch_compact = _phase82_compact_operational_view(watch_df, "一般/早期觀察｜不列正式推薦")
    exclude_compact = _phase82_compact_operational_view(exclude_df, "正式排除｜禁止新倉")

    candidate_diag = build_candidate_diagnosis(candidate_source) if callable(build_candidate_diagnosis) else candidate_source.copy()
    engine_diag = build_engine_diagnostic_table(candidate_source) if callable(build_engine_diagnostic_table) else pd.DataFrame()
    summary_df = _phase80_build_recommendation_summary(
        formal_df, a_minus_df, core_df, risk_df, exclude_df,
        len(candidate_diag) if isinstance(candidate_diag, pd.DataFrame) else len(governed),
        scan_report=report,
    )
    if isinstance(summary_df, pd.DataFrame) and not summary_df.empty:
        summary_df["最新K線研究雷達檔數"] = int(len(live_intraday_df))
        summary_df["資料待更新雷達檔數"] = int(len(stale_intraday_df))
        if len(stale_intraday_df) > 0:
            summary_df["空白/封鎖原因"] = (
                f"有 {len(stale_intraday_df)} 檔盤中雷達候選的K線不是最新交易日，已移至『資料待更新雷達』；"
                "正式與A-維持封鎖，更新K線與官方因子後必須重新推薦。"
            )
        elif not formal_usable and len(live_intraday_df) > 0:
            summary_df["空白/封鎖原因"] = (
                f"有 {len(live_intraday_df)} 檔最新K線研究雷達，但官方因子不足；"
                "只供研究，不得當成正式買進清單。"
            )
        else:
            summary_df["空白/封鎖原因"] = "本輪沒有通過資料新鮮度與操作門檻的候選。"
    scan_df = pd.DataFrame([report]) if report else pd.DataFrame([{
        "掃描品質狀態": "未知｜舊資料未記錄掃描完整性",
        "正式推薦可用": False,
        "掃描品質說明": "請重新執行完整掃描後再判斷正式推薦。",
    }])
    master_rank_df = _phase90_build_master_recommendation_rank(candidate_source, top_n=30)
    navigation_df = _phase90_navigation_table()
    _v188_cols = [c for c in [
        "股票代號", "股票名稱", "市場別", "類別", "V188股神作戰優先分",
        "SuperAI Alpha分", "SuperAI Alpha等級", "SuperAI Trade分", "SuperAI Trade等級", "SuperAI最終作戰等級",
        "SuperAI執行風報比", "SuperAI風報比來源", "V188交易許可", "V188正式推薦資格",
        "V188RR治理", "V188T+1追價治理", "V188個股資料證據", "V188市場對齊治理",
        "V188類股集中治理", "V188類股集中扣分", "SuperAI隔日上漲機率%", "SuperAI校準後隔日上漲機率%",
        "主要進場路徑", "實戰觸發價", "觸發後守價", "實戰停損參考", "第一壓力價",
    ] if c in candidate_source.columns]
    v188_trade_df = candidate_source[_v188_cols].copy() if _v188_cols else pd.DataFrame()
    try:
        v188_truth_df = pd.DataFrame(load_t1_truth_rows(limit=500) if callable(load_t1_truth_rows) else [])
    except Exception:
        v188_truth_df = pd.DataFrame()
    try:
        _v188_cal = load_probability_calibration() if callable(load_probability_calibration) else {}
        v188_calibration_df = pd.DataFrame((_v188_cal or {}).get("bins") or [])
        if isinstance(v188_calibration_df, pd.DataFrame) and not v188_calibration_df.empty:
            v188_calibration_df.insert(0, "總樣本數", int((_v188_cal or {}).get("eligible_samples") or 0))
            v188_calibration_df.insert(1, "全體Brier", (_v188_cal or {}).get("brier_score"))
    except Exception:
        v188_calibration_df = pd.DataFrame()

    # Compact ranking pages also use the full candidate pool whenever available.
    if (cat_export is None or not isinstance(cat_export, pd.DataFrame) or cat_export.empty) and isinstance(candidate_source, pd.DataFrame):
        try:
            _, cat_export, leader_export, factor_export = _build_export_views(candidate_source, pd.DataFrame(), max(50, len(candidate_source)))
        except Exception:
            pass

    sheets = [
        ("股神推薦總排名", master_rank_df, "目前沒有資料新鮮且達排名門檻的推薦/觀察候選。"),
        ("V188交易品質治理", v188_trade_df, "尚無 V188 Alpha/Trade 交易品質治理資料。"),
        ("T+1實戰真相", v188_truth_df, "尚無成熟 T+1 實戰真相；待下一交易日更新。"),
        ("AI機率校準", v188_calibration_df, "尚無足夠成熟樣本建立機率校準。"),
        ("使用導航", navigation_df, "使用導航無資料。"),
        ("股神正式推薦摘要", summary_df, "本輪沒有摘要資料。"),
        ("掃描完整性", scan_df, "缺少掃描完整性資料；請重新掃描。"),
        ("股神作戰總表", battle_df, "本輪沒有可操作推薦；若掃描不完整，請先重新掃描。"),
        ("完整推薦表", battle_df, "完整推薦表只保留真正可操作名單；本輪沒有可操作推薦。"),
        ("正式下週主推薦", formal_compact, "目前沒有正式主推薦。"),
        ("A-準主推薦小量試單", a_minus_compact, "目前沒有 A- 準主推薦。"),
        ("盤中核心雷達", core_compact, "目前沒有 R1 盤中核心雷達。"),
        ("強勢動能核心雷達", momentum_core_compact, "本輪沒有 R1-M 強勢動能核心雷達。"),
        ("強勢前兆核心雷達", prebreak_core_compact, "本輪沒有 R1-P 強勢前兆核心雷達。"),
        ("強勢動能完整雷達", momentum_full_compact, "本輪沒有強勢動能條件雷達。"),
        ("強勢前兆完整雷達", prebreak_full_compact, "本輪沒有強勢前兆條件雷達。"),
        ("盤中備援雷達", backup_compact, "目前沒有 R2 備援雷達。"),
        ("盤中低優先觀察", low_compact, "目前沒有 R3 低優先觀察。"),
        ("盤中雷達完整名單", intraday_full_compact, "目前沒有使用最新K線的盤中雷達資料。"),
        ("資料待更新雷達", stale_radar_compact, "目前沒有因K線落後而被封鎖的雷達候選。"),
        ("高風險雷達觀察", risk_compact, "目前沒有高風險觀察。"),
        ("不可直接買觀察", watch_compact, "目前沒有一般觀察候選。"),
        ("正式排除清單", exclude_compact, "目前沒有正式排除候選。"),
        ("候選診斷總表", candidate_diag, "沒有完整候選診斷資料；請重新完整掃描。"),
        ("引擎訊號診斷", engine_diag, "沒有引擎訊號診斷資料。"),
        ("類股強度榜", cat_export, "類股強度榜沒有資料。"),
        ("同類股領先榜", leader_export, "同類股領先榜沒有資料。"),
        ("自動因子榜", factor_export, "自動因子榜沒有資料。"),
    ]

    diag_rows = []
    for sheet_name, frame, empty_message in sheets:
        _write_df_to_ws(wb, sheet_name, frame, empty_message)
        diag_rows.append({
            "分頁": sheet_name,
            "用途": ("第一優先" if sheet_name == "股神推薦總排名" else "使用說明" if sheet_name == "使用導航" else "操作" if sheet_name in {"股神作戰總表", "完整推薦表", "正式下週主推薦", "A-準主推薦小量試單", "盤中核心雷達", "強勢動能核心雷達", "強勢前兆核心雷達", "強勢動能完整雷達", "強勢前兆完整雷達"} else "資料待更新/禁止操作" if sheet_name == "資料待更新雷達" else "診斷/管理"),
            "列數": len(frame) if isinstance(frame, pd.DataFrame) else 0,
            "欄數": len(frame.columns) if isinstance(frame, pd.DataFrame) else 0,
        })
    _write_df_to_ws(wb, "匯出診斷", pd.DataFrame(diag_rows), "匯出診斷無資料")

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _result_export_signature_v164(rec_df: pd.DataFrame, extra: str = "") -> str:
    stamp = _safe_str(st.session_state.get(_k("result_saved_at")))
    codes = ""
    try:
        if isinstance(rec_df, pd.DataFrame) and "股票代號" in rec_df.columns:
            codes = ",".join(rec_df["股票代號"].astype(str).map(_normalize_code).tolist())
    except Exception:
        codes = ""
    raw = f"{stamp}|{len(rec_df) if isinstance(rec_df, pd.DataFrame) else 0}|{len(rec_df.columns) if isinstance(rec_df, pd.DataFrame) else 0}|{codes}|{extra}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def _render_export_block(rec_df: pd.DataFrame, category_strength_df: pd.DataFrame, top_n: int):
    """V164：Excel 改為按需產生；一般按鈕 rerun 不再重建 20+ 工作表。"""
    if rec_df is None or rec_df.empty:
        return

    render_pro_section("Excel 匯出")
    sig = _result_export_signature_v164(rec_df, f"main|{top_n}")
    cache_key = _k("main_export_cache_v164")
    cache = st.session_state.get(cache_key, {})
    ready = isinstance(cache, dict) and cache.get("sig") == sig and isinstance(cache.get("bytes"), (bytes, bytearray))

    c1, c2 = st.columns([2, 4])
    with c1:
        if not ready:
            if st.button("準備推薦結果 Excel", use_container_width=True, key=_k("prepare_main_excel_v164")):
                with st.spinner("只在這次需要下載時建立 Excel..."):
                    full_order = _get_full_table_order_for_export(rec_df)
                    rec_export, cat_export, leader_export, factor_export = _build_export_views(
                        rec_df, category_strength_df, top_n, full_order=full_order
                    )
                    rec_export_for_excel = _format_df(rec_export.copy()) if isinstance(rec_export, pd.DataFrame) and not rec_export.empty else rec_export
                    candidate_export = st.session_state.get(_k("candidate_diagnosis_store"))
                    scan_report = st.session_state.get(_k("scan_quality_report"), {})
                    excel_bytes = _build_excel_bytes(
                        rec_export_for_excel, cat_export, leader_export, factor_export,
                        candidate_diagnosis_export=candidate_export if isinstance(candidate_export, pd.DataFrame) else None,
                        scan_report=scan_report if isinstance(scan_report, dict) else None,
                    )
                    cache = {
                        "sig": sig,
                        "bytes": excel_bytes,
                        "name": f"股神推薦_V2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    }
                    st.session_state[cache_key] = cache
                    ready = True
        if ready:
            st.download_button(
                label="匯出推薦結果 Excel",
                data=cache["bytes"],
                file_name=cache["name"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=_k("main_excel_download_v164"),
            )
    with c2:
        st.caption("V164：Excel 只在按『準備』時建立並快取；其他勾選、篩選、匯入、欄位按鈕不再重做整份活頁簿。推薦內容與所有分頁均完整保留。")


def _render_selected_export_block():
    selected_df = st.session_state.get(_k("selected_rec_snapshot"))
    if not isinstance(selected_df, pd.DataFrame) or selected_df.empty:
        return

    export_df = selected_df.copy()
    want_cols = [
        "股票代號", "股票名稱", "市場別", "類別", "類股內排名", "類股前3強",
        "推薦模式", "推薦等級", "推薦總分", "實戰品質分", "量能狀態", "趨勢狀態", "實戰降分", "夜間股神總分", "隔日實戰排序分", "隔日進場分數", "波段潛力分數",
        "進場型態_隔日", "隔日建議動作", "預估進場點", "突破確認價_隔日", "回測承接價", "停損價_隔日", "第一壓力價", "資料完整度",
        "上漲機率估計%", "上漲機率等級", "上漲機率信心", "推薦分桶", "起漲等級", "信心等級",
        "技術結構分數", "起漲前兆分數", "飆股起漲分數", "起漲等級", "起漲摘要", "交易可行分數", "類股熱度分數",
        "同類股領先幅度", "是否領先同類股", "最新價", "推薦買點_拉回", "推薦買點_突破",
        "停損價", "賣出目標1", "賣出目標2", "3日績效%", "5日績效%", "10日績效%", "20日績效%",
        "推薦標籤", "機會股說明", "股神推論邏輯", "風險說明", "推薦理由摘要",
    ]
    export_df = export_df[[c for c in want_cols if c in export_df.columns]].copy()
    sig = _result_export_signature_v164(export_df, "selected")
    cache_key = _k("selected_export_cache_v164")
    cache = st.session_state.get(cache_key, {})
    ready = isinstance(cache, dict) and cache.get("sig") == sig and isinstance(cache.get("bytes"), (bytes, bytearray))
    if not ready and st.button("準備勾選推薦股 Excel", use_container_width=True, key=_k("prepare_selected_excel_v164")):
        with st.spinner("建立勾選股票 Excel..."):
            selected_bytes = _build_excel_bytes(
                rec_export=export_df, cat_export=pd.DataFrame(), leader_export=pd.DataFrame(), factor_export=pd.DataFrame(),
                candidate_diagnosis_export=export_df, scan_report=st.session_state.get(_k("scan_quality_report"), {}),
            )
            cache = {"sig": sig, "bytes": selected_bytes, "name": f"股神推薦_勾選結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"}
            st.session_state[cache_key] = cache
            ready = True
    if ready:
        st.download_button(
            label="匯出勾選推薦股 Excel", data=cache["bytes"], file_name=cache["name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True,
            key=_k("selected_excel_download_v164"),
        )


def _build_record_export_bytes(record_rows: list[dict[str, Any]]) -> bytes:
    df = _ensure_godpick_record_columns(pd.DataFrame(record_rows))
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="股神推薦紀錄匯入", index=False)
        try:
            ws = writer.book["股神推薦紀錄匯入"]
            ws.freeze_panes = "A2"
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    cell_val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(cell_val))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 42)
        except Exception:
            pass
    output.seek(0)
    return output.getvalue()




def _render_recommendation_scoring_guide():
    st.markdown(
        """
        <style>
        .gp-guide-wrap{
            background:#f8fafc;
            border:1px solid rgba(99,102,241,.14);
            border-radius:18px;
            padding:18px 18px 14px 18px;
            margin:18px 0 10px 0;
        }
        .gp-guide-title{
            font-size:1.65rem;
            font-weight:800;
            color:#0f172a;
            margin-bottom:16px;
        }
        .gp-guide-grid{
            display:grid;
            grid-template-columns:repeat(4,minmax(0,1fr));
            gap:14px;
        }
        .gp-guide-card{
            background:#ffffff;
            border:1px solid rgba(99,102,241,.14);
            border-radius:18px;
            padding:18px 18px 14px 18px;
            box-shadow:0 1px 3px rgba(15,23,42,.04);
            height:100%;
        }
        .gp-guide-card h4{
            margin:0 0 10px 0;
            font-size:1.28rem;
            font-weight:800;
            color:#111827;
        }
        .gp-guide-card p{
            margin:0 0 10px 0;
            color:#334155;
            line-height:1.75;
            font-size:1rem;
        }
        .gp-guide-card ul{
            margin:0;
            padding-left:1.2rem;
        }
        .gp-guide-card li{
            margin:0 0 8px 0;
            color:#334155;
            line-height:1.75;
            font-size:1rem;
        }
        .gp-score-list{display:flex;flex-direction:column;gap:10px;}
        .gp-score-row{display:flex;align-items:flex-start;gap:10px;line-height:1.6;}
        .gp-badge{
            display:inline-block;
            min-width:92px;
            text-align:center;
            padding:6px 10px;
            border-radius:10px;
            font-size:.95rem;
            font-weight:800;
            border:1px solid transparent;
            white-space:nowrap;
        }
        .gp-badge.green{background:#e8f7ee;color:#15803d;border-color:#b7e4c7;}
        .gp-badge.green2{background:#eefbf3;color:#166534;border-color:#ccefd7;}
        .gp-badge.yellow{background:#fff7db;color:#b45309;border-color:#f7d98a;}
        .gp-badge.orange{background:#fff1e6;color:#c2410c;border-color:#fdc9a6;}
        .gp-badge.red{background:#feecec;color:#b91c1c;border-color:#f5b5b5;}
        .gp-guide-foot{
            margin-top:14px;
            padding-top:10px;
            border-top:1px solid rgba(99,102,241,.12);
            color:#475569;
            font-size:.98rem;
            font-weight:600;
        }
        @media (max-width: 1200px){
            .gp-guide-grid{grid-template-columns:repeat(2,minmax(0,1fr));}
        }
        @media (max-width: 760px){
            .gp-guide-grid{grid-template-columns:1fr;}
        }
        </style>
        <div class="gp-guide-wrap">
            <div class="gp-guide-title">推薦條件說明 / 分數解讀</div>
            <div class="gp-guide-grid">
                <div class="gp-guide-card">
                    <h4>評分是怎麼算的？</h4>
                    <p>系統依多個面向加總評分，推薦總分代表「候選強度」；買進分數代表「當下價格是否接近可操作」。兩者必須分開判斷。</p>
                    <ul>
                        <li><b>趨勢強度：</b>均線多頭、突破型態、是否站穩關鍵價位</li>
                        <li><b>量價結構：</b>量能放大、價量配合、是否有主力進場跡象</li>
                        <li><b>風險控管：</b>回檔風險、追高風險、破線風險、波動風險</li>
                        <li><b>交易可行：</b>進場點清楚、停損點明確、風險報酬比合理</li>
                        <li><b>類股動能：</b>所屬類股熱度、資金輪動、族群帶動性</li>
                    </ul>
                </div>
                <div class="gp-guide-card">
                    <h4>分數代表什麼？</h4>
                    <div class="gp-score-list">
                        <div class="gp-score-row"><span class="gp-badge green">90 分以上</span><div><b>高優先作戰候選：</b>條件完整，但仍需買進分數與盤中確認</div></div>
                        <div class="gp-score-row"><span class="gp-badge green2">80–89 分</span><div><b>優先追蹤候選：</b>適合找回測、突破與風險報酬比</div></div>
                        <div class="gp-score-row"><span class="gp-badge yellow">70–79 分</span><div><b>觀察等待區：</b>條件尚可，需等突破、量能或支撐確認</div></div>
                        <div class="gp-score-row"><span class="gp-badge orange">60–69 分</span><div><b>弱候選區：</b>有題材但訊號不足，只保留觀察</div></div>
                        <div class="gp-score-row"><span class="gp-badge red">60 分以下</span><div><b>不列主推薦：</b>條件不足，不建議追價</div></div>
                    </div>
                </div>
                <div class="gp-guide-card">
                    <h4>何時才接近可操作？</h4>
                    <ul>
                        <li><b>推薦總分</b>只代表值得追蹤，不等於直接買進</li>
                        <li>需同時看 <b>買進分數、風險報酬比、追價風險、盤中確認條件</b></li>
                        <li><b>高分禁買</b>代表股票可追蹤，但目前價位不適合追</li>
                        <li>盤中若跳空過高、開高走低、量能不足或跌破失效價，取消作戰</li>
                    </ul>
                </div>
                <div class="gp-guide-card">
                    <h4>使用提醒</h4>
                    <ul>
                        <li>本分數為輔助判斷，不等於保證獲利，也不是自動買進訊號</li>
                        <li>建議搭配停損、部位控管與大盤方向一起判讀</li>
                        <li>短線、波段、領頭羊模式的標準會略有不同</li>
                    </ul>
                </div>
            </div>
            <div class="gp-guide-foot">第一優先看「V188股神作戰優先分」與 Alpha×Trade；舊股神分數只代表股票/候選強度。真正進場仍須通過V188交易許可、盤中觸發與停損控管。</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _render_record_export_block(rec_df: pd.DataFrame):
    selected_df = st.session_state.get(_k("selected_rec_snapshot"))
    if not isinstance(selected_df, pd.DataFrame) or selected_df.empty:
        return
    selected_codes = [_normalize_code(x) for x in selected_df["股票代號"].astype(str).tolist() if _normalize_code(x)]
    if not selected_codes:
        return
    record_rows = _build_record_rows_from_rec_df(rec_df, selected_codes)
    if not record_rows:
        return

    render_pro_section("匯出到股神推薦紀錄")
    st.caption("這裡只做匯出，不直接串接 8_股神推薦紀錄；V164 改為需要下載時才建立檔案。")
    sig = hashlib.md5((
        _safe_str(st.session_state.get(_k("result_saved_at"))) + "|" + ",".join(selected_codes)
    ).encode("utf-8")).hexdigest()[:16]
    cache_key = _k("record_export_cache_v164")
    cache = st.session_state.get(cache_key, {})
    ready = isinstance(cache, dict) and cache.get("sig") == sig and isinstance(cache.get("bytes"), (bytes, bytearray))
    if not ready and st.button("準備股神推薦紀錄 Excel", use_container_width=True, key=_k("prepare_record_excel_v164")):
        with st.spinner("建立推薦紀錄匯入檔..."):
            cache = {
                "sig": sig,
                "bytes": _build_record_export_bytes(record_rows),
                "name": f"股神推薦紀錄匯入檔_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            }
            st.session_state[cache_key] = cache
            ready = True
    if ready:
        st.download_button(
            label="匯出股神推薦紀錄 Excel", data=cache["bytes"], file_name=cache["name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True,
            key=_k("record_excel_download_v164"),
        )


# =========================================================
# 頁面設定 / 欄位順序記憶
# =========================================================
def _ui_pref_key(name: str) -> str:
    return _k(f"ui_{name}")

def _ensure_ui_pref(name: str, default):
    pref_key = _k(name)
    ui_key = _ui_pref_key(name)
    if pref_key not in st.session_state:
        st.session_state[pref_key] = copy.deepcopy(default)
    if ui_key not in st.session_state:
        st.session_state[ui_key] = copy.deepcopy(st.session_state[pref_key])

def _sync_ui_pref_to_saved(name: str):
    pref_key = _k(name)
    ui_key = _ui_pref_key(name)
    if ui_key in st.session_state:
        st.session_state[pref_key] = copy.deepcopy(st.session_state[ui_key])

def _reset_ui_pref(name: str, default):
    pref_key = _k(name)
    ui_key = _ui_pref_key(name)
    st.session_state[pref_key] = copy.deepcopy(default)
    st.session_state[ui_key] = copy.deepcopy(default)


def _default_recommend_scan_settings(watchlist_map=None) -> dict[str, Any]:
    group_default = ""
    try:
        if isinstance(watchlist_map, dict) and watchlist_map:
            group_default = list(watchlist_map.keys())[0]
    except Exception:
        group_default = ""
    return {
        "universe_mode": "自選群組",
        "group": group_default,
        "days": 120,
        "top_n": 20,
        "manual_codes": "",
        "scan_limit": 1000,
        "selected_categories": ["全部"],
        "min_total_score": 55.0,
        "min_signal_score": -2.0,
        "min_prelaunch_score": 45.0,
        "min_trade_score": 45.0,
        "recommend_mode": "飆股模式",
        "risk_strictness": "標準",
        "pick_strategy": "結合版",
    }


def _recommend_setting_names() -> list[str]:
    return [
        "universe_mode", "group", "days", "top_n", "manual_codes", "scan_limit",
        "selected_categories", "min_total_score", "min_signal_score",
        "min_prelaunch_score", "min_trade_score",
        "recommend_mode", "risk_strictness", "pick_strategy",
    ]


def _normalize_recommend_scan_settings(raw: Any, watchlist_map=None, category_options=None) -> dict[str, Any]:
    base = _default_recommend_scan_settings(watchlist_map)
    data = raw if isinstance(raw, dict) else {}

    for k in base.keys():
        if k in data:
            base[k] = copy.deepcopy(data[k])

    universe_options = ["自選群組", "手動輸入", "全市場", "上市", "上櫃", "興櫃"]
    if base["universe_mode"] not in universe_options:
        base["universe_mode"] = "自選群組"

    group_options = list(watchlist_map.keys()) if isinstance(watchlist_map, dict) and watchlist_map else [""]
    if base["group"] not in group_options:
        base["group"] = group_options[0] if group_options else ""

    for key, options, default in [
        ("days", [60, 90, 120, 180, 240], 120),
        ("top_n", [10, 20, 30, 50], 20),
        ("scan_limit", [100, 200, 300, 500, 1000, 1500, 2000, "全部"], 1000),
    ]:
        if base[key] not in options:
            try:
                iv = int(base[key])
                base[key] = iv if iv in options else default
            except Exception:
                base[key] = default

    mode_options = ["飆股模式", "波段模式", "領頭羊模式", "綜合模式"]
    if base["recommend_mode"] not in mode_options:
        base["recommend_mode"] = "飆股模式"

    strict_options = ["寬鬆", "標準", "嚴格"]
    if base["risk_strictness"] not in strict_options:
        base["risk_strictness"] = "標準"

    pick_options = ["精準版", "結合版"]
    if base["pick_strategy"] not in pick_options:
        base["pick_strategy"] = "結合版"

    category_options = category_options or ["全部"]
    cats = base.get("selected_categories", ["全部"])
    if not isinstance(cats, list):
        cats = ["全部"]
    cats = [x for x in cats if x in category_options] or ["全部"]
    base["selected_categories"] = cats

    for k in ["min_total_score", "min_signal_score", "min_prelaunch_score", "min_trade_score"]:
        try:
            base[k] = float(base[k])
        except Exception:
            base[k] = float(_default_recommend_scan_settings(watchlist_map)[k])

    return base


def _load_persistent_recommend_scan_settings(watchlist_map=None, category_options=None) -> dict[str, Any]:
    payload = _load_persistent_settings()
    raw = payload.get("scan_settings", {}) if isinstance(payload, dict) else {}
    return _normalize_recommend_scan_settings(raw, watchlist_map, category_options)


def _save_persistent_recommend_scan_settings(settings: dict[str, Any]) -> tuple[bool, list[str]]:
    payload = _load_persistent_settings()
    if not isinstance(payload, dict):
        payload = {}
    payload["scan_settings"] = copy.deepcopy(settings)
    payload["applied_weights"] = _normalize_weight_map(payload.get("applied_weights", GODPICK_DEFAULT_SCORE_WEIGHTS))
    payload["original_default_weights"] = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()
    payload["column_orders"] = payload.get("column_orders", {}) if isinstance(payload.get("column_orders", {}), dict) else {}
    payload["updated_at"] = _now_text()
    payload["version"] = "godpick_v5_persistent_settings"
    local_ok, local_msg = _safe_json_write_local(GODPICK_SETTINGS_FILE, payload)
    github_ok, github_msg = _write_json_to_github_path(GODPICK_SETTINGS_FILE, payload)
    return (local_ok or github_ok), [local_msg, github_msg]


def _apply_recommend_scan_settings_to_state(settings: dict[str, Any], sync_widgets: bool = True):
    """
    套用推薦設定到 session_state。

    sync_widgets=True 只能在 widget 建立前使用；此時要強制把畫面 widget key
    同步成永久設定，避免換頁/重開後又吃到舊的 session 預設值。
    sync_widgets=False 用於按鈕提交後，避免 StreamlitAPIException。
    """
    settings = settings or {}
    for name in _recommend_setting_names():
        val = copy.deepcopy(settings.get(name, _default_recommend_scan_settings().get(name)))
        st.session_state[_k(name)] = val
        ui_key = _ui_pref_key(name)
        if sync_widgets:
            st.session_state[ui_key] = val


def _current_form_settings_from_values(
    form_universe_mode, form_group, form_days, form_top_n, form_manual_codes,
    form_scan_limit, form_selected_categories, form_min_total_score,
    form_min_signal_score, form_min_prelaunch_score, form_min_trade_score,
    form_recommend_mode, form_risk_strictness, form_pick_strategy,
) -> dict[str, Any]:
    return {
        "universe_mode": form_universe_mode,
        "group": form_group,
        "days": form_days,
        "top_n": form_top_n,
        "manual_codes": form_manual_codes,
        "scan_limit": form_scan_limit,
        "selected_categories": form_selected_categories if form_selected_categories else ["全部"],
        "min_total_score": float(form_min_total_score),
        "min_signal_score": float(form_min_signal_score),
        "min_prelaunch_score": float(form_min_prelaunch_score),
        "min_trade_score": float(form_min_trade_score),
        "recommend_mode": form_recommend_mode,
        "risk_strictness": form_risk_strictness,
        "pick_strategy": form_pick_strategy,
    }


def _stage_recommend_scan_settings_reset(settings: dict[str, Any], msg: str = ""):
    st.session_state[_k("scan_settings_reset_pending")] = True
    st.session_state[_k("scan_settings_reset_payload")] = copy.deepcopy(settings)
    if msg:
        st.session_state[_k("scan_settings_msg")] = msg


def _normalize_column_order(saved_order, available_cols: list[str], default_cols: list[str]) -> list[str]:
    saved = [str(x) for x in (saved_order or []) if str(x) in available_cols]
    defaults = [str(x) for x in default_cols if str(x) in available_cols]
    remain = [c for c in available_cols if c not in saved and c not in defaults]
    merged = saved + [c for c in defaults if c not in saved] + remain
    final = []
    seen = set()
    for c in merged:
        if c in available_cols and c not in seen:
            final.append(c)
            seen.add(c)
    return final


def _fixed_columns_for_order_manager(name: str) -> list[str]:
    # v72：勾選欄固定在最前，不參與欄位順序調整，避免 data_editor 位置錯亂。
    if name == "full_table":
        return ["勾選"]
    return []


def _column_order_state_key(name: str) -> str:
    return _k(f"column_order_{name}")


def _column_order_fingerprint(cols: list[str]) -> str:
    """v78：依欄位順序建立短指紋；用於重建 data_editor key，避免前端沿用舊欄位位置。"""
    try:
        raw = "|".join([str(c) for c in cols])
        return hashlib.md5(raw.encode("utf-8")).hexdigest()[:10]
    except Exception:
        return str(int(time.time()))


def _clear_full_table_editor_widget_states():
    """v78：清除完整推薦表所有舊版 / 動態 data_editor 狀態。
    Streamlit data_editor 會保留前端欄位 layout；只清固定 key 不夠，必須清所有前綴。
    """
    prefixes = [
        _k("full_table_editor"),
        _k("full_table_editor_"),
        _k("full_table_editor_code_map"),
        _k("full_table_editor_code_map_"),
    ]
    for key in list(st.session_state.keys()):
        if any(str(key).startswith(pfx) for pfx in prefixes):
            st.session_state.pop(key, None)


def _render_column_order_manager(name: str, title: str, available_cols: list[str], default_cols: list[str]) -> list[str]:
    """v81：欄位順序快速管理器。

    說明：Streamlit data_editor 的前端拖曳欄位順序不會穩定回寫 Python，
    因此這裡提供不需額外套件的快速欄位管理：多選批次移動、指定位置、
    常用版面套用；每次操作都會立即寫入 godpick_column_orders.json 並重建表格。
    """
    state_key = _column_order_state_key(name)
    applied_key = _k(f"column_order_applied_{name}")
    draft_key = _k(f"column_order_draft_{name}")
    pick_key = _k(f"column_pick_{name}")
    multi_key = _k(f"column_multi_pick_{name}")
    target_pos_key = _k(f"column_target_pos_{name}")
    preset_key = _k(f"column_preset_{name}")

    fixed_cols = [c for c in _fixed_columns_for_order_manager(name) if c in available_cols]
    managed_available_cols = [c for c in available_cols if c not in fixed_cols]
    managed_default_cols = [c for c in default_cols if c in managed_available_cols]

    persistent_order = [c for c in _load_persistent_column_order(name) if c in managed_available_cols]
    base_order = persistent_order if persistent_order else st.session_state.get(applied_key, st.session_state.get(state_key, managed_default_cols))

    applied_order = _normalize_column_order(base_order, managed_available_cols, managed_default_cols)
    draft_order = _normalize_column_order(st.session_state.get(draft_key, applied_order), managed_available_cols, managed_default_cols)

    st.session_state[applied_key] = applied_order
    st.session_state[draft_key] = draft_order
    st.session_state[state_key] = applied_order

    def _commit_column_order(new_order: list[str], toast_msg: str = "欄位順序已套用並保存。"):
        new_order = _normalize_column_order(new_order, managed_available_cols, managed_default_cols)
        st.session_state[applied_key] = new_order
        st.session_state[state_key] = new_order
        st.session_state[draft_key] = new_order
        ok, msgs = _save_persistent_column_order(name, new_order)
        if name == "full_table":
            _clear_full_table_editor_widget_states()
            st.session_state[_k("full_table_layout_version")] = (
                _column_order_fingerprint(fixed_cols + new_order)
                + "_"
                + str(int(time.time() * 1000))
            )
        st.session_state[_k(f"column_order_last_message_{name}")] = toast_msg if ok else "欄位順序已套用，但保存可能失敗。"
        st.session_state[_k(f"column_order_last_save_ok_{name}")] = bool(ok)
        if not ok:
            st.session_state[_k(f"column_order_last_save_msgs_{name}")] = msgs
        st.toast(st.session_state[_k(f"column_order_last_message_{name}")], icon="✅" if ok else "⚠️")
        st.rerun()

    def _remove_selected(order: list[str], selected: list[str]) -> list[str]:
        selected_set = set(selected)
        return [c for c in order if c not in selected_set]

    def _insert_selected(order: list[str], selected: list[str], index: int) -> list[str]:
        selected_in_order = [c for c in order if c in set(selected)]
        remain = _remove_selected(order, selected_in_order)
        index = max(0, min(int(index), len(remain)))
        return remain[:index] + selected_in_order + remain[index:]

    def _preset_columns(preset_name: str) -> list[str]:
        presets = {
            "VNext績效回饋校正版": [
                "股票代號", "股票名稱", "市場別", "類別", "推薦角色", "新買點分級", "今日決策結論",
                "候選強度分", "股神實戰總分", "Alpha選股潛力分", "Entry進場買點分", "Risk風控安全分", "Feedback績效校正分",
                "選股潛力分", "進場買點分", "風控安全分", "績效校正分", "績效校正說明",
                "建議動作", "建議倉位", "突破確認狀態", "突破確認條件", "等待突破原因", "假陰性檢討",
                "加碼條件", "失效條件", "過熱原因", "真禁買原因", "硬否決原因", "決策版本",
                "推薦總分", "推薦型態", "買點分級", "小量試單建議", "績效回饋建議",
                "失效條件_績效回饋", "最新價", "近5日漲幅%", "追價風險分", "風險報酬比",
            ],
            "夜間隔日股神版": [
                "股票代號", "股票名稱", "市場別", "類別", "推薦等級", "推薦總分", "夜間股神總分",
                "隔日實戰排序分", "隔日進場分數", "波段潛力分數", "進場型態_隔日", "隔日建議動作",
                "預估進場點", "回測承接價", "突破確認價_隔日", "停損價_隔日", "第一壓力價", "觀察週期",
                "法人籌碼分數", "大戶鎖碼分數", "基本面成長分數", "估值風險分數", "PER本益比",
                "外資近1日買賣超", "投信近1日買賣超", "三大法人近1日合計", "資料完整度",
                "夜間股神建議", "夜間風險提醒",
            ],
            "股神推薦精簡版": [
                "股票代號", "股票名稱", "市場別", "類別", "推薦等級", "推薦總分",
                "上漲機率估計%", "買點狀態", "高分禁買旗標", "高分禁買原因", "實戰買點分數",
                "最新價", "推薦理由摘要",
            ],
            "買點實戰版": [
                "股票代號", "股票名稱", "推薦等級", "推薦總分", "買點狀態", "進場型態",
                "實戰買點分數", "支撐距離%", "壓力空間%", "近5日漲幅%", "風險報酬比",
                "停損距離%", "目標報酬%", "停損價", "賣出目標1", "賣出目標2", "實戰操作建議",
            ],
            "風控檢查版": [
                "股票代號", "股票名稱", "推薦總分", "高分禁買旗標", "高分禁買原因", "追價風險分",
                "追高風險等級", "長上影風險", "假突破風險", "過熱風險", "大盤橋接風控",
                "大盤資料品質", "隔夜風險", "隔夜資料品質", "風險說明",
            ],
            "隔夜風控版": [
                "股票代號", "股票名稱", "推薦總分", "上漲機率估計%", "大盤橋接分數",
                "大盤橋接風控", "大盤交易時段", "隔夜分數", "隔夜風險", "隔夜偏向",
                "隔夜解讀", "NASDAQ漲跌%", "費半漲跌%", "台指夜盤漲跌", "台指夜盤資料來源",
                "台指夜盤備援說明",
            ],
            "完整除錯版": managed_default_cols,
        }
        cols = presets.get(preset_name, managed_default_cols)
        return [c for c in cols if c in managed_available_cols]

    with st.expander(title, expanded=False):
        fixed_msg = "；固定欄位：" + "、".join(fixed_cols) if fixed_cols else ""
        st.caption(
            "欄位順序會永久記錄；固定欄位不參與排序" + fixed_msg + "。"
            "v81：提供批次移動、指定位置與常用版面；不要只拖曳表格欄位，拖曳結果無法穩定永久保存。"
        )

        last_msg = st.session_state.get(_k(f"column_order_last_message_{name}"), "")
        if last_msg:
            if st.session_state.get(_k(f"column_order_last_save_ok_{name}"), True):
                st.success(last_msg)
            else:
                st.warning(last_msg)

        p1, p2 = st.columns([2, 1])
        with p1:
            preset = st.selectbox(
                "快速套用常用欄位版面",
                ["請選擇", "夜間隔日股神版", "股神推薦精簡版", "買點實戰版", "風控檢查版", "隔夜風控版", "完整除錯版"],
                key=preset_key,
            )
        with p2:
            if st.button("套用版面", key=_k(f"apply_preset_{name}"), use_container_width=True, type="primary"):
                if preset and preset != "請選擇":
                    _commit_column_order(_preset_columns(preset), f"已套用「{preset}」欄位版面。")
                else:
                    st.warning("請先選擇一個常用版面。")

        if pick_key not in st.session_state or st.session_state[pick_key] not in draft_order:
            st.session_state[pick_key] = draft_order[0] if draft_order else ""
        picked = st.selectbox("單一欄位快速移動", draft_order, key=pick_key) if draft_order else ""

        b1, b2, b3, b4, b5 = st.columns(5)
        if draft_order and picked:
            idx = draft_order.index(picked)
            with b1:
                if st.button("左移", key=_k(f"move_left_{name}"), use_container_width=True) and idx > 0:
                    new_order = draft_order.copy()
                    new_order[idx - 1], new_order[idx] = new_order[idx], new_order[idx - 1]
                    _commit_column_order(new_order, f"已將「{picked}」左移。")
            with b2:
                if st.button("右移", key=_k(f"move_right_{name}"), use_container_width=True) and idx < len(draft_order) - 1:
                    new_order = draft_order.copy()
                    new_order[idx + 1], new_order[idx] = new_order[idx], new_order[idx + 1]
                    _commit_column_order(new_order, f"已將「{picked}」右移。")
            with b3:
                if st.button("移到最前", key=_k(f"move_front_{name}"), use_container_width=True):
                    new_order = draft_order.copy()
                    new_order.remove(picked)
                    new_order.insert(0, picked)
                    _commit_column_order(new_order, f"已將「{picked}」移到最前。")
            with b4:
                if st.button("移到最後", key=_k(f"move_last_{name}"), use_container_width=True):
                    new_order = draft_order.copy()
                    new_order.remove(picked)
                    new_order.append(picked)
                    _commit_column_order(new_order, f"已將「{picked}」移到最後。")
            with b5:
                if st.button("恢復原始設定", key=_k(f"move_restore_default_{name}"), use_container_width=True):
                    _commit_column_order(managed_default_cols, "已恢復原始欄位順序。")

        st.divider()
        st.markdown("**v81 批次欄位快速管理**")

        selected_cols = st.multiselect(
            "一次選多個欄位",
            options=draft_order,
            default=[c for c in st.session_state.get(multi_key, []) if c in draft_order],
            key=multi_key,
            help="可一次選多個欄位，批次移到最前、最後、勾選後面或指定位置。",
        )

        target_default = int(st.session_state.get(target_pos_key, 1) or 1)
        target_pos = st.number_input(
            "指定目標位置（1 = 第一個非固定欄位；勾選固定在最前不受影響）",
            min_value=1,
            max_value=max(1, len(draft_order)),
            value=max(1, min(target_default, max(1, len(draft_order)))),
            step=1,
            key=target_pos_key,
        )

        q1, q2, q3, q4 = st.columns(4)
        with q1:
            if st.button("批次移到最前", key=_k(f"batch_front_{name}"), use_container_width=True, disabled=not selected_cols):
                new_order = _insert_selected(draft_order, selected_cols, 0)
                _commit_column_order(new_order, f"已將 {len(selected_cols)} 個欄位批次移到最前。")
        with q2:
            if st.button("批次移到最後", key=_k(f"batch_last_{name}"), use_container_width=True, disabled=not selected_cols):
                remain = _remove_selected(draft_order, selected_cols)
                selected_in_order = [c for c in draft_order if c in set(selected_cols)]
                _commit_column_order(remain + selected_in_order, f"已將 {len(selected_cols)} 個欄位批次移到最後。")
        with q3:
            if st.button("移到勾選後面", key=_k(f"batch_after_checkbox_{name}"), use_container_width=True, disabled=not selected_cols):
                new_order = _insert_selected(draft_order, selected_cols, 0)
                _commit_column_order(new_order, f"已將 {len(selected_cols)} 個欄位移到勾選後面。")
        with q4:
            if st.button("移到指定位置", key=_k(f"batch_to_pos_{name}"), use_container_width=True, disabled=not selected_cols):
                new_order = _insert_selected(draft_order, selected_cols, int(target_pos) - 1)
                _commit_column_order(new_order, f"已將 {len(selected_cols)} 個欄位移到第 {int(target_pos)} 個位置。")

        c1, c2 = st.columns([1.3, 3])
        with c1:
            if st.button("儲存目前順序", key=_k(f"apply_column_order_{name}"), use_container_width=True):
                _commit_column_order(draft_order, "目前欄位順序已重新儲存。")
        with c2:
            st.caption("提示：表格本身的滑鼠拖曳欄位只會改前端視覺，不會穩定回寫；請用本管理器保存。")

        preview_order = fixed_cols + st.session_state.get(applied_key, applied_order)
        st.caption("目前已保存欄位順序：" + " ｜ ".join(preview_order[:24]) + (" ..." if len(preview_order) > 24 else ""))

    return fixed_cols + st.session_state.get(applied_key, applied_order)
def _postprocess_dependency_signature_v164(macro_bridge: dict[str, Any], enabled: bool) -> str:
    """只在真正依賴資料改變時重算推薦後處理，不因一般 widget rerun 重算。"""
    file_parts: list[str] = []
    for name in [
        GODPICK_SETTINGS_FILE,
        MACRO_MODE_BRIDGE_FILE,
        MARKET_SNAPSHOT_FILE,
        OFFICIAL_FACTORS_CACHE_FILE,
        "godpick_records.json",
        "godpick_calibration_samples.json",
    ]:
        try:
            path = Path(name)
            stat = path.stat()
            file_parts.append(f"{name}:{stat.st_mtime_ns}:{stat.st_size}")
        except Exception:
            file_parts.append(f"{name}:missing")
    try:
        bridge_text = json.dumps(macro_bridge or {}, ensure_ascii=False, sort_keys=True, default=str)
    except Exception:
        bridge_text = str(macro_bridge)
    try:
        weight_text = json.dumps(
            st.session_state.get(_k("score_weights"), GODPICK_DEFAULT_SCORE_WEIGHTS),
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        )
    except Exception:
        weight_text = ""
    raw = "|".join([PAGE07_SPEED_FIX_VERSION, str(bool(enabled)), bridge_text, weight_text, *file_parts])
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _postprocess_recommend_result_v164(
    rec_df: pd.DataFrame,
    hot_pick_df: pd.DataFrame,
    macro_bridge: dict[str, Any],
    macro_bridge_enabled: bool,
    *,
    force: bool = False,
) -> tuple[pd.DataFrame, pd.DataFrame, bool]:
    """V164：推薦後處理快取。

    舊版每次任何按鈕 rerun 都重做進階欄位、大盤橋接、官方因子、夜間策略、
    動態資金流分流與數百欄 schema 正規化。實測 57+30 筆約需 19.5 秒。
    新版只有新掃描或依賴檔案/權重改變時重算，其餘直接取 session 快取。
    """
    dep_sig = _postprocess_dependency_signature_v164(macro_bridge, macro_bridge_enabled)
    cache_key = _k("postprocess_cache_v164")
    cache = st.session_state.get(cache_key, {})
    if (
        not force
        and isinstance(cache, dict)
        and cache.get("dep_sig") == dep_sig
        and isinstance(cache.get("rec_df"), pd.DataFrame)
        and not cache.get("rec_df").empty
    ):
        cached_hot = cache.get("hot_df") if isinstance(cache.get("hot_df"), pd.DataFrame) else pd.DataFrame()
        return cache["rec_df"].copy(deep=False), cached_hot.copy(deep=False), True

    rec = rec_df.copy() if isinstance(rec_df, pd.DataFrame) else pd.DataFrame()
    hot = hot_pick_df.copy() if isinstance(hot_pick_df, pd.DataFrame) else pd.DataFrame()
    rec = _apply_advanced_godpick_columns(rec)
    hot = _apply_advanced_godpick_columns(hot)
    rec = _apply_macro_bridge_columns(rec, macro_bridge, macro_bridge_enabled)
    hot = _apply_macro_bridge_columns(hot, macro_bridge, macro_bridge_enabled)
    rec = _apply_official_factor_cache_v109(rec)
    hot = _apply_official_factor_cache_v109(hot)
    rec = _recalc_night_strategy_after_macro_v100(rec)
    hot = _recalc_night_strategy_after_macro_v100(hot)
    rec = _apply_v139_dynamic_hot_money_breakout_rules(rec)
    hot = _apply_v139_dynamic_hot_money_breakout_rules(hot)
    try:
        if normalize_godpick_dataframe is not None:
            rec = normalize_godpick_dataframe(rec, add_missing=False)
            hot = normalize_godpick_dataframe(hot, add_missing=False)
    except Exception:
        pass
    st.session_state[cache_key] = {
        "dep_sig": dep_sig,
        "rec_df": rec.copy(deep=False),
        "hot_df": hot.copy(deep=False),
        "built_at": _now_text(),
    }
    return rec, hot, False


def _render_phase105_learning_panel(candidate_df: pd.DataFrame | None = None) -> None:
    """顯示每日學習型AI狀態；只呈現可驗證統計，不宣稱保證獲利。"""
    state = {}
    if callable(load_learning_state):
        try:
            state = load_learning_state() or {}
        except Exception:
            state = {}
    summary = state.get("last_run_summary", {}) if isinstance(state, dict) else {}
    if isinstance(candidate_df, pd.DataFrame) and not candidate_df.empty and callable(build_learning_summary):
        try:
            summary = build_learning_summary(candidate_df)
        except Exception:
            pass
    if not summary:
        st.info("每日學習型AI尚未建立第一筆決策快照；完成一次重新推薦後會開始累積。")
        return
    render_pro_section("每日學習型AI｜Champion / Challenger")
    render_pro_kpi_row([
        {"label": "學習樣本", "value": int(summary.get("eligible_samples", 0) or 0), "delta": "只採可驗證績效", "delta_class": "pro-kpi-delta-flat"},
        {"label": "本輪候選", "value": int(summary.get("candidate_count", 0) or 0), "delta": "多路召回", "delta_class": "pro-kpi-delta-flat"},
        {"label": "AI正式候選", "value": int(summary.get("formal_ai", 0) or 0), "delta": "仍需正式風控", "delta_class": "pro-kpi-delta-flat"},
        {"label": "AI A-候選", "value": int(summary.get("a_minus_ai", 0) or 0), "delta": "條件確認", "delta_class": "pro-kpi-delta-flat"},
        {"label": "好股等買點", "value": int(summary.get("quality_wait", 0) or 0), "delta": "Alpha高/Timing不足", "delta_class": "pro-kpi-delta-flat"},
        {"label": "平均AI決策", "value": f"{float(summary.get('avg_decision', 0) or 0):.1f}", "delta": GODPICK_AI_MODEL_VERSION, "delta_class": "pro-kpi-delta-flat"},
    ])
    route_counts = summary.get("route_counts", {}) if isinstance(summary, dict) else {}
    if route_counts:
        st.caption("本輪主要召回路徑：" + "｜".join(f"{k} {v}" for k, v in route_counts.items()))
    metrics = summary.get("global_metrics", {}) if isinstance(summary, dict) else {}
    if isinstance(metrics, dict) and metrics:
        probability_samples = int(metrics.get("probability_samples", 0) or 0)
        brier = metrics.get("brier_score")
        metric_text = (
            f"歷史可驗證樣本平均報酬 {float(metrics.get('mean', 0) or 0):.2f}%｜"
            f"正報酬率 {float(metrics.get('hit_rate', 0) or 0):.1f}%｜"
            f"機率校準樣本 {probability_samples}"
        )
        if brier is not None:
            metric_text += f"｜Brier {float(brier):.4f}"
        else:
            metric_text += "｜Brier 將自本版推薦開始累積"
        st.caption(metric_text)
    error_taxonomy = summary.get("error_taxonomy", {}) if isinstance(summary, dict) else {}
    if isinstance(error_taxonomy, dict) and error_taxonomy:
        top_errors = list(error_taxonomy.items())[:3]
        st.caption("歷史錯誤分類：" + "｜".join(f"{k} {v}" for k, v in top_errors))
    st.caption("Champion 影響正式排序；Challenger 只做影子比較，未經足夠樣本驗證不會自動取代正式模型。")
    msgs = st.session_state.get(_k("learning_run_messages"), [])
    if msgs:
        with st.expander("本輪AI決策快照與永久保存", expanded=False):
            for msg in msgs:
                st.write(f"- {msg}")


def _run_page07_automation_v191_h2(cfg: dict[str, Any] | None = None) -> dict[str, Any]:
    """Canonical V191 automation entrypoint owned by Module 07.

    The central scheduler only invokes this callable.  All recommendation
    orchestration, snapshot saving, Page08 record persistence, rotation,
    learning and SuperAI experience saving remain owned by Page07 so manual and
    scheduled execution cannot drift into two different recommendation paths.
    """
    cfg = dict(cfg or {})
    started_at = _now_text()
    execution_context = {
        "owner": "07_股神推薦",
        "trigger": "V191中央自動排程",
        "started_at": started_at,
        "automation_version": "V191-H5",
    }
    st.session_state[_k("recommend_execution_context_v191")] = execution_context
    notes: list[str] = []
    try:
        watchlist_map = _load_watchlist_map() or {}
        master_df = _load_master_df()
        if master_df is None or master_df.empty:
            master_df = _load_master_df_fallback_only()
        if master_df is None or master_df.empty:
            return {"ok": False, "message": "07股神推薦模組執行失敗：股票主檔為空", "execution_context": execution_context}

        category_options = ["全部"] + (_collect_all_categories(master_df, watchlist_map) or [])
        settings = _load_persistent_recommend_scan_settings(watchlist_map, category_options)
        if bool(cfg.get("force_full_market", False)):
            settings["universe_mode"] = "全市場"
            settings["scan_limit"] = "全部"
        _apply_recommend_scan_settings_to_state(settings, sync_widgets=False)

        persistent = _load_persistent_settings(local_first=True)
        weights = _normalize_weight_map(persistent.get("applied_weights") or persistent.get("score_weights"))
        macro_bridge = _read_macro_mode_bridge()
        weights = _apply_macro_bridge_to_weights(weights, macro_bridge, enabled=True)
        global GODPICK_ACTIVE_SCORE_WEIGHTS
        GODPICK_ACTIVE_SCORE_WEIGHTS = weights.copy()

        mode = _safe_str(settings.get("universe_mode")) or "全市場"
        if mode == "自選群組":
            universe_items = watchlist_map.get(_safe_str(settings.get("group")), [])
        elif mode == "手動輸入":
            universe_items = _parse_manual_codes(settings.get("manual_codes", ""), master_df)
        else:
            universe_items = _build_universe_from_market(
                master_df=master_df,
                market_mode=mode,
                limit_count=settings.get("scan_limit", 1000),
                selected_categories=settings.get("selected_categories") or ["全部"],
            )
        if not universe_items:
            return {"ok": False, "message": "07股神推薦模組執行失敗：掃描池為空", "execution_context": execution_context, "settings": settings}

        today = datetime.now(ZoneInfo("Asia/Taipei")).date()
        start_dt = today - timedelta(days=int(settings.get("days", 120) or 120))
        previous_rec_df, previous_category_df, previous_hot_df = _load_recommend_result_from_state()
        rec_df, category_strength_df, hot_pick_df = _build_recommend_df(
            universe_items=universe_items,
            master_df=master_df,
            start_dt=start_dt,
            end_dt=today,
            min_total_score=float(settings.get("min_total_score", 55)),
            min_signal_score=float(settings.get("min_signal_score", -2)),
            selected_categories=settings.get("selected_categories") or ["全部"],
            mode=_safe_str(settings.get("recommend_mode")) or "飆股模式",
            risk_strictness=_safe_str(settings.get("risk_strictness")) or "標準",
            min_prelaunch_score=float(settings.get("min_prelaunch_score", 45)),
            min_trade_score=float(settings.get("min_trade_score", 45)),
            resume_scan=False,
            reuse_finished_checkpoint=False,
        )
        rec_df, hot_pick_df, _ = _postprocess_recommend_result_v164(rec_df, hot_pick_df, macro_bridge, True, force=True)

        candidate_df = st.session_state.get(_k("candidate_diagnosis_store"))
        if rec_df is None or rec_df.empty:
            conditional_df = _conditional_reference_rows(candidate_df, max_rows=8) if isinstance(candidate_df, pd.DataFrame) else pd.DataFrame()
            if not conditional_df.empty:
                rec_df = conditional_df
                notes.append("本輪無正式推薦，依Page07規則保存條件式參考名單。")
            elif isinstance(previous_rec_df, pd.DataFrame) and not previous_rec_df.empty:
                rec_df = previous_rec_df.copy()
                category_strength_df = previous_category_df.copy() if isinstance(previous_category_df, pd.DataFrame) else pd.DataFrame()
                hot_pick_df = previous_hot_df.copy() if isinstance(previous_hot_df, pd.DataFrame) else pd.DataFrame()
                notes.append("本輪沒有可用新結果，依Page07安全網保留上一輪非空結果。")
            else:
                rec_df = pd.DataFrame()

        save_ok = bool(_save_recommend_result_to_state(rec_df, category_strength_df, hot_pick_df))
        source_df = st.session_state.get(_k("candidate_diagnosis_store"))
        if not isinstance(source_df, pd.DataFrame) or source_df.empty:
            source_df = rec_df

        # Keep the user's existing contract: Page07 executes, Page08 authority
        # continues to permanently store recommendation records.
        record_added, record_msgs = _v159_auto_record_actionable_recommendations(source_df, background_write=False)

        calibration_added = 0
        calibration_msgs: list[str] = []
        calibration_summary: dict[str, int] = {"near": 0, "missed": 0, "total": 0}
        if callable(save_calibration_samples):
            try:
                calibration_added, calibration_msgs, calibration_summary = save_calibration_samples(
                    source_df, max_near=24, max_missed=20, background_remote=False
                )
            except Exception as exc:
                calibration_msgs = [f"校正研究樣本保存例外：{exc}"]

        if callable(save_rotation_snapshot):
            try:
                rotation_source = _phase93_single_source_decision_frame(rec_df, source_df)
                rotation_ok, rotation_msg = save_rotation_snapshot(rotation_source, background_remote=False)
                notes.append(f"推薦輪動快照：{'成功' if rotation_ok else '警示'}｜{rotation_msg}")
            except Exception as exc:
                notes.append(f"推薦輪動快照例外：{exc}")
        if callable(save_learning_run):
            try:
                learning_ok, learning_msgs, _ = save_learning_run(
                    source_df, rec_df,
                    scan_report=st.session_state.get(_k("scan_quality_report"), {}),
                    metadata={**settings, "automation": "V191中央自動排程", "execution_owner": "07_股神推薦"},
                    persist_remote=True, background_remote=False, pre_scored=True,
                )
                notes.append(f"學習快照：{'成功' if learning_ok else '警示'}｜" + "；".join(str(x) for x in (learning_msgs or [])[-4:]))
            except Exception as exc:
                notes.append(f"學習快照例外：{exc}")
        if callable(save_super_ai_run):
            try:
                super_ok, super_msg, _ = save_super_ai_run(
                    source_df, rec_df,
                    metadata={"automation": "V191中央自動排程", "execution_owner": "07_股神推薦", "scan_settings": settings},
                )
                notes.append(f"SuperAI經驗：{'成功' if super_ok else '警示'}｜{super_msg}")
            except Exception as exc:
                notes.append(f"SuperAI經驗例外：{exc}")

        scan_report = st.session_state.get(_k("scan_quality_report"), {}) or {}
        display_count = len(rec_df) if isinstance(rec_df, pd.DataFrame) else 0
        candidate_count = len(source_df) if isinstance(source_df, pd.DataFrame) else 0
        _record_text_v191_h3 = "｜".join(_safe_str(x) for x in (record_msgs or []) if _safe_str(x))
        record_integrity_failure = any(token in _record_text_v191_h3 for token in [
            "推薦紀錄未寫入權威檔", "歷史救援尚未完成", "防歸零",
            "永久服務未載入", "權威增量保存：失敗", "本輪不得顯示保存成功",
        ])
        ok = bool(save_ok and isinstance(rec_df, pd.DataFrame) and not record_integrity_failure)
        no_actionable = bool(ok and display_count <= 0)
        message = (
            f"07股神推薦模組自動執行{'完成' if ok else '未完整完成'}："
            f"掃描 {len(universe_items)}／候選 {candidate_count}／顯示 {display_count}／08永久紀錄 {record_added}"
        )
        if record_integrity_failure:
            message += "｜08推薦歷史完整性/永久化未通過，本輪不得標示SUCCESS；已保留07候選診斷供修復後重試"
        elif no_actionable:
            message += "｜本輪0檔通過可操作底線，候選診斷已保存；不硬塞弱股、不清空歷史紀錄"
        return {
            "ok": ok,
            "warning": no_actionable,
            "message": message,
            "execution_context": execution_context,
            "execution_owner": "pages/7_股神推薦.py",
            "settings": settings,
            "scan_report": scan_report,
            "universe_count": len(universe_items),
            "candidate_count": candidate_count,
            "display_count": display_count,
            "record_added": int(record_added or 0),
            "record_integrity_failure": bool(record_integrity_failure),
            "record_messages": list(record_msgs or [])[-20:],
            "calibration_added": int(calibration_added or 0),
            "calibration_summary": calibration_summary,
            "calibration_messages": list(calibration_msgs or [])[-10:],
            "notes": notes[-20:],
            "changed_files": [
                "godpick_latest_recommendations.json", "godpick_latest_run_anchor.json", "godpick_records.json",
                "godpick_recommend_list.json", "godpick_rotation_history.json", "godpick_learning_state.json",
                "godpick_calibration_samples.json",
            ],
        }
    except Exception as exc:
        return {
            "ok": False,
            "message": f"07股神推薦模組自動執行例外：{type(exc).__name__}: {exc}",
            "execution_context": execution_context,
            "execution_owner": "pages/7_股神推薦.py",
        }


def main():
    st.set_page_config(page_title=PAGE_TITLE, layout="wide")

    # v40：啟用欄位管理極速模式；不再全頁攔截所有表格
    try:
        from godpick_column_manager import install_auto_column_manager
        install_auto_column_manager("7_股神推薦")
    except Exception:
        pass
    inject_pro_theme()

    # >>> GODPICK_BI_V135_REAL_CHART_STYLE
    try:
        from godpick_bi_theme import install_bi_dashboard_style
        install_bi_dashboard_style("7_recommend")
    except Exception:
        pass
    # <<< GODPICK_BI_V135_REAL_CHART_STYLE

    watchlist_map = _load_watchlist_map()
    master_df = _load_master_df()
    if master_df is None or master_df.empty:
        master_df = _load_master_df_fallback_only()
    today = date.today()

    defaults = {
        "universe_mode": "自選群組",
        "group": list(watchlist_map.keys())[0] if watchlist_map else "",
        "days": 120,
        "top_n": 20,
        "manual_codes": "",
        "scan_limit": 1000,
        "selected_categories": ["全部"],
        "min_total_score": 55.0,
        "min_signal_score": -2.0,
        "submitted_once": False,
        "focus_code": "",
        "status_msg": "",
        "status_type": "info",
        "rec_pick_group": list(watchlist_map.keys())[0] if watchlist_map else "",
        "rec_pick_codes": [],
        "rec_record_codes": [],
        "result_saved_at": "",
        "recommend_mode": "飆股模式",
        "risk_strictness": "標準",
        "min_prelaunch_score": 45.0,
        "min_trade_score": 45.0,
        "pick_strategy": "結合版",
        "score_weights": GODPICK_DEFAULT_SCORE_WEIGHTS.copy(),
        "score_weights_edit": GODPICK_DEFAULT_SCORE_WEIGHTS.copy(),
        "top_table_columns": [],
        "full_table_columns": [],
    }
    persistent_settings = _load_persistent_settings(local_first=True)
    persisted_weights = _normalize_weight_map(persistent_settings.get("applied_weights", GODPICK_DEFAULT_SCORE_WEIGHTS))

    for name, value in defaults.items():
        if _k(name) not in st.session_state:
            st.session_state[_k(name)] = value

    # v138：進入 7_股神推薦時自動偵測 14_股神權重校正是否有新套用權重。
    # 這段在任何權重 widget 建立前執行，因此可安全同步 number_input 的 widget key。
    _maybe_auto_reload_weight_settings()

    if _k("score_weights") not in st.session_state or st.session_state.get(_k("score_weights")) == GODPICK_DEFAULT_SCORE_WEIGHTS:
        st.session_state[_k("score_weights")] = persisted_weights.copy()
    if _k("score_weights_edit") not in st.session_state or st.session_state.get(_k("score_weights_edit")) == GODPICK_DEFAULT_SCORE_WEIGHTS:
        st.session_state[_k("score_weights_edit")] = persisted_weights.copy()

    if _k("selected_rec_snapshot") not in st.session_state:
        st.session_state[_k("selected_rec_snapshot")] = pd.DataFrame()

    _ensure_ui_pref("universe_mode", st.session_state.get(_k("universe_mode"), "自選群組"))
    _ensure_ui_pref("group", st.session_state.get(_k("group"), list(watchlist_map.keys())[0] if watchlist_map else ""))
    _ensure_ui_pref("days", st.session_state.get(_k("days"), 120))
    _ensure_ui_pref("top_n", st.session_state.get(_k("top_n"), 20))
    _ensure_ui_pref("manual_codes", st.session_state.get(_k("manual_codes"), ""))
    _ensure_ui_pref("scan_limit", st.session_state.get(_k("scan_limit"), 1000))
    _ensure_ui_pref("selected_categories", st.session_state.get(_k("selected_categories"), ["全部"]))
    _ensure_ui_pref("recommend_mode", st.session_state.get(_k("recommend_mode"), "飆股模式"))
    _ensure_ui_pref("risk_strictness", st.session_state.get(_k("risk_strictness"), "標準"))
    _ensure_ui_pref("pick_strategy", st.session_state.get(_k("pick_strategy"), "結合版"))
    _ensure_ui_pref("min_total_score", float(st.session_state.get(_k("min_total_score"), 55.0)))
    _ensure_ui_pref("min_signal_score", float(st.session_state.get(_k("min_signal_score"), -2.0)))
    _ensure_ui_pref("min_prelaunch_score", float(st.session_state.get(_k("min_prelaunch_score"), 45.0)))
    _ensure_ui_pref("min_trade_score", float(st.session_state.get(_k("min_trade_score"), 45.0)))

    next_pick_key = _k("rec_pick_codes_next")
    real_pick_key = _k("rec_pick_codes")
    widget_pick_key = _k("rec_pick_codes_widget")
    if next_pick_key in st.session_state:
        _next_pick_val = st.session_state.pop(next_pick_key)
        st.session_state[real_pick_key] = _next_pick_val
        # widget 尚未建立前可安全更新 widget key
        st.session_state[widget_pick_key] = _next_pick_val

    next_record_key = _k("rec_record_codes_next")
    real_record_key = _k("rec_record_codes")
    widget_record_key = _k("rec_record_codes_widget")
    if next_record_key in st.session_state:
        _next_record_val = st.session_state.pop(next_record_key)
        st.session_state[real_record_key] = _next_record_val
        # widget 尚未建立前可安全更新 widget key
        st.session_state[widget_record_key] = _next_record_val

    render_pro_hero(
        title="股神推薦｜V4 加速記憶版",
        subtitle="保留舊版完整功能 + 加速顯示 + 條件記憶 + 欄位順序可調整並保留。",
    )

    st.caption(f"目前7頁修正版：{STATE_FIX_VERSION}")
    st.caption("勾選欄位穩定版：v44 data_editor callback fix")
    st.caption(f"重複確認版：{DUPLICATE_CONFIRM_VERSION}")
    st.caption(f"7/8/9 起漲欄位版：{PRELAUNCH_789_VERSION}")
    st.caption(f"大盤串聯版：{MACRO_LINK_VERSION}")
    st.caption(f"股神決策引擎：{GOD_DECISION_ENGINE_VERSION}")
    st.caption(f"推薦設定永久記錄版：{SCAN_SETTINGS_PERSIST_VERSION}")
    st.caption(f"推薦設定Widget修正版：{SCAN_SETTINGS_WIDGET_FIX_VERSION}")
    st.caption(f"推薦設定自動保存版：{SCAN_SETTINGS_AUTOSAVE_VERSION}")
    st.caption(f"權重狀態修正版：{WEIGHT_STATE_FIX_VERSION}")
    st.caption(f"頁面加速修正版：{PAGE07_SPEED_FIX_VERSION}｜本機優先、GitHub背景同步、Excel按需產生、單功能區運算")
    st.caption(f"每日學習型AI：{LEARNING_SYSTEM_VERSION}｜Champion {GODPICK_AI_MODEL_VERSION}｜多路召回＋四引擎＋不可變決策快照")

    data_freshness_snapshot = _render_project_data_freshness_warning_v173()

    macro_ref_for_ui = _load_latest_macro_reference()
    with st.expander("大盤走勢串聯狀態", expanded=False):
        render_pro_info_card(
            "大盤已串入股神推薦評分",
            [
                ("大盤參考等級", _safe_str(macro_ref_for_ui.get("大盤參考等級")), ""),
                ("大盤可參考分數", format_number(_safe_float(macro_ref_for_ui.get("大盤可參考分數"), 0), 2), ""),
                ("推薦權重建議", _safe_str(macro_ref_for_ui.get("大盤推薦權重")), ""),
                ("操作風格", _safe_str(macro_ref_for_ui.get("大盤操作風格")), ""),
                ("風控模式", _safe_str(macro_ref_for_ui.get("大盤策略模式")) or _safe_str(macro_ref_for_ui.get("大盤風險濾網")), ""),
                ("資料日期", _safe_str(macro_ref_for_ui.get("大盤資料日期")) or "尚未儲存大盤紀錄", ""),
            ],
            chips=["大盤濾網", "輔助加權", "不硬篩"],
        )
        st.caption("大盤採輔助加權與風險降權，不會直接刪除逆勢強股，避免漏掉飆股。")


    if master_df is None or master_df.empty:
        st.warning("股票主檔暫時抓不到，已改用備援模式。若推薦結果偏少，請先到股票主檔頁更新主檔後再試。")


    status_msg = _safe_str(st.session_state.get(_k("status_msg"), ""))
    status_type = _safe_str(st.session_state.get(_k("status_type"), "info"))
    if status_msg:
        if status_type == "success":
            st.success(status_msg)
        elif status_type == "warning":
            st.warning(status_msg)
        elif status_type == "error":
            st.error(status_msg)
        else:
            st.info(status_msg)

    if st.session_state.get("watchlist_version"):
        st.caption(
            f"自選股同步狀態：watchlist_version = {st.session_state.get('watchlist_version', 0)}"
            + (
                f" / 最後更新：{_safe_str(st.session_state.get('watchlist_last_saved_at', ''))}"
                if _safe_str(st.session_state.get("watchlist_last_saved_at", ""))
                else ""
            )
        )

    all_categories = _collect_all_categories(master_df, watchlist_map)
    category_options = ["全部"] + all_categories if all_categories else ["全部"]

    saved_categories = st.session_state.get(_k("selected_categories"), ["全部"])
    saved_categories = [x for x in saved_categories if x in category_options] or ["全部"]

    # 推薦設定永久記錄：第一次進頁面先載入；按恢復原始/套用後才改變。
    if st.session_state.pop(_k("scan_settings_reset_pending"), False):
        _payload = st.session_state.pop(_k("scan_settings_reset_payload"), _default_recommend_scan_settings(watchlist_map))
        _payload = _normalize_recommend_scan_settings(_payload, watchlist_map, category_options)
        _apply_recommend_scan_settings_to_state(_payload, sync_widgets=True)

    if not st.session_state.get(_k("scan_settings_loaded_once"), False):
        _persistent_scan_settings = _load_persistent_recommend_scan_settings(watchlist_map, category_options)
        _apply_recommend_scan_settings_to_state(_persistent_scan_settings, sync_widgets=True)
        st.session_state[_k("scan_settings_loaded_once")] = True

    render_pro_section("掃描設定")
    st.caption("本頁條件會固定保留；只有按「套用設定」或「恢復原始設定」才會永久變更。推薦結果也會保留，除非你重新推薦。")
    if st.session_state.get(_k("scan_settings_msg")):
        st.success(st.session_state.pop(_k("scan_settings_msg")))

    applied_weights = _render_score_weight_panel()
    macro_bridge, macro_adjusted_weights, macro_bridge_enabled = _render_macro_bridge_panel(applied_weights)

    global GODPICK_ACTIVE_SCORE_WEIGHTS
    GODPICK_ACTIVE_SCORE_WEIGHTS = macro_adjusted_weights.copy()

    show_v2_logic = st.toggle("顯示 V2 選股邏輯 / 條件說明", value=False, key=_k("show_v2_logic"))
    _render_weight_dynamic_guide(applied_weights)

    with st.form(key=_k("recommend_form"), clear_on_submit=False):
        c1, c2, c3, c4 = st.columns([2, 2, 2, 2])

        with c1:
            universe_options = ["自選群組", "手動輸入", "全市場", "上市", "上櫃", "興櫃"]
            saved_universe = st.session_state.get(_k("universe_mode"), "自選群組")
            if saved_universe not in universe_options:
                saved_universe = "自選群組"
            if st.session_state.get(_ui_pref_key("universe_mode")) not in universe_options:
                st.session_state[_ui_pref_key("universe_mode")] = saved_universe
            form_universe_mode = st.selectbox("掃描範圍", universe_options, key=_ui_pref_key("universe_mode"))

        with c2:
            group_options = list(watchlist_map.keys()) if watchlist_map else [""]
            saved_group = st.session_state.get(_k("group"), "")
            if saved_group not in group_options:
                saved_group = group_options[0] if group_options else ""
            if st.session_state.get(_ui_pref_key("group")) not in group_options:
                st.session_state[_ui_pref_key("group")] = saved_group
            form_group = st.selectbox("自選群組", group_options, key=_ui_pref_key("group"))

        with c3:
            day_options = [60, 90, 120, 180, 240]
            saved_days = int(st.session_state.get(_k("days"), 120))
            if saved_days not in day_options:
                saved_days = 120
            if st.session_state.get(_ui_pref_key("days")) not in day_options:
                st.session_state[_ui_pref_key("days")] = saved_days
            form_days = st.selectbox("觀察天數", day_options, key=_ui_pref_key("days"))

        with c4:
            topn_options = [10, 20, 30, 50]
            saved_topn = int(st.session_state.get(_k("top_n"), 20))
            if saved_topn not in topn_options:
                saved_topn = 20
            if st.session_state.get(_ui_pref_key("top_n")) not in topn_options:
                st.session_state[_ui_pref_key("top_n")] = saved_topn
            form_top_n = st.selectbox("輸出 Top N", topn_options, key=_ui_pref_key("top_n"))

        d1, d2 = st.columns([2, 2])
        with d1:
            limit_options = [100, 200, 300, 500, 1000, 1500, 2000, "全部"]
            saved_limit = st.session_state.get(_k("scan_limit"), 1000)
            if saved_limit not in limit_options:
                saved_limit = 1000
            if st.session_state.get(_ui_pref_key("scan_limit")) not in limit_options:
                st.session_state[_ui_pref_key("scan_limit")] = saved_limit
            form_scan_limit = st.selectbox(
                "掃描上限筆數",
                limit_options,
                key=_ui_pref_key("scan_limit"),
                help="選『全部』時，會把目前市場範圍內的股票全部納入掃描，不做截斷。",
            )

        with d2:
            form_manual_codes = st.text_area(
                "手動輸入股票（可代碼 / 名稱，一行一檔）",
                key=_ui_pref_key("manual_codes"),
                height=110,
                placeholder="2330\n2454\n3548\n台積電",
            )

        render_pro_section("模式 / 類型篩選")
        m1, m2, m3 = st.columns([2, 2, 2])
        with m1:
            mode_options = ["飆股模式", "波段模式", "領頭羊模式", "綜合模式", "低檔轉強模式", "拉回承接模式", "回測支撐模式", "低檔拉回綜合模式", "保守低風險模式"]
            if st.session_state.get(_ui_pref_key("recommend_mode")) not in mode_options:
                st.session_state[_ui_pref_key("recommend_mode")] = st.session_state.get(_k("recommend_mode"), "飆股模式")
            form_recommend_mode = st.selectbox("推薦模式", mode_options, key=_ui_pref_key("recommend_mode"))
        with m2:
            strict_options = ["寬鬆", "標準", "嚴格"]
            if st.session_state.get(_ui_pref_key("risk_strictness")) not in strict_options:
                st.session_state[_ui_pref_key("risk_strictness")] = st.session_state.get(_k("risk_strictness"), "標準")
            form_risk_strictness = st.selectbox("風險過濾強度", strict_options, key=_ui_pref_key("risk_strictness"))
        with m3:
            pick_options = ["精準版", "結合版"]
            if st.session_state.get(_ui_pref_key("pick_strategy")) not in pick_options:
                st.session_state[_ui_pref_key("pick_strategy")] = st.session_state.get(_k("pick_strategy"), "結合版")
            form_pick_strategy = st.selectbox(
                "推薦策略",
                pick_options,
                key=_ui_pref_key("pick_strategy"),
                help="精準版=只看主名單；結合版=主名單外另顯示飆股補抓名單，不混入主名單排序。",
            )

        valid_saved_categories = [x for x in st.session_state.get(_ui_pref_key("selected_categories"), saved_categories) if x in category_options] or ["全部"]
        st.session_state[_ui_pref_key("selected_categories")] = valid_saved_categories
        form_selected_categories = st.multiselect(
            "選擇類型（可多選）",
            options=category_options,
            key=_ui_pref_key("selected_categories"),
            help="已細分為 IC設計、晶圓代工、封測、AI伺服器、散熱、金控、銀行等。",
        )

        render_pro_section("推薦門檻")
        f1, f2, f3, f4 = st.columns(4)
        with f1:
            form_min_total_score = st.number_input("推薦總分下限", key=_ui_pref_key("min_total_score"), step=1.0)
        with f2:
            form_min_signal_score = st.number_input("訊號分數下限", key=_ui_pref_key("min_signal_score"), step=1.0)
        with f3:
            form_min_prelaunch_score = st.number_input("起漲前兆分數下限", key=_ui_pref_key("min_prelaunch_score"), step=1.0)
        with f4:
            form_min_trade_score = st.number_input("交易可行分數下限", key=_ui_pref_key("min_trade_score"), step=1.0)

        btn1, btn2, btn3, btn4, btn5 = st.columns([2, 2, 2, 2, 2])
        with btn1:
            submit_recommend = st.form_submit_button("開始推薦", use_container_width=True, type="primary")
        with btn2:
            submit_refresh = st.form_submit_button("重新推薦", use_container_width=True)
        with btn3:
            submit_apply_settings = st.form_submit_button("套用設定", use_container_width=True)
        with btn4:
            submit_restore_default = st.form_submit_button("恢復原始設定", use_container_width=True)
        with btn5:
            submit_clear = st.form_submit_button("清空條件", use_container_width=True)

    # 頁首的「立即重新推薦」沿用目前永久保存條件，等同按下本表單的重新推薦。
    submit_refresh = bool(
        submit_refresh
        or (isinstance(data_freshness_snapshot, dict) and data_freshness_snapshot.get("request_rescan"))
    )

    render_pro_section("V34 高速掃描優化與斷點續掃")
    cache_stat = get_history_disk_cache_stats() if callable(get_history_disk_cache_stats) else {}
    cp_stat = _v22_checkpoint_status()
    ccache1, ccache2, ccache3, ccache4 = st.columns([1.2, 1.2, 1.2, 2.2])
    with ccache1:
        clear_cache_btn = st.button("清除推薦快取", use_container_width=True)
    with ccache2:
        resume_scan_btn = st.button("接續上次掃描", use_container_width=True)
    with ccache3:
        clear_checkpoint_btn = st.button("清除斷點檔", use_container_width=True)
    with ccache4:
        st.caption(
            f"歷史快取：{int(cache_stat.get('files', 0) or 0)} 檔 / {cache_stat.get('size_mb', 0)} MB"
            f"｜斷點：{int(cp_stat.get('processed_count', 0) or 0)}/{int(cp_stat.get('total_count', 0) or 0)}"
            f"｜更新：{cp_stat.get('updated_at', '') or cache_stat.get('latest_update', '') or '—'}"
        )

    if clear_cache_btn:
        try:
            _get_history_smart.clear()
        except Exception:
            pass
        try:
            _analyze_stock_bundle.clear()
        except Exception:
            pass
        try:
            _load_master_df.clear()
        except Exception:
            pass
        try:
            _build_excel_bytes.clear()
        except Exception:
            pass
        try:
            n, msg = clear_history_disk_cache()
            st.success(f"推薦快取已清除；{msg}")
        except Exception:
            st.success("推薦快取已清除")

    if clear_checkpoint_btn:
        ok, msg = _v22_clear_checkpoint()
        if ok:
            st.success(msg)
        else:
            st.error(msg)

    current_form_settings = _current_form_settings_from_values(
        form_universe_mode, form_group, form_days, form_top_n, form_manual_codes,
        form_scan_limit, form_selected_categories, form_min_total_score,
        form_min_signal_score, form_min_prelaunch_score, form_min_trade_score,
        form_recommend_mode, form_risk_strictness, form_pick_strategy,
    )

    if submit_apply_settings:
        normalized_settings = _normalize_recommend_scan_settings(current_form_settings, watchlist_map, category_options)
        _apply_recommend_scan_settings_to_state(normalized_settings, sync_widgets=False)
        ok, msgs = _save_persistent_recommend_scan_settings(normalized_settings)
        _stage_recommend_scan_settings_reset(
            normalized_settings,
            "推薦設定已套用並永久記錄，換頁或重新開啟後會沿用此設定。" if ok else "推薦設定已套用，但永久記錄失敗；請展開保存明細檢查 GitHub 寫入狀態。"
        )
        st.session_state[_k("scan_settings_save_msgs")] = msgs
        st.rerun()

    if submit_restore_default:
        default_settings = _normalize_recommend_scan_settings(_default_recommend_scan_settings(watchlist_map), watchlist_map, category_options)
        ok, msgs = _save_persistent_recommend_scan_settings(default_settings)
        _stage_recommend_scan_settings_reset(default_settings, "已恢復原始推薦設定並永久記錄。" if ok else "已恢復原始推薦設定，但永久記錄失敗。")
        st.session_state[_k("scan_settings_save_msgs")] = msgs
        st.rerun()

    if submit_clear:
        default_settings = _normalize_recommend_scan_settings(_default_recommend_scan_settings(watchlist_map), watchlist_map, category_options)
        ok, msgs = _save_persistent_recommend_scan_settings(default_settings)
        _stage_recommend_scan_settings_reset(default_settings, "已清空條件、恢復原始推薦設定並永久記錄。" if ok else "已清空條件、恢復原始推薦設定，但永久記錄失敗。")
        st.session_state[_k("scan_settings_save_msgs")] = msgs
        st.session_state[_k("score_weights")] = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()
        st.session_state[_k("score_weights_edit")] = GODPICK_DEFAULT_SCORE_WEIGHTS.copy()
        st.session_state[_k("submitted_once")] = False
        st.session_state[_k("focus_code")] = ""
        st.session_state[_k("rec_df_store")] = pd.DataFrame()
        st.session_state[_k("category_strength_store")] = pd.DataFrame()
        st.session_state[_k("rec_pick_codes_next")] = []
        st.session_state[_k("rec_record_codes_next")] = []
        st.session_state[_k("rec_pick_codes_widget")] = []
        st.session_state[_k("rec_record_codes_widget")] = []
        st.session_state[_k("selected_rec_snapshot")] = pd.DataFrame()
        st.session_state["godpick_rec_selected_df"] = pd.DataFrame()
        st.session_state[_k("top_table_selected_codes")] = []
        st.session_state[_k("full_table_selected_codes")] = []
        st.rerun()

    if st.session_state.get(_k("scan_settings_save_msgs")):
        with st.expander("推薦設定保存明細", expanded=False):
            for msg in st.session_state.pop(_k("scan_settings_save_msgs"), []):
                st.write(f"- {msg}")

    if submit_recommend or submit_refresh:
        # 開始推薦 / 重新推薦時，同步把目前條件永久記錄；
        # 這樣不用另外按套用，換頁或關閉後也不會恢復原始值。
        normalized_settings = _normalize_recommend_scan_settings(current_form_settings, watchlist_map, category_options)
        _apply_recommend_scan_settings_to_state(normalized_settings, sync_widgets=False)
        ok, msgs = _save_persistent_recommend_scan_settings(normalized_settings)
        st.session_state[_k("scan_settings_save_msgs")] = msgs
        if ok:
            st.session_state[_k("scan_settings_msg")] = "目前推薦條件已自動永久記錄。"
        else:
            st.session_state[_k("scan_settings_msg")] = "目前推薦條件已套用，但永久記錄失敗；請展開保存明細檢查。"
        for pref_name in [
            "universe_mode", "group", "days", "top_n", "manual_codes", "scan_limit",
            "selected_categories", "min_total_score", "min_signal_score", "min_prelaunch_score", "min_trade_score", "pick_strategy",
            "recommend_mode", "risk_strictness",
        ]:
            _sync_ui_pref_to_saved(pref_name)
        st.session_state[_k("universe_mode")] = form_universe_mode
        st.session_state[_k("group")] = form_group
        st.session_state[_k("days")] = form_days
        st.session_state[_k("top_n")] = form_top_n
        st.session_state[_k("manual_codes")] = form_manual_codes
        st.session_state[_k("scan_limit")] = form_scan_limit
        st.session_state[_k("selected_categories")] = form_selected_categories if form_selected_categories else ["全部"]
        st.session_state[_k("min_total_score")] = float(form_min_total_score)
        st.session_state[_k("min_signal_score")] = float(form_min_signal_score)
        st.session_state[_k("min_prelaunch_score")] = float(form_min_prelaunch_score)
        st.session_state[_k("min_trade_score")] = float(form_min_trade_score)
        st.session_state[_k("pick_strategy")] = form_pick_strategy
        st.session_state[_k("recommend_mode")] = form_recommend_mode
        st.session_state[_k("risk_strictness")] = form_risk_strictness
        st.session_state[_k("submitted_once")] = True

    if show_v2_logic:
        render_pro_info_card(
            "V2 選股邏輯",
            [
                ("推薦模式", "保留飆股/波段/領頭羊/綜合，新增低檔轉強、拉回承接、回測支撐、低檔拉回綜合、保守低風險。", ""),
                ("推薦策略", "新增 精準版 / 結合版；結合版會另外列出飆股補抓，不混入主名單。", ""),
                ("起漲前兆", "新增均線轉強、量能啟動、突破準備、動能翻多、支撐防守。", ""),
                ("風險淘汰", "新增風險過濾強度：寬鬆 / 標準 / 嚴格。", ""),
                ("交易可行", "新增交易可行分數、追價風險、拉回買點、突破買點、風險報酬評級。", ""),
                ("類股強度", "保留類股熱度，新增類股加速度與熱度排名。", ""),
                ("匯出", "新增 Excel 匯出，不重算目前結果。", ""),
                ("推薦紀錄", "新增可勾選後直接寫入 8_股神推薦紀錄。", ""),
                ("勾選快照", "本輪精華推薦表可直接勾選，並同步到自選股/推薦紀錄/勾選匯出。", ""),
            ],
            chips=["V2", "功能不刪", "顯示加速", "精準度升級", "Excel匯出", "推薦紀錄串接"],
        )


    if show_v2_logic:
        _render_recommendation_scoring_guide()

    if resume_scan_btn:
        st.session_state[_k("submitted_once")] = True

    if not st.session_state.get(_k("submitted_once"), False):
        saved_rec_df, saved_cat_df, saved_hot_df = _load_recommend_result_from_state()
        loaded_candidate_df = st.session_state.get(_k("candidate_diagnosis_store"))
        loaded_saved_at = _safe_str(st.session_state.get(_k("loaded_snapshot_saved_at_v191_h3")))
        loaded_trigger = _safe_str(st.session_state.get(_k("loaded_snapshot_execution_trigger_v191_h3")))
        if isinstance(saved_rec_df, pd.DataFrame) and not saved_rec_df.empty:
            st.session_state[_k("submitted_once")] = True
            if "V191" in loaded_trigger:
                st.success(f"已自動載入 V191 中央排程產生的 07 股神推薦結果｜{loaded_saved_at or '時間未取得'}。")
            else:
                st.info("已載入上一次推薦結果；資料會保留到下一次按「開始推薦 / 重新推薦」才覆蓋。")
        elif isinstance(loaded_candidate_df, pd.DataFrame) and not loaded_candidate_df.empty and loaded_saved_at:
            # Automation did run, but zero stocks crossed the actionable floor.
            # Continue into the result area so the user sees the diagnosis instead
            # of the misleading "請先按開始推薦" message.
            st.session_state[_k("submitted_once")] = True
            st.warning(
                f"已載入 V191 自動推薦執行結果｜{loaded_saved_at}｜本輪可操作推薦 0 檔；"
                f"完整候選診斷 {len(loaded_candidate_df)} 檔。系統不會為了湊名單硬塞弱股。"
            )
        else:
            st.info("目前尚無可驗證的手動或V191自動推薦執行結果；請先設定條件，再按「開始推薦」。")
            return

    selected_categories = st.session_state.get(_k("selected_categories"), ["全部"])
    universe_mode = _safe_str(st.session_state.get(_k("universe_mode"), ""))

    if universe_mode == "自選群組":
        universe_items = watchlist_map.get(_safe_str(st.session_state.get(_k("group"), "")), [])
    elif universe_mode == "手動輸入":
        universe_items = _parse_manual_codes(st.session_state.get(_k("manual_codes"), ""), master_df)
    else:
        universe_items = _build_universe_from_market(
            master_df=master_df,
            market_mode=universe_mode,
            limit_count=st.session_state.get(_k("scan_limit"), 1000),
            selected_categories=selected_categories,
        )

    if not universe_items:
        st.warning("目前掃描池沒有股票。")
        return

    start_dt = today - timedelta(days=int(st.session_state.get(_k("days"), 120)))
    end_dt = today

    rec_df = pd.DataFrame()
    category_strength_df = pd.DataFrame()
    hot_pick_df = pd.DataFrame()

    if submit_recommend or submit_refresh or resume_scan_btn:
        st.session_state[_k("recommend_execution_context_v191")] = {
            "owner": "07_股神推薦",
            "trigger": "手動斷點續掃" if resume_scan_btn else ("手動重新推薦" if submit_refresh else "手動開始推薦"),
            "started_at": _now_text(),
            "automation_version": "V191-H5",
        }
        previous_rec_df, previous_category_df, previous_hot_df = _load_recommend_result_from_state()
        rec_df, category_strength_df, hot_pick_df = _build_recommend_df(
            universe_items=universe_items,
            master_df=master_df,
            start_dt=start_dt,
            end_dt=end_dt,
            min_total_score=float(st.session_state.get(_k("min_total_score"), 55.0)),
            min_signal_score=float(st.session_state.get(_k("min_signal_score"), -2.0)),
            selected_categories=selected_categories,
            mode=_safe_str(st.session_state.get(_k("recommend_mode"), "飆股模式")),
            risk_strictness=_safe_str(st.session_state.get(_k("risk_strictness"), "標準")),
            min_prelaunch_score=float(st.session_state.get(_k("min_prelaunch_score"), 45.0)),
            min_trade_score=float(st.session_state.get(_k("min_trade_score"), 45.0)),
            resume_scan=bool(resume_scan_btn),
            reuse_finished_checkpoint=bool(submit_recommend and not submit_refresh and not resume_scan_btn),
        )
        rec_df, hot_pick_df, _ = _postprocess_recommend_result_v164(
            rec_df, hot_pick_df, macro_bridge, macro_bridge_enabled, force=True
        )

        # Final safety net: first try the complete candidate diagnosis, then keep
        # the previous non-empty result.  Never replace a useful list with 0 rows.
        if rec_df.empty:
            candidate_source = st.session_state.get(_k("candidate_diagnosis_store"))
            conditional_df = _conditional_reference_rows(candidate_source, max_rows=8) if isinstance(candidate_source, pd.DataFrame) else pd.DataFrame()
            if not conditional_df.empty:
                rec_df = conditional_df
                st.session_state[_k("result_fallback_notice")] = "本輪沒有正式推薦，改顯示通過流動性、買點與風控底線的條件式參考名單；未觸發不可買。"
            elif isinstance(previous_rec_df, pd.DataFrame) and not previous_rec_df.empty:
                rec_df = previous_rec_df.copy()
                category_strength_df = previous_category_df.copy() if isinstance(previous_category_df, pd.DataFrame) else pd.DataFrame()
                hot_pick_df = previous_hot_df.copy() if isinstance(previous_hot_df, pd.DataFrame) else pd.DataFrame()
                st.session_state[_k("result_fallback_notice")] = "本輪掃描沒有可用結果，已保留上一輪非空推薦；本輪未覆蓋 JSON。請依保存時間判斷資料新鮮度。"
            else:
                st.session_state[_k("result_fallback_notice")] = "本輪沒有任何通過資料、流動性與風控底線的股票；系統未硬塞弱股，也未覆蓋既有 JSON。"
        else:
            if "推薦用途" in rec_df.columns and rec_df["推薦用途"].astype(str).eq("條件式參考名單").any():
                st.session_state[_k("result_fallback_notice")] = "本輪沒有正式推薦，顯示條件式參考名單；未觸發不可買，且建議倉位上限為 0%。"
            else:
                st.session_state[_k("result_fallback_notice")] = ""
        _save_recommend_result_to_state(rec_df, category_strength_df, hot_pick_df)
        _latest_pack_error_v185 = _safe_str(st.session_state.get(_k("latest_pack_permanent_error")))
        if _latest_pack_error_v185:
            st.error(_latest_pack_error_v185)
        # 每個交易日只保存一份輕量排名快照，供下一輪辨識真正續強與
        # 「結構分數黏著、但今日沒有新訊號」的重複推薦。此檔不是績效權威檔。
        if callable(save_rotation_snapshot):
            try:
                rotation_source = st.session_state.get(_k("candidate_diagnosis_store"))
                if not isinstance(rotation_source, pd.DataFrame) or rotation_source.empty:
                    rotation_source = rec_df
                rotation_source = _phase93_single_source_decision_frame(rec_df, rotation_source)
                rotation_ok, rotation_msg = save_rotation_snapshot(rotation_source, background_remote=True)
                st.session_state[_k("rotation_snapshot_message")] = rotation_msg
            except Exception as rotation_error:
                st.session_state[_k("rotation_snapshot_message")] = f"推薦輪動快照未保存：{rotation_error}"
        # Phase105：每次完整掃描保存不可變決策快照。正式/A-為0也照樣保存，供日後漏選與模型校準。
        if callable(save_learning_run):
            try:
                learning_source = st.session_state.get(_k("candidate_diagnosis_store"))
                if not isinstance(learning_source, pd.DataFrame) or learning_source.empty:
                    learning_source = rec_df
                learning_ok, learning_msgs, learning_state = save_learning_run(
                    learning_source,
                    rec_df,
                    scan_report=st.session_state.get(_k("scan_quality_report"), {}),
                    metadata={
                        "recommend_mode": _safe_str(st.session_state.get(_k("recommend_mode"))),
                        "risk_strictness": _safe_str(st.session_state.get(_k("risk_strictness"))),
                        "pick_strategy": _safe_str(st.session_state.get(_k("pick_strategy"))),
                        "universe_mode": _safe_str(st.session_state.get(_k("universe_mode"))),
                    },
                    persist_remote=True,
                    background_remote=True,
                    pre_scored=True,
                )
                st.session_state[_k("learning_run_messages")] = learning_msgs
                st.session_state[_k("learning_state")] = learning_state
                st.session_state[_k("learning_run_ok")] = bool(learning_ok)
            except Exception as learning_save_error:
                st.session_state[_k("learning_run_messages")] = [f"每日學習決策快照保存例外：{learning_save_error}"]
                st.session_state[_k("learning_run_ok")] = False
        # V183：保存「全候選特徵 + SuperAI情境/進出場決策」不可變經驗快照。
        # 後續 Page8 推薦後績效成熟後，experience profile 只做小幅有界校準，
        # 避免模型只學自己推薦過的股票造成 selection bias。
        if callable(save_super_ai_run):
            try:
                super_source = st.session_state.get(_k("candidate_diagnosis_store"))
                if not isinstance(super_source, pd.DataFrame) or super_source.empty:
                    super_source = rec_df
                super_ok, super_msg, super_meta = save_super_ai_run(
                    super_source, rec_df,
                    metadata={
                        "recommend_mode": _safe_str(st.session_state.get(_k("recommend_mode"))),
                        "risk_strictness": _safe_str(st.session_state.get(_k("risk_strictness"))),
                        "universe_mode": _safe_str(st.session_state.get(_k("universe_mode"))),
                        "scan_quality": st.session_state.get(_k("scan_quality_report"), {}),
                    },
                )
                st.session_state[_k("super_ai_run_message")] = super_msg
                st.session_state[_k("super_ai_run_meta")] = super_meta
            except Exception as super_save_error:
                st.session_state[_k("super_ai_run_message")] = f"SuperAI經驗快照保存例外：{super_save_error}"
        # V188：主掃描完成後只排入「已成熟舊推薦」的 T+1 真相更新。
        # 網路行情抓取不阻塞本輪 1,700+ 檔 AI 最終結果。
        if callable(refresh_t1_truth_async):
            try:
                _truth_ok, _truth_msg = refresh_t1_truth_async(max_records=160, max_workers=8)
                st.session_state[_k("v188_t1_truth_async_message")] = _truth_msg
            except Exception as _truth_schedule_error:
                st.session_state[_k("v188_t1_truth_async_message")] = f"V188 T+1真相背景排程失敗：{_truth_schedule_error}"
    else:
        rec_df, category_strength_df, hot_pick_df = _load_recommend_result_from_state()
        rec_df, hot_pick_df, _postprocess_cache_hit_v164 = _postprocess_recommend_result_v164(
            rec_df, hot_pick_df, macro_bridge, macro_bridge_enabled, force=False
        )
        if _postprocess_cache_hit_v164:
            st.session_state[_k("postprocess_cache_last_hit_v164")] = _now_text()

    # V159：只有按下開始/重新推薦/斷點續掃完成的新結果才自動記錄；
    # 一般換頁 rerun 不重寫。使用同日 business key，因此重跑只更新同一筆，不會重複膨脹。
    if submit_recommend or submit_refresh or resume_scan_btn:
        try:
            auto_source = st.session_state.get(_k("candidate_diagnosis_store"))
            if not isinstance(auto_source, pd.DataFrame) or auto_source.empty:
                auto_source = rec_df
            auto_added, auto_msgs = _v159_auto_record_actionable_recommendations(auto_source, background_write=False)  # V185：推薦紀錄需同步通過永久權威驗證後才算完成
            calibration_added = 0
            calibration_msgs: list[str] = []
            calibration_summary: dict[str, int] = {"near": 0, "missed": 0, "total": 0}
            if callable(save_calibration_samples):
                calibration_added, calibration_msgs, calibration_summary = save_calibration_samples(
                    auto_source, max_near=24, max_missed=20, background_remote=True
                )
            else:
                calibration_msgs = ["校正研究樣本服務未載入。"]
            st.session_state[_k("auto_record_detail")] = [
                f"正式/A-/雷達紀錄永久權威已處理：{auto_added} 筆",
                *[str(x) for x in (auto_msgs or [])],
                f"校正研究樣本新增：{calibration_added} 筆｜近門檻 {calibration_summary.get('near', 0)}｜市場漏選強勢 {calibration_summary.get('missed', 0)}",
                *[str(x) for x in (calibration_msgs or [])],
                "正式推薦績效與校正研究樣本分檔保存，不會把觀察股冒充正式推薦。",
            ]
        except Exception as e:
            st.session_state[_k("auto_record_detail")] = [f"推薦紀錄自動寫入例外：{e}"]

    # V185：頁首新鮮度是在掃描前先渲染；本輪保存完成後重新選舉永久權威並只刷新一次。
    # 同一個 Streamlit run 會同時看到「舊7/9」與「本輪8/11」的矛盾文字。
    # 完成所有本機保存/背景排程後只 rerun 一次，刷新頁首，不會重跑掃描。
    if (submit_recommend or submit_refresh or resume_scan_btn) and bool(st.session_state.get(_k("latest_pack_permanent_ok"), False)):
        st.session_state[_k("v184_post_scan_ui_refresh")] = True
    if st.session_state.pop(_k("v184_post_scan_ui_refresh"), False):
        st.rerun()

    rotation_snapshot_message = _safe_str(st.session_state.get(_k("rotation_snapshot_message")))
    if rotation_snapshot_message and (submit_recommend or submit_refresh or resume_scan_btn):
        st.caption(f"推薦輪動紀錄：{rotation_snapshot_message}")

    _render_debug_scan_summary()
    _render_recommend_status_panel(rec_df)
    _render_vnext_performance_feedback_panel()
    _learning_candidate_now = st.session_state.get(_k("candidate_diagnosis_store"))
    _render_phase105_learning_panel(_learning_candidate_now if isinstance(_learning_candidate_now, pd.DataFrame) else rec_df)
    _super_msg = _safe_str(st.session_state.get(_k("super_ai_run_message")))
    if _super_msg:
        st.caption(f"SuperAI經驗永久化：{_super_msg}")
    _truth_async_msg = _safe_str(st.session_state.get(_k("v188_t1_truth_async_message")))
    if _truth_async_msg:
        st.caption(f"V188 T+1實戰真相：{_truth_async_msg}")

    # V188：把「選股是否跑贏大盤」與「是否真的觸發交易」分開顯示。
    # 未觸發雷達永遠不計交易勝負，避免把候選上漲冒充可執行績效。
    if callable(load_t1_truth_summary):
        try:
            _truth_summary = load_t1_truth_summary() or {}
        except Exception:
            _truth_summary = {}
        with st.expander("V188｜T+1實戰真相、Alpha/Trade分離與機率校準", expanded=False):
            _matured = int(_truth_summary.get("matured_t1_samples") or 0)
            _exec_n = int(_truth_summary.get("executable_samples") or 0)
            _trigger = _truth_summary.get("trigger_rate_pct")
            _win = _truth_summary.get("executable_win_rate_pct")
            _alpha = _truth_summary.get("avg_selection_alpha_pct")
            _brier = _truth_summary.get("brier_score")
            render_pro_kpi_row([
                {"label": "T+1成熟樣本", "value": _matured, "delta": "Selection Alpha樣本", "delta_class": "pro-kpi-delta-flat"},
                {"label": "真正觸發交易", "value": _exec_n, "delta": f"觸發率 {float(_trigger or 0):.1f}%", "delta_class": "pro-kpi-delta-flat"},
                {"label": "可執行勝率", "value": f"{float(_win or 0):.1f}%" if _win is not None else "待累積", "delta": "未觸發不計勝負", "delta_class": "pro-kpi-delta-flat"},
                {"label": "平均Selection Alpha", "value": f"{float(_alpha or 0):+.2f}%" if _alpha is not None else "待累積", "delta": "個股隔日－市場基準", "delta_class": "pro-kpi-delta-flat"},
                {"label": "Brier Score", "value": f"{float(_brier):.4f}" if _brier is not None else "待累積", "delta": f"校準樣本 {int(_truth_summary.get('calibration_samples') or 0)}", "delta_class": "pro-kpi-delta-flat"},
            ])
            st.caption("V188 將 Alpha Grade（股票品質）與 Trade Grade（現在是否值得買）分離；RR<1 禁止正式新倉、RR 1.0~1.3只列雷達，前日大漲/過熱自動切換只准回測。")
            if callable(refresh_t1_trade_truth) and st.button("更新 T+1 實戰真相與 AI 機率校準", key=_k("v188_refresh_t1_truth"), use_container_width=True):
                with st.spinner("正在回放成熟推薦：判斷觸發、MFE/MAE、Selection Alpha、Entry/Risk Alpha..."):
                    try:
                        _truth_result = refresh_t1_trade_truth(max_records=200, max_workers=8, persist=True)
                        if callable(refresh_super_ai_experience_profile):
                            try:
                                refresh_super_ai_experience_profile()
                            except Exception:
                                pass
                        _truth_msg = (
                            f"V188 T+1真相已更新：本輪 {_truth_result.get('processed_this_run',0)} 筆｜"
                            f"成熟 {_truth_result.get('matured_t1_samples',0)}｜可執行 {_truth_result.get('executable_samples',0)}。"
                        )
                        if bool(_truth_result.get("persistence_ok", True)):
                            st.success(_truth_msg + "｜T+1真相/機率校準已完成永久化確認。")
                        else:
                            st.warning(_truth_msg + "｜本輪計算完成，但永久化未確認；請至第17頁重試永久化後再Reboot。")
                            for _pm in (_truth_result.get("persistence_messages") or [])[:4]:
                                st.caption(str(_pm))
                    except Exception as _truth_manual_error:
                        st.error(f"V188 T+1真相更新失敗：{_truth_manual_error}")

    render_pro_info_card(
        "股神交易決策升級",
        [
            ("推薦分桶", "把結果分為立即觀察、等拉回、等突破、高分但過熱、假突破風險等交易情境。", ""),
            ("分層績效資料庫", "正式/A-/R1維持推薦紀錄；近門檻與漏選強勢另存校正研究樣本，避免每天零樣本與選擇偏誤。", ""),
            ("信心等級", "依總分、起漲、交易可行、類股熱度、過熱與假突破風險綜合分級。", ""),
            ("買點劇本", "自動整理現價、拉回買點、突破買點、停損、目標價。", ""),
            ("失效條件", "明確標示跌破何處或量價不延續時應降級。", ""),
            ("SuperAI情境", "新增隔日開高走高/開高走低/開低走高/開低走低/震盪機率、本週進場適合度、條件式進場與動態出場。", ""),
            ("經驗永久化", "全候選決策快照保存，後續以1/3/5/10/20日真實績效做有界校準，不只學已推薦股票。", ""),
            ("追蹤預留", "保留 3/5/10/20 日追蹤欄位，後續可做推薦勝率回測。", ""),
        ],
        chips=["交易決策", "風控", "回測預留"],
    )

    if st.session_state.get(_k("latest_recommendation_sync_msgs")):
        with st.expander("本輪推薦永久保存明細", expanded=False):
            for msg in st.session_state.get(_k("latest_recommendation_sync_msgs"), []):
                st.write(f"- {msg}")

    fallback_notice = _safe_str(st.session_state.get(_k("result_fallback_notice"), ""))
    if fallback_notice:
        st.warning(fallback_notice)

    if rec_df.empty:
        diagnosis_now = st.session_state.get(_k("candidate_diagnosis_store"))
        auto_saved_at = _safe_str(st.session_state.get(_k("loaded_snapshot_saved_at_v191_h3"))) or _safe_str(st.session_state.get(_k("result_saved_at")))
        auto_trigger = _safe_str(st.session_state.get(_k("loaded_snapshot_execution_trigger_v191_h3")))
        if isinstance(diagnosis_now, pd.DataFrame) and not diagnosis_now.empty:
            st.warning(
                f"07｜股神推薦已完成{'（V191中央自動排程）' if 'V191' in auto_trigger else ''}，"
                f"但本輪 0 檔通過『資料完整＋流動性＋買點＋風控＋風報比』可操作底線。"
                "這不是沒有執行；系統選擇不硬塞弱股。下表是本輪候選診斷，非買進清單。"
            )
            if auto_saved_at:
                st.caption(f"本輪執行/保存時間：{auto_saved_at}｜候選診斷：{len(diagnosis_now)} 檔")
            diag = diagnosis_now.copy()
            sort_col = next((c for c in ["V188股神作戰優先分", "股神推薦優先分", "候選強度分", "推薦總分", "股神實戰總分"] if c in diag.columns), None)
            if sort_col:
                diag[sort_col] = pd.to_numeric(diag[sort_col], errors="coerce")
                diag = diag.sort_values(sort_col, ascending=False, na_position="last")
            diag_cols = [c for c in [
                "股票代號", "股票名稱", "市場別", "類別", "正式推薦分區", "盤中雷達優先級",
                "V188股神作戰優先分", "股神推薦優先分", "股神實戰總分", "推薦總分",
                "起漲前兆分數", "交易可行分數", "Risk風控安全分", "風險報酬比",
                "V188交易許可", "正式推薦阻擋原因", "風險說明", "推薦理由摘要"
            ] if c in diag.columns]
            if diag_cols:
                st.dataframe(_format_df(diag[diag_cols].head(30)), use_container_width=True, hide_index=True)
            st.info("若要產生正式/A-/R1名單，應先改善官方因子、行情新鮮度與可操作買點，而不是單純調低推薦門檻。")
        elif submit_recommend or submit_refresh:
            st.warning("本輪沒有任何股票通過資料、流動性、買點與風控的最低參考底線。系統不會硬塞弱股，也不會用 0 檔覆蓋歷史推薦紀錄。")
            st.info("先看上方『推薦除錯摘要』：若抓不到歷史資料或分析錯誤很多，先修資料模組，不要只調低門檻。")
        else:
            st.error("目前沒有已保存的推薦結果或候選診斷；請確認 V191 07工作是否實際執行成功。")
        return

    saved_at = _safe_str(st.session_state.get(_k("result_saved_at"), ""))
    if saved_at:
        strategy_label = _safe_str(st.session_state.get('pick_strategy', '結合版'))
        st.caption(f"目前顯示的是已保存推薦結果｜保存時間：{saved_at}｜策略：{strategy_label}")
        try:
            latest_authority_v185, _authority_detail_v185 = _load_latest_recommendation_authority_v185()
            if isinstance(latest_authority_v185, dict) and latest_authority_v185.get("full_snapshot_pending_or_older"):
                st.warning("V185 已從永久推薦錨點恢復本輪日期與可操作名單；大型完整候選快照仍較舊或尚在遠端同步，但系統不會再回退顯示 2026-07-09。")
        except Exception:
            pass

    readiness_v171 = _load_recommendation_readiness_v171()
    readiness_status = _safe_str(readiness_v171.get("status"))
    freshness_after_scan = _project_data_freshness_snapshot_v173()
    if readiness_status.startswith("RESCAN") and not bool(freshness_after_scan.get("ready")):
        st.warning(
            f"系統健康檢查的一鍵更新已刷新股神前置資料，但目前畫面仍是舊掃描結果。"
            f"就緒度 {readiness_v171.get('score', 0)}/{readiness_v171.get('full_score', 100)}；請按『重新推薦』後再作正式判斷。"
        )
    elif readiness_status.startswith("RESCAN") and bool(freshness_after_scan.get("ready")):
        st.success("本頁已使用最新前置資料完成重新推薦；第17頁的就緒度檔會在下次健康檢查時同步更新。")
    elif readiness_status.startswith("BLOCK") and not bool(freshness_after_scan.get("ready")):
        st.error(
            f"股神前置資料尚未達正式推薦標準：{readiness_status}。"
            f"{_safe_str(readiness_v171.get('recommended_action'))}"
        )

    try:
        rec_df = _phase41_apply_week_battle_columns(rec_df)
    except Exception:
        pass

    top_n = int(st.session_state.get(_k("top_n"), 20))
    top_df = rec_df.iloc[:top_n].copy()

    avg_score = _avg_safe([_safe_float(x) for x in rec_df.get("股神實戰總分", rec_df.get("推薦總分", pd.Series([0] * len(rec_df), index=rec_df.index))).tolist()], 0)
    bucket_series = rec_df.get("主流作戰分區", rec_df.get("下週作戰分區", pd.Series([""] * len(rec_df), index=rec_df.index))).astype(str)
    attack_count = int(bucket_series.eq("主流攻擊候選").sum())
    breakout_count = int(bucket_series.eq("主流突破追蹤").sum())
    early_count = int(bucket_series.eq("早期潛伏觀察").sum())
    cold_count = int(bucket_series.eq("冷門潛伏觀察").sum())
    weak_count = int(bucket_series.eq("弱勢觀察").sum())
    exclude_count = int(bucket_series.isin(["低流動性排除", "禁止買進排除"]).sum())
    radar_bucket = rec_df.get("飆股雷達分區", pd.Series([""] * len(rec_df), index=rec_df.index)).astype(str)
    radar_count = int(radar_bucket.eq("飆股雷達").sum())
    radar_risk_count = int(radar_bucket.eq("高風險爆發觀察").sum())

    scan_report_now = st.session_state.get(_k("scan_quality_report"), {})
    candidate_store_now = st.session_state.get(_k("candidate_diagnosis_store"))
    candidate_count_now = len(candidate_store_now) if isinstance(candidate_store_now, pd.DataFrame) else len(rec_df)
    render_pro_kpi_row(
        [
            {"label": "預計掃描", "value": int(scan_report_now.get("預計掃描數", len(universe_items)) or len(universe_items)), "delta": universe_mode, "delta_class": "pro-kpi-delta-flat"},
            {"label": "成功分析", "value": int(scan_report_now.get("成功分析數", candidate_count_now) or candidate_count_now), "delta": f"有效K線 {float(scan_report_now.get('有效K線資料率%', 0) or 0):.1f}%", "delta_class": "pro-kpi-delta-flat"},
            {"label": "流動性覆蓋", "value": f"{float(scan_report_now.get('流動性資料覆蓋率%', 0) or 0):.1f}%", "delta": _safe_str(scan_report_now.get("推薦適用範圍")), "delta_class": "pro-kpi-delta-flat"},
            {"label": "官方有效覆蓋", "value": f"{float(scan_report_now.get('官方有效因子覆蓋率%', scan_report_now.get('官方因子覆蓋率%', 0)) or 0):.1f}%", "delta": f"最新可信 {float(scan_report_now.get('官方最新可信覆蓋率%', 0) or 0):.1f}%｜來源可信 {float(scan_report_now.get('官方來源可信覆蓋率%', 0) or 0):.1f}%｜T-1內 {float(scan_report_now.get('官方日期T-1內覆蓋率%', 0) or 0):.1f}%", "delta_class": "pro-kpi-delta-flat"},
            {"label": "完整候選池", "value": candidate_count_now, "delta": "非買進清單", "delta_class": "pro-kpi-delta-flat"},
            {"label": "作戰候選", "value": len(rec_df), "delta": "完成最終分流", "delta_class": "pro-kpi-delta-flat"},
            {"label": "主流攻擊/突破", "value": attack_count + breakout_count, "delta": "仍需最終操作許可", "delta_class": "pro-kpi-delta-flat"},
            {"label": "排除/弱勢", "value": exclude_count + weak_count + cold_count, "delta": "不買/冷門隔離", "delta_class": "pro-kpi-delta-flat"},
        ]
    )
    if isinstance(scan_report_now, dict) and scan_report_now:
        st.caption(
            "V187 官方因子治理｜"
            f"有效 {float(scan_report_now.get('官方有效因子覆蓋率%', 0) or 0):.1f}%｜"
            f"日期T-1內 {float(scan_report_now.get('官方日期T-1內覆蓋率%', 0) or 0):.1f}%｜"
            f"來源可信 {float(scan_report_now.get('官方來源可信覆蓋率%', 0) or 0):.1f}%｜"
            f"最終最新可信 {float(scan_report_now.get('官方最新可信覆蓋率%', 0) or 0):.1f}%"
        )
    if isinstance(scan_report_now, dict) and scan_report_now and not bool(scan_report_now.get("正式推薦可用", False)):
        _quality_text = _safe_str(scan_report_now.get("掃描品質說明")) or "本輪掃描或資料品質不足；目前僅作條件式參考。"
        if _safe_str(scan_report_now.get("掃描品質等級")) == "legacy_cache":
            st.warning(_quality_text)
        else:
            st.error(_quality_text)
    elif _safe_str(scan_report_now.get("掃描品質等級")) in {"limited", "warning"}:
        st.warning(f"本輪只代表『{_safe_str(scan_report_now.get('推薦適用範圍'))}』，建議倉位已自動乘上 {float(scan_report_now.get('倉位折減係數', 1) or 1):.1f}。")
    if attack_count <= 0 and radar_count <= 0:
        st.warning("本輪沒有『主流攻擊候選』或『飆股雷達』。完整推薦表仍可能有 B/C/R 或冷門觀察股，但不是直接買進名單；請優先看『飆股雷達分區』、主流作戰分區與盤中觸發價。")

    render_pro_section("推薦股票加入自選股中心")
    st.caption("本輪推薦完成後已同步寫入 godpick_recommend_list.json，10_推薦清單.py 可直接讀取。下次重新推薦會覆蓋本輪清單。")
    auto_record_detail = st.session_state.get(_k("auto_record_detail"), [])
    if auto_record_detail:
        with st.expander("推薦紀錄＋校正研究樣本｜本輪自動寫入明細", expanded=False):
            for line in auto_record_detail:
                st.write(f"- {line}")
    watchlist_map = _load_watchlist_map()

    g1, g2, g3 = st.columns([3, 2, 1])
    with g1:
        new_group_name = st.text_input("新增群組名稱", key=_k("new_group_name"), placeholder="例如：0422股神推薦")
    with g2:
        st.write("")
        st.write("")
        create_group_btn = st.button("新增群組", key=_k("create_group_btn"), use_container_width=True)
    with g3:
        st.write("")
        st.write("")
        refresh_group_btn = st.button("重新載入群組", key=_k("refresh_group_btn"), use_container_width=True)

    if create_group_btn:
        ok, msg = _create_watchlist_group(new_group_name)
        if ok:
            st.success(msg)
            watchlist_map = _load_watchlist_map()
            st.session_state[_k("rec_pick_group")] = _safe_str(new_group_name)
            st.rerun()
        else:
            st.warning(msg)

    if refresh_group_btn:
        watchlist_map = _load_watchlist_map()
        st.rerun()

    rec_group_options = list(watchlist_map.keys()) if watchlist_map else [""]
    saved_pick_group = st.session_state.get(_k("rec_pick_group"), "")
    if saved_pick_group not in rec_group_options:
        saved_pick_group = rec_group_options[0] if rec_group_options else ""
        st.session_state[_k("rec_pick_group")] = saved_pick_group

    rec_code_to_label = {
        str(r["股票代號"]): f"{r['股票代號']} {r['股票名稱']}｜{_safe_str(r.get('新買點分級')) or r['推薦等級']}｜實戰{format_number(r.get('股神實戰總分', r.get('推薦總分')),1)}"
        for _, r in rec_df.iterrows()
    }
    rec_all_codes = rec_df["股票代號"].astype(str).tolist()

    p1, p2, p3 = st.columns([2, 4, 2])
    with p1:
        if rec_group_options and rec_group_options != [""]:
            pick_group = st.selectbox(
                "加入群組",
                options=rec_group_options,
                index=rec_group_options.index(saved_pick_group) if saved_pick_group in rec_group_options else 0,
                key=_k("rec_pick_group"),
            )
        else:
            pick_group = ""
            st.info("目前尚無群組，請先新增群組名稱。")
    with p2:
        current_pick_codes = [x for x in st.session_state.get(_k("rec_pick_codes"), []) if x in rec_all_codes]
        if _k("rec_pick_codes_widget") not in st.session_state:
            st.session_state[_k("rec_pick_codes_widget")] = current_pick_codes
        selected_pick_widget = st.multiselect(
            "勾選推薦股",
            options=rec_all_codes,
            default=current_pick_codes,
            format_func=lambda x: rec_code_to_label.get(str(x), str(x)),
            key=_k("rec_pick_codes_widget"),
        )
        # rec_pick_codes 不是 widget key，可以安全同步資料狀態
        st.session_state[_k("rec_pick_codes")] = selected_pick_widget
    with p3:
        st.write("")
        st.write("")
        add_selected_btn = st.button("加入勾選股票到自選股中心", use_container_width=True, type="primary")

    q1, q2 = st.columns([1, 1])
    with q1:
        if st.button("全選本輪推薦", use_container_width=True):
            st.session_state[_k("rec_pick_codes_next")] = rec_all_codes
            st.session_state[_k("rec_record_codes_next")] = rec_all_codes
            st.session_state[_k("top_pick_codes_next")] = rec_all_codes
            st.rerun()
    with q2:
        if st.button("清空勾選", use_container_width=True):
            st.session_state[_k("rec_pick_codes_next")] = []
            st.session_state[_k("rec_record_codes_next")] = []
            st.session_state[_k("top_pick_codes_next")] = []
            st.rerun()

    if add_selected_btn:
        selected_codes = [_normalize_code(x) for x in st.session_state.get(_k("rec_pick_codes"), []) if _normalize_code(x)]
        if not selected_codes:
            snap = st.session_state.get(_k("selected_rec_snapshot"))
            if isinstance(snap, pd.DataFrame) and not snap.empty and "股票代號" in snap.columns:
                selected_codes = [_normalize_code(x) for x in snap["股票代號"].astype(str).tolist() if _normalize_code(x)]
        if not selected_codes:
            st.warning("請先勾選推薦股票。可在『本輪精華推薦』表格勾選後，直接按表格下方加入自選股。")
        else:
            picked_rows = []
            work = rec_df[rec_df["股票代號"].astype(str).isin(selected_codes)].copy()
            for _, r in work.iterrows():
                picked_rows.append(
                    {
                        "code": _normalize_code(r.get("股票代號")),
                        "name": _safe_str(r.get("股票名稱")),
                        "market": _safe_str(r.get("市場別")) or "上市",
                        "category": _normalize_category(r.get("類別")),
                    }
                )

            duplicate_codes = _find_existing_watchlist_codes(pick_group, selected_codes)
            if duplicate_codes and not st.session_state.get(_k("confirm_watchlist_duplicate"), False):
                st.warning(
                    f"自選股中心群組「{pick_group}」已存在 {len(duplicate_codes)} 檔："
                    + "、".join(duplicate_codes[:20])
                    + ("..." if len(duplicate_codes) > 20 else "")
                )
                st.info("請確認是否仍要繼續加入；已存在的股票會略過，只加入未重複股票。")
                st.session_state[_k("pending_watchlist_rows")] = picked_rows
                st.session_state[_k("pending_watchlist_group")] = pick_group
                if st.button("確認：繼續加入未重複股票", use_container_width=True, key=_k("confirm_watchlist_duplicate_btn")):
                    st.session_state[_k("confirm_watchlist_duplicate")] = True
                    st.rerun()
            else:
                if st.session_state.get(_k("confirm_watchlist_duplicate"), False):
                    picked_rows = st.session_state.get(_k("pending_watchlist_rows"), picked_rows)
                    pick_group = st.session_state.get(_k("pending_watchlist_group"), pick_group)

                added, messages = _append_multiple_stocks_to_watchlist(pick_group, picked_rows)
                st.session_state[_k("confirm_watchlist_duplicate")] = False
                st.session_state[_k("pending_watchlist_rows")] = []
                st.session_state[_k("pending_watchlist_group")] = ""

                if added > 0:
                    st.success(f"已加入 {added} 檔到 {pick_group}")
                    watchlist_map = _load_watchlist_map()
                else:
                    st.warning("沒有新增成功，可能勾選股票都已存在。")

                if messages:
                    with st.expander("加入結果明細", expanded=True):
                        for msg in messages:
                            st.write(f"- {msg}")

    detail_lines = st.session_state.get(_k("last_dual_write_detail"), [])
    if detail_lines:
        with st.expander("雙寫狀態明細", expanded=False):
            for line in detail_lines:
                st.write(f"- {line}")

    render_pro_section("寫入 8_股神推薦紀錄")
    record_code_to_label = {
        str(r["股票代號"]): f"{r['股票代號']} {r['股票名稱']}｜{_safe_str(r.get('新買點分級')) or r['推薦等級']}｜實戰{format_number(r.get('股神實戰總分', r.get('推薦總分')),1)}"
        for _, r in rec_df.iterrows()
    }
    record_all_codes = rec_df["股票代號"].astype(str).tolist()

    rr1, rr2 = st.columns([4, 2])
    with rr1:
        current_record_codes = [x for x in st.session_state.get(_k("rec_record_codes"), []) if x in record_all_codes]
        if _k("rec_record_codes_widget") not in st.session_state:
            st.session_state[_k("rec_record_codes_widget")] = current_record_codes
        selected_record_widget = st.multiselect(
            "勾選要記錄到 8_股神推薦紀錄 的股票",
            options=record_all_codes,
            default=current_record_codes,
            format_func=lambda x: record_code_to_label.get(str(x), str(x)),
            key=_k("rec_record_codes_widget"),
        )
        # rec_record_codes 不是 widget key，可以安全同步資料狀態
        st.session_state[_k("rec_record_codes")] = selected_record_widget

    with rr2:
        st.write("")
        st.write("")
        record_to_log_btn = st.button("記錄到 8_股神推薦紀錄", use_container_width=True, type="primary")

    rr3, rr4 = st.columns([1, 1])
    with rr3:
        if st.button("全選本輪推薦做紀錄", use_container_width=True):
            st.session_state[_k("rec_record_codes_next")] = record_all_codes
            st.session_state[_k("rec_pick_codes_next")] = record_all_codes
            st.session_state[_k("top_pick_codes_next")] = record_all_codes
            st.rerun()
    with rr4:
        if st.button("清空紀錄勾選", use_container_width=True):
            st.session_state[_k("rec_record_codes_next")] = []
            st.session_state[_k("rec_pick_codes_next")] = []
            st.session_state[_k("top_pick_codes_next")] = []
            st.rerun()

    selected_snapshot_df = rec_df[
        rec_df["股票代號"].astype(str).isin([_normalize_code(x) for x in st.session_state.get(_k("rec_record_codes"), []) if _normalize_code(x)])
    ].copy()
    st.session_state[_k("selected_rec_snapshot")] = selected_snapshot_df
    st.session_state["godpick_rec_selected_df"] = selected_snapshot_df

    if record_to_log_btn:
        selected_record_codes = [_normalize_code(x) for x in st.session_state.get(_k("rec_record_codes"), []) if _normalize_code(x)]
        if not selected_record_codes:
            st.warning("請先勾選要記錄的推薦股票。")
        else:
            record_rows = _build_record_rows_from_rec_df(rec_df, selected_record_codes)
            dup_codes, dup_keys = _find_existing_godpick_record_codes(record_rows)

            if dup_codes and not st.session_state.get(_k("confirm_record_duplicate"), False):
                st.warning(
                    f"8_股神推薦紀錄已存在 {len(dup_codes)} 檔相同推薦紀錄："
                    + "、".join(dup_codes[:20])
                    + ("..." if len(dup_codes) > 20 else "")
                )
                st.info("請確認是否仍要重複紀錄。若確認，會保留舊紀錄並新增一筆新紀錄，備註會標記『使用者確認重複紀錄』。")
                st.session_state[_k("pending_record_rows")] = record_rows
                if st.button("確認：仍要重複寫入股神推薦紀錄", use_container_width=True, key=_k("confirm_record_duplicate_btn")):
                    st.session_state[_k("confirm_record_duplicate")] = True
                    st.rerun()
            else:
                force_duplicate = bool(st.session_state.get(_k("confirm_record_duplicate"), False))
                if force_duplicate:
                    record_rows = st.session_state.get(_k("pending_record_rows"), record_rows)

                added_count, record_msgs = _append_godpick_records(record_rows, force_duplicate=force_duplicate)
                st.session_state[_k("confirm_record_duplicate")] = False
                st.session_state[_k("pending_record_rows")] = []

                if added_count > 0:
                    if force_duplicate:
                        st.success(f"已重複寫入 {added_count} 筆到 8_股神推薦紀錄")
                    else:
                        st.success(f"已寫入 {added_count} 筆到 8_股神推薦紀錄")
                else:
                    st.warning("沒有新增任何推薦紀錄，可能已存在或寫入失敗。")
                if record_msgs:
                    with st.expander("推薦紀錄寫入明細", expanded=True):
                        for msg in record_msgs:
                            st.write(f"- {msg}")

    record_detail_lines = st.session_state.get(_k("last_record_write_detail"), [])
    if record_detail_lines:
        with st.expander("8_股神推薦紀錄 同步明細", expanded=False):
            for line in record_detail_lines:
                st.write(f"- {line}")

    _render_export_block(rec_df=rec_df, category_strength_df=category_strength_df, top_n=top_n)
    _render_selected_export_block()
    _render_record_export_block(rec_df)

    render_pro_section("本輪精華推薦")

    top_selected_codes = st.session_state.pop(_k("top_pick_codes_next"), None)
    if top_selected_codes is None:
        top_selected_codes = st.session_state.get(
            _k("top_table_selected_codes"),
            st.session_state.get(_k("rec_record_codes"), st.session_state.get(_k("rec_pick_codes"), [])),
        )
    top_selected_codes = {_normalize_code(x) for x in top_selected_codes if _normalize_code(x)}

    top_df = top_df.copy()
    if "勾選" not in top_df.columns:
        top_df.insert(0, "勾選", False)
    top_df["勾選"] = top_df["股票代號"].astype(str).map(lambda x: _normalize_code(x) in top_selected_codes)

    # V91 hotfix：舊版快取 / 舊推薦紀錄可能沒有 V90 新欄位。
    # pandas 直接用 df[[cols]] 遇到缺欄會 KeyError，導致「本輪精華推薦」整頁掛掉。
    # 這裡補齊缺欄，保留原本顯示順序與舊資料相容性，不重跑推薦、不拖慢速度。
    top_show_cols = [
        "勾選",
        "股票代號",
        "股票名稱",
        "市場別",
        "類別",
        "類股內排名",
        "類股前3強",
        "最終操作結論",
        "正式推薦分區",
        "操作許可",
        "正式推薦等級",
        "可操作分",
        "實戰操作品質分",
        "推薦可信度分",
        "推薦模式",
        "推薦等級",
        "推薦總分",
        "夜間股神總分",
        "隔日進場分數",
        "進場型態_隔日",
        "隔日建議動作",
        "預估進場點",
        "實戰觸發價",
        "觸發後守價",
        "實戰停損參考",
        "實戰停損距離%",
        "實戰壓力空間%",
        "實戰風險報酬比",
        "正式推薦動作",
        "停損價_隔日",
        "資料完整度",
        "買點分級",
        "信心等級",
        "推薦分桶",
        "市場環境分數",
        "型態名稱",
        "型態突破分數",
        "爆發力分數",
        "起漲前兆分數",
        "交易可行分數",
        "類股熱度分數",
        "是否領先同類股",
        "起漲判斷",
        "最新價",
        "推薦買點_拉回",
        "推薦買點_突破",
        "停損價",
        "賣出目標1",
        "賣出目標2",
        "股神推論邏輯", "風險說明", "買點劇本", "失效條件", "假突破風險", "過熱風險", "推薦理由摘要",
    ]
    for _missing_col in top_show_cols:
        if _missing_col not in top_df.columns:
            if _missing_col == "勾選":
                top_df[_missing_col] = False
            elif _missing_col in {"夜間股神總分", "隔日進場分數", "型態突破分數", "爆發力分數", "起漲前兆分數", "交易可行分數", "類股熱度分數", "最新價", "推薦買點_拉回", "推薦買點_突破", "停損價", "賣出目標1", "賣出目標2", "可操作分", "實戰操作品質分", "推薦可信度分", "實戰觸發價", "觸發後守價", "實戰停損參考", "實戰停損距離%", "實戰壓力空間%", "實戰風險報酬比"}:
                top_df[_missing_col] = pd.NA
            elif _missing_col == "資料完整度":
                top_df[_missing_col] = "舊版快取，未含夜間欄位；請按一次重新推薦更新"
            else:
                top_df[_missing_col] = ""
    top_show_df = top_df[top_show_cols].copy()

    # v44：先建立 row_index → 股票代號 對照表，讓 on_change callback 可在 rerun 前保存勾選狀態。
    top_editor_key = _k("top_pick_editor")
    top_editor_code_map_key = _k("top_pick_editor_code_map")
    st.session_state[top_editor_code_map_key] = [
        _normalize_code(x) for x in top_show_df["股票代號"].astype(str).tolist()
    ]

    # v87：勾選免跳列模式。
    # 原本 data_editor 每點一次 checkbox 就會觸發整頁 rerun，表格視窗會回到第一列。
    # 放進 form 後，勾選時不重跑；按「套用勾選」才更新 session_state。
    with st.form(_k("top_pick_editor_form_v87"), clear_on_submit=False):
        edited_top_df = st.data_editor(
            _format_df(top_show_df),
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            key=top_editor_key,
            disabled=[c for c in top_show_df.columns if c != "勾選"],
            column_config={
                "勾選": st.column_config.CheckboxColumn("勾選", help="v87：可連續勾選，表格不會每點一次就跳回第一列；勾完請按下方套用。"),
                "股神推論邏輯": st.column_config.TextColumn("股神推論邏輯", width="large"),
                "風險說明": st.column_config.TextColumn("風險說明", width="large"),
                "推薦理由摘要": st.column_config.TextColumn("推薦理由摘要", width="large"),
            },
        )
        top_apply_selection = st.form_submit_button(
            "✅ 套用本輪精華推薦勾選",
            use_container_width=True,
        )

    if top_apply_selection:
        picked_codes_from_top = _extract_checked_codes_from_editor_state(
            _k("top_pick_editor"),
            edited_top_df,
            _k("top_table_selected_codes"),
        )
        st.session_state[_k("top_table_selected_codes")] = picked_codes_from_top
        st.session_state[_k("top_pick_codes_next")] = picked_codes_from_top
        st.success(f"已套用本輪精華推薦勾選：{len(picked_codes_from_top)} 檔。")
    else:
        picked_codes_from_top = [
            _normalize_code(x)
            for x in st.session_state.get(_k("top_table_selected_codes"), [])
            if _normalize_code(x)
        ]

    current_pick_codes = [_normalize_code(x) for x in st.session_state.get(_k("rec_pick_codes"), []) if _normalize_code(x)]
    current_record_codes = [_normalize_code(x) for x in st.session_state.get(_k("rec_record_codes"), []) if _normalize_code(x)]

    # 注意：rec_pick_codes / rec_record_codes 是 multiselect widget key。
    # Streamlit 不允許 widget 建立後在同一次 rerun 直接寫入該 key，
    # 所以只能寫到 *_next，下一次 rerun 開頭再套用，避免 StreamlitAPIException。
    if picked_codes_from_top != current_pick_codes:
        st.session_state[_k("rec_pick_codes_next")] = picked_codes_from_top
    if picked_codes_from_top != current_record_codes:
        st.session_state[_k("rec_record_codes_next")] = picked_codes_from_top

    selected_snapshot_top = rec_df[rec_df["股票代號"].astype(str).isin([str(x) for x in picked_codes_from_top])].copy()
    st.session_state[_k("selected_rec_snapshot")] = selected_snapshot_top
    st.session_state["godpick_rec_selected_df"] = selected_snapshot_top

    if picked_codes_from_top:
        st.success(f"已勾選 {len(picked_codes_from_top)} 檔：可直接加入自選股或寫入股神推薦紀錄。")

        fast_a1, fast_a2, fast_a3 = st.columns([1.4, 1.4, 2.2])
        with fast_a1:
            quick_add_watchlist = st.button("將勾選股票加入自選股中心", use_container_width=True, type="primary", key=_k("quick_add_watchlist_from_editor"))
        with fast_a2:
            quick_add_record = st.button("將勾選股票寫入股神推薦紀錄", use_container_width=True, key=_k("quick_add_record_from_editor"))
        with fast_a3:
            st.caption("此處直接使用表格勾選結果，不需要再到上方多選一次。")

        if quick_add_watchlist:
            pick_group = _safe_str(st.session_state.get(_k("rec_pick_group"), ""))
            if not pick_group:
                st.warning("請先在上方選擇或新增自選股群組。")
            else:
                work = rec_df[rec_df["股票代號"].astype(str).isin([str(x) for x in picked_codes_from_top])].copy()
                picked_rows = []
                for _, r in work.iterrows():
                    picked_rows.append(
                        {
                            "code": _normalize_code(r.get("股票代號")),
                            "name": _safe_str(r.get("股票名稱")),
                            "market": _safe_str(r.get("市場別")) or "上市",
                            "category": _normalize_category(r.get("類別")),
                        }
                    )
                added, messages = _append_multiple_stocks_to_watchlist(pick_group, picked_rows)
                if added > 0:
                    st.success(f"已加入 {added} 檔到自選股中心：{pick_group}")
                    st.session_state[_k("rec_pick_codes_next")] = picked_codes_from_top
                    st.rerun()
                else:
                    st.warning("沒有新增成功，可能已存在或寫入失敗。")
                with st.expander("加入自選股明細", expanded=True):
                    for msg in messages:
                        st.write(f"- {msg}")

        if quick_add_record:
            record_rows = _build_record_rows_from_rec_df(rec_df, picked_codes_from_top)
            added_count, record_msgs = _append_godpick_records(record_rows)
            if added_count > 0:
                st.success(f"已寫入 {added_count} 筆到 8_股神推薦紀錄")
                st.session_state[_k("rec_record_codes_next")] = picked_codes_from_top
            else:
                st.warning("沒有新增任何推薦紀錄，可能已存在或寫入失敗。")
            with st.expander("推薦紀錄寫入明細", expanded=True):
                for msg in record_msgs:
                    st.write(f"- {msg}")

    pick_options = top_df["股票代號"].astype(str).tolist()
    if pick_options and st.session_state.get(_k("focus_code"), "") not in pick_options:
        st.session_state[_k("focus_code")] = pick_options[0]

    code_to_row = {str(r["股票代號"]): r for _, r in rec_df.iterrows()}

    render_pro_section("單股股神劇本")
    selected_code = st.selectbox(
        "選擇推薦股",
        options=pick_options,
        format_func=lambda x: f"{x} {code_to_row.get(str(x), {}).get('股票名稱', '')}",
        key=_k("focus_code"),
    )

    focus_row = code_to_row.get(str(selected_code))
    if focus_row is not None:
        render_pro_info_card(
            "股神推薦結論",
            [
                ("股票", f"{_safe_str(focus_row.get('股票代號'))} {_safe_str(focus_row.get('股票名稱'))}", ""),
                ("類別", _safe_str(focus_row.get("類別")), ""),
                ("類股內排名", _safe_str(focus_row.get("類股內排名")), ""),
                ("類股前3強", _safe_str(focus_row.get("類股前3強")), ""),
                ("推薦模式", _safe_str(focus_row.get("推薦模式")), ""),
                ("推薦等級", _safe_str(focus_row.get("推薦等級")), ""),
                ("推薦總分", format_number(focus_row.get("推薦總分"), 1), ""),
                ("上漲機率估計", f"{format_number(focus_row.get('上漲機率估計%'), 1)}%", _safe_str(focus_row.get("上漲機率等級"))),
                ("市場環境", _safe_str(focus_row.get("市場環境")), ""),
                ("市場環境分數", format_number(focus_row.get("市場環境分數"), 1), ""),
                ("型態名稱", _safe_str(focus_row.get("型態名稱")), ""),
                ("型態突破分數", format_number(focus_row.get("型態突破分數"), 1), ""),
                ("爆發力分數", format_number(focus_row.get("爆發力分數"), 1), ""),
                ("起漲前兆分數", format_number(focus_row.get("起漲前兆分數"), 1), ""),
                ("交易可行分數", format_number(focus_row.get("交易可行分數"), 1), ""),
                ("類股熱度分數", format_number(focus_row.get("類股熱度分數"), 1), ""),
                ("是否領先同類股", _safe_str(focus_row.get("是否領先同類股")), ""),
                ("起漲判斷", _safe_str(focus_row.get("起漲判斷")), ""),
                ("建議切入區", _safe_str(focus_row.get("建議切入區")), ""),
                ("推薦買點（拉回）", format_number(focus_row.get("推薦買點_拉回"), 2), ""),
                ("推薦買點（突破）", format_number(focus_row.get("推薦買點_突破"), 2), ""),
                ("停損價", format_number(focus_row.get("停損價"), 2), ""),
                ("賣出目標1", format_number(focus_row.get("賣出目標1"), 2), ""),
                ("賣出目標2", format_number(focus_row.get("賣出目標2"), 2), ""),
                ("風險報酬（拉回）", _safe_str(focus_row.get("風險報酬_拉回")), ""),
                ("風險報酬（突破）", _safe_str(focus_row.get("風險報酬_突破")), ""),
                ("股神推論邏輯", "風險說明", "推薦理由摘要", _safe_str(focus_row.get("推薦理由摘要")), ""),
            ],
            chips=[_safe_str(focus_row.get("推薦等級")), _safe_str(focus_row.get("類別")), _safe_str(focus_row.get("推薦標籤"))],
        )


    if _safe_str(st.session_state.get(_k("pick_strategy"), "結合版")) == "結合版" and isinstance(hot_pick_df, pd.DataFrame) and not hot_pick_df.empty:
        render_pro_section("飆股補抓名單")
        st.caption("這份名單不影響主名單排序；用途是補抓接近門檻、但具起漲結構與類股熱度的股票。")
        hot_show_cols = [
            "股票代號", "股票名稱", "市場別", "類別", "推薦模式", "推薦總分", "上漲機率估計%", "上漲機率等級",
            "市場環境分數", "型態名稱", "型態突破分數", "爆發等級", "爆發力分數",
            "起漲前兆分數", "交易可行分數", "類股熱度分數", "訊號分數",
            "起漲判斷", "建議切入區", "股神推論邏輯", "風險說明", "推薦理由摘要", "補抓原因"
        ]
        st.dataframe(_format_df(hot_pick_df[[c for c in hot_show_cols if c in hot_pick_df.columns]].head(max(top_n, 20))), use_container_width=True, hide_index=True)

    leader_df = rec_df.sort_values(["是否領先同類股", "推薦總分", "類股熱度分數"], ascending=[False, False, False]).reset_index(drop=True)
    factor_rank = rec_df.sort_values(["自動因子總分", "EPS代理分數", "營收動能代理分數", "獲利代理分數"], ascending=[False, False, False, False]).reset_index(drop=True)

    _phase80_render_actionable_panel(rec_df)

    diagnosis_df = st.session_state.get(_k("candidate_diagnosis_store"))
    if not isinstance(diagnosis_df, pd.DataFrame) or diagnosis_df.empty:
        try:
            diagnosis_df = build_candidate_diagnosis(rec_df) if callable(build_candidate_diagnosis) else rec_df.copy()
        except Exception:
            diagnosis_df = rec_df.copy()
    diagnosis_df = diagnosis_df.reset_index(drop=True)
    # 排行榜以完整候選診斷池計算，不再只看最終少量作戰名單。
    leader_df = _safe_sort_export_df(
        diagnosis_df,
        ["是否領先同類股", "候選強度分", "類股熱度分數", "同類股領先幅度"],
        [False, False, False, False],
    )
    factor_rank = _safe_sort_export_df(
        diagnosis_df,
        ["自動因子總分", "EPS代理分數", "營收動能代理分數", "獲利代理分數", "候選強度分"],
        [False, False, False, False, False],
    )

    detail_sections_v164 = ["候選診斷總表（非買進清單）", "類股強度榜", "同類股領先榜", "自動因子榜", "飆股補抓", "操作說明"]
    active_detail_section_v164 = st.radio(
        "推薦結果功能區｜V164 單區運算",
        detail_sections_v164,
        horizontal=True,
        key=_k("active_detail_section_v164"),
    )
    st.caption("V164：只執行目前選取的功能區，不再讓 st.tabs 在每次 rerun 同時建立六個區塊。")

    if active_detail_section_v164 == "候選診斷總表（非買進清單）":
        # v26 欄位統一：完整推薦表使用與 8_股神推薦紀錄 / 10_推薦清單 / 12_股神管理中心一致的標準欄位順序。
        full_default_cols = [c for c in (UNIFIED_RECOMMEND_DISPLAY_COLUMNS or list(diagnosis_df.columns)) if c in diagnosis_df.columns]
        if not full_default_cols:
            full_default_cols = [c for c in list(diagnosis_df.columns) if c != "勾選"]
        # v48：完整推薦表改用與 12_股神管理中心相同的欄位管理樣式。
        full_for_manager = diagnosis_df.head(1).copy()
        if "勾選" not in full_for_manager.columns:
            full_for_manager.insert(0, "勾選", False)
        try:
            from godpick_column_manager import render_column_manager
            full_order = render_column_manager(
                "page07_godpick_recommend_full",
                "候選診斷總表（非買進清單）",
                full_for_manager,
                ["勾選"] + full_default_cols,
            )
        except Exception:
            full_order = ["勾選"] + full_default_cols
        full_show_cols = [c for c in full_order if c in diagnosis_df.columns and c != "勾選"]
        if not full_show_cols:
            full_show_cols = full_default_cols

        # v78：確保完整推薦表的 DataFrame 實體欄位順序完全依 full_show_cols 建立。
        # 注意：直接在表格前端拖曳欄位不會寫回 Python；需使用上方欄位順序設定後按「套用」。

        # v25.6：完整推薦表直接勾選，並可匯入 05_自選股中心 / 09_股神推薦紀錄。
        full_selected_codes_prev = {
            _normalize_code(x)
            for x in st.session_state.get(_k("full_table_selected_codes"), [])
            if _normalize_code(x)
        }

        full_opt1, full_opt2, full_opt3 = st.columns([1.2, 1.2, 3.6])
        with full_opt1:
            full_fast_mode_v164 = st.toggle("候選表快速模式", value=True, key=_k("full_table_fast_mode_v164"))
        with full_opt2:
            full_visible_limit_v164 = st.number_input("候選表顯示上限", min_value=50, max_value=3000, value=150, step=50, key=_k("full_table_visible_limit_v164"))
        with full_opt3:
            st.caption("只限制畫面 data_editor 的渲染筆數；完整候選資料、排序、匯出與同步來源都不會被截斷。關閉快速模式即可顯示全部。")
        full_display_source_v164 = diagnosis_df.head(int(full_visible_limit_v164)) if full_fast_mode_v164 else diagnosis_df
        if full_fast_mode_v164 and len(diagnosis_df) > len(full_display_source_v164):
            st.info(f"快速模式：畫面先顯示 {len(full_display_source_v164)} / {len(diagnosis_df)} 筆；完整資料仍保留。")

        full_work_df = full_display_source_v164[full_show_cols].copy()
        if "勾選" not in full_work_df.columns:
            full_work_df.insert(0, "勾選", False)
        full_work_df["勾選"] = full_work_df["股票代號"].astype(str).map(lambda x: _normalize_code(x) in full_selected_codes_prev)

        # v78：完整推薦表 key 依欄位順序指紋重建。
        # 原因：Streamlit data_editor 會保留前端 column layout；若 key 固定，即使 Python 欄位順序改了，畫面仍可能沿用舊位置。
        full_order_hash = _column_order_fingerprint(list(full_work_df.columns))
        full_layout_version = st.session_state.get(_k("full_table_layout_version"), full_order_hash)
        full_editor_key = _k(f"full_table_editor_{full_layout_version}")
        full_editor_code_map_key = _k(f"full_table_editor_code_map_{full_order_hash}")
        st.session_state[full_editor_code_map_key] = [
            _normalize_code(x) for x in full_work_df["股票代號"].astype(str).tolist()
        ]
        st.caption(f"候選診斷總表欄位順序版本：{full_order_hash}｜此表保留所有候選與排除原因，不等於正式買進清單。")

        # v87：完整推薦表勾選免跳列模式。
        # 放進 form 後，點 checkbox 不會即時 rerun，因此不會跳回第一列。
        with st.form(_k("full_table_editor_form_v87"), clear_on_submit=False):
            full_editor_df = st.data_editor(
                _format_df(full_work_df),
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                key=full_editor_key,
                disabled=[c for c in full_work_df.columns if c != "勾選"],
                column_config={
                    "勾選": st.column_config.CheckboxColumn("勾選", help="v87：可連續勾選，表格不會每點一次就跳回第一列；勾完請按下方套用。"),
                    "推薦理由摘要": st.column_config.TextColumn("推薦理由摘要", width="large"),
                    "股神推論邏輯": st.column_config.TextColumn("股神推論邏輯", width="large"),
                    "風險說明": st.column_config.TextColumn("風險說明", width="large"),
                },
            )
            full_apply_selection = st.form_submit_button(
                "✅ 套用候選診斷表勾選",
                use_container_width=True,
            )

        # v25.8：同時讀取 data_editor 回傳值與 widget edited_rows，避免勾選要點兩次才生效。
        if full_apply_selection:
            full_picked_codes = _extract_checked_codes_from_editor_state(
                full_editor_key,
                full_editor_df,
                _k("full_table_selected_codes"),
            )
            # 去重但保留表格順序。
            full_picked_codes = list(dict.fromkeys(full_picked_codes))
            st.session_state[_k("full_table_selected_codes")] = full_picked_codes
            st.success(f"已套用候選診斷表勾選：{len(full_picked_codes)} 檔。")
        else:
            full_picked_codes = [
                _normalize_code(x)
                for x in st.session_state.get(_k("full_table_selected_codes"), [])
                if _normalize_code(x)
            ]

        selected_snapshot_full = diagnosis_df[diagnosis_df["股票代號"].astype(str).isin([str(x) for x in full_picked_codes])].copy()
        if full_picked_codes:
            st.session_state[_k("selected_rec_snapshot")] = selected_snapshot_full
            st.session_state["godpick_rec_selected_df"] = selected_snapshot_full

        st.caption(f"候選診斷表目前勾選：{len(full_picked_codes)} 檔。05 自選股可收觀察股；8 推薦紀錄接受正式/A-/R1核心雷達（含 R1-M、R1-P）並保留分層；10 推薦清單接受正式/A-/R1核心雷達。")

        full_a1, full_a2, full_a3, full_a4 = st.columns([1.4, 1.3, 1.3, 1.6])
        with full_a1:
            group_options_full = list(watchlist_map.keys()) if isinstance(watchlist_map, dict) and watchlist_map else ["預設"]
            default_full_group = st.session_state.get(_k("full_table_pick_group"), st.session_state.get(_k("rec_pick_group"), group_options_full[0]))
            if default_full_group not in group_options_full:
                default_full_group = group_options_full[0]
            full_target_group = st.selectbox(
                "匯入自選股群組",
                options=group_options_full,
                index=group_options_full.index(default_full_group),
                key=_k("full_table_pick_group"),
            )
        with full_a2:
            full_add_watchlist = st.button(
                "匯入 05_自選股中心",
                use_container_width=True,
                type="primary",
                disabled=(len(full_picked_codes) == 0),
                key=_k("full_table_add_watchlist"),
            )
        with full_a3:
            full_add_record = st.button(
                "匯入 09_股神推薦紀錄",
                use_container_width=True,
                disabled=(len(full_picked_codes) == 0),
                key=_k("full_table_add_record"),
            )
        with full_a4:
            full_add_list = st.button(
                "匯入 10_推薦清單",
                use_container_width=True,
                disabled=(len(full_picked_codes) == 0),
                key=_k("full_table_add_recommend_list"),
            )

        full_b1, full_b2, full_b3 = st.columns([1.5, 1.35, 3.0])
        with full_b1:
            full_sync_all = st.button(
                "一鍵同步 05 + 09 + 10",
                use_container_width=True,
                disabled=(len(full_picked_codes) == 0),
                key=_k("full_table_sync_all"),
            )
        with full_b2:
            # v25.7：完整推薦表直接匯出 Excel。
            export_target_df = selected_snapshot_full.copy() if len(full_picked_codes) > 0 else diagnosis_df.copy()
            export_target_cols = ["勾選"] + [c for c in full_show_cols if c != "勾選"]
            if "勾選" not in export_target_df.columns:
                if "股票代號" in export_target_df.columns:
                    export_target_df.insert(0, "勾選", export_target_df["股票代號"].astype(str).map(lambda x: _normalize_code(x) in set(full_picked_codes)))
                else:
                    export_target_df.insert(0, "勾選", False)
            export_target_df = export_target_df[[c for c in export_target_cols if c in export_target_df.columns]].copy()
            export_target_for_excel = _format_df(export_target_df.copy()) if isinstance(export_target_df, pd.DataFrame) and not export_target_df.empty else export_target_df
            # v72：若欄位管理狀態異常導致只剩「勾選」或空表，直接回退用完整 diagnosis_df，避免 Excel 完整推薦表空白。
            try:
                _real_export_cols = [c for c in export_target_for_excel.columns if str(c) != "勾選"] if isinstance(export_target_for_excel, pd.DataFrame) else []
                if (not isinstance(export_target_for_excel, pd.DataFrame)) or export_target_for_excel.empty or len(_real_export_cols) == 0:
                    export_target_for_excel = _format_df((selected_snapshot_full.copy() if len(full_picked_codes) > 0 else diagnosis_df.copy()))
            except Exception:
                export_target_for_excel = _format_df((selected_snapshot_full.copy() if len(full_picked_codes) > 0 else diagnosis_df.copy()))

            # V164：完整/勾選 Excel 改為按需建立並依結果與勾選指紋快取。
            export_source_for_rank = selected_snapshot_full.copy() if len(full_picked_codes) > 0 else diagnosis_df.copy()
            if isinstance(export_source_for_rank, pd.DataFrame) and "勾選" in export_source_for_rank.columns:
                export_source_for_rank = export_source_for_rank.drop(columns=["勾選"], errors="ignore")
            export_sig_v164 = _result_export_signature_v164(
                export_source_for_rank,
                "full-table|" + ",".join(sorted(full_picked_codes)) + "|" + _column_order_fingerprint(full_show_cols),
            )
            export_cache_key_v164 = _k("full_table_export_cache_v164")
            export_cache_v164 = st.session_state.get(export_cache_key_v164, {})
            export_ready_v164 = (
                isinstance(export_cache_v164, dict)
                and export_cache_v164.get("sig") == export_sig_v164
                and isinstance(export_cache_v164.get("bytes"), (bytes, bytearray))
            )
            export_label = "匯出勾選 Excel" if len(full_picked_codes) > 0 else "匯出完整 Excel"
            if not export_ready_v164:
                if st.button(
                    f"準備{export_label}",
                    use_container_width=True,
                    key=_k("prepare_full_table_excel_v164"),
                ):
                    with st.spinner("只在需要下載時建立完整 Excel 分頁..."):
                        try:
                            full_export_order = [c for c in export_target_for_excel.columns if c != "勾選"] if isinstance(export_target_for_excel, pd.DataFrame) else None
                            _, cat_export_full, leader_export_full, factor_export_full = _build_export_views(
                                export_source_for_rank,
                                category_strength_df if len(full_picked_codes) == 0 else pd.DataFrame(),
                                top_n=max(int(top_n or 200), len(export_source_for_rank) if isinstance(export_source_for_rank, pd.DataFrame) else 200),
                                full_order=full_export_order,
                            )
                        except Exception:
                            cat_export_full, leader_export_full, factor_export_full = pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
                        candidate_store_for_export = st.session_state.get(_k("candidate_diagnosis_store"))
                        if len(full_picked_codes) > 0 and isinstance(candidate_store_for_export, pd.DataFrame) and "股票代號" in candidate_store_for_export.columns:
                            candidate_store_for_export = candidate_store_for_export[
                                candidate_store_for_export["股票代號"].astype(str).map(_normalize_code).isin(set(full_picked_codes))
                            ].copy()
                        export_bytes_full_table = _build_excel_bytes(
                            rec_export=(selected_snapshot_full.copy() if len(full_picked_codes) > 0 else rec_df.copy()),
                            cat_export=cat_export_full,
                            leader_export=leader_export_full,
                            factor_export=factor_export_full,
                            candidate_diagnosis_export=(candidate_store_for_export if isinstance(candidate_store_for_export, pd.DataFrame) else export_target_for_excel),
                            scan_report=st.session_state.get(_k("scan_quality_report"), {}),
                        )
                        export_cache_v164 = {
                            "sig": export_sig_v164,
                            "bytes": export_bytes_full_table,
                            "name": f"股神正式推薦作戰表_{'勾選' if len(full_picked_codes) > 0 else '全部'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        }
                        st.session_state[export_cache_key_v164] = export_cache_v164
                        export_ready_v164 = True
            if export_ready_v164:
                st.download_button(
                    export_label,
                    data=export_cache_v164["bytes"],
                    file_name=export_cache_v164["name"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key=_k("full_table_excel_download_v164"),
                )
        with full_b3:
            st.caption("Phase 8：候選診斷表可加入 05 自選股；09/10 會自動攔截正式排除與高風險觀察，避免把非推薦股寫成正式推薦。")

        if full_add_watchlist:
            work = diagnosis_df[diagnosis_df["股票代號"].astype(str).isin([str(x) for x in full_picked_codes])].copy()
            picked_rows = []
            for _, r in work.iterrows():
                picked_rows.append(
                    {
                        "code": _normalize_code(r.get("股票代號")),
                        "name": _safe_str(r.get("股票名稱")),
                        "market": _safe_str(r.get("市場別")) or "上市",
                        "category": _normalize_category(r.get("類別")),
                    }
                )
            added, messages = _append_multiple_stocks_to_watchlist(full_target_group, picked_rows)
            _show_import_result_notice(
                title=f"匯入 05_自選股中心（{full_target_group}）",
                added_count=added,
                selected_count=len(full_picked_codes),
                messages=messages,
                module_name="05_自選股中心",
            )
            if added > 0:
                st.session_state[_k("rec_pick_codes_next")] = full_picked_codes

        if full_add_record:
            record_allowed_codes, record_rejected_codes = _phase80_allowed_codes(diagnosis_df, full_picked_codes, "record")
            if record_rejected_codes:
                st.warning(f"已攔截 {len(record_rejected_codes)} 檔非正式/A-/R1推薦，未寫入推薦紀錄：{', '.join(record_rejected_codes[:20])}")
            record_rows = _build_record_rows_from_rec_df(diagnosis_df, record_allowed_codes)
            # v25.9：完整推薦表匯入推薦紀錄加入防呆。
            # 同一天 + 同股票代號 + 同推薦模式 已存在時，不再重複新增。
            added_count, record_msgs = _append_godpick_records(record_rows, force_duplicate=False)
            _show_import_result_notice(
                title="匯入 09_股神推薦紀錄",
                added_count=added_count,
                selected_count=len(record_allowed_codes),
                messages=record_msgs,
                module_name="09_股神推薦紀錄",
            )
            if added_count > 0:
                st.session_state[_k("rec_record_codes_next")] = full_picked_codes
                st.session_state[_k("full_table_selected_codes")] = full_picked_codes


        if full_add_list:
            list_allowed_codes, list_rejected_codes = _phase80_allowed_codes(diagnosis_df, full_picked_codes, "list")
            if list_rejected_codes:
                st.warning(f"已攔截 {len(list_rejected_codes)} 檔非正式/A-/R1核心雷達，未寫入推薦清單：{', '.join(list_rejected_codes[:20])}")
            added_list_count, list_msgs = _append_recommend_list_from_full_table(diagnosis_df, list_allowed_codes)
            _show_import_result_notice(
                title="匯入 10_推薦清單",
                added_count=added_list_count,
                selected_count=len(list_allowed_codes),
                messages=list_msgs,
                module_name="10_推薦清單",
            )

        if full_sync_all:
            work = diagnosis_df[diagnosis_df["股票代號"].astype(str).isin([str(x) for x in full_picked_codes])].copy()
            picked_rows = []
            for _, r in work.iterrows():
                picked_rows.append(
                    {
                        "code": _normalize_code(r.get("股票代號")),
                        "name": _safe_str(r.get("股票名稱")),
                        "market": _safe_str(r.get("市場別")) or "上市",
                        "category": _normalize_category(r.get("類別")),
                    }
                )

            added_wl, msg_wl = _append_multiple_stocks_to_watchlist(full_target_group, picked_rows)
            record_allowed_codes, record_rejected_codes = _phase80_allowed_codes(diagnosis_df, full_picked_codes, "record")
            list_allowed_codes, list_rejected_codes = _phase80_allowed_codes(diagnosis_df, full_picked_codes, "list")
            record_rows = _build_record_rows_from_rec_df(diagnosis_df, record_allowed_codes)
            added_rec, msg_rec = _append_godpick_records(record_rows, force_duplicate=False)
            added_list, msg_list = _append_recommend_list_from_full_table(diagnosis_df, list_allowed_codes)
            blocked_union = list(dict.fromkeys(record_rejected_codes + list_rejected_codes))
            if blocked_union:
                st.warning(f"一鍵同步已攔截 {len(blocked_union)} 檔非正式推薦資料寫入 09/10；仍可保留在 05 自選股觀察。")

            st.success(f"一鍵同步完成：05自選股新增 {added_wl} 檔｜09紀錄新增 {added_rec} 筆｜10清單新增 {added_list} 筆")
            _show_import_result_notice("一鍵同步｜05_自選股中心", added_wl, len(full_picked_codes), msg_wl, "05_自選股中心")
            _show_import_result_notice("一鍵同步｜09_股神推薦紀錄", added_rec, len(full_picked_codes), msg_rec, "09_股神推薦紀錄")
            _show_import_result_notice("一鍵同步｜10_推薦清單", added_list, len(full_picked_codes), msg_list, "10_推薦清單")

    if active_detail_section_v164 == "類股強度榜":
        category_show = category_strength_df.copy()
        for c in ["類股平均總分", "類股平均訊號", "類股平均漲幅", "類股平均雷達", "類股平均自動因子", "類股平均起漲前兆", "類股平均交易可行", "類股熱度分數", "類股加速度"]:
            if c in category_show.columns:
                if c == "類股平均漲幅":
                    category_show[c] = category_show[c].apply(lambda x: _format_percent_value(x, 2))
                else:
                    category_show[c] = category_show[c].apply(lambda x: format_number(x, 1) if pd.notna(x) else "")
        st.dataframe(category_show, use_container_width=True, hide_index=True)

    if active_detail_section_v164 == "同類股領先榜":
        st.dataframe(
            _format_df(
                leader_df[[c for c in [
                    "股票代號", "股票名稱", "類別", "類股內排名", "類股前3強",
                    "是否領先同類股", "同類股領先幅度", "市場環境分數", "型態名稱", "型態突破分數", "爆發力分數", "個股原始總分",
                    "類股平均總分", "類股熱度分數", "族群資金流分數", "強勢族群等級", "候選強度分", "推薦總分", "股神推論邏輯", "風險說明", "推薦理由摘要",
                ] if c in leader_df.columns]].head(top_n)
            ),
            use_container_width=True,
            hide_index=True,
        )

    if active_detail_section_v164 == "自動因子榜":
        st.dataframe(
            _format_df(
                factor_rank[[c for c in [
                    "股票代號", "股票名稱", "類別", "市場環境分數", "型態名稱", "型態突破分數", "爆發等級", "爆發力分數", "自動因子總分", "EPS代理分數",
                    "營收動能代理分數", "獲利代理分數", "大戶鎖碼代理分數",
                    "法人連買代理分數", "自動因子摘要", "候選強度分",
                ] if c in factor_rank.columns]].head(top_n)
            ),
            use_container_width=True,
            hide_index=True,
        )

    if active_detail_section_v164 == "飆股補抓":
        if _safe_str(st.session_state.get(_k("pick_strategy"), "結合版")) == "結合版" and isinstance(hot_pick_df, pd.DataFrame) and not hot_pick_df.empty:
            st.dataframe(_format_df(hot_pick_df), use_container_width=True, hide_index=True)
        else:
            st.info("目前未啟用結合版，或本輪沒有補抓名單。")

    if active_detail_section_v164 == "操作說明":
        render_pro_info_card(
            "V2 模組邏輯",
            [
                ("按鈕觸發", "調整條件不會自動重算，按下開始推薦才會跑；條件會自動記住。", ""),
                ("類型更細分", "已由大類擴充成 IC設計、晶圓代工、封測、AI伺服器、散熱、金控、銀行等。", ""),
                ("推薦模式", "保留飆股/波段/領頭羊/綜合，新增低檔轉強、拉回承接、回測支撐、低檔拉回綜合、保守低風險。", ""),
                ("市場環境分數", "新增市場順風/逆風分數，讓同樣條件下順風盤優先。", ""),
                ("型態 / 爆發", "新增型態突破分數、爆發力分數，讓起漲股更容易被拉出。", ""),
            ("推薦策略", "新增 精準版 / 結合版；結合版會另外列出飆股補抓，不混入主名單。", ""),
                ("風險過濾", "新增 寬鬆 / 標準 / 嚴格，先淘汰不合格股票。", ""),
                ("起漲前兆", "新增均線轉強、量能啟動、突破準備、動能翻多、支撐防守。", ""),
                ("交易可行", "新增交易可行分數、追價風險、拉回買點、突破買點、風險報酬評級。", ""),
                ("類股強度", "每個類別都會算平均總分、平均訊號、平均漲幅、類股熱度與類股加速度。", ""),
                ("個股領先", "若個股原始總分高於同類股平均，視為領先股。", ""),
                ("推薦表勾選", "本輪精華推薦表可直接勾選，且欄位順序可調整並記住。", ""),
                ("類股內排名", "新增每個類別內部排名，快速找該族群最強個股。", ""),
                ("類股前3強", "若個股在該類別內排名 1~3，會標記為類股前3強。", ""),
                ("理由升級", "推薦理由已改成更偏交易決策語言，不只是分數描述。", ""),
                ("績效預留", "已預留 3日 / 5日 / 10日 / 20日績效欄位，供下一版自動回填。", ""),
                ("推薦加入自選股", "可直接勾選推薦結果並批次加入指定群組。", ""),
                ("寫入推薦紀錄", "可直接勾選推薦結果並批次寫入 8_股神推薦紀錄。", ""),
                ("雙寫同步", "自選股新增/刪除/批次加入時，同步寫回 GitHub watchlist.json + Firestore。", ""),
                ("Excel 匯出", "可匯出完整推薦表、類股強度榜、同類股領先榜、自動因子榜。", ""),
                ("加速與 ETA", "歷史資料與單股分析保留快取，整批推薦改成併發並顯示剩餘時間。", ""),
                ("推薦結果保留", "推薦結果會存到 session_state，切頁後回來不會立刻消失，條件也會一起記住。", ""),
                ("掃描上限", "已支援 1000 / 1500 / 2000 / 全部掃描。", ""),
                ("7/8 對齊", "record_id、推薦日期、推薦時間、推薦欄位已正式對齊 8 頁。", ""),
            ],
            chips=["V2", "功能不刪", "顯示加速", "三模式", "起漲前兆", "風險過濾", "Excel匯出", "推薦紀錄串接"],
        )


# =========================================================
# v71：讀取 01 大盤趨勢一鍵更新 / 寫入狀態
# 07 股神推薦不連外，只顯示 01 是否已完成一鍵更新與橋接寫入。
# =========================================================
MACRO_V70_STATUS_FILE = "macro_v70_one_click_status.json"
GODPICK_V71_MACRO_STATUS_VERSION = "v71_godpick_macro_status_sync_20260430"


def _read_macro_one_click_status_v71() -> dict[str, Any]:
    try:
        p = Path(MACRO_V70_STATUS_FILE)
        if not p.exists():
            return {}
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _render_macro_one_click_status_v71():
    status = _read_macro_one_click_status_v71()
    if not status:
        st.info("01 大盤趨勢尚未建立一鍵更新狀態檔；建議先到 01 大盤趨勢按『一鍵更新全部並寫入』。")
        return
    all_update = bool(status.get("all_required_updated"))
    all_write = bool(status.get("all_required_written"))
    finished = _safe_str(status.get("finished_at"))
    failed = status.get("failed_items") if isinstance(status.get("failed_items"), list) else []
    if all_update and all_write:
        st.success(f"01 大盤趨勢：全部資料已更新，橋接檔已完整寫入。時間：{finished}")
    elif all_write:
        st.warning(f"01 大盤趨勢：橋接檔已寫入，但仍有資料源未完全更新。時間：{finished}｜未完成：{', '.join(map(str, failed)) if failed else '—'}")
    else:
        st.error(f"01 大盤趨勢：尚未完成完整寫入，建議回 01 大盤趨勢重新按一鍵更新。時間：{finished}｜未完成：{', '.join(map(str, failed)) if failed else '—'}")
    with st.expander("v71 一鍵更新狀態明細", expanded=False):
        st.json(status)

# v71：包裝既有大盤橋接顯示函式，補上 01 一鍵更新完成通知。
try:
    _v71_old_render_market_bridge = _render_market_bridge_v37
    def _render_market_bridge_v37(*args, **kwargs):
        _v71_old_render_market_bridge(*args, **kwargs)
        _render_macro_one_click_status_v71()
except Exception:
    pass


if __name__ == "__main__":
    main()
