from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = "1E88E5"
SUB_HEADER_FILL = "EAF4FF"
TITLE_FILL = "0F172A"
NOTE_FILL = "FFF7D6"
BORDER_COLOR = "B7C9DD"


def _style_sheet(ws, freeze: str | None = None) -> None:
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if cell.row == 1:
                cell.font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
            elif cell.row == 2:
                cell.font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
            else:
                cell.font = Font(name="Microsoft JhengHei", size=10)
    if freeze:
        ws.freeze_panes = freeze


def _auto_width(ws, max_width: int = 28) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        width = 10
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            # Chinese characters are wider than ASCII in Excel.
            width = max(width, min(max_width, int(len(text) * 1.35) + 2))
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 22


def _add_readme(wb: Workbook, rows: Iterable[tuple[str, str]]) -> None:
    ws = wb.active
    ws.title = "填寫說明"
    ws.append(["匯入範例說明", "內容"])
    for title, desc in rows:
        ws.append([title, desc])
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["B1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A1"].font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=12)
    ws["B1"].font = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=12)
    for row in ws.iter_rows(min_row=2):
        row[0].fill = PatternFill("solid", fgColor=SUB_HEADER_FILL)
        row[0].font = Font(name="Microsoft JhengHei", bold=True)
        row[1].font = Font(name="Microsoft JhengHei")
        row[1].alignment = Alignment(vertical="top", wrap_text=True)
    _style_sheet(ws)
    _auto_width(ws, 68)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 82


def _add_list_validation(ws, cell_range: str, values: list[str]) -> None:
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _save_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_org_import_template(year: int = 2026) -> bytes:
    """Build the organization workbook import template used by Page 10.

    This template follows the organization import layout read by
    excel_parser.py: employee and dispatch sheets use row 2 as the header;
    columns may include 到職日, 離職日, 職稱, 課別, 部門, 工段 and vendor fields.
    """
    wb = Workbook()
    _add_readme(
        wb,
        [
            ("用途", "這份範例是給 10. 資料匯入與版本管理上傳使用，包含 01. 超慧員工名單與 02. 派遣名單。"),
            ("工作表名稱", f"系統會依工作表名稱判斷年份，例如 {year}員工名單、{year}派遣名單。"),
            ("欄位位置", "員工名單與派遣名單請保留第 2 列為欄位標題；第 3 列開始填資料。已支援離職日欄位，空白代表仍在職。"),
            ("累積年資", "累積年資可以留空或用 Excel 公式；匯入系統後仍會依到職日自動計算。"),
            ("離職日", "離職日空白代表仍在職；離職日之後不再列入人力計算。"),
            ("工段 / 機型", "若要讓 03. 製造部組織圖與 04. 產能人力計算正確，請務必填寫工段；機型可後續在 01/02 系統表格補充。"),
        ],
    )

    ws = wb.create_sheet(f"{year}員工名單")
    ws.append(["員工名冊", None, None, None, None, None, None, None, "=COUNTA(B3:B200)"])
    ws.append(["員工編號", "姓名", "到職日", "離職日", "累計年資", "職 稱", "課別", "部 門", "工段"])
    sample_rows = [
        ["SPT001", "王小明", date(year - 4, 3, 1), None, "=DATEDIF(C3,TODAY(),\"y\")&\"年\"&DATEDIF(C3,TODAY(),\"YM\")&\"月\"&DATEDIF(C3,TODAY(),\"MD\")&\"天\"", "課長", "製造一課", "製造部", "組裝"],
        ["SPT002", "林小華", date(year - 2, 7, 15), None, "=DATEDIF(C4,TODAY(),\"y\")&\"年\"&DATEDIF(C4,TODAY(),\"YM\")&\"月\"&DATEDIF(C4,TODAY(),\"MD\")&\"天\"", "工程師", "製造一課", "製造部", "配電"],
        ["SPT003", "陳小美", date(year - 1, 10, 8), None, "=DATEDIF(C5,TODAY(),\"y\")&\"年\"&DATEDIF(C5,TODAY(),\"YM\")&\"月\"&DATEDIF(C5,TODAY(),\"MD\")&\"天\"", "技術員", "製造二課", "製造部", "GPTC"],
    ]
    for row in sample_rows:
        ws.append(row)
    _style_sheet(ws, "A3")
    _auto_width(ws)
    _add_list_validation(ws, "F3:F200", ["經理", "課長", "主任", "組長", "資深工程師", "高級工程師", "工程師", "助理工程師", "技術員"])
    _add_list_validation(ws, "G3:G200", ["製造部", "製造一課", "製造二課"])
    _add_list_validation(ws, "I3:I200", ["組裝", "配電", "調機", "GPTC", "包裝", "倉儲"])

    ws = wb.create_sheet(f"{year}派遣名單")
    ws.append(["派遣名冊", None, None, None, None, None, None, None, None, "=COUNTA(B3:B200)"])
    ws.append(["員工編號", "姓名", "到職日", "離職日", "累計年資", "職 稱", "課別", "部 門", "外包商年資", "工段"])
    sample_rows = [
        ["D001", "派遣甲", date(year - 2, 6, 1), None, "=DATEDIF(C3,TODAY(),\"y\")&\"年\"&DATEDIF(C3,TODAY(),\"YM\")&\"月\"&DATEDIF(C3,TODAY(),\"MD\")&\"天\"", "技術員", "製造一課", "德興", "2年", "組裝"],
        ["D002", "派遣乙", date(year - 1, 8, 16), None, "=DATEDIF(C4,TODAY(),\"y\")&\"年\"&DATEDIF(C4,TODAY(),\"YM\")&\"月\"&DATEDIF(C4,TODAY(),\"MD\")&\"天\"", "助理工程師", "製造二課", "晟銘", "1年", "配電"],
        ["D003", "外包丙", date(year, 1, 20), None, "=DATEDIF(C5,TODAY(),\"y\")&\"年\"&DATEDIF(C5,TODAY(),\"YM\")&\"月\"&DATEDIF(C5,TODAY(),\"MD\")&\"天\"", "工程師", "製造二課", "協力廠商", "新進", "GPTC"],
    ]
    for row in sample_rows:
        ws.append(row)
    _style_sheet(ws, "A3")
    _auto_width(ws)
    _add_list_validation(ws, "F3:F200", ["組長", "工程師", "助理工程師", "技術員", "派遣", "外包"] )
    _add_list_validation(ws, "G3:G200", ["製造部", "製造一課", "製造二課"])
    _add_list_validation(ws, "J3:J200", ["組裝", "配電", "調機", "GPTC", "包裝", "倉儲"])
    return _save_bytes(wb)




def build_employee_import_template(year: int = 2026) -> bytes:
    """Build a single-module import template for 01. 超慧員工名單."""
    wb = Workbook()
    _add_readme(
        wb,
        [
            ("用途", "這份範例只匯入 01. 超慧員工名單。適合只更新正職員工資料，不影響其他模組。"),
            ("工作表名稱", f"請保留工作表名稱為 {year}員工名單；其他年度可改為 2025員工名單、2026員工名單。"),
            ("欄位位置", "請保留第 2 列為欄位標題，第 3 列開始填資料；系統支援到職日、離職日與職稱欄位。"),
            ("直接人力", "匯入後系統會依工段、職稱與啟用狀態判斷是否直接人力；若要人工確認，可匯入後在 01 表格中檢查。"),
            ("累積年資", "累積年資可留空；系統會依到職日自動重算。"),
            ("離職日", "離職日空白代表仍在職；離職日之後不再列入人力計算。"),
        ],
    )
    ws = wb.create_sheet(f"{year}員工名單")
    ws.append(["員工名冊", None, None, None, None, None, None, None, "=COUNTA(B3:B200)"])
    ws.append(["員工編號", "姓名", "到職日", "離職日", "累計年資", "職 稱", "課別", "部 門", "工段"])
    rows = [
        ["SPT001", "王小明", date(year - 4, 3, 1), None, "=DATEDIF(C3,TODAY(),\"y\")&\"年\"&DATEDIF(C3,TODAY(),\"YM\")&\"月\"&DATEDIF(C3,TODAY(),\"MD\")&\"天\"", "課長", "製造一課", "製造部", "組裝"],
        ["SPT002", "林小華", date(year - 2, 7, 15), None, "=DATEDIF(C4,TODAY(),\"y\")&\"年\"&DATEDIF(C4,TODAY(),\"YM\")&\"月\"&DATEDIF(C4,TODAY(),\"MD\")&\"天\"", "工程師", "製造一課", "製造部", "配電"],
        ["SPT003", "陳小美", date(year - 1, 10, 8), None, "=DATEDIF(C5,TODAY(),\"y\")&\"年\"&DATEDIF(C5,TODAY(),\"YM\")&\"月\"&DATEDIF(C5,TODAY(),\"MD\")&\"天\"", "技術員", "製造二課", "製造部", "GPTC"],
    ]
    for row in rows:
        ws.append(row)
    _style_sheet(ws, "A3")
    _auto_width(ws)
    _add_list_validation(ws, "F3:F500", ["經理", "課長", "主任", "組長", "資深工程師", "高級工程師", "工程師", "助理工程師", "技術員"])
    _add_list_validation(ws, "G3:G500", ["製造部", "製造一課", "製造二課"])
    _add_list_validation(ws, "I3:I500", ["組裝", "配電", "調機", "GPTC", "包裝", "倉儲"])
    return _save_bytes(wb)


def build_dispatch_import_template(year: int = 2026) -> bytes:
    """Build a single-module import template for 02. 派遣名單."""
    wb = Workbook()
    _add_readme(
        wb,
        [
            ("用途", "這份範例只匯入 02. 派遣名單。適合只更新派遣/外包人力，不影響其他模組。"),
            ("工作表名稱", f"請保留工作表名稱為 {year}派遣名單；其他年度可改為 2025派遣名單、2026派遣名單。"),
            ("欄位位置", "請保留第 2 列為欄位標題，第 3 列開始填資料；系統支援到職日、離職日與職稱欄位。"),
            ("人力來源", "部門欄可填外包商 / 派遣廠商名稱；工段會影響組織圖與直接人力判斷。"),
            ("累積年資", "累積年資可留空；系統會依到職日自動重算。"),
            ("離職日", "離職日空白代表仍在職；離職日之後不再列入人力計算。"),
        ],
    )
    ws = wb.create_sheet(f"{year}派遣名單")
    ws.append(["派遣名冊", None, None, None, None, None, None, None, "=COUNTA(B3:B200)"])
    ws.append(["員工編號", "姓名", "到職日", "離職日", "累計年資", "職 稱", "課別", "部 門", "外包商年資", "工段"])
    rows = [
        ["D001", "派遣甲", date(year - 2, 6, 1), None, "=DATEDIF(C3,TODAY(),\"y\")&\"年\"&DATEDIF(C3,TODAY(),\"YM\")&\"月\"&DATEDIF(C3,TODAY(),\"MD\")&\"天\"", "技術員", "製造一課", "德興", "2年", "組裝"],
        ["D002", "派遣乙", date(year - 1, 8, 16), None, "=DATEDIF(C4,TODAY(),\"y\")&\"年\"&DATEDIF(C4,TODAY(),\"YM\")&\"月\"&DATEDIF(C4,TODAY(),\"MD\")&\"天\"", "助理工程師", "製造二課", "晟銘", "1年", "配電"],
        ["D003", "外包丙", date(year, 1, 20), None, "=DATEDIF(C5,TODAY(),\"y\")&\"年\"&DATEDIF(C5,TODAY(),\"YM\")&\"月\"&DATEDIF(C5,TODAY(),\"MD\")&\"天\"", "工程師", "製造二課", "協力廠商", "新進", "GPTC"],
    ]
    for row in rows:
        ws.append(row)
    _style_sheet(ws, "A3")
    _auto_width(ws)
    _add_list_validation(ws, "F3:F500", ["組長", "工程師", "助理工程師", "技術員", "派遣", "外包"] )
    _add_list_validation(ws, "G3:G500", ["製造部", "製造一課", "製造二課"])
    _add_list_validation(ws, "J3:J500", ["組裝", "配電", "調機", "GPTC", "包裝", "倉儲"])
    return _save_bytes(wb)


def build_schedule_import_template(year: int = 2026) -> bytes:
    """Build a single-module import template for 05. 排程表."""
    wb = Workbook()
    _add_readme(
        wb,
        [
            ("用途", "這份範例只匯入 05. 排程表。儲存後會串接 04 產能負荷表與 09 情境模擬。"),
            ("工作表名稱", f"請保留工作表名稱為 {year}排程表；其他年度可改為 2025排程表、2026排程表。"),
            ("月份 / 機台計數", "新版範例已提供『月份』與『機台計數』欄。大量台數可直接在同一列填機台計數，例如 30；系統會以台數 × 標準工時計算原始需求工時。"),
            ("舊版相容", "若舊 Excel 仍把 1月～12月填在『台數』欄，系統仍會相容並將每個非空白月份標記視為 1 台；若同時填『機台計數』，會優先採用機台計數。"),
            ("需求工時欄位", "原始需求工時、產能計算排除工時、排除後需求工時與需求總工時皆為系統計算欄位：原始 = 台數 × 標準工時；排除後 = 原始 − 排除組立地點工時；04 需求總 = 排除後 + 月別調整工時。匯入值不會取代系統重算。"),
            ("生產廠區", "有一廠／二廠時請填『生產廠區』，例如一廠、二廠。若空白，12 場地週轉會依分廠規則判斷；仍無法判斷時標記為待分廠，不會自動塞入任一廠。"),
            ("日期欄位", "機台入庫日代表機台完成日；12 場地週轉會依各廠區格位天往前推算占用起始日。MOVE IN 可作其他日期查詢。"),
        ],
    )
    ws = wb.create_sheet(f"{year}排程表")
    headers = ["年份", "WO", "客戶", "P/N", "Type", "Category", "組立地點", "生產廠區", "機台入庫日", "MOVE IN", "月份", "台數", "機台計數", "PO", "工期", "標準工時", "需求工時", "狀態", "備註"]
    ws.append(headers)
    rows = [
        [year, "26A0001", "超慧客戶A", "PN-1001", "Type-A", "一般機", "竹東", "一廠", date(year, 1, 10), None, "1月", 30, 30, "PO-001", 10, 56, "=IFERROR(M2*P2,0)", "未開工", "可在機台計數一次填大量台數"],
        [year, "26A0002", "超慧客戶B", "PN-2001", "Type-B", "大型機", "竹東", "二廠", date(year, 2, 12), None, "2月", 12, 12, "PO-002", 15, 88, "=IFERROR(M3*P3,0)", "排程中", ""],
        [year, "26A0003", "超慧客戶C", "PN-3001", "Type-C", "改造機", "竹東", "", None, date(year, 3, 5), "3月", 5, 5, "PO-003", 12, 40, "=IFERROR(M4*P4,0)", "待確認", "空白時由 12 分廠規則判斷"],
    ]
    for row in rows:
        ws.append(row)
    _style_sheet(ws, "A2")
    _auto_width(ws, 34)
    _add_list_validation(ws, "H2:H1000", ["一廠", "二廠"])
    _add_list_validation(ws, "K2:K1000", [f"{i}月" for i in range(1, 13)])
    _add_list_validation(ws, "R2:R1000", ["未開工", "排程中", "已完工", "待確認", "取消"])
    return _save_bytes(wb)


def build_standard_hours_import_template(year: int = 2026) -> bytes:
    """Build a single-module import template for 06. 標準工時."""
    wb = Workbook()
    _add_readme(
        wb,
        [
            ("用途", "這份範例只匯入 06. 標準工時。05 排程表會依客戶 / P/N / Type / Category 對應標準工時。"),
            ("工作表名稱", f"請保留工作表名稱為 {year}標準工時(超)；也可使用 {year}標準工時。"),
            ("對應規則", "建議至少填客戶、P/N、Type、Category、標準工時。系統會做多階段比對，資料越完整越準確。"),
            ("標準天數", "可用公式 =ROUNDUP(標準工時/8,1) 或直接填數字。"),
            ("格位數", "填入該 P/N / Type 會佔用幾個場地格位；12. 場地週轉會優先引用此欄位做預排。"),
            ("是否啟用", "填是/否；若未使用此欄，仍可匯入但建議保留。"),
        ],
    )
    ws = wb.create_sheet(f"{year}標準工時(超)")
    ws.append(["年份", "客戶", "P/N", "Type", "Category", "標準工時", "標準天數", "格位數", "版本", "是否啟用", "備註"])
    rows = [
        [year, "超慧客戶A", "PN-1001", "Type-A", "一般機", 56, "=ROUNDUP(F2/8,1)", 1, "V1", "是", "排程與場地週轉會依 P/N / Type / Category 對應"],
        [year, "超慧客戶B", "PN-2001", "Type-B", "大型機", 88, "=ROUNDUP(F3/8,1)", 2, "V1", "是", ""],
        [year, "超慧客戶C", "PN-3001", "Type-C", "改造機", 40, "=ROUNDUP(F4/8,1)", 1, "V1", "是", ""],
    ]
    for row in rows:
        ws.append(row)
    _style_sheet(ws, "A2")
    _auto_width(ws, 34)
    _add_list_validation(ws, "J2:J1000", ["是", "否"])
    return _save_bytes(wb)

def build_capacity_import_template(year: int = 2026) -> bytes:
    """Build the capacity workbook import template used by Page 10."""
    wb = Workbook()
    _add_readme(
        wb,
        [
            ("用途", "這份範例是給 10. 資料匯入與版本管理上傳使用，包含 05. 排程表、06. 標準工時、07. 工作天數設定與 04. Excel 原始彙整參考。"),
            ("建議流程", "先維護 06. 標準工時與產能計算排除組立地點，再維護 05. 排程表。05 儲存後會重算原始／排除／排除後需求工時；04 再以排除後需求工時 + 月別調整工時計算需求總工時，並串聯 09。"),
            ("年份判斷", f"系統會優先依工作表名稱判斷年份，例如 {year}排程表、{year}標準工時(超)、{year}工作天數。"),
            ("必要工作表", "若只要更新排程，可只保留排程表與標準工時；若要更新工作日，請同時填工作天數。彙整表只是保留 Excel 原始參考，不是 04 的主要計算來源。"),
            ("欄位提醒", "排程表的月份請使用 1月～12月；大量台數請填『機台計數』；有多廠區時請填『生產廠區』。06 標準工時提供預設工時／天數／格位數，12 廠區補充規則可再按廠區覆寫格位天與格位數。"),
        ],
    )

    ws = wb.create_sheet(f"{year}排程表")
    headers = ["年份", "WO", "客戶", "P/N", "Type", "Category", "組立地點", "生產廠區", "機台入庫日", "MOVE IN", "月份", "台數", "機台計數", "PO", "工期", "標準工時", "需求工時", "狀態", "備註"]
    ws.append(headers)
    rows = [
        [year, "26A0001", "超慧客戶A", "PN-1001", "Type-A", "一般機", "竹東", "一廠", date(year, 1, 10), None, "1月", 30, 30, "PO-001", 10, 56, "=IFERROR(M2*P2,0)", "未開工", "可在機台計數一次填大量台數"],
        [year, "26A0002", "超慧客戶B", "PN-2001", "Type-B", "大型機", "竹東", "二廠", date(year, 2, 12), None, "2月", 12, 12, "PO-002", 15, 88, "=IFERROR(M3*P3,0)", "排程中", ""],
        [year, "26A0003", "超慧客戶C", "PN-3001", "Type-C", "改造機", "竹東", "", None, date(year, 3, 5), "3月", 5, 5, "PO-003", 12, 40, "=IFERROR(M4*P4,0)", "待確認", "空白時由 12 分廠規則判斷"],
    ]
    for row in rows:
        ws.append(row)
    _style_sheet(ws, "A2")
    _auto_width(ws, 32)
    _add_list_validation(ws, "H2:H500", ["一廠", "二廠"])
    _add_list_validation(ws, "K2:K500", [f"{i}月" for i in range(1, 13)])
    _add_list_validation(ws, "R2:R500", ["未開工", "排程中", "已完工", "待確認", "取消"])

    ws = wb.create_sheet(f"{year}標準工時(超)")
    ws.append(["年份", "客戶", "P/N", "Type", "Category", "標準工時", "標準天數", "格位數", "版本", "是否啟用", "備註"])
    rows = [
        [year, "超慧客戶A", "PN-1001", "Type-A", "一般機", 56, "=ROUNDUP(F2/8,1)", 1, "V1", "是", "排程與場地週轉會依 P/N / Type / Category 對應"],
        [year, "超慧客戶B", "PN-2001", "Type-B", "大型機", 88, "=ROUNDUP(F3/8,1)", 2, "V1", "是", ""],
        [year, "超慧客戶C", "PN-3001", "Type-C", "改造機", 40, "=ROUNDUP(F4/8,1)", 1, "V1", "是", ""],
    ]
    for row in rows:
        ws.append(row)
    _style_sheet(ws, "A2")
    _auto_width(ws, 34)
    _add_list_validation(ws, "J2:J500", ["是", "否"])

    ws = wb.create_sheet(f"{year}工作天數")
    # Keep this layout compatible with excel_parser._load_work_calendar_sheet.
    ws["B1"] = "起始年"
    ws["D1"] = "終止(年)"
    ws["E1"] = "今天日期"
    ws["B2"] = date(year, 1, 1)
    ws["D2"] = date(year, 12, 31)
    ws["E2"] = "=TODAY()"
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["月份", "月起日", None, "月迄日", None, "六日天數", "週六天數", "週日天數", "法定假日", "扣除六日工作日", "正常工作日"])
    month_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    normal_days = [21, 18, 21, 21, 20, 21, 23, 21, 21, 22, 20, 23]
    for i in range(1, 13):
        start = date(year, i, 1)
        end = date(year, i, month_days[i - 1])
        weekend = month_days[i - 1] - normal_days[i - 1]
        ws.append([i, start, None, end, None, weekend, weekend // 2, weekend - weekend // 2, 0, normal_days[i - 1], normal_days[i - 1]])
    _style_sheet(ws, "A8")
    _auto_width(ws, 26)

    ws = wb.create_sheet(f"{year}彙整表")
    months = [f"{i}月" for i in range(1, 13)]
    ws.append(["產能負荷表"])
    ws.append(["月份", *months])
    ws.append(["每月機台數", 2, 1, 3, 2, 2, 1, 2, 3, 1, 2, 2, 1])
    ws.append(["月份天數", *normal_days])
    ws.append(["原始需求工時", 140, 100, 150, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    ws.append(["產能計算排除工時", 28, 12, 30, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    ws.append(["排除後需求工時", 112, 88, 120, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    ws.append(["調整工時", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    ws.append(["需求總工時", 112, 88, 120, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    ws.append(["可用工時", 900, 800, 920, 900, 880, 900, 920, 900, 900, 920, 880, 920])
    ws.append(["產能負荷率", 12, 11, 13, 0, 0, 0, 0, 0, 0, 0, 0, 0])
    ws.append(["產能餘額", 788, 712, 800, 900, 880, 900, 920, 900, 900, 920, 880, 920])
    _style_sheet(ws, "B3")
    _auto_width(ws, 18)
    return _save_bytes(wb)


def build_factory_turnover_import_template(year: int = 2026) -> bytes:
    """Build a multi-factory import template for Page 10 and Module 12.

    Every factory-owned sheet contains a 廠區 column.  The importer performs
    key-based upsert: matching keys are updated, new keys are added, and rows
    omitted from the workbook are kept.  This protects existing factory data
    while still allowing 一廠／二廠 to be maintained independently.
    """
    wb = Workbook()
    _add_readme(
        wb,
        [
            ("用途", "匯入 12. 超慧科技場地週轉的一廠、二廠及後續廠區設定。可在同一本活頁簿同時維護多個廠區。"),
            ("資料隔離原則", "每個設定工作表都必須填『廠區』。格位、座標、Category 管制、格位天、格位數、場地可用率與畫布設定都按廠區獨立保存，不會共用。"),
            ("建議匯入順序", "1 廠區主檔 → 2 格位主檔 → 3 區域Category管制 → 4 P_N_Type格位規則 → 5 生產廠區分派。可只保留要更新的工作表，未提供的工作表不會被清空。"),
            ("匯入模式", "採『依主鍵新增／更新』。相同廠區與主鍵會更新；新主鍵會新增；Excel 未列出的既有資料保留。匯入前系統會自動建立版本快照。"),
            ("廠區主檔", "設定各廠區場地可用率、畫布寬高、縮放、預設年度、日期條件與組立地點。納入組立地點可用逗號、頓號或換行分隔。"),
            ("格位主檔", "主鍵為 廠區＋格位ID。x/y/w/h 必須與系統位置圖座標一致；格位類型可填機台格位、區域標示或通道。"),
            ("區域Category管制", "主鍵為 廠區＋區域。允許Category可填多個值，以逗號、頓號或換行分隔；空白代表該區域不限制 Category。"),
            ("P_N_Type格位規則", "主鍵為 廠區＋P/N＋Type＋Type代碼。DAYS 是製程天數，格位天才是場地實際占用天數；12 預排以格位天計算。"),
            ("生產廠區分派", "當 05 排程表的生產廠區空白時，系統依優先順序及客戶/P/N/Type/Category/組立地點判斷。數字越小越優先；正式製令仍建議直接在 05 指定生產廠區。"),
            ("05 排程表", "05 匯入範例已加入『生產廠區』欄。已明確指定的資料優先於 12 分廠規則；兩個以上廠區仍無法判斷時會標記為待分廠。"),
            ("刪除規則", "本範本不會因空白列刪除既有資料。需要刪除廠區或格位時，請回 12 模組使用刪除功能，避免誤刪。"),
        ],
    )

    ws = wb.create_sheet("廠區主檔")
    ws.append(["廠區", "場地可用率(%)", "畫布寬度", "畫布高度", "預設縮放(%)", "預設年份", "篩選模式", "完成日期欄位", "預設月份", "開始日期", "結束日期", "納入組立地點", "備註"])
    ws.append(["一廠", 95, 2200, 1030, 100, year, "依月份", "機台入庫日", "全部", date(year, 1, 1), date(year, 12, 31), "竹東,宏田", "一廠參數獨立保存"])
    ws.append(["二廠", 90, 1800, 900, 100, year, "依月份", "機台入庫日", "全部", date(year, 1, 1), date(year, 12, 31), "二廠", "二廠參數不與一廠共用"])
    _style_sheet(ws, "A2")
    _auto_width(ws, 34)
    _add_list_validation(ws, "G2:G200", ["依月份", "依日期區間"])
    _add_list_validation(ws, "H2:H200", ["機台入庫日", "MOVE IN"])
    _add_list_validation(ws, "I2:I200", ["全部", *[f"{i}月" for i in range(1, 13)]])
    ws.column_dimensions["L"].width = 30

    ws = wb.create_sheet("格位主檔")
    ws.append(["廠區", "格位ID", "區域", "列", "欄", "x", "y", "w", "h", "格位類型", "啟用", "備註", "顯示標題", "顯示副標題", "字體大小", "旋轉角度"])
    slot_rows = [
        ["一廠", "A-01-01", "一廠組裝區A", 1, 1, 40, 80, 70, 52, "機台格位", "是", "", "", "", 0, 0],
        ["一廠", "ZONE-A", "一廠組裝區A", 0, 0, 30, 35, 300, 28, "區域標示", "是", "區域標示圖卡", "一廠組裝區A", "", 14, 0],
        ["二廠", "B-01-01", "二廠組裝區B", 1, 1, 40, 80, 70, 52, "機台格位", "是", "", "", "", 0, 0],
        ["二廠", "ZONE-B", "二廠組裝區B", 0, 0, 30, 35, 300, 28, "區域標示", "是", "區域標示圖卡", "二廠組裝區B", "", 14, 0],
    ]
    for row in slot_rows:
        ws.append(row)
    _style_sheet(ws, "A2")
    _auto_width(ws, 26)
    _add_list_validation(ws, "J2:J2000", ["機台格位", "區域標示", "通道"])
    _add_list_validation(ws, "K2:K2000", ["是", "否"])

    ws = wb.create_sheet("區域Category管制")
    ws.append(["廠區", "區域", "允許Category", "啟用", "備註"])
    ws.append(["一廠", "一廠組裝區A", "Sorter,EFEM", "是", "空白代表不限制；多個 Category 以逗號分隔"])
    ws.append(["二廠", "二廠組裝區B", "GPTC,BWBS", "是", "二廠獨立設定，不會影響一廠"])
    _style_sheet(ws, "A2")
    _auto_width(ws, 42)
    _add_list_validation(ws, "D2:D1000", ["是", "否"])

    ws = wb.create_sheet("P_N_Type格位規則")
    ws.append(["廠區", "客戶", "P/N", "Type", "Type代碼", "HOURS", "DAYS", "格位天", "格位數", "適用區域", "啟用", "備註"])
    ws.append(["一廠", "客戶A", "PN-1001", "Sorter", "RSC141", 80, 10, 6, 2, "一廠組裝區A", "是", "DAYS 是製程天；格位天是場地占用天"])
    ws.append(["二廠", "客戶B", "PN-2001", "BWBS", "BWBS", 160, 20, 12, 4, "二廠組裝區B", "是", "同一 P/N 可在不同廠區設定不同格位天與格位數"])
    _style_sheet(ws, "A2")
    _auto_width(ws, 34)
    _add_list_validation(ws, "K2:K2000", ["是", "否"])

    ws = wb.create_sheet("生產廠區分派")
    ws.append(["優先順序", "廠區", "客戶", "P/N", "Type", "Category", "組立地點", "啟用", "備註"])
    ws.append([10, "一廠", "", "PN-1001", "", "", "", "是", "P/N 精準規則優先"])
    ws.append([20, "二廠", "", "", "BWBS", "", "", "是", "Type 規則"])
    ws.append([30, "二廠", "", "", "", "GPTC", "二廠", "是", "Category＋組立地點規則"])
    _style_sheet(ws, "A2")
    _auto_width(ws, 34)
    _add_list_validation(ws, "H2:H2000", ["是", "否"])

    return _save_bytes(wb)
